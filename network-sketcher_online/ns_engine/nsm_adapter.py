"""
nsm_adapter.py - NSM (Network Sketcher Master) format adapter.

Provides headless-mode infrastructure for the web engine:
  1. Tkinter stub classes (_Noop, Entry, StringVar, etc.)
  2. Bootstrap: sys.path setup, tkinter module injection, openpyxl .nsm patch
  3. In-process CLI runner (run_cli / run_cli_isolated)

Merged from: _tk_stub.py, _bootstrap.py, runner.py
"""
import sys
import os
import io
import types
import tempfile
import atexit
import shutil
import uuid
import threading
from pathlib import Path
from dataclasses import dataclass
from typing import Optional


# ---------------------------------------------------------------------------
# Section 1: Tkinter Stubs (formerly _tk_stub.py)
# ---------------------------------------------------------------------------

class _Noop:
    """Catch-all object that silently absorbs any attribute access or call."""
    def __init__(self, *a, **kw): pass
    def __call__(self, *a, **kw): return self
    def __getattr__(self, name): return self
    def __bool__(self): return False
    def __str__(self): return ''
    def __repr__(self): return '_Noop()'
    def __iter__(self): return iter([])


class Toplevel(_Noop): pass
class Frame(_Noop): pass
class Label(_Noop): pass
class Entry(_Noop):
    def __init__(self, *a, **kw):
        super().__init__(*a, **kw)
        self._value = ''
    def get(self):
        return self._value
    def insert(self, index, text):
        if str(index) == 'end' or index == 'insert':
            self._value += str(text)
        else:
            try:
                pos = int(index)
                self._value = self._value[:pos] + str(text) + self._value[pos:]
            except (ValueError, TypeError):
                self._value += str(text)
    def delete(self, first, last=None):
        if str(first) == '0' and (last is None or str(last) == 'end'):
            self._value = ''
        elif last is not None:
            try:
                f = int(first)
                l = len(self._value) if str(last) == 'end' else int(last)
                self._value = self._value[:f] + self._value[l:]
            except (ValueError, TypeError):
                self._value = ''
        else:
            try:
                f = int(first)
                self._value = self._value[:f] + self._value[f+1:]
            except (ValueError, TypeError):
                pass
class Text(_Noop):
    def __init__(self, *a, **kw):
        super().__init__(*a, **kw)
        self._value = ''
    def get(self, *a):
        return self._value
    def insert(self, index, text):
        self._value += str(text)
    def delete(self, *a):
        self._value = ''
class StringVar(_Noop):
    def __init__(self, *a, **kw): self._val = ''
    def get(self): return self._val
    def set(self, v): self._val = str(v)
class IntVar(_Noop):
    def __init__(self, *a, **kw): self._val = 0
    def get(self): return self._val
    def set(self, v): self._val = int(v)
class BooleanVar(_Noop):
    def __init__(self, *a, **kw): self._val = False
    def get(self): return self._val
    def set(self, v): self._val = bool(v)
class Tk(_Noop): pass
class Listbox(_Noop): pass
class Button(_Noop): pass
class Checkbutton(_Noop): pass
class Radiobutton(_Noop): pass
class Canvas(_Noop): pass
class Scrollbar(_Noop): pass
class Menu(_Noop): pass
class OptionMenu(_Noop): pass
class LabelFrame(_Noop): pass
class PanedWindow(_Noop): pass
class Spinbox(_Noop): pass
class Scale(_Noop): pass

END = 'end'
INSERT = 'insert'
NORMAL = 'normal'
DISABLED = 'disabled'
ACTIVE = 'active'
HORIZONTAL = 'horizontal'
VERTICAL = 'vertical'
BOTH = 'both'
LEFT = 'left'
RIGHT = 'right'
TOP = 'top'
BOTTOM = 'bottom'
CENTER = 'center'
X = 'x'
Y = 'y'
NW = 'nw'
NE = 'ne'
SW = 'sw'
SE = 'se'
N = 'n'
S = 's'
E = 'e'
W = 'w'
NSEW = 'nsew'
RAISED = 'raised'
SUNKEN = 'sunken'
FLAT = 'flat'
RIDGE = 'ridge'
GROOVE = 'groove'
SOLID = 'solid'
BROWSE = 'browse'
MULTIPLE = 'multiple'
EXTENDED = 'extended'
SINGLE = 'single'
WORD = 'word'
CHAR = 'char'
NONE = 'none'
TRUE = True
FALSE = False
YES = True
NO = False


class _TtkStub:
    Treeview = _Noop
    Combobox = _Noop
    Notebook = _Noop
    Progressbar = _Noop
    Separator = _Noop
    Sizegrip = _Noop
    Style = _Noop
    Button = _Noop
    Label = _Noop
    Entry = _Noop
    Frame = _Noop
    LabelFrame = _Noop
    Scrollbar = _Noop
    Checkbutton = _Noop
    Radiobutton = _Noop
    Scale = _Noop
    Spinbox = _Noop
    OptionMenu = _Noop
    PanedWindow = _Noop
    Menubutton = _Noop


ttk = _TtkStub()


class _FileDialogStub:
    @staticmethod
    def askopenfilename(*a, **kw): return ''
    @staticmethod
    def askopenfilenames(*a, **kw): return ()
    @staticmethod
    def asksaveasfilename(*a, **kw): return ''
    @staticmethod
    def askdirectory(*a, **kw): return ''

filedialog = _FileDialogStub()


class _MessageBoxStub:
    @staticmethod
    def showinfo(*a, **kw): pass
    @staticmethod
    def showwarning(*a, **kw): pass
    @staticmethod
    def showerror(*a, **kw): pass
    @staticmethod
    def askquestion(*a, **kw): return 'no'
    @staticmethod
    def askokcancel(*a, **kw): return False
    @staticmethod
    def askyesno(*a, **kw): return False
    @staticmethod
    def askretrycancel(*a, **kw): return False

messagebox = _MessageBoxStub()


# ---------------------------------------------------------------------------
# Section 2: Bootstrap (formerly _bootstrap.py)
# ---------------------------------------------------------------------------

_bootstrapped = False
_nsm_temp_dirs = []


def _cleanup_nsm_temps():
    for d in _nsm_temp_dirs:
        shutil.rmtree(d, ignore_errors=True)

atexit.register(_cleanup_nsm_temps)


_nsm_xlsx_cache = {}


def invalidate_nsm_cache(nsm_path: str):
    """Invalidate the cached temp xlsx for a modified nsm file."""
    path_str = str(nsm_path)
    cached = _nsm_xlsx_cache.pop(path_str, None)
    if cached and os.path.exists(cached):
        try:
            os.remove(cached)
        except OSError:
            pass


def _patch_openpyxl_for_nsm():
    """Monkey-patch openpyxl.load_workbook to handle .nsm files transparently.

    When a .nsm file is passed, xlsx is reconstructed from Parquet data
    and cached for the session lifetime.
    """
    import openpyxl as _openpyxl

    _original_load = _openpyxl.load_workbook

    def _patched_load_workbook(filename, *args, **kwargs):
        path_str = str(filename)
        if path_str.lower().endswith('.nsm'):
            xlsx_path = _nsm_xlsx_cache.get(path_str)
            if xlsx_path is None or not os.path.exists(xlsx_path):
                from ns_engine.nsm_io import nsm_to_xlsx
                tmp_dir = tempfile.mkdtemp(prefix='nsm_')
                _nsm_temp_dirs.append(tmp_dir)
                xlsx_path = os.path.join(
                    tmp_dir,
                    os.path.basename(path_str).rsplit('.', 1)[0] + '.xlsx'
                )
                nsm_to_xlsx(path_str, xlsx_path)
                _nsm_xlsx_cache[path_str] = xlsx_path
            return _original_load(xlsx_path, *args, **kwargs)
        return _original_load(filename, *args, **kwargs)

    _openpyxl.load_workbook = _patched_load_workbook


def bootstrap():
    """Call once before importing any ns_engine module."""
    global _bootstrapped
    if _bootstrapped:
        return
    _bootstrapped = True

    _engine_dir = os.path.dirname(os.path.abspath(__file__))

    if _engine_dir not in sys.path:
        sys.path.insert(0, _engine_dir)

    _this = sys.modules[__name__]

    sys.modules.setdefault('tkinter', _this)
    sys.modules.setdefault('tkinter.ttk', types.ModuleType('tkinter.ttk'))
    sys.modules['tkinter.ttk'].Treeview = _Noop
    sys.modules['tkinter.ttk'].Combobox = _Noop
    sys.modules['tkinter.ttk'].Notebook = _Noop
    sys.modules['tkinter.ttk'].Progressbar = _Noop
    sys.modules['tkinter.ttk'].Separator = _Noop
    sys.modules['tkinter.ttk'].Style = _Noop
    sys.modules['tkinter.ttk'].Button = _Noop
    sys.modules['tkinter.ttk'].Label = _Noop
    sys.modules['tkinter.ttk'].Entry = _Noop
    sys.modules['tkinter.ttk'].Frame = _Noop
    sys.modules['tkinter.ttk'].LabelFrame = _Noop
    sys.modules['tkinter.ttk'].Scrollbar = _Noop
    sys.modules['tkinter.ttk'].Checkbutton = _Noop
    sys.modules['tkinter.ttk'].Radiobutton = _Noop
    sys.modules['tkinter.ttk'].Scale = _Noop
    sys.modules['tkinter.ttk'].Spinbox = _Noop
    sys.modules['tkinter.ttk'].OptionMenu = _Noop
    sys.modules['tkinter.ttk'].PanedWindow = _Noop
    sys.modules['tkinter.ttk'].Menubutton = _Noop

    _filedialog = types.ModuleType('tkinter.filedialog')
    _filedialog.askopenfilename = _FileDialogStub.askopenfilename
    _filedialog.askopenfilenames = _FileDialogStub.askopenfilenames
    _filedialog.asksaveasfilename = _FileDialogStub.asksaveasfilename
    _filedialog.askdirectory = _FileDialogStub.askdirectory
    sys.modules.setdefault('tkinter.filedialog', _filedialog)

    _messagebox_mod = types.ModuleType('tkinter.messagebox')
    _messagebox_mod.showinfo = _MessageBoxStub.showinfo
    _messagebox_mod.showwarning = _MessageBoxStub.showwarning
    _messagebox_mod.showerror = _MessageBoxStub.showerror
    _messagebox_mod.askquestion = _MessageBoxStub.askquestion
    _messagebox_mod.askokcancel = _MessageBoxStub.askokcancel
    _messagebox_mod.askyesno = _MessageBoxStub.askyesno
    _messagebox_mod.askretrycancel = _MessageBoxStub.askretrycancel
    sys.modules.setdefault('tkinter.messagebox', _messagebox_mod)

    _ns_stub = types.ModuleType('network_sketcher')
    _ns_stub.__dict__['__file__'] = os.path.join(_engine_dir, 'network_sketcher.py')
    _ns_stub.ns_front_run = _Noop()
    sys.modules.setdefault('network_sketcher', _ns_stub)

    _ns_dev_stub = types.ModuleType('ns_dev')
    _ns_dev_stub.ns_front_run = _Noop()
    sys.modules.setdefault('ns_dev', _ns_dev_stub)

    _patch_openpyxl_for_nsm()


# ---------------------------------------------------------------------------
# Section 3: In-process CLI Runner (formerly runner.py)
# ---------------------------------------------------------------------------

_cli_lock = threading.Lock()


@dataclass
class RunResult:
    """Drop-in replacement for subprocess.CompletedProcess."""
    returncode: int = 0
    stdout: str = ''
    stderr: str = ''


def run_cli(args: list[str], cwd: Optional[str] = None) -> RunResult:
    """Run an NS CLI command in-process.

    Args:
        args: argument list (same format as subprocess argv after script name),
              e.g. ['show', 'area', '--master', '/path/to/master.xlsx']
        cwd:  working directory (defaults to ns_engine package dir)

    Returns:
        RunResult with captured stdout/stderr and returncode.
    """
    bootstrap()

    engine_dir = os.path.dirname(os.path.abspath(__file__))
    result = RunResult()
    captured_out = io.StringIO()
    captured_err = io.StringIO()

    with _cli_lock:
        original_cwd = os.getcwd()
        original_argv = sys.argv[:]
        original_stdout = sys.stdout
        original_stderr = sys.stderr

        try:
            work_dir = cwd or engine_dir
            os.chdir(work_dir)

            sys.argv = ['ns_engine'] + list(args)
            sys.stdout = captured_out
            sys.stderr = captured_err

            import nsm_cli
            nsm_cli.ns_cli_run(args)

        except SystemExit as e:
            result.returncode = int(e.code) if e.code is not None else 0
        except Exception as e:
            result.returncode = 1
            captured_err.write(str(e))
        finally:
            os.chdir(original_cwd)
            sys.argv = original_argv
            sys.stdout = original_stdout
            sys.stderr = original_stderr

    result.stdout = captured_out.getvalue()
    result.stderr = captured_err.getvalue()
    return result


def run_cli_isolated(args: list[str], work_dir: Path,
                     master_filename: str) -> RunResult:
    """Run CLI with an isolated copy of the master file.

    Creates a temporary subdirectory, copies the master file there,
    runs the command, then moves generated outputs back to work_dir.
    """
    task_dir = work_dir / f'_task_{uuid.uuid4().hex[:8]}'
    task_dir.mkdir(exist_ok=True)

    try:
        task_master = task_dir / master_filename
        shutil.copy2(str(work_dir / master_filename), str(task_master))

        isolated_args = []
        replace_next = False
        for a in args:
            if replace_next:
                isolated_args.append(str(task_master))
                replace_next = False
            elif a == '--master':
                isolated_args.append(a)
                replace_next = True
            else:
                isolated_args.append(a)

        result = run_cli(isolated_args, cwd=str(task_dir))

        for f in task_dir.iterdir():
            if f.name == master_filename:
                continue
            dest = work_dir / f.name
            try:
                shutil.move(str(f), str(dest))
            except Exception:
                pass

        return result
    finally:
        shutil.rmtree(str(task_dir), ignore_errors=True)
