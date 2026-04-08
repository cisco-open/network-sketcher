"""
nsm_io.py - I/O adapter for Network Sketcher master files.

Supports both .xlsx (Excel) and .nsm (ZIP + Parquet) formats.
Provides a unified interface for reading/writing section data,
with automatic format detection based on file extension.

.nsm file structure (ZIP containing only Parquet files):
    ROOT_FOLDER.parquet
    POSITION_FOLDER.parquet
    STYLE_FOLDER.parquet
    POSITION_SHAPE.parquet
    STYLE_SHAPE.parquet
    POSITION_LINE.parquet
    POSITION_TAG.parquet
    ATTRIBUTE.parquet
    L2_TABLE.parquet
    L3_TABLE.parquet
    END_MARK.parquet
"""
import io
import os
import zipfile
from typing import Optional

import openpyxl
import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq

SECTION_SHEET_MAP = {
    'ROOT_FOLDER':     ('Master_Data',    '<<ROOT_FOLDER>>'),
    'POSITION_FOLDER': ('Master_Data',    '<<POSITION_FOLDER>>'),
    'STYLE_FOLDER':    ('Master_Data',    '<<STYLE_FOLDER>>'),
    'POSITION_SHAPE':  ('Master_Data',    '<<POSITION_SHAPE>>'),
    'STYLE_SHAPE':     ('Master_Data',    '<<STYLE_SHAPE>>'),
    'POSITION_LINE':   ('Master_Data',    '<<POSITION_LINE>>'),
    'POSITION_TAG':    ('Master_Data',    '<<POSITION_TAG>>'),
    'ATTRIBUTE':       ('Master_Data',    '<<ATTRIBUTE>>'),
    'END_MARK':        ('Master_Data',    '<<END_MARK>>'),
    'L2_TABLE':        ('Master_Data_L2', '<<L2_TABLE>>'),
    'L3_TABLE':        ('Master_Data_L3', '<<L3_TABLE>>'),
}

ALL_SECTIONS = list(SECTION_SHEET_MAP.keys())

MASTER_DATA_SECTIONS = [
    'ROOT_FOLDER', 'POSITION_FOLDER', 'STYLE_FOLDER',
    'POSITION_SHAPE', 'STYLE_SHAPE', 'POSITION_LINE',
    'POSITION_TAG', 'ATTRIBUTE', 'END_MARK',
]


def _section_key(section_name: str) -> str:
    """Normalize <<SECTION>> or SECTION to bare key."""
    return section_name.strip('<>').strip()


def _parquet_filename(section_key: str) -> str:
    return f'{section_key}.parquet'


# ========== Type-preserving encoding for Parquet storage ==========
# Parquet columns require uniform types, so all values are stored as strings.
# A single-byte prefix distinguishes original Python types:
#   \x00I = int, \x00F = float, \x00B = bool
#   No prefix = str (or empty string for None)

_TP = '\x00'


def _encode_cell(value):
    """Encode a cell value with type prefix for Parquet storage."""
    if value is None or value == '':
        return ''
    if isinstance(value, bool):
        return f'{_TP}B{value}'
    if isinstance(value, int):
        return f'{_TP}I{value}'
    if isinstance(value, float):
        return f'{_TP}F{repr(value)}'
    return str(value)


def _decode_cell(v):
    """Decode a Parquet-stored value back to its original Python type."""
    if not isinstance(v, str) or v == '':
        return '' if (v is None or v == '') else v
    if len(v) >= 2 and v[0] == _TP:
        tc, payload = v[1], v[2:]
        if tc == 'I':
            return int(payload)
        if tc == 'F':
            return float(payload)
        if tc == 'B':
            return payload == 'True'
    return v


# ========== xlsx helpers ==========

def _find_section_rows(ws, section_tag: str):
    """Find start_row and end_row for a section in a worksheet.

    Returns (start_row, end_row) where start_row is the tag row
    and end_row is the last data row before the next tag or empty gap.
    """
    start_row = None
    empty_count = 0

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=100000, min_col=1, max_col=1, values_only=True),
        start=1
    ):
        val = row[0]
        if val == section_tag:
            start_row = row_idx
            empty_count = 0
            continue

        if start_row is not None:
            if isinstance(val, str) and val.startswith('<<') and val.endswith('>>'):
                return start_row, row_idx - 1

            if val is None:
                empty_count += 1
                if empty_count >= 3000:
                    return start_row, row_idx - empty_count
            else:
                empty_count = 0

    if start_row is not None:
        max_row = ws.max_row or start_row
        return start_row, max_row

    return None, None


def _read_section_from_xlsx(xlsx_path: str, ws_name: str,
                            section_tag: str) -> pd.DataFrame:
    """Read a section from an xlsx file and return as DataFrame."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if ws_name not in wb.sheetnames:
            return pd.DataFrame()
        ws = wb[ws_name]
        start_row, end_row = _find_section_rows(ws, section_tag)
        if start_row is None:
            return pd.DataFrame()

        rows = []
        for row in ws.iter_rows(min_row=start_row, max_row=end_row,
                                values_only=True):
            cleaned = [_encode_cell(cell) for cell in row]
            trailing_empty = 0
            for v in reversed(cleaned):
                if v == '':
                    trailing_empty += 1
                else:
                    break
            if trailing_empty > 0 and trailing_empty < len(cleaned):
                cleaned = cleaned[:len(cleaned) - trailing_empty]
            rows.append(cleaned)

        if not rows:
            return pd.DataFrame()

        max_cols = max(len(r) for r in rows)
        for r in rows:
            while len(r) < max_cols:
                r.append('')

        cols = [f'col_{i}' for i in range(max_cols)]
        df = pd.DataFrame(rows, columns=cols)
        return df
    finally:
        wb.close()


def _read_flow_data_from_xlsx(xlsx_path: str) -> pd.DataFrame:
    """Read the Flow_Data sheet (if present) as a flat DataFrame.

    Unlike _read_section_from_xlsx, this reads the entire sheet without
    requiring <<SECTION>> tags, since Flow_Data is a flat table.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if 'Flow_Data' not in wb.sheetnames:
            return pd.DataFrame()
        ws = wb['Flow_Data']
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([_encode_cell(c) for c in row])
        if not rows:
            return pd.DataFrame()
        max_cols = max(len(r) for r in rows)
        for r in rows:
            while len(r) < max_cols:
                r.append('')
        return pd.DataFrame(rows, columns=[f'col_{i}' for i in range(max_cols)])
    finally:
        wb.close()


def _write_section_to_xlsx_all(xlsx_path: str,
                               sections_dict: dict[str, pd.DataFrame]):
    """Write all sections to an xlsx file (creates new file)."""
    wb = openpyxl.Workbook()

    sheets_data: dict[str, list[tuple[str, pd.DataFrame]]] = {}
    for section_key, df in sections_dict.items():
        if section_key in SECTION_SHEET_MAP:
            ws_name, tag = SECTION_SHEET_MAP[section_key]
        else:
            ws_name = 'Master_Data'
            tag = f'<<{section_key}>>'
        if ws_name not in sheets_data:
            sheets_data[ws_name] = []
        sheets_data[ws_name].append((tag, df))

    first_sheet = True
    for ws_name in ['Master_Data', 'Master_Data_L2', 'Master_Data_L3']:
        if first_sheet:
            ws = wb.active
            ws.title = ws_name
            first_sheet = False
        else:
            ws = wb.create_sheet(ws_name)
        if ws_name not in sheets_data:
            continue

        current_row = 1
        for tag, df in sheets_data[ws_name]:
            for _, row_data in df.iterrows():
                for col_idx, value in enumerate(row_data, start=1):
                    decoded = _decode_cell(value)
                    cell_value = None if decoded == '' else decoded
                    ws.cell(row=current_row, column=col_idx, value=cell_value)
                current_row += 1
            current_row += 2

    # Restore Flow_Data sheet if present in nsm
    if 'FLOW_DATA' in sections_dict:
        flow_df = sections_dict['FLOW_DATA']
        ws_flow = wb.create_sheet('Flow_Data')
        for _, row_data in flow_df.iterrows():
            row_values = []
            for value in row_data:
                decoded = _decode_cell(value)
                row_values.append(None if decoded == '' else decoded)
            ws_flow.append(row_values)

    wb.save(xlsx_path)
    wb.close()


# ========== Parquet / .nsm helpers ==========

def _read_section_from_nsm(nsm_path: str, section_key: str) -> pd.DataFrame:
    """Read a single section from a .nsm (ZIP+Parquet) file."""
    pq_name = _parquet_filename(section_key)
    with zipfile.ZipFile(nsm_path, 'r') as zf:
        if pq_name not in zf.namelist():
            return pd.DataFrame()
        with zf.open(pq_name) as f:
            table = pq.read_table(io.BytesIO(f.read()))
            return table.to_pandas()


def _write_section_to_nsm(nsm_path: str, section_key: str,
                           df: pd.DataFrame):
    """Write a single section to a .nsm file (updates in place)."""
    pq_name = _parquet_filename(section_key)

    existing = {}
    if os.path.exists(nsm_path):
        with zipfile.ZipFile(nsm_path, 'r') as zf:
            for name in zf.namelist():
                if name != pq_name and not name.endswith('.xlsx'):
                    existing[name] = zf.read(name)

    with zipfile.ZipFile(nsm_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for name, data in existing.items():
            zf.writestr(name, data)
        buf = io.BytesIO()
        table = pa.Table.from_pandas(df.astype(str), preserve_index=False)
        pq.write_table(table, buf)
        zf.writestr(pq_name, buf.getvalue())

    try:
        from ns_engine.nsm_adapter import invalidate_nsm_cache
        invalidate_nsm_cache(nsm_path)
    except ImportError:
        pass


def _read_all_from_nsm(nsm_path: str) -> dict[str, pd.DataFrame]:
    """Read all sections from a .nsm file."""
    result = {}
    with zipfile.ZipFile(nsm_path, 'r') as zf:
        for name in zf.namelist():
            if name.endswith('.parquet'):
                section_key = name[:-len('.parquet')]
                with zf.open(name) as f:
                    table = pq.read_table(io.BytesIO(f.read()))
                    result[section_key] = table.to_pandas()
    return result


def _write_all_to_nsm(nsm_path: str,
                      sections: dict[str, pd.DataFrame]):
    """Write all sections to a new .nsm file."""
    with zipfile.ZipFile(nsm_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for section_key, df in sections.items():
            buf = io.BytesIO()
            table = pa.Table.from_pandas(df.astype(str), preserve_index=False)
            pq.write_table(table, buf)
            zf.writestr(_parquet_filename(section_key), buf.getvalue())


def load_sections_bulk_nsm(nsm_path: str, section_keys: list[str]) -> dict[str, pd.DataFrame]:
    """Read multiple sections from a .nsm file in a single ZIP open."""
    results = {}
    wanted = {_parquet_filename(k): k for k in section_keys}
    with zipfile.ZipFile(nsm_path, 'r') as zf:
        names = set(zf.namelist())
        for pq_name, key in wanted.items():
            if pq_name in names:
                with zf.open(pq_name) as f:
                    table = pq.read_table(io.BytesIO(f.read()))
                    results[key] = table.to_pandas()
            else:
                results[key] = pd.DataFrame()
    return results


# ========== Public API ==========

def is_nsm(file_path: str) -> bool:
    """Check if file is .nsm format."""
    return str(file_path).lower().endswith('.nsm')


def is_xlsx(file_path: str) -> bool:
    """Check if file is .xlsx format."""
    return str(file_path).lower().endswith('.xlsx')


def load_section(file_path: str, section_name: str) -> pd.DataFrame:
    """Load a single section from a master file (.xlsx or .nsm).

    Args:
        file_path: Path to .xlsx or .nsm file
        section_name: Section name, e.g. '<<POSITION_SHAPE>>' or 'POSITION_SHAPE'

    Returns:
        DataFrame with section data (all columns as strings)
    """
    key = _section_key(section_name)

    if is_nsm(file_path):
        return _read_section_from_nsm(file_path, key)
    else:
        if key in SECTION_SHEET_MAP:
            ws_name, tag = SECTION_SHEET_MAP[key]
        else:
            ws_name = 'Master_Data'
            tag = f'<<{key}>>'
        return _read_section_from_xlsx(file_path, ws_name, tag)


def save_section(file_path: str, section_name: str, df: pd.DataFrame):
    """Save a single section to a master file (.xlsx or .nsm).

    For .nsm: updates the specific Parquet file within the ZIP.
    For .xlsx: currently not supported for partial writes (use save_all).
    """
    key = _section_key(section_name)

    if is_nsm(file_path):
        _write_section_to_nsm(file_path, key, df)
    else:
        raise NotImplementedError(
            'Partial section writes to .xlsx are not supported via nsm_io. '
            'Use the existing openpyxl-based write functions or save_all().'
        )


def load_all_sections(file_path: str) -> dict[str, pd.DataFrame]:
    """Load all sections from a master file.

    Returns:
        dict mapping section key (e.g. 'POSITION_SHAPE') to DataFrame.
    """
    if is_nsm(file_path):
        return _read_all_from_nsm(file_path)
    else:
        result = {}
        for key, (ws_name, tag) in SECTION_SHEET_MAP.items():
            df = _read_section_from_xlsx(file_path, ws_name, tag)
            if not df.empty:
                result[key] = df
        return result


def save_all_sections(file_path: str, sections: dict[str, pd.DataFrame]):
    """Save all sections to a master file.

    Args:
        file_path: Path to .xlsx or .nsm file
        sections: dict mapping section key to DataFrame
    """
    if is_nsm(file_path):
        _write_all_to_nsm(file_path, sections)
    else:
        _write_section_to_xlsx_all(file_path, sections)


def xlsx_to_nsm(xlsx_path: str, nsm_path: str):
    """Convert an xlsx master file to .nsm (ZIP + Parquet) format."""
    sections = load_all_sections(xlsx_path)
    flow_df = _read_flow_data_from_xlsx(xlsx_path)

    with zipfile.ZipFile(nsm_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for section_key, df in sections.items():
            buf = io.BytesIO()
            table = pa.Table.from_pandas(df.astype(str), preserve_index=False)
            pq.write_table(table, buf)
            zf.writestr(_parquet_filename(section_key), buf.getvalue())
        if not flow_df.empty:
            buf = io.BytesIO()
            table = pa.Table.from_pandas(flow_df.astype(str), preserve_index=False)
            pq.write_table(table, buf)
            zf.writestr('FLOW_DATA.parquet', buf.getvalue())


def extract_original_xlsx(nsm_path: str, xlsx_path: str) -> bool:
    """Legacy stub — always returns False.

    The _original.xlsx embedding was removed to reduce nsm file size.
    Use nsm_to_xlsx() to reconstruct xlsx from Parquet data instead.
    """
    return False


def nsm_to_xlsx(nsm_path: str, xlsx_path: str):
    """Convert an .nsm file to xlsx format (for export/Offline compatibility).

    Reconstructs xlsx from Parquet data.
    """
    sections = _read_all_from_nsm(nsm_path)
    _write_section_to_xlsx_all(xlsx_path, sections)
