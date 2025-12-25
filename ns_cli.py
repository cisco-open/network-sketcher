'''
SPDX-License-Identifier: Apache-2.0

Copyright 2023 Cisco Systems, Inc. and its affiliates

Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.
'''

import tkinter as tk ,tkinter.ttk as ttk,tkinter.filedialog, tkinter.messagebox
import sys, os, subprocess ,webbrowser ,openpyxl

class ns_cli_run():
    def __init__(self,argv_array):
        # GET file path of master
        master_file_path = get_next_arg(argv_array, '--master')
        if master_file_path == None:
            print('[ERROR] Master file must be selected --master <filepath>.')
            exit()

        #only cli, the flag changed to True
        self.cli_flag_no_export = True

        # run show command
        if 'show' in argv_array:
            print_type(self,argv_array, ns_cli_run.cli_show(self, master_file_path,argv_array))
            exit()
        elif 'add' in argv_array:
            print_type(self,argv_array, ns_cli_run.cli_add(self, master_file_path,argv_array))
            exit()
        elif 'rename' in argv_array:
            print_type(self,argv_array, ns_cli_run.cli_rename(self, master_file_path,argv_array))
            exit()
        elif 'delete' in argv_array:
            print_type(self, argv_array, ns_cli_run.cli_delete(self, master_file_path, argv_array))
            exit()
        # ★★★ Add export command ★★★
        elif 'export' in argv_array:
            print_type(self, argv_array, ns_cli_run.cli_export(self, master_file_path, argv_array))
            exit()
        # ★★★ End of addition ★★★
        else:
            print('[ERROR] Supported commands are as follows', 'add', 'delete', 'rename', 'show', 'export', sep='\n')

    def cli_export(self, master_file_path, argv_array):
        """
        Handle export commands

        Args:
            master_file_path: Path to the master Excel file
            argv_array: Command line arguments array

        Returns:
            List with status message or error
        """
        import ns_def
        import tkinter as tk

        next_arg = get_next_arg(argv_array, 'export')
        export_command_list = [
            'export ai_context_file',
            'export master_file_backup',
            'export master_file_nodata',
            'export device_file',
        ]

        if next_arg == None or '--' in next_arg or str('export ' + next_arg) not in export_command_list:
            print('[ERROR] Supported commands are as follows')
            for tmp_export_command_list in export_command_list:
                print(tmp_export_command_list)
            exit()


        if next_arg == 'ai_context_file':
            try:
                # Validate master file exists
                if not os.path.isfile(master_file_path):
                    return ([f'[ERROR] Master file not found: {master_file_path}'])

                # Validate master file format
                filename = os.path.basename(master_file_path)
                if not filename.startswith('[MASTER]'):
                    return ([f'[ERROR] Master file must start with "[MASTER]". Current filename: {filename}'])

                # ★★★ Check for --accept-security-risk flag ★★★
                confirm_risk = '--accept-security-risk' in argv_array

                if not confirm_risk:
                    # ★★★ Interactive confirmation (shows tkinter window - acceptable) ★★★
                    warning_msg = '''
        [WARNING] AI Context File Export
        ================================================================================
        The AI Context file contains ALL configuration information from the master file,
        including:
        - Network topology and device configurations
        - IP addresses and network segments
        - Device names and connections
        - All technical details about your network

        SECURITY RISKS:
        - Data leakage risk if loaded into Large Language Models (LLMs)
        - Sensitive network information may be exposed
        - This file should be treated as confidential

        Do you understand and accept these risks?
        Type 'YES' to proceed or 'NO' to cancel: '''

                    print(warning_msg, end='', flush=True)

                    # Get user confirmation
                    try:
                        user_response = input().strip().upper()
                    except (EOFError, KeyboardInterrupt):
                        return (['\n[Info] Export cancelled by user'])

                    if user_response != 'YES':
                        return ([f'[Info] Export cancelled. User response: {user_response}'])
                    # ★★★ End of interactive confirmation ★★★
                else:
                    # ★★★ Flag-based confirmation (no interaction needed) ★★★
                    print('[Info] Security risk acknowledged via --accept-security-risk flag')
                    # ★★★ End of flag-based confirmation ★★★

                # Generate AI context file path (cross-platform compatible)
                iDir = os.path.dirname(master_file_path)
                if not iDir:
                    iDir = os.getcwd()

                basename_without_ext = os.path.splitext(os.path.basename(master_file_path))[0]
                basename_without_ext = basename_without_ext.replace('[MASTER]', '')

                ai_context_filename = f'[AI_Context]{basename_without_ext}.txt'
                ai_context_file_path = os.path.join(iDir, ai_context_filename)

                # Check if AI context file already exists
                file_was_removed = False
                if os.path.isfile(ai_context_file_path):
                    try:
                        os.remove(ai_context_file_path)
                        file_was_removed = True
                    except Exception as e:
                        return ([f'[ERROR] Failed to remove existing AI context file: {str(e)}'])

                # ★★★ Use Mock Entry objects (no tkinter windows) ★★★
                class MockEntry:
                    """Mock Entry widget that doesn't require tkinter"""

                    def __init__(self, value=''):
                        self.value = value

                    def get(self):
                        return self.value

                    def delete(self, start, end=None):
                        pass

                    def insert(self, index, value):
                        self.value = value

                # Create mock Entry objects
                self.inFileTxt_L2_3_1 = MockEntry(master_file_path)
                self.outFileTxt_11_3 = MockEntry(ai_context_file_path)
                # ★★★ End of mock setup ★★★

                # Get show_l3_interface data for AI context
                argv_show = ['show', 'l3_interface']
                self.show_l3_interface = ns_cli_run.cli_show(self, master_file_path, argv_show)

                # Call AI context export function
                try:
                    import ns_extensions
                    ns_extensions.ai_context.export_ai_context_file(self, 'dummy')
                except Exception as e:
                    import traceback
                    error_detail = traceback.format_exc()
                    return ([f'[ERROR] Failed during AI context export: {str(e)}\n\nDetails:\n{error_detail}'])

                # Verify AI context file was created
                import time
                time.sleep(0.3)

                if os.path.isfile(ai_context_file_path):
                    ai_context_file_size = os.path.getsize(ai_context_file_path)

                    # Build simple return message
                    return_text = ''

                    if file_was_removed:
                        return_text += '[Info] Removed existing AI context file\n'

                    return_text += '--- AI Context file created successfully ---\n'
                    return_text += f'AI Context file: {ai_context_file_path}\n'
                    return_text += f'  Size: {ai_context_file_size:,} bytes\n'
                    return_text += '\n[IMPORTANT] This file contains sensitive network information.\n'
                    return_text += 'Please handle it securely and do not share it without proper authorization.'

                    return ([return_text])
                else:
                    return ([f'[ERROR] Failed to create AI context file'])

            except Exception as e:
                import traceback
                error_detail = traceback.format_exc()
                return ([f'[ERROR] Failed to export AI context file: {str(e)}\n\nDetails:\n{error_detail}'])

        if next_arg == 'device_file':
            try:
                # Validate master file exists
                if not os.path.isfile(master_file_path):
                    return ([f'[ERROR] Master file not found: {master_file_path}'])

                # Validate master file format
                filename = os.path.basename(master_file_path)
                if not filename.startswith('[MASTER]'):
                    return ([f'[ERROR] Master file must start with "[MASTER]". Current filename: {filename}'])

                # ★★★ Generate device file path (cross-platform compatible) ★★★
                iDir = os.path.dirname(master_file_path)
                if not iDir:
                    iDir = os.getcwd()

                basename_without_ext = os.path.splitext(os.path.basename(master_file_path))[0]
                # Remove [MASTER] prefix from basename
                basename_without_ext = basename_without_ext.replace('[MASTER]', '')

                device_filename = f'[DEVICE]{basename_without_ext}.xlsx'

                # Use os.path.join for cross-platform compatibility
                device_file_path = os.path.join(iDir, device_filename)
                # ★★★ End of path generation ★★★

                # ★★★ Track if existing file was removed ★★★
                file_was_removed = False
                # ★★★ End of tracking variable ★★★

                # Check if device file already exists
                if os.path.isfile(device_file_path):
                    # Check if file is open
                    if ns_def.check_file_open(device_file_path):
                        return ([f'[ERROR] Device file is currently open. Please close it first: {device_file_path}'])

                    # Remove existing file
                    try:
                        os.remove(device_file_path)
                        file_was_removed = True  # ★★★ Set flag ★★★
                    except Exception as e:
                        return ([f'[ERROR] Failed to remove existing device file: {str(e)}'])

                # Set up parameters for device file creation
                import tkinter as tk
                dummy_tk = tk.Toplevel()
                dummy_tk.withdraw()

                self.inFileTxt_11_1 = tk.Entry(dummy_tk)
                self.inFileTxt_11_1.delete(0, tkinter.END)
                self.inFileTxt_11_1.insert(tk.END, master_file_path)

                self.outFileTxt_11_2 = tk.Entry(dummy_tk)
                self.outFileTxt_11_2.delete(0, tkinter.END)
                self.outFileTxt_11_2.insert(tk.END, device_file_path)

                self.click_value = '11-4'

                # Step 1: Create device file and L1 Table
                try:
                    import ns_l1_table_from_master
                    ns_l1_table_from_master.ns_l1_table_from_master.__init__(self)
                except Exception as e:
                    dummy_tk.destroy()
                    import traceback
                    error_detail = traceback.format_exc()
                    return ([f'[ERROR] Failed to create L1 Table: {str(e)}\n\nDetails:\n{error_detail}'])

                # Step 2: Create L2 Table
                try:
                    self.inFileTxt_L2_1_1 = tk.Entry(dummy_tk)
                    self.inFileTxt_L2_1_1.delete(0, tkinter.END)
                    self.inFileTxt_L2_1_1.insert(tk.END, master_file_path)

                    self.click_value = 'L2-1-2'

                    # Check if Master_Data_L2 sheet exists
                    input_excel_master_data = openpyxl.load_workbook(master_file_path)
                    ws_list = input_excel_master_data.sheetnames

                    if 'Master_Data_L2' not in ws_list:
                        import ns_l2_table_from_master
                        ns_l2_table_from_master.ns_l2_table_from_master.__init__(self)
                    else:
                        import ns_l2_table_from_master
                        ns_l2_table_from_master.ns_l2_table_from_master_l2_sheet.__init__(self)

                    input_excel_master_data.close()
                except Exception as e:
                    dummy_tk.destroy()
                    import traceback
                    error_detail = traceback.format_exc()
                    return ([f'[ERROR] Failed to create L2 Table: {str(e)}\n\nDetails:\n{error_detail}'])

                # Step 3: Create L3 Table
                try:
                    self.inFileTxt_L3_1_1 = tk.Entry(dummy_tk)
                    self.inFileTxt_L3_1_1.delete(0, tkinter.END)
                    self.inFileTxt_L3_1_1.insert(tk.END, master_file_path)

                    self.click_value = 'L3-1-2'

                    input_excel_master_data = openpyxl.load_workbook(master_file_path)
                    ws_list = input_excel_master_data.sheetnames

                    if 'Master_Data_L3' not in ws_list:
                        import ns_l3_table_from_master
                        ns_l3_table_from_master.ns_l3_table_from_master.__init__(self)
                    else:
                        import ns_l3_table_from_master
                        ns_l3_table_from_master.ns_l3_table_from_master_l3_sheet.__init__(self)

                    input_excel_master_data.close()
                except Exception as e:
                    dummy_tk.destroy()
                    import traceback
                    error_detail = traceback.format_exc()
                    return ([f'[ERROR] Failed to create L3 Table: {str(e)}\n\nDetails:\n{error_detail}'])

                # Step 4: Create Attribute Table
                try:
                    self.click_value = 'ATTR-1-1'

                    if not ns_def.check_file_open(device_file_path):
                        import ns_attribute_table_from_master
                        ns_attribute_table_from_master.ns_attribute_table_from_master.__init__(self)
                except Exception as e:
                    dummy_tk.destroy()
                    import traceback
                    error_detail = traceback.format_exc()
                    return ([f'[ERROR] Failed to create Attribute Table: {str(e)}\n\nDetails:\n{error_detail}'])

                dummy_tk.destroy()

                # Final verification
                import time
                time.sleep(0.5)  # Give OS time to flush file system

                if os.path.isfile(device_file_path):
                    # ★★★ Build simple return message ★★★
                    return_text = ''

                    # Add info message if existing file was removed
                    if file_was_removed:
                        return_text += '[Info] Removed existing device file\n'

                    return_text += '--- Device file created successfully ---'

                    return ([return_text])
                    # ★★★ End of simple message ★★★
                else:
                    return ([f'[ERROR] Device file not found at expected path: {device_file_path}'])

            except Exception as e:
                import traceback
                error_detail = traceback.format_exc()
                return ([f'[ERROR] Failed to export device file: {str(e)}\n\nDetails:\n{error_detail}'])

        if next_arg == 'master_file_backup':
            try:
                # Check if master file exists
                if not os.path.isfile(master_file_path):
                    return ([f'[ERROR] Master file not found: {master_file_path}'])

                # Call ns_def.get_backup_filename to create backup
                backup_file_path = ns_def.get_backup_filename(master_file_path)

                # Verify backup was created
                if os.path.isfile(backup_file_path):
                    # Get file sizes for verification
                    original_size = os.path.getsize(master_file_path)
                    backup_size = os.path.getsize(backup_file_path)

                    return_text = '--- Master file backup created successfully ---\n'
                    #return_text += f'Original file: {master_file_path}\n'
                    #return_text += f'  Size: {original_size:,} bytes\n'
                    return_text += f'Backup file: {backup_file_path}\n'
                    #return_text += f'  Size: {backup_size:,} bytes\n'

                    if original_size != backup_size:
                        return_text += f'Warning: File sizes differ (original: {original_size}, backup: {backup_size})'

                    return ([return_text])
                else:
                    return ([f'[ERROR] Failed to create backup file'])

            except Exception as e:
                import traceback
                traceback.print_exc()
                return ([f'[ERROR] Failed to create backup: {str(e)}'])

        # ★★★ Modified: Use --master for output path ★★★
        if next_arg == 'master_file_nodata':
            try:
                # Use master_file_path as output file path
                output_file_path = master_file_path

                # Validate file path
                # Check if filename starts with [MASTER]
                filename = os.path.basename(output_file_path)
                if not filename.startswith('[MASTER]'):
                    return ([f'[ERROR] Filename must start with "[MASTER]". Current filename: {filename}'])

                # Check if extension is .xlsx
                if not output_file_path.endswith('.xlsx'):
                    return ([f'[ERROR] File extension must be ".xlsx". Current file: {output_file_path}'])

                # Check if directory exists
                output_dir = os.path.dirname(output_file_path)
                if output_dir and not os.path.isdir(output_dir):
                    return ([f'[ERROR] Directory does not exist: {output_dir}'])

                # Check if file already exists
                if os.path.isfile(output_file_path):
                    return ([f'[ERROR] File already exists: {output_file_path}\n  Please specify a different filename or delete the existing file.'])

                # Create empty master file
                result = def_common.create_empty_master_file(output_file_path)

                if result['status'] == 'success':
                    return_text = '--- Empty master file created successfully ---\n'
                    return_text += f'File: {output_file_path}\n'
                    return_text += f'Size: {os.path.getsize(output_file_path):,} bytes\n'
                    return_text += 'Sheets: Master_Data, Master_Data_L2, Master_Data_L3\n'
                    return_text += 'The file contains no areas or devices.'
                    return ([return_text])
                else:
                    return ([f"[ERROR] {result['message']}"])

            except Exception as e:
                import traceback
                traceback.print_exc()
                return ([f'[ERROR] Failed to create empty master file: {str(e)}'])
        # ★★★ End of modification ★★★



    def cli_rename(self, master_file_path, argv_array): # add at ver 2.5.4
        import ns_def
        import tkinter as tk
        next_arg = get_next_arg(argv_array, 'rename')
        rename_command_list = [ \
            'rename area', \
            'rename device', \
            'rename l3_instance', \
            'rename port', \
            ]

        if next_arg == None or '--' in next_arg or str('rename ' + next_arg) not in rename_command_list:
            print('[ERROR] Supported commands are as follows')
            for tmp_add_command_list in rename_command_list:
                print(tmp_add_command_list)
            exit()

        if next_arg == 'area':
            idx = argv_array.index('area')
            updated_name_array = [str(argv_array[idx + 1]), str(argv_array[idx + 2]).replace(' ', '').replace('　', '')]

            if isinstance(updated_name_array[1], str) and updated_name_array[1].strip(' \u3000') == '':
                return_text = '[ERROR] This area name is invalid.  --- ' + ' ' + updated_name_array[1]
                return ([return_text])
            if  '--' in str(argv_array[idx + 2]):
                return_text = '[ERROR] This area name is invalid.  --- ' + ' ' + updated_name_array[1]
                return ([return_text])

            '''rename area name in Master_Data.'''
            ws_name = 'Master_Data'
            ppt_meta_file = master_file_path

            position_folder_array = ns_def.convert_master_to_array(ws_name, ppt_meta_file, '<<POSITION_FOLDER>>')
            style_folder_array = ns_def.convert_master_to_array(ws_name, ppt_meta_file, '<<STYLE_FOLDER>>')
            position_shape_array = ns_def.convert_master_to_array(ws_name, ppt_meta_file, '<<POSITION_SHAPE>>')

            # Replace updated_name_array[0] with updated_name_array[1] (no def).
            src = updated_name_array[0]
            dst = updated_name_array[1]
            sentinels = {"<SET_WIDTH>", "<DEFAULT>", "<END>","N/A","<EMPTY>"}

            # position_folder_array and style_folder_array: target all elements in the inner list
            # (exclude sentinels; exclude items whose leading number is 1).
            for arr in (position_folder_array, style_folder_array):
                for item in arr:
                    if not isinstance(item, list) or len(item) < 2:
                        continue
                    idx, values = item[0], item[1]
                    if idx == 1 or not isinstance(values, list):
                        continue
                    for j, val in enumerate(values):
                        if isinstance(val, str) and val == src and val not in sentinels:
                            values[j] = dst

            # position_shape_array: only the first item of the second element (the inner list) is targeted
            # (exclude sentinels; exclude items whose leading number is 1).
            for item in position_shape_array:
                if not isinstance(item, list) or len(item) < 2:
                    continue
                idx, values = item[0], item[1]
                if idx == 1 or not isinstance(values, list) or len(values) == 0:
                    continue
                if updated_name_array[1] == item[1][0]:
                    return_text = '[ERROR] This Area name is already exist.  --- ' + updated_name_array[1]
                    return ([return_text])
                first = values[0]
                if isinstance(first, str) and first == src and first not in sentinels:
                    values[0] = dst

            '''rename area name in Master_Data_L2 and Master_Data_L3'''
            ws_name_l2 = 'Master_Data_L2'
            l2_table_array = ns_def.convert_master_to_array(ws_name_l2, ppt_meta_file, '<<L2_TABLE>>')
            ws_name_l3 = 'Master_Data_L3'
            l3_table_array = ns_def.convert_master_to_array(ws_name_l3, ppt_meta_file, '<<L3_TABLE>>')

            # l2_table_array: only replace the first item of the inner list when it matches src and is not a sentinel.
            for item in l2_table_array:
                if not isinstance(item, list) or len(item) < 2:
                    continue
                idx, values = item[0], item[1]
                if idx in (1, 2) or not isinstance(values, list) or len(values) == 0:
                    continue
                first = values[0]
                if isinstance(first, str) and first == src and first not in sentinels:
                    values[0] = dst

            # l3_table_array: only replace the first item of the inner list when it matches src and is not a sentinel.
            for item in l3_table_array:
                if not isinstance(item, list) or len(item) < 2:
                    continue
                idx, values = item[0], item[1]
                if idx in (1, 2) or not isinstance(values, list) or len(values) == 0:
                    continue
                first = values[0]
                if isinstance(first, str) and first == src and first not in sentinels:
                    values[0] = dst

            self.position_folder_tuple = ns_def.convert_array_to_tuple(position_folder_array)
            self.style_folder_tuple = ns_def.convert_array_to_tuple(style_folder_array)
            self.position_shape_tuple = ns_def.convert_array_to_tuple(position_shape_array)
            self.l2_table_tuple = ns_def.convert_array_to_tuple(l2_table_array)
            self.l3_table_tuple = ns_def.convert_array_to_tuple(l3_table_array)

            excel_file_path = ppt_meta_file
            worksheet_name = ws_name
            offset_row = 0
            offset_column = 0

            master_excel_meta = self.position_folder_tuple
            section_write_to = '<<POSITION_FOLDER>>'
            ns_def.overwrite_excel_meta(master_excel_meta, excel_file_path, worksheet_name, section_write_to, offset_row, offset_column)

            master_excel_meta = self.style_folder_tuple
            section_write_to = '<<STYLE_FOLDER>>'
            ns_def.overwrite_excel_meta(master_excel_meta, excel_file_path, worksheet_name, section_write_to, offset_row, offset_column)

            master_excel_meta = self.position_shape_tuple
            section_write_to = '<<POSITION_SHAPE>>'
            ns_def.overwrite_excel_meta(master_excel_meta, excel_file_path, worksheet_name, section_write_to, offset_row, offset_column)

            worksheet_name = ws_name_l2
            master_excel_meta = self.l2_table_tuple
            section_write_to = '<<L2_TABLE>>'
            ns_def.overwrite_excel_meta(master_excel_meta, excel_file_path, worksheet_name, section_write_to, offset_row, offset_column)

            worksheet_name = ws_name_l3
            master_excel_meta = self.l3_table_tuple
            section_write_to = '<<L3_TABLE>>'
            ns_def.overwrite_excel_meta(master_excel_meta, excel_file_path, worksheet_name, section_write_to, offset_row, offset_column)

            return_text = '--- Area renamed --- ' + str(updated_name_array[0]) + ' -> ' + str(updated_name_array[1])
            return ([return_text])


        if next_arg == 'port':
            import ns_def
            idx = argv_array.index('port')
            updated_name_array = [str(argv_array[idx + 1]), str(argv_array[idx + 2]),str(argv_array[idx + 3])]
            updated_name_array.append(ns_def.adjust_portname(updated_name_array[1]))
            updated_name_array.append(ns_def.adjust_portname(updated_name_array[2]))
            flag_l1_port_name = False
            flag_l2_port_name = False

            if isinstance(updated_name_array[0], str) and updated_name_array[0].strip(' \u3000') == '':
                return_text = '[ERROR] This device name is invalid.  --- ' + ' ' + updated_name_array[0]
                return ([return_text])

            if ns_def.get_if_value(str(updated_name_array[1])) == -1 or ns_def.get_if_value(str(updated_name_array[2])) == -1:
                return_text = '[ERROR] The port name is invalid.  --- ' + ' ' + updated_name_array[1] + '  or  ' + updated_name_array[2]
                return ([return_text])


            '''
            rename L1 port name in Master_Data
            '''
            ws_name = 'Master_Data'
            ppt_meta_file = master_file_path

            position_line_array = ns_def.convert_master_to_array(ws_name, ppt_meta_file, '<<POSITION_LINE>>')

            # position_line_array: The first and second lines are not included. Only the first (0) and second (1) are replaced.
            for row in position_line_array:
                if isinstance(row, list) and len(row) == 2:
                    line_no, fields = row
                    if line_no not in (1, 2) and isinstance(fields, list):
                        if fields[0] == updated_name_array[0]:
                            if updated_name_array[4][1] == fields[12] and updated_name_array[4][2] == str(ns_def.split_portname(fields[2])[1]):
                                return_text = '[ERROR] This port name is already exist.  --- ' + updated_name_array[0]  + ' ' + updated_name_array[2]
                                return ([return_text])
                            if updated_name_array[3][1] == fields[12] and updated_name_array[3][2] == str(ns_def.split_portname(fields[2])[1]):
                                fields[2] = updated_name_array[4][0] + ' ' + updated_name_array[4][2]
                                fields[12] = updated_name_array[4][1]
                                flag_l1_port_name = True

                        if fields[1] == updated_name_array[0]:
                            if updated_name_array[4][1] == fields[16] and updated_name_array[4][2] == str(ns_def.split_portname(fields[3])[1]):
                                return_text = '[ERROR] This port name is already exist.  --- ' + updated_name_array[0] + ' ' + updated_name_array[2]
                                return ([return_text])

                            if updated_name_array[3][1] == fields[16] and updated_name_array[3][2] == str(ns_def.split_portname(fields[3])[1]):
                                fields[3] = updated_name_array[4][0] + ' ' + updated_name_array[4][2]
                                fields[16] = updated_name_array[4][1]
                                flag_l1_port_name = True

            if flag_l1_port_name == True:
                excel_file_path = ppt_meta_file
                worksheet_name = ws_name
                offset_row = 0
                offset_column = 0

                self.position_line_tuple = ns_def.convert_array_to_tuple(position_line_array)
                master_excel_meta = self.position_line_tuple
                section_write_to = '<<POSITION_LINE>>'
                ns_def.overwrite_excel_meta(master_excel_meta, excel_file_path, worksheet_name, section_write_to, offset_row, offset_column)

                '''rename portname in Master_Data_L2 and Master_Data_L3.'''
                self.full_filepath = master_file_path
                self.update_port_num_array = [[updated_name_array[0],updated_name_array[1],updated_name_array[2]]]
                import ns_sync_between_layers
                ns_sync_between_layers.l1_device_port_name_sync_with_l2l3_master(self)

                return_text = '--- Physical Port Name renamed --- ' + updated_name_array[0] + ' ' + updated_name_array[1] + ' -> ' + updated_name_array[2]
                return ([return_text])

            '''
            rename L2 port name in Master_Data
            '''
            ws_name = 'Master_Data_L2'
            l2_table_array = ns_def.convert_master_to_array(ws_name, ppt_meta_file, '<<L2_TABLE>>')
            #print(l2_table_array)

            # l2_table_array: The first and second lines are not included. Only the first (0) and second (1) are replaced.
            for row in l2_table_array:
                if isinstance(row, list) and len(row) == 2:
                    line_no, fields = row
                    if line_no not in (1, 2) and isinstance(fields, list):
                        #print(fields)

                        if fields[1] == updated_name_array[0]:
                            #print(updated_name_array)
                            if updated_name_array[2] == fields[5]:
                                return_text = '[ERROR] This port name is already exist.  --- ' + updated_name_array[0]  + ' ' + updated_name_array[2]
                                return ([return_text])

                            #print(updated_name_array)
                            if updated_name_array[1] == fields[5]:
                                fields[5] = updated_name_array[2]
                                flag_l2_port_name = True

            if flag_l2_port_name == True:
                excel_file_path = ppt_meta_file
                worksheet_name = ws_name
                offset_row = 0
                offset_column = 0

                self.l2_table_tuple = ns_def.convert_array_to_tuple(l2_table_array)
                master_excel_meta = self.l2_table_tuple
                section_write_to = '<<L2_TABLE>>'
                ns_def.overwrite_excel_meta(master_excel_meta, excel_file_path, worksheet_name, section_write_to, offset_row, offset_column)

                '''rename Virtual port name in Master_Data_L3.'''
                ws_name = 'Master_Data_L3'
                l3_table_array = ns_def.convert_master_to_array(ws_name, ppt_meta_file, '<<L3_TABLE>>')
                #print(l3_table_array)

                # l2_table_array: The first and second lines are not included. Only the first (0) and second (1) are replaced.
                for row in l3_table_array:
                    if isinstance(row, list) and len(row) == 2:
                        line_no, fields = row
                        if line_no not in (1, 2) and isinstance(fields, list):
                            #print(fields)

                            if fields[1] == updated_name_array[0]:
                                #print(updated_name_array)
                                if updated_name_array[1] == fields[2]:
                                    fields[2] = updated_name_array[2]
                                    break

                self.l3_table_tuple = ns_def.convert_array_to_tuple(l3_table_array)
                master_excel_meta = self.l3_table_tuple
                section_write_to = '<<L3_TABLE>>'
                worksheet_name = 'Master_Data_L3'
                ns_def.overwrite_excel_meta(master_excel_meta, excel_file_path, worksheet_name, section_write_to, offset_row, offset_column)

                return_text = '--- Virtual Port Name renamed --- ' + updated_name_array[0] + ' ' + updated_name_array[1] + ' -> ' + updated_name_array[2]
                return ([return_text])

            return_text = '[ERROR]The port name did not exist. --- ' + updated_name_array[0] + ' ' + updated_name_array[1]
            return ([return_text])

        if next_arg == 'device':
            idx = argv_array.index('device')
            updated_name_array = [[str(argv_array[idx + 1]), str(argv_array[idx + 2]).replace(' ', '').replace('　', '')]]

            if isinstance(updated_name_array[0][1], str) and updated_name_array[0][1].strip(' \u3000') == '':
                return_text = '[ERROR] This device name is invalid.  --- ' + ' ' + updated_name_array[0][1]
                return ([return_text])

            self.updated_name_array = updated_name_array

            '''rename device name in Master_Data.'''
            ws_name = 'Master_Data'
            ppt_meta_file = master_file_path

            position_shape_array = ns_def.convert_master_to_array(ws_name, ppt_meta_file, '<<POSITION_SHAPE>>')
            position_line_array = ns_def.convert_master_to_array(ws_name, ppt_meta_file, '<<POSITION_LINE>>')
            position_style_shape_array = ns_def.convert_master_to_array(ws_name, ppt_meta_file, '<<STYLE_SHAPE>>')
            position_tag_array = ns_def.convert_master_to_array(ws_name, ppt_meta_file, '<<POSITION_TAG>>')
            attribute_array = ns_def.convert_master_to_array(ws_name, ppt_meta_file, '<<ATTRIBUTE>>')

            mapping = {}
            for pair in updated_name_array:
                if isinstance(pair, (list, tuple)) and len(pair) == 2:
                    old, new = pair
                    if isinstance(old, str) and isinstance(new, str) and old != '':
                        mapping[old] = new

            # position_shape_array: The first line is ignored, and all other string elements in the line are replaced.
            for row in position_shape_array:
                if isinstance(row, list) and len(row) == 2:
                    line_no, fields = row
                    if line_no != 1 and isinstance(fields, list):
                        for i in range(len(fields)):
                            v = fields[i]
                            if isinstance(v, str) and v in mapping:
                                fields[i] = mapping[v]

            # position_line_array: The first and second lines are not included. Only the first (0) and second (1) are replaced.
            for row in position_line_array:
                if isinstance(row, list) and len(row) == 2:
                    line_no, fields = row
                    if line_no not in (1, 2) and isinstance(fields, list):
                        if len(fields) > 0 and isinstance(fields[0], str) and fields[0] in mapping:
                            fields[0] = mapping[fields[0]]
                        if len(fields) > 1 and isinstance(fields[1], str) and fields[1] in mapping:
                            fields[1] = mapping[fields[1]]

            # position_style_shape_array: Lines 1, 2, and 3 are not included. For the others, only the first (0) is replaced.
            flag_source_exist = False
            for row in position_style_shape_array:
                if isinstance(row, list) and len(row) == 2:
                    line_no, fields = row
                    if line_no not in (1, 2, 3) and isinstance(fields, list) and len(fields) > 0:
                        if isinstance(fields[0], str) and fields[0] == updated_name_array[0][1]:
                            return_text = '[ERROR] The device name already exists.  --- ' + ' ' + updated_name_array[0][1]
                            return ([return_text])

                        if isinstance(fields[0], str) and fields[0] in mapping:
                            fields[0] = mapping[fields[0]]
                            flag_source_exist = True
            if flag_source_exist == False:
                return_text = '[ERROR] This device name does not exist.  --- ' + ' ' + updated_name_array[0][0]
                return ([return_text])

            # position_tag_array: The first and second lines are not included. For the others, only the first (0) is replaced.
            for row in position_tag_array:
                if isinstance(row, list) and len(row) == 2:
                    line_no, fields = row
                    if line_no not in (1, 2) and isinstance(fields, list) and len(fields) > 0:
                        if isinstance(fields[0], str) and fields[0] in mapping:
                            fields[0] = mapping[fields[0]]

            # attribute_array: The first line is not included. For the others, only the first (0) is replaced.
            for row in attribute_array:
                if isinstance(row, list) and len(row) == 2:
                    line_no, fields = row
                    if line_no != 1 and isinstance(fields, list) and len(fields) > 0:
                        if isinstance(fields[0], str) and fields[0] in mapping:
                            fields[0] = mapping[fields[0]]

            self.position_shape_tuple = ns_def.convert_array_to_tuple(position_shape_array)
            self.position_line_tuple = ns_def.convert_array_to_tuple(position_line_array)
            self.position_style_shape_tuple = ns_def.convert_array_to_tuple(position_style_shape_array)
            self.position_tag_tuple = ns_def.convert_array_to_tuple(position_tag_array)
            self.attribute_tuple = ns_def.convert_array_to_tuple(attribute_array)

            excel_file_path = ppt_meta_file
            worksheet_name = ws_name
            offset_row = 0
            offset_column = 0

            master_excel_meta = self.position_shape_tuple
            section_write_to = '<<POSITION_SHAPE>>'
            ns_def.overwrite_excel_meta(master_excel_meta, excel_file_path, worksheet_name, section_write_to, offset_row, offset_column)
            master_excel_meta = self.position_line_tuple
            section_write_to = '<<POSITION_LINE>>'
            ns_def.overwrite_excel_meta(master_excel_meta, excel_file_path, worksheet_name, section_write_to, offset_row, offset_column)
            master_excel_meta = self.position_style_shape_tuple
            section_write_to = '<<STYLE_SHAPE>>'
            ns_def.overwrite_excel_meta(master_excel_meta, excel_file_path, worksheet_name, section_write_to, offset_row, offset_column)
            master_excel_meta = self.position_tag_tuple
            section_write_to = '<<POSITION_TAG>>'
            ns_def.overwrite_excel_meta(master_excel_meta, excel_file_path, worksheet_name, section_write_to, offset_row, offset_column)
            master_excel_meta = self.attribute_tuple
            section_write_to = '<<ATTRIBUTE>>'
            ns_def.overwrite_excel_meta(master_excel_meta, excel_file_path, worksheet_name, section_write_to, offset_row, offset_column)

            '''rename device name in Master_Data_L2 and Master_Data_L3.'''
            self.full_filepath = master_file_path
            import ns_sync_between_layers
            ns_sync_between_layers.l1_sketch_device_name_sync_with_l2l3_master(self)

            return_text = '--- Device Name renamed --- ' + ' ' + updated_name_array[0][0] + ' -> ' + updated_name_array[0][1]
            return ([return_text])

        if next_arg == 'l3_instance':
            import ns_def
            l3_attribute_array = ns_def.convert_master_to_array('Master_Data_L3', master_file_path,'<<L3_TABLE>>')

            TARGET_LEN = 7
            for row in l3_attribute_array:
                if isinstance(row, list) and len(row) >= 2 and isinstance(row[1], list):
                    vals = row[1]
                    if len(vals) < TARGET_LEN:
                        vals += [''] * (TARGET_LEN - len(vals))

            idx = argv_array.index('l3_instance')

            target = idx + 3
            if target >= len(argv_array):
                argv_array.extend([None] * (target - len(argv_array) + 1))
                argv_array[target] = '__default__'
            elif '--' in argv_array[target]:
                argv_array.insert(target, '__default__')

            try:
                # The element right after 'l3_instance' is the hostname
                hostname = argv_array[idx + 1]
                # The next element is the portname
                portname = argv_array[idx + 2]
                # The next element is the l3 instance name
                add_l3instance_name = argv_array[idx + 3]
            except IndexError:
                # If there are not enough elements after 'l3_instance', print an error
                return ([f"[ERROR] hostname or portname or l3_instance is missing"])

            match_found = False
            for row in l3_attribute_array[2:]:
                values = row[1]
                if len(values) > 1 and values[1] == hostname:
                    if (len(values) > 3 and values[2] == portname):
                        match_found = True
                        break
            if not match_found:
                return ([f"Error: No matching entry found for hostname: {hostname} and portname: {portname}"])

            add_l3instance_name = add_l3instance_name.replace(' ', '')

            for entry in l3_attribute_array:
                if entry == row:
                    # Get the 4th element (index 3)
                    entry[1][3] = add_l3instance_name

                    if add_l3instance_name == '__default__':
                        entry[1][3] = ''
                        add_l3instance_name = ''

                    break


            #write to Master file
            excel_maseter_file = master_file_path
            excel_master_ws_name_l3 = 'Master_Data_L3'
            last_l3_table_tuple = {}
            last_l3_table_tuple = ns_def.convert_array_to_tuple(l3_attribute_array)

            # delete L2 Table sheet
            ns_def.remove_excel_sheet(excel_maseter_file, excel_master_ws_name_l3)
            # create L2 Table sheet
            ns_def.create_excel_sheet(excel_maseter_file, excel_master_ws_name_l3)
            # write tuple to excel master data
            ns_def.write_excel_meta(last_l3_table_tuple, excel_maseter_file, excel_master_ws_name_l3, '_template_', 0,0)

            return_text = '--- l3 instance renamed --- ' + ' ' + hostname + ',' + portname + ',' + add_l3instance_name
            return ([return_text])

    def cli_add(self, master_file_path, argv_array): # add at ver 2.5.3
        import ns_def
        import tkinter as tk
        next_arg = get_next_arg(argv_array, 'add')
        add_command_list = [ \
            'add area_location', # Add this line at ver 2.6.1
            'add device', # Add this line at ver 2.6.1
            'add device_location',  # Add this line at ver 2.6.1
            'add ip_address', \
            'add l1_link', # Add this line at ver 2.6.1
            'add l1_link_bulk',  # Add this line at ver 2.6.1
            'add l2_segment', \
            'add portchannel', \
            'add virtual_port', \
            'add vport_l1if_direct_binding', \
            'add vport_l2_direct_binding', \
            'add waypoint', # Add this line at ver 2.6.1
            ]

        if next_arg == None or '--' in next_arg or str('add ' + next_arg) not in add_command_list:
            print('[ERROR] Supported commands are as follows')
            for tmp_add_command_list in add_command_list:
                print(tmp_add_command_list)
            exit()

        if next_arg == 'l1_link_bulk':
            import ast
            import ns_def

            # Get the array argument from command line
            if 'l1_link_bulk' in argv_array:
                idx = argv_array.index('l1_link_bulk')
                try:
                    # Parse the array argument
                    l1_link_bulk_array_str = argv_array[idx + 1]

                    # ★★★ Validate that all elements are properly quoted ★★★
                    def validate_quoted_format(input_str):
                        """
                        Validate that hostnames and port names are properly quoted
                        Returns: (is_valid, error_message)
                        """
                        try:
                            # Check if input starts and ends correctly
                            input_str = input_str.strip()
                            if not input_str.startswith('[[') or not input_str.endswith(']]'):
                                return (False, "Input must start with [[ and end with ]]")

                            # Remove outer [[ and ]]
                            content = input_str[2:-2].strip()

                            # Check for unquoted elements (words not surrounded by quotes)
                            # Pattern: Look for sequences like ", word," or ", word]" or "[word,"
                            # where word is not quoted
                            import re

                            # Find all comma-separated values
                            # Split by '], [' first to get individual link definitions
                            link_defs = content.split('], [')

                            for link_idx, link_def in enumerate(link_defs):
                                # Clean up brackets
                                link_def = link_def.strip('[]').strip()

                                # Split by comma, but we need to check if elements are quoted
                                # Use regex to find all elements (quoted or unquoted)
                                element_pattern = r'''
                                    (?:
                                        '(?:[^'\\]|\\.)*'|    # Single-quoted string
                                        "(?:[^"\\]|\\.)*"|    # Double-quoted string
                                        [^,\[\]]+             # Unquoted element
                                    )
                                '''
                                elements = re.findall(element_pattern, link_def, re.VERBOSE)
                                elements = [e.strip() for e in elements if e.strip()]

                                # Check each element
                                for elem_idx, element in enumerate(elements):
                                    element = element.strip()

                                    # Check if element is quoted
                                    is_quoted = (element.startswith("'") and element.endswith("'")) or \
                                                (element.startswith('"') and element.endswith('"'))

                                    if not is_quoted:
                                        # Determine what this element should be
                                        if elem_idx in [0, 1]:
                                            elem_type = "hostname"
                                        else:
                                            elem_type = "port name"

                                        error_msg = f"[ERROR] All elements must be quoted with single or double quotes.\n"
                                        error_msg += f"  Link {link_idx + 1}, Element {elem_idx + 1}: '{element}' ({elem_type}) is not quoted.\n"
                                        error_msg += f"  Correct format: [['host1', 'host2', 'port1', 'port2'], ...]\n"
                                        error_msg += f"  Example: [[\"SW-1\", \"SW-2\", \"'GigabitEthernet 0/0'\", \"'GigabitEthernet 0/1'\"]]"
                                        return (False, error_msg)

                            return (True, "Format is valid")

                        except Exception as e:
                            return (False, f"Validation error: {str(e)}")

                    # Validate format first
                    is_valid, validation_msg = validate_quoted_format(l1_link_bulk_array_str)
                    if not is_valid:
                        return ([validation_msg])
                    # ★★★ End of validation ★★★

                    # Parse the array using ast.literal_eval
                    l1_link_bulk_array = ast.literal_eval(l1_link_bulk_array_str)

                except (IndexError, ValueError, SyntaxError) as e:
                    return ([f'[ERROR] Invalid l1_link_bulk array format: {str(e)}'])

            # Validate input format
            if not isinstance(l1_link_bulk_array, list):
                return ([f'[ERROR] l1_link_bulk must be a list of link definitions'])

            # Call bulk add function
            return def_common.add_l1_links_bulk(self, l1_link_bulk_array, master_file_path)



        if next_arg == 'device_location':
            import ast

            # Get the array argument from command line
            if 'device_location' in argv_array:
                idx = argv_array.index('device_location')
                try:
                    # Parse the array argument
                    device_location_array_str = argv_array[idx + 1]
                    device_location_array = ast.literal_eval(device_location_array_str)
                except (IndexError, ValueError, SyntaxError) as e:
                    return ([f'[ERROR] Invalid device_location array format: {str(e)}'])

            # Validate input format
            if not isinstance(device_location_array, list) or len(device_location_array) != 2:
                return ([f'[ERROR] device_location must be [area_name, [[device_grid]]]'])

            area_name = device_location_array[0]
            device_grid = device_location_array[1]

            if not isinstance(area_name, str) or not isinstance(device_grid, list):
                return ([f'[ERROR] Invalid device_location format. Expected: [area_name, [[device_grid]]]'])

            # ★★★ OPTIMIZED: Call batch-optimized function ★★★
            return def_common.update_device_location_batch(self, area_name, device_grid, master_file_path)

        if next_arg == 'area_location':
            import ast

            # Get the array argument from command line
            if 'area_location' in argv_array:
                idx = argv_array.index('area_location')
                try:
                    # Parse the array argument
                    area_location_array_str = argv_array[idx + 1]
                    area_location_array = ast.literal_eval(area_location_array_str)
                except (IndexError, ValueError, SyntaxError) as e:
                    return ([f'[ERROR] Invalid area_location array format: {str(e)}'])

            # Call common function
            return def_common.update_area_with_map(self, area_location_array, master_file_path)

        # Add device command (modified to exclude waypoints)
        if next_arg == 'device':
            import ns_def

            # Check if the input data exists in argv_array
            if 'device' in argv_array:
                idx = argv_array.index('device')
                try:
                    device_name = argv_array[idx + 1]
                    reference_device = argv_array[idx + 2]
                    direction = argv_array[idx + 3].upper()
                except IndexError:
                    return (['[Error] device name, reference device, or direction is missing'])

            # Check if device is a waypoint (should not be added as device)
            position_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_SHAPE>>')
            for item in position_shape_array:
                if isinstance(item, list) and len(item) >= 2 and isinstance(item[1], list):
                    if len(item[1]) > 0 and '_wp_' in str(item[1][0]):
                        # This is a waypoint group
                        if device_name in item[1]:
                            return ([f"[Error] '{device_name}' is a waypoint. Please use 'add waypoint' command instead."])

            # Call common add function with device type
            return def_common.add_device_or_waypoint_common(
                self,
                device_name,
                reference_device,
                direction,
                master_file_path,
                device_type='device'
            )
        # Add waypoint command (modified)
        if next_arg == 'waypoint':
            import ns_def

            # Check if the input data exists in argv_array
            if 'waypoint' in argv_array:
                idx = argv_array.index('waypoint')
                try:
                    waypoint_name = argv_array[idx + 1]
                    reference_device = argv_array[idx + 2]
                    direction = argv_array[idx + 3].upper()
                except IndexError:
                    return (['[Error] waypoint name, reference device, or direction is missing'])

            # Check if waypoint is actually a device (should not be added as waypoint)
            position_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_SHAPE>>')
            for item in position_shape_array:
                if isinstance(item, list) and len(item) >= 2 and isinstance(item[1], list):
                    if len(item[1]) > 0 and '_wp_' not in str(item[1][0]):
                        # This is a device group
                        if waypoint_name in item[1]:
                            return ([f"[Error] '{waypoint_name}' is a device. Please use 'add device' command instead."])

            # Validate direction for waypoint (exclude _WITH_GRID options)
            valid_directions_waypoint = ['UP', 'DOWN', 'LEFT', 'RIGHT']

            if direction not in valid_directions_waypoint:
                return ([f"[Error] Invalid direction '{direction}' for waypoint. Valid directions: {', '.join(valid_directions_waypoint)}"])

            # Check folder configuration to determine allowed directions
            position_folder_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_FOLDER>>')

            # Find the folder containing the reference device (use position_shape_array)
            # Need to track the current folder name as we iterate
            reference_folder = None
            current_folder = None

            for item in position_shape_array:
                if isinstance(item, list) and len(item) >= 2 and isinstance(item[1], list):
                    if len(item[1]) > 0:
                        # Check if this row starts a new folder
                        if item[1][0] not in ['', '<END>', '<<POSITION_SHAPE>>'] and item[1][0] != '<END>':
                            current_folder = item[1][0]

                        # If first element is empty but we have a current folder, this row belongs to that folder
                        if item[1][0] == '' and current_folder:
                            # This row continues the current folder
                            pass
                        elif item[1][0] == '<END>':
                            # End of current folder
                            current_folder = None
                            continue

                        # Check if reference device is in this row
                        if current_folder and reference_device in item[1]:
                            reference_folder = current_folder
                            break

            if reference_folder is None:
                return ([f"[Error] Reference device '{reference_device}' not found"])

            # Check if reference folder is a waypoint folder
            if '_wp_' not in reference_folder:
                return ([f"[Error] Reference device '{reference_device}' must be in a waypoint folder (_wp_) to add a waypoint"])

            # Find the position of reference_folder in POSITION_FOLDER
            folder_row = None
            folder_col = None
            for item in position_folder_array:
                if len(item) >= 2 and isinstance(item[1], list):
                    for col_idx, folder_name in enumerate(item[1]):
                        # Skip numeric values and special markers
                        if not isinstance(folder_name, str):
                            continue
                        if folder_name in ['', '<SET_WIDTH>', '<<POSITION_FOLDER>>']:
                            continue

                        # Direct match
                        if folder_name == reference_folder:
                            folder_row = item[0]
                            folder_col = col_idx
                            break
                    if folder_row is not None:
                        break

            if folder_row is None or folder_col is None:
                return ([f"[Error] Could not find folder position for '{reference_folder}'"])

            # Check adjacent folders (left and right)
            left_folder_has_normal = False  # Has normal (non-_wp_) folder
            right_folder_has_normal = False  # Has normal (non-_wp_) folder

            # Check left folder
            if folder_col > 0:
                for item in position_folder_array:
                    if item[0] == folder_row and len(item[1]) > folder_col - 1:
                        left_folder = item[1][folder_col - 1]
                        # Check if it's a non-empty string and NOT a waypoint folder
                        if isinstance(left_folder, str) and left_folder != '' and '_wp_' not in left_folder:
                            left_folder_has_normal = True
                        break

            # Check right folder
            if folder_col < 100:  # Reasonable max column check
                for item in position_folder_array:
                    if item[0] == folder_row and len(item[1]) > folder_col + 1:
                        right_folder = item[1][folder_col + 1]
                        # Check if it's a non-empty string and NOT a waypoint folder
                        if isinstance(right_folder, str) and right_folder != '' and '_wp_' not in right_folder:
                            right_folder_has_normal = True
                        break

            # Determine allowed directions based on adjacent folders
            allowed_directions = []

            if left_folder_has_normal or right_folder_has_normal:
                # If at least one adjacent folder is a normal (non-_wp_) folder, only UP/DOWN are allowed
                allowed_directions = ['UP', 'DOWN']
                if direction not in allowed_directions:
                    return ([f"[Error] When adjacent folder is a normal folder (without _wp_), only UP/DOWN directions are allowed. Current direction: '{direction}'"])
            else:
                # If both adjacent folders are waypoint folders or empty, only LEFT/RIGHT are allowed
                allowed_directions = ['LEFT', 'RIGHT']
                if direction not in allowed_directions:
                    return ([f"[Error] When adjacent folders are waypoint folders (_wp_) or empty, only LEFT/RIGHT directions are allowed. Current direction: '{direction}'"])

            # Call common add function with waypoint type
            return def_common.add_device_or_waypoint_common(
                self,
                waypoint_name,
                reference_device,
                direction,
                master_file_path,
                device_type='waypoint'
            )

        if next_arg == 'l1_link':  # add at Ver 2.6.1
            import ns_def
            position_folder_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_FOLDER>>')
            position_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_SHAPE>>')
            style_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<STYLE_SHAPE>>')
            position_line_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_LINE>>')
            position_tag_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_TAG>>')
            ori_position_line_tuple = ns_def.convert_array_to_tuple(position_line_array)

            # Check if the input data exists in argv_array
            if 'l1_link' in argv_array:
                idx = argv_array.index('l1_link')
                try:
                    from_hostname = argv_array[idx + 1]
                    to_hostname = argv_array[idx + 2]
                    from_portname = argv_array[idx + 3]
                    to_portname = argv_array[idx + 4]
                except IndexError:
                    return (['[Error] hostname or portname is missing'])

                # Swap if from_hostname and to_hostname are in descending order
                if from_hostname > to_hostname:
                    from_hostname, to_hostname = to_hostname, from_hostname
                    from_portname, to_portname = to_portname, from_portname

            def check_duplicate_port(position_line_array, from_hostname, to_hostname, from_port, to_port):
                """
                Check if the combination of hostname and port already exists in position_line_array
                Compares ports in their stored format without any normalization
                """

                for item in position_line_array:
                    if len(item) < 2 or not isinstance(item[1], list):
                        continue

                    row_data = item[1]

                    # Skip header rows
                    if (len(row_data) > 0 and (row_data[0] == '<<POSITION_LINE>>' or row_data[0] == 'From_Name')):
                        continue

                    if len(row_data) < 4:
                        continue

                    existing_from_hostname = row_data[0]
                    existing_to_hostname = row_data[1]
                    existing_from_port = row_data[2]  # Use as-is from storage (e.g., "GE 0/0")
                    existing_to_port = row_data[3]  # Use as-is from storage (e.g., "GE 0/0")

                    # Direct comparison without normalization
                    if existing_from_hostname == from_hostname and existing_from_port == from_port:
                        return f"Error: Port '{from_port}' on device '{from_hostname}' is already used in an existing link"
                    if existing_to_hostname == to_hostname and existing_to_port == to_port:
                        return f"Error: Port '{to_port}' on device '{to_hostname}' is already used in an existing link"
                    if existing_from_hostname == to_hostname and existing_from_port == to_port:
                        return f"Error: Port '{to_port}' on device '{to_hostname}' is already used in an existing link"
                    if existing_to_hostname == from_hostname and existing_to_port == from_port:
                        return f"Error: Port '{from_port}' on device '{from_hostname}' is already used in an existing link"

                return None



            # Helper function for adding element
            def add_element_to_position_line_array(self, position_line_array, from_hostname, to_hostname, from_port,
                                                   to_port, result_check_hostnames_in_same_element, from_full_portname,
                                                   to_full_portname):
                """
                Add a new element to position_line_array
                """
                if position_line_array:
                    number = max(item[0] for item in position_line_array) + 1
                else:
                    number = 1

                if result_check_hostnames_in_same_element == 'RIGHT_LEFT':
                    new_element = [number, [from_hostname, to_hostname, from_port, to_port, 'RIGHT', 'LEFT', '', '', '', '', '',
                                            '', from_full_portname, 'N/A', 'N/A', 'N/A', to_full_portname, 'N/A', 'N/A', 'N/A']]
                elif result_check_hostnames_in_same_element == 'LEFT_RIGHT':
                    new_element = [number, [from_hostname, to_hostname, from_port, to_port, 'LEFT', 'RIGHT', '', '', '', '', '',
                                            '', from_full_portname, 'N/A', 'N/A', 'N/A', to_full_portname, 'N/A', 'N/A', 'N/A']]
                elif result_check_hostnames_in_same_element in ['UP_DOWN', 'DOWN_UP']:
                    new_element = [number, [from_hostname, to_hostname, from_port, to_port, '', '', '', '', '', '', '', '',
                                            from_full_portname, 'N/A', 'N/A', 'N/A', to_full_portname, 'N/A', 'N/A', 'N/A']]
                else:
                    return None

                position_line_array.append(new_element)

                return position_line_array

            # Validate port names
            if ns_def.split_portname(from_portname)[0] == '' or ns_def.split_portname(from_portname)[1] == '':
                error_msg = f"Error: invalid from_portname '{from_portname}'. \n [Tips] When specifying the portname, be sure to use single quotes, such as 'port 0'."
                return ([error_msg])
            elif ns_def.split_portname(to_portname)[0] == '' or ns_def.split_portname(to_portname)[1] == '':
                error_msg = f"Error: invalid to_portname '{to_portname}'. \n [Tips] When specifying the portname, be sure to use single quotes, such as 'port 0'."
                return ([error_msg])

            # Prepare port information
            from_port = str(ns_def.adjust_portname(from_portname)[0]) + ' ' + str(ns_def.split_portname(from_portname)[1])
            to_port = str(ns_def.adjust_portname(to_portname)[0]) + ' ' + str(ns_def.split_portname(to_portname)[1])
            from_full_portname = ns_def.split_portname(from_portname)[0]
            to_full_portname = ns_def.split_portname(to_portname)[0]

            # Check hostnames are in same element
            result_check_hostnames_in_same_element = def_common.check_hostnames_in_same_element_static(
                position_shape_array, from_hostname, to_hostname, position_folder_array
            )


            if isinstance(result_check_hostnames_in_same_element, str) and \
                    result_check_hostnames_in_same_element.startswith("Error:"):
                return ([result_check_hostnames_in_same_element])

            # Check for duplicate ports
            duplicate_check_result = check_duplicate_port(position_line_array, from_hostname, to_hostname, from_port, to_port)
            if duplicate_check_result is not None:
                return ([duplicate_check_result])

            # Add the connection (once only)
            result = add_element_to_position_line_array(self, position_line_array, from_hostname, to_hostname,
                                                        from_port, to_port, result_check_hostnames_in_same_element,
                                                        from_full_portname, to_full_portname)

            # Check if the addition was successful
            if result is None:
                return (['[Error] Failed to add connection. Please check hostname and direction.'])

            position_line_array = result

            affected_hostnames = {from_hostname, to_hostname}

            # Call common processing for ALL devices
            position_line_array, style_shape_array = def_common.process_l1_link_common(
                self, position_shape_array, style_shape_array, position_line_array,
                position_tag_array, master_file_path, ori_position_line_tuple,
                allow_shrink=False, position_folder_array=position_folder_array
            )

            # update_positon_folder_tuple
            position_folder_tuple = ns_def.convert_array_to_tuple(position_folder_array)
            position_shape_tuple = ns_def.convert_array_to_tuple(position_shape_array)
            style_shape_tuple = ns_def.convert_array_to_tuple(style_shape_array)
            update_position_folder_tuple = def_common.make_position_folder_tuple(position_folder_tuple, style_shape_tuple, position_shape_tuple)
            ns_def.overwrite_excel_meta(update_position_folder_tuple, master_file_path, worksheet_name='Master_Data', section_write_to='<<POSITION_FOLDER>>', offset_row=0, offset_column=0)

            # ★★★ ADD: Call recalculate_folder_sizes to remove unnecessary empty cells ★★★
            result = def_common.recalculate_folder_sizes(master_file_path)
            if result['status'] != 'success':
                return ([f"[ERROR] Failed to recalculate folder sizes: {result['message']}"])
            # ★★★ End of addition ★★★

            return_text = '--- Added Layer 1 link --- ' + from_hostname + ' ' + to_hostname + ' ' + from_portname + ' ' + to_portname
            return ([return_text])

        if next_arg == 'ip_address':
            import ns_def
            l3_attribute_array = ns_def.convert_master_to_array('Master_Data_L3', master_file_path,'<<L3_TABLE>>')

            TARGET_LEN = 7
            for row in l3_attribute_array:
                if isinstance(row, list) and len(row) >= 2 and isinstance(row[1], list):
                    vals = row[1]
                    if len(vals) < TARGET_LEN:
                        vals += [''] * (TARGET_LEN - len(vals))

            idx = argv_array.index('ip_address')
            try:
                # The element right after 'l2_segment' is the hostname
                hostname = argv_array[idx + 1]
                # The next element is the portname
                portname = argv_array[idx + 2]
                add_ipaddress_name = argv_array[idx + 3]
            except IndexError:
                return ([f"[ERROR] hostname or portname or ip address is missing"])

            # check IP Addresses
            if ns_def.check_ip_format(add_ipaddress_name) != 'IPv4':
                return ([f"Error: IP Address format is invalid: {add_ipaddress_name}"])

            match_found = False
            for row in l3_attribute_array[2:]:
                values = row[1]
                if len(values) > 1 and values[1] == hostname:
                    if (len(values) > 3 and values[2] == portname):
                        match_found = True
                        break
            if not match_found:
                return ([f"Error: No matching entry found for hostname: {hostname} and portname: {portname}"])

            add_ipaddress_name = add_ipaddress_name.replace(' ', '')

            for entry in l3_attribute_array:
                if entry == row:
                    # Get the 5th element (index 4)
                    current_value = entry[1][4]
                    # Split by comma, remove extra spaces
                    ips = [seg.strip() for seg in current_value.split(',')] if current_value else []
                    # Only add if not already present
                    if add_ipaddress_name not in ips:
                        if current_value.strip() == '':
                            entry[1][4] = add_ipaddress_name
                        else:
                            entry[1][4] = current_value + ',' + add_ipaddress_name
                    else:
                        return ([add_ipaddress_name + ' already exists in the ipaddress. No change made'])
                    break


            #write to Master file
            excel_maseter_file = master_file_path
            excel_master_ws_name_l3 = 'Master_Data_L3'
            last_l3_table_tuple = {}
            last_l3_table_tuple = ns_def.convert_array_to_tuple(l3_attribute_array)

            # delete L2 Table sheet
            ns_def.remove_excel_sheet(excel_maseter_file, excel_master_ws_name_l3)
            # create L2 Table sheet
            ns_def.create_excel_sheet(excel_maseter_file, excel_master_ws_name_l3)
            # write tuple to excel master data
            ns_def.write_excel_meta(last_l3_table_tuple, excel_maseter_file, excel_master_ws_name_l3, '_template_', 0,0)

            return_text = '--- IP Address added --- ' + ' ' + hostname + ',' + portname + ',' + add_ipaddress_name
            return ([return_text])

        if next_arg == 'l2_segment' or 'virtual_port' or 'portchannel' or 'vport_l2_direct_binding' or 'vport_l1if_direct_binding':
            import ns_def
            l2_attribute_array = ns_def.convert_master_to_array('Master_Data_L2', master_file_path,'<<L2_TABLE>>')

            TARGET_LEN = 9
            for row in l2_attribute_array:
                if isinstance(row, list) and len(row) >= 2 and isinstance(row[1], list):
                    vals = row[1]
                    if len(vals) < TARGET_LEN:
                        vals += [''] * (TARGET_LEN - len(vals))

            #print(argv_array)
            # Check if the input data exists in argv_array
            if 'l2_segment' in argv_array:
                idx = argv_array.index('l2_segment')
                try:
                    # The element right after 'l2_segment' is the hostname
                    hostname = argv_array[idx + 1]
                    # The next element is the portname
                    portname = argv_array[idx + 2]
                    add_l2seg_name = argv_array[idx + 3]
                except IndexError:
                    # If there are not enough elements after 'l2_segment', print an error
                    return (['[Error] hostname or portname is missing'])

            elif 'virtual_port' in argv_array:
                idx = argv_array.index('virtual_port')
                try:
                    hostname = argv_array[idx + 1]
                    add_vport_name = argv_array[idx + 2]
                except IndexError:
                    return (['[Error] hostname is missing'])

            elif 'portchannel' in argv_array:
                idx = argv_array.index('portchannel')
                try:
                    # The element right after 'portchannel' is the hostname
                    hostname = argv_array[idx + 1]
                    # The next element is the portname
                    portname = argv_array[idx + 2]
                    add_portchannel_name = argv_array[idx + 3]
                except IndexError:
                    # If there are not enough elements after 'portchannel', print an error
                    return (['[Error] hostname or portname is missing'])

            elif 'vport_l2_direct_binding' in argv_array:
                idx = argv_array.index('vport_l2_direct_binding')
                try:
                    # The element right after 'vport_l2_direct_binding' is the hostname
                    hostname = argv_array[idx + 1]
                    # The next element is the portname
                    portname = argv_array[idx + 2]
                    add_l2seg_name = argv_array[idx + 3]
                except IndexError:
                    # If there are not enough elements after 'vport_l2_direct_binding', print an error
                    return (['[Error] hostname or portname is missing'])

            elif 'vport_l1if_direct_binding' in argv_array:
                idx = argv_array.index('vport_l1if_direct_binding')
                try:
                    # The element right after 'vport_l1if_direct_binding' is the hostname
                    hostname = argv_array[idx + 1]
                    # The next element is the portname
                    portname = argv_array[idx + 2]
                    add_vport_name = argv_array[idx + 3]
                except IndexError:
                    # If there are not enough elements after 'vport_l1if_direct_binding', print an error
                    return (['[Error] hostname or portname is missing'])

            else:
                print("Not found in arguments")

            match_found = False
            row_array = []
            flag_l2_segment_vport = False
            if 'l2_segment' in argv_array:
                for row in l2_attribute_array[2:]:
                    values = row[1]
                    if len(values) > 1 and values[1] == hostname:
                        if len(values) > 3 and values[3] == portname:
                            if values[7] != '':
                                return ([f"[ERROR] Any vport_l2_direct_binding is already configured for hostname: {hostname} and virtual portname: {portname}"])
                            match_found = True
                            break
                        elif len(values) > 5 and values[5] == portname:
                            if values[7] != '':
                                return ([f"[ERROR] Any vport_l2_direct_binding is already configured for hostname: {hostname} and virtual portname: {portname}"])
                            #for virtual port
                            match_found = True
                            flag_l2_segment_vport = True
                            row_array.append(row)
                if flag_l2_segment_vport == True:
                    row = []

            elif 'virtual_port' in argv_array:
                for row in l2_attribute_array[2:]:
                    values = row[1]
                    if len(values) > 1 and values[1] == hostname:
                        match_found = True
                        break

                #check to L1 interface duplicate
                for row3 in l2_attribute_array[2:]:
                    values2 = row3[1]
                    if len(values2) > 3 and values2[1] == hostname and values2[3] == add_vport_name :
                        return (["[ERROR] There is a Layer 1 interface with the same name as the virtual port."])

            elif 'portchannel' in argv_array:
                for row in l2_attribute_array[2:]:
                    values = row[1]
                    if len(values) > 1 and values[1] == hostname:
                        if (len(values) > 3 and values[3] == portname) or (len(values) > 5 and values[5] == portname):
                            match_found = True
                            break

            elif 'vport_l2_direct_binding' in argv_array:
                for row in l2_attribute_array[2:]:
                    values = row[1]
                    if len(values) > 1 and values[1] == hostname:
                        if len(values) > 5 and values[5] == portname:
                            if values[6] != '':
                                return ([f"[ERROR] Any L2 segment is already configured for hostname: {hostname} and virtual portname: {portname}"])
                            #for virtual port
                            match_found = True
                            flag_l2_segment_vport = True
                            row_array.append(row)
                if flag_l2_segment_vport == True:
                    row = []

            elif 'vport_l1if_direct_binding' in argv_array:
                flag_vport_l1if_direct_binding_exists = False

                for row in l2_attribute_array[2:]:
                    values = row[1]
                    if len(values) > 1 and values[1] == hostname:
                        if (len(values) > 3 and values[3] == portname) and (len(values) > 5 and values[5] == add_vport_name):
                            return ([f"[ERROR] {add_vport_name} is already configured for hostname: {hostname} and portname: {portname}"])

                for row in l2_attribute_array[2:]:
                    values = row[1]
                    if len(values) > 1 and values[1] == hostname:
                        if (len(values) > 3 and values[3] == portname):
                            match_found = True
                            if (len(values) > 5 and values[5] != ''):
                                flag_vport_l1if_direct_binding_exists = True
                            break

                #check to L1 interface duplicate
                for row3 in l2_attribute_array[2:]:
                    values2 = row3[1]
                    if len(values2) > 3 and values2[1] == hostname and values2[3] == add_vport_name :
                        return (["[ERROR] There is a Layer 1 interface with the same name as the virtual port."])

            if not match_found:
                if 'l2_segment' in argv_array:
                    return ([f"No matching entry found for hostname: {hostname} and portname: {portname}"])
                elif 'virtual_port' in argv_array:
                    return ([f"No matching entry found for hostname: {hostname}"])
                elif 'portchannel' in argv_array:
                    return ([f"No matching entry found for hostname: {hostname} and portname: {portname}"])
                elif 'vport_l2_direct_binding' in argv_array:
                    return ([f"No matching entry found for hostname: {hostname} and virtual portname: {portname}"])
                elif 'vport_l1if_direct_binding' in argv_array:
                    return ([f"No matching entry found for hostname: {hostname} and portname: {portname}"])
                exit()

            if 'l2_segment' in argv_array:
                add_l2seg_name = add_l2seg_name.replace(' ', '')
                for entry in l2_attribute_array:
                    if entry == row:
                        # Get the 7th element (index 6)
                        current_value = entry[1][6]
                        # Split by comma, remove extra spaces
                        segs = [seg.strip() for seg in current_value.split(',')] if current_value else []
                        # Only add if not already present
                        if add_l2seg_name not in segs:
                            if current_value.strip() == '':
                                entry[1][6] = add_l2seg_name
                                entry[1][6] = ','.join(sorted(s.strip() for s in entry[1][6].split(',')))
                            else:
                                entry[1][6] = current_value + ',' + add_l2seg_name
                                entry[1][6] = ','.join(sorted(s.strip() for s in entry[1][6].split(',')))
                        else:
                            #print(f"'{add_l2seg_name}' already exists in the l2 segments. No change made.")
                            return ([add_l2seg_name + ' already exists in the l2 segments. No change made'])
                        break

                    #for virtual port
                    if entry in row_array:
                        # Get the 7th element (index 6)
                        current_value = entry[1][6]
                        # Split by comma, remove extra spaces
                        segs = [seg.strip() for seg in current_value.split(',')] if current_value else []
                        # Only add if not already present
                        if add_l2seg_name not in segs:
                            if current_value.strip() == '':
                                entry[1][6] = add_l2seg_name
                                entry[1][6] = ','.join(sorted(s.strip() for s in entry[1][6].split(',')))
                            else:
                                entry[1][6] = current_value + ',' + add_l2seg_name
                                entry[1][6] = ','.join(sorted(s.strip() for s in entry[1][6].split(',')))
                        else:
                            #print(f"'{add_l2seg_name}' already exists in the l2 segments. No change made.")
                            return ([add_l2seg_name + ' already exists in the l2 segments. No change made'])

            elif 'virtual_port' in argv_array:
                if ns_def.get_if_value(add_vport_name) == -1:
                    return ([f"Invalid virtual port name: {add_vport_name}"])
                # Check if the virtual port name already exists for this hostname
                for entry in l2_attribute_array:
                    if entry[0] > 2 and entry[1][1] == hostname and entry[1][5] == add_vport_name:
                        return ([f"{add_vport_name} already exists as virtual port for hostname: {hostname}"])

                # get insert number
                tmp_if_value_array = []
                for row in l2_attribute_array[2:]:
                    values = row[1]
                    if len(values) > 1 and values[1] == hostname:
                        if (len(values) > 5 and values[5] != '') and (len(values) > 3 and values[3] == '') :
                            tmp_if_value_array.append(ns_def.get_if_value(values[5]))
                target_if_value = ns_def.get_if_value(add_vport_name)
                from bisect import bisect_left
                insert_idx = bisect_left(tmp_if_value_array, target_if_value)

                # Find the index of the first matching hostname row
                insert_index = -1
                for i, entry in enumerate(l2_attribute_array):
                    if entry[0] > 2 and entry[1][1] == hostname:
                        insert_index = i
                        break

                if insert_index != -1:
                    # Create new entry with the same area and hostname
                    new_entry = [0, [l2_attribute_array[insert_index][1][0], hostname, '', '', '', add_vport_name, '', '']]
                    # Insert the new entry at the found index (before the existing entry)
                    l2_attribute_array.insert(insert_index + insert_idx, new_entry)

                    # Renumber all entries
                    for i, entry in enumerate(l2_attribute_array):
                        entry[0] = i + 1

                return_text = '--- Virtual Port added --- ' + ' ' + hostname + ',' + add_vport_name

            elif 'portchannel' in argv_array:
                if ns_def.get_if_value(add_portchannel_name) == -1:
                    return ([f"Invalid portchannel name: {add_portchannel_name}"])

                for entry in l2_attribute_array:
                    if entry == row:
                        entry[1][5] = add_portchannel_name
                        return_text = '--- portchannel added --- ' + ' ' + hostname + ',' + portname + ',' + add_portchannel_name
                        break

            elif 'vport_l2_direct_binding' in argv_array:
                add_l2seg_name = add_l2seg_name.replace(' ', '')
                for entry in l2_attribute_array:
                    #for virtual port
                    if entry in row_array:
                        # Get the 8th element (index 7)
                        current_value = entry[1][7]
                        # Split by comma, remove extra spaces
                        segs = [seg.strip() for seg in current_value.split(',')] if current_value else []
                        # Only add if not already present
                        if add_l2seg_name not in segs:
                            if current_value.strip() == '':
                                entry[1][7] = add_l2seg_name
                                entry[1][7] = ','.join(sorted(s.strip() for s in entry[1][7].split(',')))
                            else:
                                entry[1][7] = current_value + ',' + add_l2seg_name
                                entry[1][7] = ','.join(sorted(s.strip() for s in entry[1][7].split(',')))
                        else:
                            #print(f"'{add_l2seg_name}' already exists in the l2 segments. No change made.")
                            return ([add_l2seg_name + ' already exists in the vport_l2_direct_binding. No change made'])

            elif 'vport_l1if_direct_binding' in argv_array:
                if ns_def.get_if_value(add_vport_name) == -1:
                    return ([f"Invalid virtual port name: {add_vport_name}"])

                if flag_vport_l1if_direct_binding_exists == False:
                    for entry in l2_attribute_array:
                        if entry == row:
                            entry[1][5] = add_vport_name
                            return_text = '--- vport_l1if_direct_binding added --- ' + ' ' + hostname + ',' + portname + ',' + add_vport_name
                            break
                elif flag_vport_l1if_direct_binding_exists == True:
                    #get insert number
                    tmp_if_value_array = []
                    for row in l2_attribute_array[2:]:
                        values = row[1]
                        if len(values) > 1 and values[1] == hostname:
                            if (len(values) > 3 and values[3] == portname) and (len(values) > 5 and values[5] != ''):
                                tmp_if_value_array.append(ns_def.get_if_value(values[5]))
                    target_if_value = ns_def.get_if_value(add_vport_name)
                    from bisect import bisect_left
                    insert_idx = bisect_left(tmp_if_value_array, target_if_value)

                    # Find the index of the first matching hostname row
                    insert_index = -1
                    for i, entry in enumerate(l2_attribute_array):
                        if entry[0] > 2 and entry[1][1] == hostname:
                            insert_index = i
                            break

                    if insert_index != -1:
                        # Create new entry with the same area and hostname
                        new_entry = [0, [l2_attribute_array[insert_index][1][0], hostname, '', portname, '', add_vport_name, '', '']]
                        # Insert the new entry at the found index (before the existing entry)
                        l2_attribute_array.insert(insert_index + 1 + insert_idx, new_entry)

                        # Renumber all entries
                        for i, entry in enumerate(l2_attribute_array):
                            entry[0] = i + 1
                    return_text = '--- vport_l1if_direct_binding added --- ' + ' ' + hostname + ',' + portname + ',' + add_vport_name

            #write to Master file
            excel_maseter_file = master_file_path
            excel_master_ws_name_l2 = 'Master_Data_L2'
            last_l2_table_tuple = {}
            last_l2_table_tuple = ns_def.convert_array_to_tuple(l2_attribute_array)

            # delete L2 Table sheet
            ns_def.remove_excel_sheet(excel_maseter_file, excel_master_ws_name_l2)
            # create L2 Table sheet
            ns_def.create_excel_sheet(excel_maseter_file, excel_master_ws_name_l2)
            # write tuple to excel master data
            ns_def.write_excel_meta(last_l2_table_tuple, excel_maseter_file, excel_master_ws_name_l2, '_template_', 0,0)

            if 'virtual_port' in argv_array or 'portchannel' in argv_array or 'vport_l1if_direct_binding' in argv_array:
                # sync l2 sheet of Master file to L3 sheet
                dummy_tk = tk.Toplevel()
                self.inFileTxt_L3_1_1 = tk.Entry(dummy_tk )
                self.inFileTxt_L3_1_1 .delete(0, tkinter.END)
                self.inFileTxt_L3_1_1 .insert(tk.END, master_file_path)

                self.outFileTxt_11_2 = tk.Entry(dummy_tk )
                self.outFileTxt_11_2.delete(0, tkinter.END)
                self.outFileTxt_11_2.insert(tk.END, master_file_path)

                self.inFileTxt_L2_1_1 = tk.Entry(dummy_tk )
                self.inFileTxt_L2_1_1.delete(0, tkinter.END)
                self.inFileTxt_L2_1_1.insert(tk.END, master_file_path)

                tmp_delete_excel_name = self.inFileTxt_L2_1_1.get().replace('[MASTER]', '[L3_TABLE]')

                import ns_l3_table_from_master
                ns_l3_table_from_master.ns_l3_table_from_master.__init__(self)

                if os.path.isfile(tmp_delete_excel_name) == True:
                    os.remove(tmp_delete_excel_name)

                return ([return_text])

            if 'vport_l2_direct_binding' in argv_array:
                return_text = '--- vport_l2_direct_binding added --- ' + ' ' + hostname + ',' + portname + ',' + add_l2seg_name
                return ([return_text])

            else:
                # sync l2 sheet of Master file to L3 sheet
                dummy_tk = tk.Toplevel()
                self.inFileTxt_L3_1_1 = tk.Entry(dummy_tk )
                self.inFileTxt_L3_1_1 .delete(0, tkinter.END)
                self.inFileTxt_L3_1_1 .insert(tk.END, master_file_path)

                self.outFileTxt_11_2 = tk.Entry(dummy_tk )
                self.outFileTxt_11_2.delete(0, tkinter.END)
                self.outFileTxt_11_2.insert(tk.END, master_file_path)

                self.inFileTxt_L2_1_1 = tk.Entry(dummy_tk )
                self.inFileTxt_L2_1_1.delete(0, tkinter.END)
                self.inFileTxt_L2_1_1.insert(tk.END, master_file_path)

                tmp_delete_excel_name = self.inFileTxt_L2_1_1.get().replace('[MASTER]', '[L3_TABLE]')

                import ns_l3_table_from_master
                ns_l3_table_from_master.ns_l3_table_from_master.__init__(self)

                if os.path.isfile(tmp_delete_excel_name) == True:
                    os.remove(tmp_delete_excel_name)
                return_text = '--- l2 Segment added --- ' + ' ' + hostname + ',' + portname + ',' + add_l2seg_name
                return ([return_text])

    def cli_delete(self, master_file_path, argv_array):
        import ns_def
        import tkinter as tk
        next_arg = get_next_arg(argv_array, 'delete')
        delete_command_list = [
            'delete area',  # Add this line at ver 2.6.1
            'delete device',  # Add this line at ver 2.6.1
            'delete ip_address', \
            'delete l1_link', \
            'delete l2_segment', \
            'delete portchannel', \
            'delete virtual_port', \
            'delete vport_l1if_direct_binding', \
            'delete vport_l2_direct_binding', \
            'delete waypoint', # Add this line at ver 2.6.1
            ]

        if next_arg is None or '--' in next_arg or str('delete ' + next_arg) not in delete_command_list:
            print(next_arg)
            print('[ERROR] Supported commands are as follows')
            for tmp_delete_command_list in delete_command_list:
                print(tmp_delete_command_list)
            exit()

        # Delete area command
        if next_arg == 'area':
            import ns_def

            # Get area name from arguments
            if 'area' in argv_array:
                idx = argv_array.index('area')
                try:
                    area_name = argv_array[idx + 1]
                except IndexError:
                    return (['[Error] Area name is missing'])

            # ★★★ Check if this is the last area ★★★
            # Get all existing areas from STYLE_FOLDER
            style_folder_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<STYLE_FOLDER>>')
            existing_areas = []
            for item in style_folder_array:
                if item[0] not in [1, 2, 3]:
                    existing_areas.append(item[1][0])

            # Check if trying to delete the last area
            if len(existing_areas) <= 1:
                return ([f"[Error] Cannot delete area '{area_name}' as it is the last area in the layout. At least one area must remain."])

            # Check if the specified area exists
            if area_name not in existing_areas:
                return ([f"[Error] Area '{area_name}' not found"])
            # ★★★ End of check ★★★

            # ========== Step 1: Get all devices/waypoints in this area ==========
            area_devices_result = ns_cli_run.cli_show(self, master_file_path, ['show', 'area_device'])
            area_waypoints_result = ns_cli_run.cli_show(self, master_file_path, ['show', 'area_waypoint'])

            # Find devices in target area
            devices_to_delete = []
            for area_device_pair in area_devices_result:
                if area_device_pair[0] == area_name:
                    devices_to_delete = area_device_pair[1]
                    break

            # Find waypoints in target area
            waypoints_to_delete = []
            for area_waypoint_pair in area_waypoints_result:
                if area_waypoint_pair[0] == area_name:
                    waypoints_to_delete = area_waypoint_pair[1]
                    break

            # Combine all elements to delete
            all_elements_to_delete = devices_to_delete + waypoints_to_delete

            if not all_elements_to_delete:
                return ([f"[Error] Area '{area_name}' not found or has no devices/waypoints"])

            # ========== Step 2: Delete from <<STYLE_FOLDER>> ==========
            style_folder_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<STYLE_FOLDER>>')
            ori_style_folder_tuple = ns_def.convert_array_to_tuple(style_folder_array)

            # Remove the area from style_folder_array
            style_folder_array = [item for item in style_folder_array
                                  if not (len(item) >= 2 and isinstance(item[1], list)
                                          and len(item[1]) > 0 and item[1][0] == area_name)]

            # Renumber
            for i, item in enumerate(style_folder_array):
                item[0] = i + 1

            # Write
            style_folder_tuple = ns_def.convert_array_to_tuple(style_folder_array)
            ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_style_folder_tuple)
            ns_def.write_excel_meta(style_folder_tuple, master_file_path,
                                    'Master_Data', '<<STYLE_FOLDER>>', 0, 0)

            # ========== Step 3: Delete from <<POSITION_SHAPE>> ==========
            position_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_SHAPE>>')
            ori_position_shape_tuple = ns_def.convert_array_to_tuple(position_shape_array)

            # Remove area and its devices
            filtered_position_shape_array = []
            skip_until_end = False

            for item in position_shape_array:
                if isinstance(item, list) and len(item) >= 2 and isinstance(item[1], list):
                    row_data = item[1]

                    # Check if this row starts with the target area
                    if len(row_data) > 0 and row_data[0] == area_name:
                        skip_until_end = True
                        continue

                    # Check if this is the end marker
                    if len(row_data) == 1 and row_data[0] == '<END>' and skip_until_end:
                        skip_until_end = False
                        continue

                    # Keep rows that are not part of deleted area
                    if not skip_until_end:
                        filtered_position_shape_array.append(item)

            # Renumber
            for i, item in enumerate(filtered_position_shape_array):
                item[0] = i + 1

            # Write
            position_shape_tuple = ns_def.convert_array_to_tuple(filtered_position_shape_array)
            ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_position_shape_tuple)
            ns_def.write_excel_meta(position_shape_tuple, master_file_path,
                                    'Master_Data', '<<POSITION_SHAPE>>', 0, 0)

            # ========== Step 4: Delete from <<STYLE_SHAPE>>, <<POSITION_TAG>>, <<ATTRIBUTE>> ==========

            # STYLE_SHAPE
            style_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<STYLE_SHAPE>>')
            ori_style_shape_tuple = ns_def.convert_array_to_tuple(style_shape_array)

            style_shape_array = [item for item in style_shape_array
                                 if not (len(item) >= 2 and isinstance(item[1], list)
                                         and len(item[1]) > 0 and item[1][0] in all_elements_to_delete)]

            for i, item in enumerate(style_shape_array):
                item[0] = i + 1

            style_shape_tuple = ns_def.convert_array_to_tuple(style_shape_array)
            ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_style_shape_tuple)
            ns_def.write_excel_meta(style_shape_tuple, master_file_path,
                                    'Master_Data', '<<STYLE_SHAPE>>', 0, 0)

            # POSITION_TAG
            position_tag_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_TAG>>')
            ori_position_tag_tuple = ns_def.convert_array_to_tuple(position_tag_array)

            position_tag_array = [item for item in position_tag_array
                                  if not (len(item) >= 2 and isinstance(item[1], list)
                                          and len(item[1]) > 0 and item[1][0] in all_elements_to_delete)]

            for i, item in enumerate(position_tag_array):
                item[0] = i + 1

            position_tag_tuple = ns_def.convert_array_to_tuple(position_tag_array)
            ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_position_tag_tuple)
            ns_def.write_excel_meta(position_tag_tuple, master_file_path,
                                    'Master_Data', '<<POSITION_TAG>>', 0, 0)

            # ATTRIBUTE
            attribute_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<ATTRIBUTE>>')
            ori_attribute_tuple = ns_def.convert_array_to_tuple(attribute_array)

            attribute_array = [item for item in attribute_array
                               if not (len(item) >= 2 and isinstance(item[1], list)
                                       and len(item[1]) > 0 and item[1][0] in all_elements_to_delete)]

            for i, item in enumerate(attribute_array):
                item[0] = i + 1

            attribute_tuple = ns_def.convert_array_to_tuple(attribute_array)
            ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_attribute_tuple)
            ns_def.write_excel_meta(attribute_tuple, master_file_path,
                                    'Master_Data', '<<ATTRIBUTE>>', 0, 0)

            # ========== Step 5: Delete from <<POSITION_LINE>> ==========
            position_line_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_LINE>>')
            ori_position_line_tuple = ns_def.convert_array_to_tuple(position_line_array)

            # Keep header rows, filter data rows
            header_rows = position_line_array[:2]
            data_rows = [item for item in position_line_array[2:]
                         if not (len(item) >= 2 and isinstance(item[1], list)
                                 and len(item[1]) >= 2
                                 and (item[1][0] in all_elements_to_delete or item[1][1] in all_elements_to_delete))]

            position_line_array = header_rows + data_rows

            # Renumber
            for i, item in enumerate(position_line_array):
                item[0] = i + 1

            # Write
            position_line_tuple = ns_def.convert_array_to_tuple(position_line_array)
            ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_position_line_tuple)
            ns_def.write_excel_meta(position_line_tuple, master_file_path,
                                    'Master_Data', '<<POSITION_LINE>>', 0, 0)

            # ========== Step 6: Update POSITION_FOLDER ==========
            area_location_result = ns_cli_run.cli_show(self, master_file_path, ['show', 'area_location'])

            # Remove target area from layout
            updated_area_location_array = []
            for row in area_location_result:
                if isinstance(row, list):
                    # Filter out the target area
                    filtered_row = [area for area in row if area != area_name]
                    # Only add non-empty rows
                    if filtered_row:
                        updated_area_location_array.append(filtered_row)

            # If no areas left, return error
            if not updated_area_location_array:
                return ([f"[Error] Cannot delete area '{area_name}' as it would result in an empty layout"])

            # ★★★ Add: Validate waypoint layout after deletion ★★★
            validation_result = def_common.validate_waypoint_layout(updated_area_location_array)
            if validation_result['status'] == 'error':
                error_msg = f"[Error] Cannot delete area '{area_name}' as it would result in an invalid layout:\n"
                error_msg += validation_result['message']
                return ([error_msg])
            # ★★★ End of waypoint validation ★★★

            # ========== Step 7: Rebuild POSITION_FOLDER ==========
            result = def_common.rebuild_position_folder(updated_area_location_array, master_file_path)
            if result['status'] != 'success':
                return ([f"[Error] Failed to rebuild POSITION_FOLDER: {result['message']}"])

            # Recalculate folder sizes
            result = def_common.recalculate_folder_sizes(master_file_path)
            if result['status'] != 'success':
                return ([f"[Error] Failed to recalculate folder sizes: {result['message']}"])

            # ========== Sync with L2/L3 ==========
            import tkinter as tk
            dummy_tk = tk.Toplevel()
            dummy_tk.withdraw()

            self.full_filepath = master_file_path
            self.main1_1_entry_1 = tk.Entry(dummy_tk)
            self.main1_1_entry_1.insert(tk.END, master_file_path)

            # ★★★ Add missing attributes ★★★
            self.inFileTxt_L3_1_1 = tk.Entry(dummy_tk)
            self.inFileTxt_L3_1_1.delete(0, tkinter.END)
            self.inFileTxt_L3_1_1.insert(tk.END, master_file_path)

            self.outFileTxt_11_2 = tk.Entry(dummy_tk)
            self.outFileTxt_11_2.delete(0, tkinter.END)
            self.outFileTxt_11_2.insert(tk.END, master_file_path)

            self.inFileTxt_L2_1_1 = tk.Entry(dummy_tk)
            self.inFileTxt_L2_1_1.delete(0, tkinter.END)
            self.inFileTxt_L2_1_1.insert(tk.END, master_file_path)

            import ns_sync_between_layers
            ns_sync_between_layers.l1_master_device_and_line_sync_with_l2l3_master(self)

            # ★★★ Delete L2_TABLE file if exists ★★★
            tmp_delete_excel_name = master_file_path.replace('[MASTER]', '[L2_TABLE]')
            if os.path.isfile(tmp_delete_excel_name):
                os.remove(tmp_delete_excel_name)

            dummy_tk.destroy()

            # Return message
            return_text = f"--- Area deleted --- {area_name}\n"
            if devices_to_delete:
                return_text += f"  Deleted {len(devices_to_delete)} devices\n"
            if waypoints_to_delete:
                return_text += f"  Deleted {len(waypoints_to_delete)} waypoints\n"
            return_text += "Updated area layout"

            return ([return_text])


        # Delete device command
        if next_arg == 'device':
            # Get device name from arguments
            if 'device' in argv_array:
                idx = argv_array.index('device')
                try:
                    device_name = argv_array[idx + 1]
                except IndexError:
                    return (['[Error] device name is missing'])

            # Call common delete function
            return def_common.delete_device_or_waypoint_common(
                self,
                device_name,
                master_file_path,
                element_type='device'
            )

        # Delete waypoint command (separate if block)
        if next_arg == 'waypoint':
            # Get waypoint name from arguments
            if 'waypoint' in argv_array:
                idx = argv_array.index('waypoint')
                try:
                    waypoint_name = argv_array[idx + 1]
                except IndexError:
                    return (['[Error] waypoint name is missing'])

            # Call common delete function
            return def_common.delete_device_or_waypoint_common(
                self,
                waypoint_name,
                master_file_path,
                element_type='waypoint'
            )

        if next_arg == 'l1_link':  # in cli_delete
            import ns_def
            position_folder_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_FOLDER>>')
            position_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_SHAPE>>')
            style_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<STYLE_SHAPE>>')
            position_line_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_LINE>>')
            position_tag_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_TAG>>')
            ori_position_line_tuple = ns_def.convert_array_to_tuple(position_line_array)

            # Check if the input data exists in argv_array
            if 'l1_link' in argv_array:
                idx = argv_array.index('l1_link')
                try:
                    hostname = argv_array[idx + 1]
                    portname = argv_array[idx + 2]
                except IndexError:
                    return (['[Error] hostname or portname is missing'])

            # Validate port name
            if ns_def.split_portname(portname)[0] == '' or ns_def.split_portname(portname)[1] == '':
                error_msg = f"Error: invalid portname '{portname}'. \n [Tips] When specifying the portname, be sure to use single quotes, such as 'port 0'."
                return ([error_msg])

            port = str(ns_def.adjust_portname(portname)[0]) + ' ' + str(ns_def.split_portname(portname)[1])

            # Find and remove the connection
            deleted_connection = None
            for item in position_line_array[2:]:
                connection_data = item[1]
                from_h = connection_data[0]
                to_h = connection_data[1]
                from_port = connection_data[2]
                to_port = connection_data[3]

                if (from_h == hostname and from_port == port) or (to_h == hostname and to_port == port):
                    deleted_connection = item
                    position_line_array.remove(item)
                    break

            if deleted_connection is None:
                return ([f"Error: No connection found for {hostname} with port {port}"])

            # Call common processing with shrinking enabled
            position_line_array, style_shape_array = def_common.process_l1_link_common(
                self, position_shape_array, style_shape_array, position_line_array,
                position_tag_array, master_file_path, ori_position_line_tuple,
                allow_shrink=True, position_folder_array=position_folder_array
            )

            # ★★★ ADD: Call recalculate_folder_sizes to remove unnecessary empty cells ★★★
            result = def_common.recalculate_folder_sizes(master_file_path)
            if result['status'] != 'success':
                return ([f"[ERROR] Failed to recalculate folder sizes: {result['message']}"])
            # ★★★ End of addition ★★★

            deleted_from_h = deleted_connection[1][0]
            deleted_to_h = deleted_connection[1][1]
            deleted_from_port = deleted_connection[1][2]
            deleted_to_port = deleted_connection[1][3]

            return_text = f'--- Deleted Layer 1 link --- {deleted_from_h}({deleted_from_port}) ↔ {deleted_to_h}({deleted_to_port})'
            return ([return_text])



        if next_arg == 'ip_address':
            import ns_def
            l3_attribute_array = ns_def.convert_master_to_array('Master_Data_L3', master_file_path,'<<L3_TABLE>>')

            TARGET_LEN = 7
            for row in l3_attribute_array:
                if isinstance(row, list) and len(row) >= 2 and isinstance(row[1], list):
                    vals = row[1]
                    if len(vals) < TARGET_LEN:
                        vals += [''] * (TARGET_LEN - len(vals))

            idx = argv_array.index('ip_address')
            try:
                # The element right after 'ip address' is the hostname
                hostname = argv_array[idx + 1]
                # The next element is the portname
                portname = argv_array[idx + 2]
                delete_ipaddress_name = argv_array[idx + 3]
            except IndexError:
                # If there are not enough elements after 'ip address', print an error
                return ([f"[ERROR] hostname or portname or ip address is missing"])

            # check IP Addresses
            if ns_def.check_ip_format(delete_ipaddress_name) != 'IPv4':
                return ([f"Error: IP Address format is invalid: {delete_ipaddress_name}"])

            match_found = False
            target_row = None
            for row in l3_attribute_array[2:]:
                values = row[1]
                if len(values) > 1 and values[1] == hostname:
                    if (len(values) > 3 and values[2] == portname):
                        match_found = True
                        target_row = row
                        break
            if not match_found:
                return ([f"Error: No matching entry found for hostname: {hostname} and portname: {portname}"])

            delete_ipaddress_name = delete_ipaddress_name.replace(' ', '')

            for entry in l3_attribute_array:
                if entry == target_row:
                    current_value = entry[1][4]
                    segs = [seg.strip() for seg in current_value.split(',')] if current_value else []
                    if delete_ipaddress_name in segs:
                        segs.remove(delete_ipaddress_name)
                        entry[1][4] = ','.join(segs)
                    else:
                        return ([f"{delete_ipaddress_name} does not exist in the ip address. No change made"])
                    break

            #write to Master file
            excel_maseter_file = master_file_path
            excel_master_ws_name_l3 = 'Master_Data_L3'
            last_l3_table_tuple = {}
            last_l3_table_tuple = ns_def.convert_array_to_tuple(l3_attribute_array)

            # delete L2 Table sheet
            ns_def.remove_excel_sheet(excel_maseter_file, excel_master_ws_name_l3)
            # create L2 Table sheet
            ns_def.create_excel_sheet(excel_maseter_file, excel_master_ws_name_l3)
            # write tuple to excel master data
            ns_def.write_excel_meta(last_l3_table_tuple, excel_maseter_file, excel_master_ws_name_l3, '_template_', 0,0)

            return_text = '--- IP Address deleted--- ' + ' ' + hostname + ',' + portname + ',' + delete_ipaddress_name

            return ([return_text])

        if next_arg == 'l2_segment' or 'virtual_port' or 'portchannel' or 'vport_l2_direct_binding' or 'vport_l1if_direct_binding':
            import ns_def
            l2_attribute_array = ns_def.convert_master_to_array('Master_Data_L2', master_file_path, '<<L2_TABLE>>')

            TARGET_LEN = 9
            for row in l2_attribute_array:
                if isinstance(row, list) and len(row) >= 2 and isinstance(row[1], list):
                    vals = row[1]
                    if len(vals) < TARGET_LEN:
                        vals += [''] * (TARGET_LEN - len(vals))

            if 'l2_segment' in argv_array:
                idx = argv_array.index('l2_segment')
                try:
                    hostname = argv_array[idx + 1]
                    portname = argv_array[idx + 2]
                    del_l2seg_name = argv_array[idx + 3]
                except IndexError:
                    print("Error: hostname or portname is missing")
                    return
            elif 'virtual_port' in argv_array:
                idx = argv_array.index('virtual_port')
                try:
                    hostname = argv_array[idx + 1]
                    del_vport_name = argv_array[idx + 2]
                except IndexError:
                    print("Error: hostname or virtual_port name is missing")
                    return
            elif 'portchannel' in argv_array:
                idx = argv_array.index('portchannel')
                try:
                    hostname = argv_array[idx + 1]
                    portname = argv_array[idx + 2]
                    #del_portchannel_name = argv_array[idx + 3]
                except IndexError:
                    print("Error: hostname or portname is missing")
                    return
            elif 'vport_l2_direct_binding' in argv_array:
                idx = argv_array.index('vport_l2_direct_binding')
                try:
                    hostname = argv_array[idx + 1]
                    portname = argv_array[idx + 2]
                    del_l2seg_name = argv_array[idx + 3]
                except IndexError:
                    print("Error: hostname or virtual portname is missing")
                    return
            elif 'vport_l1if_direct_binding' in argv_array:
                idx = argv_array.index('vport_l1if_direct_binding')
                try:
                    hostname = argv_array[idx + 1]
                    del_vport_name = argv_array[idx + 2]
                except IndexError:
                    print("Error: hostname or portname is missing")
                    return
            else:
                print("Not found in the argument array")
                return

            match_found = False
            target_row = None
            row_array = []
            flag_l2_segment_vport = False
            if 'l2_segment' in argv_array:
                for row in l2_attribute_array[2:]:
                    values = row[1]
                    if len(values) > 1 and values[1] == hostname:
                        if len(values) > 3 and values[3] == portname:
                            match_found = True
                            target_row = row
                            break
                        elif len(values) > 5 and values[5] == portname:
                            # for virtual port
                            match_found = True
                            flag_l2_segment_vport = True
                            row_array.append(row)
                if flag_l2_segment_vport == True:
                    row = []

            elif 'virtual_port' in argv_array:
                for row in l2_attribute_array[2:]:
                    values = row[1]
                    if len(values) > 1 and values[1] == hostname and values[5] == del_vport_name:
                        if values[3] != '':
                            return (['[ERROR] To delete a virtual port bound to a Layer 1 interface, use the delete vport_l1if_direct_binding command.'])

                        match_found = True
                        target_row = row
                        break
            elif 'portchannel' in argv_array:
                for row in l2_attribute_array[2:]:
                    values = row[1]
                    if len(values) > 1 and values[1] == hostname:
                        if (len(values) > 3 and values[3] == portname) or (len(values) > 5 and values[5] == portname):
                            match_found = True
                            target_row = row
                            break
            elif 'vport_l2_direct_binding' in argv_array:
                for row in l2_attribute_array[2:]:
                    values = row[1]
                    if len(values) > 1 and values[1] == hostname:
                        if len(values) > 5 and values[5] == portname:
                            # for virtual port
                            match_found = True
                            flag_l2_segment_vport = True
                            row_array.append(row)
                if flag_l2_segment_vport == True:
                    row = []
            elif 'vport_l1if_direct_binding' in argv_array:
                match_l1_if = ''
                for row in l2_attribute_array[2:]:
                    values = row[1]
                    if len(values) > 1 and values[1] == hostname:
                        if len(values) > 5 and values[5] == del_vport_name:
                            match_found = True
                            target_row = row
                            #count match l1 interface
                            match_l1_if = values[3]
                            break

                count_match_l1if = 0
                for tmp_l2_attribute_array in l2_attribute_array[2:]:
                    value_l2_attribute_array = tmp_l2_attribute_array[1]
                    if len(value_l2_attribute_array) > 1 and value_l2_attribute_array[1] == hostname:
                        if len(value_l2_attribute_array) > 3 and value_l2_attribute_array[3] == match_l1_if:
                            count_match_l1if += 1

            if not match_found:
                if 'l2_segment' in argv_array:
                    print(f"No matching entry found for hostname: {hostname} and portname: {portname}")
                elif 'portchannel' in argv_array:
                    print(f"No matching entry found for hostname: {hostname} and portname: {portname}")
                elif 'vport_l2_direct_binding' in argv_array:
                    print(f"No matching entry found for hostname: {hostname} and virtual portname: {portname}")
                else:
                    print(f"No matching entry found for hostname: {hostname} and virtual port: {del_vport_name}")
                exit()

            if 'l2_segment' in argv_array:
                del_l2seg_name = del_l2seg_name.replace(' ', '')
                for entry in l2_attribute_array:
                    if entry == target_row:
                        current_value = entry[1][6]
                        segs = [seg.strip() for seg in current_value.split(',')] if current_value else []
                        if del_l2seg_name in segs:
                            segs.remove(del_l2seg_name)
                            entry[1][6] = ','.join(segs)
                        else:
                            return ([f"{del_l2seg_name} does not exist in the l2 segments. No change made"])
                        break

                    #for virtual port
                    if entry in row_array:
                        # Get the 7th element (index 6)
                        current_value = entry[1][6]
                        # Split by comma, remove extra spaces
                        segs = [seg.strip() for seg in current_value.split(',')] if current_value else []
                        # Only add if not already present
                        if del_l2seg_name in segs:
                            segs.remove(del_l2seg_name)
                            entry[1][6] = ','.join(segs)
                        else:
                            return ([del_l2seg_name + ' does not exist in the l2 segments. No change made'])


            elif 'virtual_port' in argv_array:
                # Remove the entire row for virtual port
                l2_attribute_array.remove(target_row)
                # Renumber all entries
                for i, entry in enumerate(l2_attribute_array):
                    entry[0] = i + 1
                    return_text = '--- Virtual Port deleted--- ' + ' ' + hostname + ',' + del_vport_name

            elif 'portchannel' in argv_array:
                for entry in l2_attribute_array:
                    if entry == target_row:
                        del_portchannel_name = entry[1][5]
                        entry[1][5] = ''
                        entry[1][6] = ''
                        return_text = '--- portchannel deleted --- ' + ' ' + hostname + ',' + portname + ',' + del_portchannel_name
                        break

            elif 'vport_l2_direct_binding' in argv_array:
                del_l2seg_name = del_l2seg_name.replace(' ', '')
                for entry in l2_attribute_array:
                    #for virtual port
                    if entry in row_array:
                        # Get the 7th element (index 6)
                        current_value = entry[1][7]
                        # Split by comma, remove extra spaces
                        segs = [seg.strip() for seg in current_value.split(',')] if current_value else []
                        # Only add if not already present
                        #print(del_l2seg_name,segs)
                        if del_l2seg_name in segs:
                            segs.remove(del_l2seg_name)
                            entry[1][7] = ','.join(segs)
                        else:
                            return ([del_l2seg_name + ' does not exist in the vport_l2_direct_binding. No change made'])

            elif 'vport_l1if_direct_binding' in argv_array:
                if count_match_l1if >= 2:
                    # Remove the entire row for virtual port
                    l2_attribute_array.remove(target_row)
                    # Renumber all entries
                    for i, entry in enumerate(l2_attribute_array):
                        entry[0] = i + 1
                        return_text = '--- vport_l1if_direct_binding deleted--- ' + ' ' + hostname + ',' + del_vport_name
                elif count_match_l1if == 1:
                    for entry in l2_attribute_array:
                        if entry == target_row:
                            del_vport_name = entry[1][5]
                            entry[1][5] = ''
                            entry[1][6] = ''
                            entry[1][7] = ''
                            return_text = '--- vport_l1if_direct_binding deleted --- ' + ' ' + hostname + ',' + del_vport_name
                            break

            #write to Master file
            excel_maseter_file = master_file_path
            excel_master_ws_name_l2 = 'Master_Data_L2'
            last_l2_table_tuple = ns_def.convert_array_to_tuple(l2_attribute_array)
            ns_def.remove_excel_sheet(excel_maseter_file, excel_master_ws_name_l2)
            ns_def.create_excel_sheet(excel_maseter_file, excel_master_ws_name_l2)
            ns_def.write_excel_meta(last_l2_table_tuple, excel_maseter_file, excel_master_ws_name_l2, '_template_', 0, 0)

            if 'virtual_port' in argv_array or 'portchannel' in argv_array or 'vport_l1if_direct_binding' in argv_array:
                #sync l2 sheet of Master file to L3 sheet
                dummy_tk = tk.Toplevel()
                self.inFileTxt_L3_1_1 = tk.Entry(dummy_tk)
                self.inFileTxt_L3_1_1.delete(0, tkinter.END)
                self.inFileTxt_L3_1_1.insert(tk.END, master_file_path)

                self.outFileTxt_11_2 = tk.Entry(dummy_tk)
                self.outFileTxt_11_2.delete(0, tkinter.END)
                self.outFileTxt_11_2.insert(tk.END, master_file_path)

                self.inFileTxt_L2_1_1 = tk.Entry(dummy_tk)
                self.inFileTxt_L2_1_1.delete(0, tkinter.END)
                self.inFileTxt_L2_1_1.insert(tk.END, master_file_path)

                tmp_delete_excel_name = self.inFileTxt_L2_1_1.get().replace('[MASTER]', '[L3_TABLE]')

                import ns_l3_table_from_master
                ns_l3_table_from_master.ns_l3_table_from_master.__init__(self)

                if os.path.isfile(tmp_delete_excel_name) == True:
                    os.remove(tmp_delete_excel_name)
            elif 'vport_l2_direct_binding' in argv_array:
                return_text = '--- vport_l2_direct_binding deleted --- ' + ' ' + hostname + ',' + portname + ',' + del_l2seg_name

            else:
                #sync l2 sheet of Master file to L3 sheet
                dummy_tk = tk.Toplevel()
                self.inFileTxt_L3_1_1 = tk.Entry(dummy_tk)
                self.inFileTxt_L3_1_1.delete(0, tkinter.END)
                self.inFileTxt_L3_1_1.insert(tk.END, master_file_path)

                self.outFileTxt_11_2 = tk.Entry(dummy_tk)
                self.outFileTxt_11_2.delete(0, tkinter.END)
                self.outFileTxt_11_2.insert(tk.END, master_file_path)

                self.inFileTxt_L2_1_1 = tk.Entry(dummy_tk)
                self.inFileTxt_L2_1_1.delete(0, tkinter.END)
                self.inFileTxt_L2_1_1.insert(tk.END, master_file_path)

                tmp_delete_excel_name = self.inFileTxt_L2_1_1.get().replace('[MASTER]', '[L3_TABLE]')

                import ns_l3_table_from_master
                ns_l3_table_from_master.ns_l3_table_from_master.__init__(self)

                if os.path.isfile(tmp_delete_excel_name) == True:
                    os.remove(tmp_delete_excel_name)
                return_text = '--- l2 Segment deleted --- ' + ' ' + hostname + ',' + portname + ',' + del_l2seg_name

            return ([return_text])

    def cli_show(self, master_file_path, argv_array):
        next_arg = get_next_arg(argv_array, 'show')
        show_command_list = [
            'show area',
            'show area_device',
            'show area_location',
            'show area_waypoint',
            'show attribute',
            'show attribute_color',
            'show device',
            'show device_interface',
            'show device_location',
            'show l1_interface',
            'show l1_link',
            'show l2_broadcast_domain',
            'show l2_interface',
            'show l3_broadcast_domain',
            'show l3_interface',
            'show waypoint',
            'show waypoint_interface',
            'show waypoint_location',
        ]

        if next_arg == None or '--' in next_arg or str('show ' + next_arg) not in show_command_list:
            print(next_arg)
            print('[ERROR] Supported commands are as follows')
            for tmp_show_command_list in show_command_list:
                print(tmp_show_command_list)
            exit()

        if next_arg == 'device_location':
            import ns_def
            position_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_SHAPE>>')
            update_position_shape_array = []

            current_folder_name = ''
            current_folder_shape = []
            flag_1st = True
            for tmp_position_folder_array in position_shape_array:
                if tmp_position_folder_array[1][0] != '<<POSITION_SHAPE>>':
                    if tmp_position_folder_array[1][0] != '<END>' and tmp_position_folder_array[0] != '':
                        if flag_1st == True and tmp_position_folder_array[1][0] != '<END>' and tmp_position_folder_array[1][0] != '':
                            current_folder_name = tmp_position_folder_array[1][0]

                        if flag_1st == False:
                            # Only include non-waypoint folders (without _wp_)
                            if '_wp_' not in str(current_folder_name):
                                update_position_shape_array.append([current_folder_name, current_folder_shape])

                            current_folder_name = tmp_position_folder_array[1][0]
                            current_folder_shape = []
                            flag_1st = True

                    if tmp_position_folder_array[1][0] != '<END>':
                        # ★★★ Modified: Keep _AIR_ in output (do not filter) ★★★
                        # Extract all elements except <END>
                        row_data = tmp_position_folder_array[1][1:]
                        if '<END>' in row_data:
                            end_idx = row_data.index('<END>')
                            row_data = row_data[:end_idx]

                        # Add row with _AIR_ included
                        current_folder_shape.append(row_data)
                        # ★★★ End of modification ★★★

                    if tmp_position_folder_array[1][0] == '<END>':
                        flag_1st = False
            else:
                # Only include non-waypoint folders (without _wp_)
                if '_wp_' not in str(current_folder_name):
                    update_position_shape_array.append([current_folder_name, current_folder_shape])

            return (update_position_shape_array)

        if next_arg == 'waypoint_location':
            import ns_def
            position_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_SHAPE>>')
            update_position_shape_array = []

            current_folder_name = ''
            current_folder_shape = []
            flag_1st = True
            for tmp_position_folder_array in position_shape_array:
                if tmp_position_folder_array[1][0] != '<<POSITION_SHAPE>>':
                    if tmp_position_folder_array[1][0] != '<END>' and tmp_position_folder_array[0] != '':
                        if flag_1st == True and tmp_position_folder_array[1][0] != '<END>' and tmp_position_folder_array[1][0] != '':
                            current_folder_name = tmp_position_folder_array[1][0]

                        if flag_1st == False:
                            # Only include waypoint folders (with _wp_)
                            if '_wp_' in str(current_folder_name):
                                update_position_shape_array.append([current_folder_name, current_folder_shape])

                            current_folder_name = tmp_position_folder_array[1][0]
                            current_folder_shape = []
                            flag_1st = True

                    if tmp_position_folder_array[1][0] != '<END>':
                        # ★★★ Modified: Keep _AIR_ in output (do not filter) ★★★
                        # Extract all elements except <END>
                        row_data = tmp_position_folder_array[1][1:]
                        if '<END>' in row_data:
                            end_idx = row_data.index('<END>')
                            row_data = row_data[:end_idx]

                        # Add row with _AIR_ included
                        current_folder_shape.append(row_data)
                        # ★★★ End of modification ★★★

                    if tmp_position_folder_array[1][0] == '<END>':
                        flag_1st = False
            else:
                # Only include waypoint folders (with _wp_)
                if '_wp_' in str(current_folder_name):
                    update_position_shape_array.append([current_folder_name, current_folder_shape])

            return (update_position_shape_array)


        if next_arg == 'area' or 'area_location':
            import ns_def
            position_folder_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<STYLE_FOLDER>>')

            folder_name_list = []
            area_wp_name_list = []
            area_name_list = []
            for item in position_folder_array:
                if item[0] not in [1, 2, 3]:
                    folder_name_list.append(item[1][0])

            for folder_name in folder_name_list:
                if "_wp_" in folder_name:
                    area_wp_name_list.append(folder_name)
                else:
                    area_name_list.append(folder_name)

            # Modified: Include both normal areas and waypoint areas
            if next_arg == 'area':
                all_areas = area_name_list + area_wp_name_list
                all_areas_sorted = sorted(all_areas, reverse=False)
                return (all_areas_sorted)

        if next_arg == 'area_location':
            import ns_def
            position_folder_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_FOLDER>>')

            # Get all areas (including waypoint areas)
            style_folder_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<STYLE_FOLDER>>')
            all_area_names = []
            for item in style_folder_array:
                if item[0] not in [1, 2, 3]:
                    all_area_names.append(item[1][0])

            update_position_folder_array = []
            for tmp_position_folder_array in position_folder_array:
                update_position_folder_array.append(tmp_position_folder_array[1])

            tmp_return = update_position_folder_array

            for sublist in tmp_return:
                for i in range(len(sublist)):
                    if sublist[i] not in all_area_names:
                        sublist[i] = ''

            tmp_return = [sublist for sublist in tmp_return if any(item != '' for item in sublist)]
            tmp_return = [[item for item in sublist if item != ''] for sublist in tmp_return]

            return (tmp_return)

        if next_arg == 'area_waypoint':
            import ns_def

            # Get waypoints directly from POSITION_SHAPE
            position_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_SHAPE>>')

            area_waypoint_dict = {}
            current_folder = None

            for item in position_shape_array:
                if isinstance(item, list) and len(item) >= 2 and isinstance(item[1], list):
                    row = item[1]

                    # Check folder name
                    if len(row) > 0 and row[0] and row[0] not in ['', '<END>', '<<POSITION_SHAPE>>', '_AIR_']:
                        current_folder = row[0]
                        # Only track waypoint folders
                        if '_wp_' in current_folder:
                            if current_folder not in area_waypoint_dict:
                                area_waypoint_dict[current_folder] = []

                    # Collect waypoints in current folder
                    if current_folder and '_wp_' in current_folder:
                        if len(row) > 0 and row[0] != '<END>':
                            for val in row[1:]:  # Skip first element
                                if val not in ['', '<END>', '_AIR_', '<<POSITION_SHAPE>>']:
                                    if val not in area_waypoint_dict[current_folder]:
                                        area_waypoint_dict[current_folder].append(val)

                    # Reset when reaching folder end
                    if len(row) == 1 and row[0] == '<END>':
                        current_folder = None

            # Convert to sorted list format
            result = [[area, sorted(waypoints)] for area, waypoints in sorted(area_waypoint_dict.items())]
            return result


        if next_arg == 'area_device':
            import ns_def

            # Get devices directly from POSITION_SHAPE instead of Master_Data_L2
            position_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_SHAPE>>')

            area_device_dict = {}
            current_folder = None

            for item in position_shape_array:
                if isinstance(item, list) and len(item) >= 2 and isinstance(item[1], list):
                    row = item[1]

                    # Check folder name (area name)
                    if len(row) > 0 and row[0] and row[0] not in ['', '<END>', '<<POSITION_SHAPE>>', '_AIR_']:
                        current_folder = row[0]
                        # Only track non-waypoint folders
                        if '_wp_' not in current_folder:
                            if current_folder not in area_device_dict:
                                area_device_dict[current_folder] = []

                    # Collect devices in current folder
                    if current_folder and '_wp_' not in current_folder:
                        if len(row) > 0 and row[0] != '<END>':
                            for val in row[1:]:  # Skip first element (folder name or empty)
                                if val not in ['', '<END>', '_AIR_', '<<POSITION_SHAPE>>']:
                                    if val not in area_device_dict[current_folder]:
                                        area_device_dict[current_folder].append(val)

                    # Reset when reaching folder end
                    if len(row) == 1 and row[0] == '<END>':
                        current_folder = None

            # Convert to sorted list format
            result = [[area, sorted(devices)] for area, devices in sorted(area_device_dict.items())]
            return result


        if next_arg == 'device_interface' or 'waypoint_interface' or 'l1_link' or 'l1_interface':
            import ns_def
            position_line_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_LINE>>')
            position_line_tuple = ns_def.convert_array_to_tuple(position_line_array)

            line_list = []
            device_interface_list = []
            waypoint_interface_list = []
            interface_detail_list = []
            device_waypoint_array = get_device_waypoint_array(master_file_path)
            for item in position_line_array:

                if item[0] not in [1,2]:
                    # get full if name
                    if_full_name_src = ns_def.get_full_name_from_tag_name(item[1][0], item[1][2],position_line_tuple)
                    if_full_name_tar = ns_def.get_full_name_from_tag_name(item[1][1], item[1][3],position_line_tuple)

                    line_list.append([[item[1][0], if_full_name_src], [item[1][1], if_full_name_tar]])
                    if str(item[1][0]) in device_waypoint_array[0]:
                        device_interface_list.append([item[1][0],if_full_name_src])
                    else:
                        waypoint_interface_list.append([item[1][0],if_full_name_src])

                    if str(item[1][1]) in device_waypoint_array[0]:
                        device_interface_list.append([item[1][1],if_full_name_tar])
                    else:
                        waypoint_interface_list.append([item[1][1],if_full_name_src])

                    interface_detail_list.append([item[1][0], item[1][2], if_full_name_src, item[1][13], item[1][14], item[1][15]])
                    interface_detail_list.append([item[1][1], item[1][3], if_full_name_tar, item[1][17], item[1][18], item[1][19]])

            if next_arg == 'l1_link':
                line_list = sorted(line_list, key=lambda x: (x[0][0],x[0][1]), reverse=False)
                return (line_list)

            elif next_arg == 'device_interface':
                device_interface_list = sorted(device_interface_list, key=lambda x: (x[0][0],x[0][1]), reverse=False)
                #print(interface_list)

                from collections import defaultdict
                grouped_data = defaultdict(list)
                for device, interface in device_interface_list:
                    grouped_data[device].append(interface, )

                result = [[device, interfaces] for device, interfaces in grouped_data.items()]
                return (result)

            elif next_arg == 'waypoint_interface':
                waypoint_interface_list = sorted(waypoint_interface_list, key=lambda x: (x[0][0],x[0][1]), reverse=False)
                #print(interface_list)

                from collections import defaultdict
                grouped_data = defaultdict(list)
                for device, interface in waypoint_interface_list:
                    grouped_data[device].append(interface, )

                result = [[device, interfaces] for device, interfaces in grouped_data.items()]
                return (result)

            elif next_arg == 'l1_interface':
                interface_detail_list = sorted(interface_detail_list, key=lambda x: (x[0], x[1]), reverse=False)
                return (interface_detail_list)

        if next_arg == 'device' or 'waypoint':
            if next_arg == 'device':
                # Get devices directly from Master_Data instead of Master_Data_L2
                import ns_def
                position_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_SHAPE>>')
                attribute_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<ATTRIBUTE>>')

                device_list_array = []

                # Method 1: Identify devices from ATTRIBUTE (not WayPoint)
                device_from_attribute = set()
                for item in attribute_array:
                    if item[0] > 1 and len(item[1]) > 1:
                        device_name = item[1][0]
                        default_attr = item[1][1]
                        # Check if it's a DEVICE (not WayPoint)
                        if 'DEVICE' in str(default_attr) or 'WayPoint' not in str(default_attr):
                            if 'DEVICE' in str(default_attr):  # Only add if explicitly DEVICE
                                device_from_attribute.add(device_name)

                # Method 2: Get devices from non-_wp_ folders in POSITION_SHAPE
                current_folder = None
                for item in position_shape_array:
                    if isinstance(item, list) and len(item) >= 2 and isinstance(item[1], list):
                        row = item[1]

                        # Check folder name
                        if len(row) > 0 and row[0] and row[0] not in ['', '<END>', '<<POSITION_SHAPE>>', '_AIR_']:
                            current_folder = row[0]

                        # Collect devices in non-_wp_ folders
                        if current_folder and '_wp_' not in current_folder:
                            for val in row:
                                if val not in ['', '<END>', '_AIR_', current_folder, '<<POSITION_SHAPE>>']:
                                    device_list_array.append(val)

                        # End of folder
                        if len(row) == 1 and row[0] == '<END>':
                            current_folder = None

                # Combine devices found by both methods (remove duplicates)
                device_list_array = list(set(device_list_array) | device_from_attribute)
                device_list_array = sorted(device_list_array, reverse=False)
                return (device_list_array)

            if next_arg == 'waypoint':
                # Get waypoints directly from Master_Data
                import ns_def
                position_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_SHAPE>>')
                attribute_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<ATTRIBUTE>>')

                wp_list_array = []

                # Method 1: Identify WayPoints from ATTRIBUTE
                waypoint_from_attribute = set()
                for item in attribute_array:
                    if item[0] > 1 and len(item[1]) > 1:
                        device_name = item[1][0]
                        default_attr = item[1][1]
                        # Check if it's a WayPoint
                        if 'WayPoint' in str(default_attr):
                            waypoint_from_attribute.add(device_name)

                # Method 2: Get devices from _wp_ folders in POSITION_SHAPE
                current_folder = None
                for item in position_shape_array:
                    if isinstance(item, list) and len(item) >= 2 and isinstance(item[1], list):
                        row = item[1]

                        # Check folder name
                        if len(row) > 0 and row[0] and row[0] not in ['', '<END>', '<<POSITION_SHAPE>>', '_AIR_']:
                            current_folder = row[0]

                        # Collect devices in _wp_ folders
                        if current_folder and '_wp_' in current_folder:
                            for val in row:
                                if val not in ['', '<END>', '_AIR_', current_folder, '<<POSITION_SHAPE>>']:
                                    wp_list_array.append(val)

                        # End of folder
                        if len(row) == 1 and row[0] == '<END>':
                            current_folder = None

                # Combine waypoints found by both methods (remove duplicates)
                wp_list_array = list(set(wp_list_array) | waypoint_from_attribute)
                wp_list_array = sorted(wp_list_array, reverse=False)
                return (wp_list_array)

        if next_arg == 'area_device':
            import ns_def
            l2_table_array = ns_def.convert_master_to_array('Master_Data_L2', master_file_path,   '<<L2_TABLE>>')
            new_l2_table_array = []
            for tmp_l2_table_array in l2_table_array:
                if tmp_l2_table_array[0] != 1 and tmp_l2_table_array[0] != 2:
                    tmp_l2_table_array[1].extend(['', '', '', '', '', '', '', ''])
                    del tmp_l2_table_array[1][8:]
                    new_l2_table_array.append(tmp_l2_table_array)

            device_list_array = []
            area_device_list_array = []
            wp_list_array = []
            for tmp_new_l2_table_array in new_l2_table_array:
                if tmp_new_l2_table_array[1][1] not in device_list_array and tmp_new_l2_table_array[1][1] not in wp_list_array:
                    if tmp_new_l2_table_array[1][0] == 'N/A':
                        wp_list_array.append(tmp_new_l2_table_array[1][1])
                    else:
                        device_list_array.append(tmp_new_l2_table_array[1][1])
                        area_device_list_array.append([tmp_new_l2_table_array[1][0], tmp_new_l2_table_array[1][1]])
            area_device_list_array = sorted(area_device_list_array, key=lambda x: (x[0],x[1]), reverse=False)

            from collections import defaultdict
            grouped_data = defaultdict(list)
            for area, device in area_device_list_array:
                grouped_data[area].append(device)
            result = [[area, devices] for area, devices in grouped_data.items()]
            return (result)

        if next_arg == 'l3_interface':
            import ns_def
            l3_interface_array = ns_def.convert_master_to_array('Master_Data_L3', master_file_path,'<<L3_TABLE>>')

            update_l3_interface_array = []
            for tmp_l3_interface_array in l3_interface_array:
                if tmp_l3_interface_array[0] != 1 and tmp_l3_interface_array[0] != 2:
                    update_l3_interface_array.append(tmp_l3_interface_array[1][1:])
            padded_data = [sublist + [''] * (4 - len(sublist)) for sublist in update_l3_interface_array]
            padded_data = sorted(padded_data, key=lambda x: (x[0]), reverse=False)
            return (padded_data)

        if next_arg == 'l2_interface':
            import ns_def
            l2_attribute_array = ns_def.convert_master_to_array('Master_Data_L2', master_file_path,'<<L2_TABLE>>')

            update_attribute_array = []
            for tmp_attribute_array in l2_attribute_array:
                if tmp_attribute_array[0] != 1 and tmp_attribute_array[0] != 2:
                    update_attribute_array.append(tmp_attribute_array[1][1:])

            padded_data = [sublist + [''] * (7 - len(sublist)) for sublist in update_attribute_array]
            update_padded_data = []

            for tmp_padded_data in padded_data:
                del tmp_padded_data[3]
                del tmp_padded_data[1]
                update_padded_data.append(tmp_padded_data)

            update_padded_data = sorted(padded_data, key=lambda x: (x[0]), reverse=False)
            return (update_padded_data)

        if next_arg == 'l2_broadcast_domain' or 'l3_broadcast_domain':
            import ns_def
            return_get_l2_broadcast_domains = []
            result_get_l2_broadcast_domains = ns_def.get_l2_broadcast_domains.run(self, master_file_path)  ## '[0] self.update_l2_table_array, [1] device_l2_boradcast_domain_array, [2] device_l2_directly_l3vport_array, [3] device_l2_other_array, [4] marged_l2_broadcast_group_array'
            if next_arg == 'l2_broadcast_domain':
                for tmp4_result_get_l2_broadcast_domains in result_get_l2_broadcast_domains[4]:
                    kari_l2_array = []
                    for tmp1_result_get_l2_broadcast_domains in result_get_l2_broadcast_domains[1]:
                        if tmp1_result_get_l2_broadcast_domains != []:
                            for tmp1_tmp1_result_get_l2_broadcast_domains in tmp1_result_get_l2_broadcast_domains:
                                if tmp1_tmp1_result_get_l2_broadcast_domains[0] in tmp4_result_get_l2_broadcast_domains:
                                    kari_l2_array.append([tmp1_tmp1_result_get_l2_broadcast_domains[1],tmp1_tmp1_result_get_l2_broadcast_domains[2]])
                                    break
                    return_get_l2_broadcast_domains.append([tmp4_result_get_l2_broadcast_domains,kari_l2_array])

                return (return_get_l2_broadcast_domains)

            if next_arg == 'l3_broadcast_domain':
                retrun_target_l2_broadcast_group_array = []
                for tmp_target_l2_broadcast_group_array in self.target_l2_broadcast_group_array:
                    if tmp_target_l2_broadcast_group_array[1] != []:
                        retrun_target_l2_broadcast_group_array.append(tmp_target_l2_broadcast_group_array)

                return (self.target_l2_broadcast_group_array)

        if next_arg == 'attribute' or 'attribute_color':
            import ns_def
            l2_attribute_array = ns_def.convert_master_to_array('Master_Data', master_file_path,'<<ATTRIBUTE>>')

            update_attribute_array = []
            for tmp_attribute_array in l2_attribute_array:
                if tmp_attribute_array[0] != 1:
                    update_attribute_array.append(tmp_attribute_array[1][1:])

            update_padded_data = []

            for tmp_padded_data in update_attribute_array:
                del tmp_padded_data[-1]
                update_padded_data.append(tmp_padded_data)

            update_padded_data = sorted(update_attribute_array, key=lambda x: (x[0]), reverse=False)

            if next_arg == 'attribute_color':
                kari_return_attribute = []
                for row in update_padded_data:
                    transformed_row = [item.replace('<EMPTY>', '') for item in row]
                    kari_return_attribute.append(transformed_row)
                return (kari_return_attribute)

            output_data = []
            for line in update_padded_data:
                if isinstance(line[0], str) and not line[0].startswith('['):
                    # Add header directly to output
                    output_data.append(line)
                else:
                    # Extract the first element from each formatted string and replace '<EMPTY>' with ''
                    extracted_elements = [eval(element)[0] if eval(element)[0] != '<EMPTY>' else '' for element in line]
                    output_data.append(extracted_elements)

            if next_arg == 'attribute':
                return (output_data)


class def_common():

    @staticmethod
    def process_l1_link_common_batch(self, position_shape_array, style_shape_array, position_line_array,
                                     position_tag_array, master_file_path, ori_position_line_tuple,
                                     allow_shrink=False, position_folder_array=None, affected_hostnames=None):
        """
        ★★★ OPTIMIZED BATCH VERSION ★★★
        Process L1 links in batch mode (no Excel I/O during processing)
        - All calculations done in memory
        - No intermediate Excel writes

        Args:
            affected_hostnames: Set of hostnames that were affected by link additions

        Returns:
            Tuple of (updated position_line_array, updated style_shape_array)
        """
        import ns_def

        # Define helper functions (same as original)
        def count_interface_directions(hostname, position_shape_array, position_line_array):
            """Count the number of interfaces in each direction"""
            up_count = down_count = left_count = right_count = 0

            for item in position_line_array[2:]:
                connection_data = item[1]
                from_hostname = connection_data[0]
                to_hostname = connection_data[1]

                from_side = connection_data[4] if len(connection_data) > 4 else ''
                to_side = connection_data[5] if len(connection_data) > 5 else ''

                if from_side in ['RIGHT', 'LEFT'] and to_side in ['RIGHT', 'LEFT']:
                    direction = 'RIGHT_LEFT' if from_side == 'RIGHT' else 'LEFT_RIGHT'
                else:
                    direction = def_common.check_hostnames_in_same_element_static(
                        position_shape_array, from_hostname, to_hostname, position_folder_array
                    )

                if hostname == from_hostname:
                    if direction == 'RIGHT_LEFT':
                        right_count += 1
                    elif direction == 'LEFT_RIGHT':
                        left_count += 1
                    elif direction == 'UP_DOWN':
                        down_count += 1
                    elif direction == 'DOWN_UP':
                        up_count += 1

                elif hostname == to_hostname:
                    if direction == 'RIGHT_LEFT':
                        left_count += 1
                    elif direction == 'LEFT_RIGHT':
                        right_count += 1
                    elif direction == 'UP_DOWN':
                        up_count += 1
                    elif direction == 'DOWN_UP':
                        down_count += 1

            return [up_count, down_count, left_count, right_count]

        def calculate_offset_values(line_count, line_distance):
            """Calculate offset values for lines based on count"""
            if line_count == 0:
                return []
            elif line_count == 1:
                return [0]
            elif line_count % 2 == 0:
                half_distance = line_distance / 2
                offsets = []
                for i in range(line_count // 2):
                    offsets.insert(0, -(half_distance + i * line_distance))
                for i in range(line_count // 2):
                    offsets.append(half_distance + i * line_distance)
                return offsets
            else:
                offsets = [0]
                for i in range(1, (line_count + 1) // 2):
                    offsets.insert(0, -i * line_distance)
                    offsets.append(i * line_distance)
                return offsets

        def update_shape_dimensions(style_shape_array, hostname, position_shape_array, position_line_array,
                                    line_distance=0.2, margin=0.1, allow_shrink=False):
            """Update Width or Height in style_shape_array for a specific hostname"""
            shape_width_min = 0.4
            shape_hight_min = 0.2
            shae_font_size = 6.0
            normal_shape_width_margin = 0.1
            waypoint_shape_width_margin = 0.2
            normal_shape_height_margin = 0.1
            waypoint_shape_height_margin = 0.2
            waypoint_width_min = 0.5

            # Get interface direction counts
            direction_counts = count_interface_directions(hostname, position_shape_array, position_line_array)
            up_count, down_count, left_count, right_count = direction_counts

            # Find hostname in style_shape_array
            hostname_row_index = None
            is_waypoint = False

            for idx, item in enumerate(style_shape_array):
                if len(item) >= 2 and isinstance(item[1], list) and len(item[1]) > 0:
                    if item[1][0] == hostname:
                        hostname_row_index = idx
                        if len(item[1]) > 4 and item[1][4] == 'BLUE':
                            is_waypoint = True
                        break

            if hostname_row_index is None:
                for idx, item in enumerate(style_shape_array):
                    if len(item) >= 2 and isinstance(item[1], list) and len(item[1]) > 0:
                        if item[1][0] == '<DEFAULT>':
                            default_data = item[1]
                            new_number = max(row[0] for row in style_shape_array) + 1
                            new_entry = [new_number, [hostname, default_data[1], default_data[2], default_data[3], default_data[4]]]
                            style_shape_array.append(new_entry)
                            hostname_row_index = len(style_shape_array) - 1
                            break

            if hostname_row_index is None:
                return style_shape_array

            # Get current dimensions
            hostname_data = style_shape_array[hostname_row_index][1]

            try:
                current_width = float(hostname_data[1]) if hostname_data[1] != 'N/A' else shape_width_min
                current_height = float(hostname_data[2]) if hostname_data[2] != 'N/A' else shape_hight_min
            except (ValueError, IndexError):
                current_width = shape_width_min
                current_height = shape_hight_min

            # Calculate minimum dimensions based on hostname text length
            text_dimensions = ns_def.get_description_width_hight(shae_font_size, hostname)
            num_char_width = text_dimensions[0]
            num_char_height = text_dimensions[1]

            if is_waypoint:
                text_width_with_margin = num_char_width + waypoint_shape_width_margin
                text_height_with_margin = num_char_height + waypoint_shape_height_margin
            else:
                text_width_with_margin = num_char_width + normal_shape_width_margin
                text_height_with_margin = num_char_height + normal_shape_height_margin

            if is_waypoint:
                absolute_min_width = max(text_width_with_margin, waypoint_width_min)
            else:
                absolute_min_width = max(text_width_with_margin, shape_width_min)

            absolute_min_height = max(text_height_with_margin, shape_hight_min)

            # For LEFT/RIGHT connections, update HEIGHT
            max_lr_count = max(left_count, right_count)
            if max_lr_count > 0:
                offset_values = calculate_offset_values(max_lr_count, line_distance)
                if len(offset_values) > 0:
                    span = max(offset_values) - min(offset_values)
                    required_height = span + (waypoint_shape_height_margin if is_waypoint else normal_shape_height_margin)
                else:
                    required_height = waypoint_shape_height_margin if is_waypoint else normal_shape_height_margin

                final_height = max(required_height, absolute_min_height)

                if allow_shrink:
                    hostname_data[2] = final_height
                else:
                    if current_height < final_height:
                        hostname_data[2] = final_height
            else:
                if allow_shrink:
                    hostname_data[2] = absolute_min_height

            # For UP/DOWN connections, update WIDTH
            max_ud_count = max(up_count, down_count)
            if max_ud_count > 0:
                offset_values = calculate_offset_values(max_ud_count, line_distance)
                if len(offset_values) > 0:
                    span = max(offset_values) - min(offset_values)
                    required_width = span + (waypoint_shape_width_margin if is_waypoint else normal_shape_width_margin)
                else:
                    required_width = waypoint_shape_width_margin if is_waypoint else normal_shape_width_margin

                final_width = max(required_width, absolute_min_width)

                if allow_shrink:
                    hostname_data[1] = final_width
                else:
                    if current_width < final_width:
                        hostname_data[1] = final_width
            else:
                if allow_shrink:
                    hostname_data[1] = absolute_min_width

            return style_shape_array

        def determine_line_order(position_line_array, position_shape_array):
            """Determine the order of lines based on device positions"""
            hostname_position = {}
            for row_idx, item in enumerate(position_shape_array):
                if len(item) >= 2 and isinstance(item[1], list):
                    for col_idx, hostname in enumerate(item[1]):
                        if hostname not in ['_AIR_', '<END>', '', '<<POSITION_SHAPE>>']:
                            if hostname not in hostname_position:
                                hostname_position[hostname] = (row_idx, col_idx)

            header_rows = position_line_array[:2]
            data_rows = position_line_array[2:]

            def sort_key(row):
                if len(row) < 2 or not isinstance(row[1], list):
                    return (999, 999, 999, 999, 999, 999)

                data = row[1]
                from_hostname = data[0] if len(data) > 0 else ''
                to_hostname = data[1] if len(data) > 1 else ''

                from_pos = hostname_position.get(from_hostname, (999, 999))
                to_pos = hostname_position.get(to_hostname, (999, 999))

                direction = 'UNKNOWN'
                if from_hostname and to_hostname:
                    direction = def_common.check_hostnames_in_same_element_static(
                        position_shape_array, from_hostname, to_hostname, position_folder_array
                    )

                if direction in ['UP_DOWN', 'DOWN_UP']:
                    if from_pos[0] < to_pos[0]:
                        key = (to_pos[0], to_pos[1], 0, from_pos[1], row[0])
                    else:
                        key = (from_pos[0], from_pos[1], 1, to_pos[1], to_pos[0], row[0])
                elif direction in ['LEFT_RIGHT', 'RIGHT_LEFT']:
                    if from_pos[1] < to_pos[1]:
                        key = (from_pos[0], from_pos[1], 2, to_pos[1], row[0])
                    else:
                        key = (from_pos[0], to_pos[1], 3, from_pos[1], row[0])
                else:
                    key = (from_pos[0], from_pos[1], 4, to_pos[0], to_pos[1], row[0])

                return key

            sorted_data_rows = sorted(data_rows, key=sort_key)
            for idx, row in enumerate(sorted_data_rows, start=3):
                row[0] = idx

            return header_rows + sorted_data_rows

        def get_tag_offset(hostname, master_line_tuple, tag_offet_inche):
            """Get tag offset value based on maximum tag name length"""
            max_len_num = 1
            for tmp in master_line_tuple:
                if master_line_tuple.get(tmp) == hostname and tmp[1] in [1, 2]:
                    tmp_len_char = len(master_line_tuple[tmp[0], 3 if tmp[1] == 1 else 4])
                    if max_len_num < tmp_len_char:
                        max_len_num = tmp_len_char
            return max_len_num * tag_offet_inche

        def update_position_tag_array_single(position_tag_array, hostname, offset_value):
            """Update Offset_LINE value for a single hostname"""
            for item in position_tag_array:
                if len(item) >= 2 and isinstance(item[1], list) and len(item[1]) >= 5:
                    row_data = item[1]
                    if row_data[0] == hostname:
                        try:
                            current_offset = float(row_data[4]) if row_data[4] != '' else 0.0
                            if current_offset < offset_value:
                                row_data[4] = offset_value
                        except (ValueError, TypeError):
                            row_data[4] = offset_value
            return position_tag_array

        # ========== MAIN PROCESSING (ALL IN MEMORY) ==========

        # Sort connections
        position_line_array = determine_line_order(position_line_array, position_shape_array)

        # Determine which hostnames to process
        if affected_hostnames is not None:
            hostnames_to_process = affected_hostnames

            # Clear offsets only for affected hostnames
            for item in position_line_array[2:]:
                connection_data = item[1]
                if connection_data[0] in hostnames_to_process or connection_data[1] in hostnames_to_process:
                    connection_data[6] = ''
                    connection_data[7] = ''
                    connection_data[8] = ''
                    connection_data[9] = ''
        else:
            # Process all devices
            unique_hostnames = set()
            for item in position_line_array[2:]:
                connection_data = item[1]
                unique_hostnames.add(connection_data[0])
                unique_hostnames.add(connection_data[1])
            hostnames_to_process = unique_hostnames

            # Clear all offsets
            for item in position_line_array[2:]:
                connection_data = item[1]
                connection_data[6] = ''
                connection_data[7] = ''
                connection_data[8] = ''
                connection_data[9] = ''

        # Recalculate offsets
        line_distance = 0.2

        print(f'[Info] Processing {len(hostnames_to_process)} affected devices...')

        for progress_idx, current_hostname in enumerate(sorted(hostnames_to_process)):
            # Progress indicator
            if (progress_idx + 1) % 20 == 0:
                print(f'[Info] Processed {progress_idx + 1}/{len(hostnames_to_process)} devices...')

            # Build hostname position mapping
            hostname_position = {}
            for row_idx, item in enumerate(position_shape_array):
                if len(item) >= 2 and isinstance(item[1], list):
                    for col_idx, hostname_iter in enumerate(item[1]):
                        if hostname_iter not in ['_AIR_', '<END>', '', '<<POSITION_SHAPE>>']:
                            if hostname_iter not in hostname_position:
                                hostname_position[hostname_iter] = (row_idx, col_idx)

            # Group connections by direction
            direction_groups = {'UP': [], 'DOWN': [], 'LEFT': [], 'RIGHT': []}

            for sort_index, item in enumerate(position_line_array[2:]):
                connection_data = item[1]
                from_h = connection_data[0]
                to_h = connection_data[1]

                if current_hostname == from_h or current_hostname == to_h:
                    from_side = connection_data[4] if len(connection_data) > 4 else ''
                    to_side = connection_data[5] if len(connection_data) > 5 else ''

                    if from_side in ['RIGHT', 'LEFT'] and to_side in ['RIGHT', 'LEFT']:
                        conn_direction = 'RIGHT_LEFT' if from_side == 'RIGHT' else 'LEFT_RIGHT'
                    else:
                        conn_direction = def_common.check_hostnames_in_same_element_static(
                            position_shape_array, from_h, to_h, position_folder_array
                        )

                    other_hostname = to_h if current_hostname == from_h else from_h
                    other_pos = hostname_position.get(other_hostname, (999, 999))

                    # Classify into direction group
                    if current_hostname == from_h:
                        if conn_direction == 'RIGHT_LEFT':
                            direction_groups['RIGHT'].append((other_pos[0], other_pos[1], sort_index, item))
                        elif conn_direction == 'LEFT_RIGHT':
                            direction_groups['LEFT'].append((other_pos[0], other_pos[1], sort_index, item))
                        elif conn_direction == 'UP_DOWN':
                            direction_groups['DOWN'].append((other_pos[0], other_pos[1], sort_index, item))
                        elif conn_direction == 'DOWN_UP':
                            direction_groups['UP'].append((other_pos[0], other_pos[1], sort_index, item))
                    elif current_hostname == to_h:
                        if conn_direction == 'RIGHT_LEFT':
                            direction_groups['LEFT'].append((other_pos[0], other_pos[1], sort_index, item))
                        elif conn_direction == 'LEFT_RIGHT':
                            direction_groups['RIGHT'].append((other_pos[0], other_pos[1], sort_index, item))
                        elif conn_direction == 'UP_DOWN':
                            direction_groups['UP'].append((other_pos[0], other_pos[1], sort_index, item))
                        elif conn_direction == 'DOWN_UP':
                            direction_groups['DOWN'].append((other_pos[0], other_pos[1], sort_index, item))

            # Calculate and assign offsets for EACH direction independently
            for direction, indexed_items in direction_groups.items():
                if len(indexed_items) == 0:
                    continue

                indexed_items.sort(key=lambda x: x[1])
                items = [x[3] for x in indexed_items]
                offset_values = calculate_offset_values(len(items), line_distance)

                for idx, item in enumerate(items):
                    if idx < len(offset_values):
                        connection_data = item[1]
                        from_h = connection_data[0]
                        offset_val = offset_values[idx]

                        if direction in ['LEFT', 'RIGHT']:
                            index_to_set = 7 if current_hostname == from_h else 9
                            connection_data[index_to_set] = offset_val
                        else:
                            index_to_set = 6 if current_hostname == from_h else 8
                            connection_data[index_to_set] = offset_val

        # Update shape dimensions
        shape_margin = 0.1
        for affected_host in hostnames_to_process:
            style_shape_array = update_shape_dimensions(
                style_shape_array, affected_host, position_shape_array,
                position_line_array, line_distance, shape_margin, allow_shrink
            )

        # Update position_tag_array
        tag_offet_inche = 0.02
        position_line_tuple = ns_def.convert_array_to_tuple(position_line_array)

        for target_hostname in hostnames_to_process:
            offset_value = get_tag_offset(target_hostname, position_line_tuple, tag_offet_inche)
            update_position_tag_array_single(position_tag_array, target_hostname, offset_value)

        print('[Info] Offset calculation completed.')

        return position_line_array, style_shape_array



    @staticmethod
    def create_empty_master_file(output_file_path):
        """
        Create an empty master Excel file with no areas or devices
        Contains 3 sheets: Master_Data, Master_Data_L2, Master_Data_L3

        Args:
            output_file_path: Path where the empty master file will be created

        Returns:
            dict: Status and message of the operation
        """
        import openpyxl
        from openpyxl import Workbook

        try:
            # Create new workbook
            wb = Workbook()

            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # ========== Create Master_Data sheet ==========
            ws = wb.create_sheet('Master_Data', 0)

            # Define empty master data structure
            empty_master_data = [
                # <<ROOT_FOLDER>>
                ['<<ROOT_FOLDER>>', 'Title Text', 'ratio_x(0.1-1.00)', 'ratio_y(0.1-1.00)', 'left(inches)', 'top(inches)', 'width(inches)', 'hight(inches)'],
                ['', '[L1] All Areas', 1, 1, 1, 1, 1, 1],
                [''],
                [''],
                [''],

                # <<POSITION_FOLDER>>
                ['<<POSITION_FOLDER>>'],
                [''],
                [''],
                [''],
                [''],

                # <<STYLE_FOLDER>>
                ['<<STYLE_FOLDER>>', 'Outline(YES/NO)', 'Text(NO/UP/DOWN)', 'Offset Upside Margin(<AUTO> or inches)', 'Offset Downside Margin(<AUTO> or inches)'],
                ['<DEFAULT>', 'YES', 'N/A', 'N/A', 'N/A'],
                ['<EMPTY>', 'NO', 'N/A', 'N/A', 'N/A'],
                [''],
                [''],
                [''],

                # <<POSITION_SHAPE>>
                ['<<POSITION_SHAPE>>'],
                [''],
                [''],
                [''],

                # <<STYLE_SHAPE>>
                ['<<STYLE_SHAPE>>', 'Width(Inches)', 'Hight(Inches)', 'Roundness(0.0-1.0)', 'Color(ORANGE/BLUE/GREEN/GRAY'],
                ['<DEFAULT>', 0.5, 0.3, 0, 'N/A'],
                ['<EMPTY>', 0.5, 0.3, 0, 'N/A'],
                [''],
                [''],
                [''],

                # <<POSITION_LINE>>
                ['<<POSITION_LINE>>'],
                ['From_Name', 'To_Name', 'From_Tag_Name', 'To_Tag_Name', 'From_Side(RIGHT/LEFT)', 'To_Side(RIGHT/LEFT)',
                 'Offset From_X (inches)', 'Offset From_Y (inches)', 'Offset To_X (inches)', 'Offset To_Y (inches)',
                 'Channel(inches)', 'Color(No) only)', 'From_Port_Name', 'From_Speed', 'From_Duplex', 'From_Port_Type',
                 'To__Port_Name', 'To_Speed', 'To_From_Duplex', 'To_From_Port_Type'],
                [''],
                [''],
                [''],

                # <<POSITION_TAG>>
                ['<<POSITION_TAG>>', 'Type(SHAPE/LINE)', 'Offset_SHAPE_X', 'Offset_SHAPE_Y', 'Offset_LINE(inches)', 'Adjust_LINE_Angle(YES/NO)'],
                ['<DEFAULT>', 'SHAPE', 0, 0, 0.3, 'YES'],
                [''],
                [''],
                [''],

                # <<ATTRIBUTE>>
                ['<<ATTRIBUTE>>'],
                ['Device Name', 'Default', 'Attribute-A', 'Attribute-B', 'Attribute-C', 'Attribute-D',
                 'Attribute-E', 'Attribute-F', 'Attribute-G', 'Attribute-H', '<END>'],
                [''],
                [''],
                [''],

                # <<END_MARK>>
                ['<<END_MARK>>'],
            ]

            # Write data to Master_Data worksheet
            for row_idx, row_data in enumerate(empty_master_data, start=1):
                for col_idx, cell_value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=cell_value)

            # ========== Create Master_Data_L2 sheet ==========
            ws_l2 = wb.create_sheet('Master_Data_L2')
            empty_l2_data = [
                ['<<L2_TABLE>>'],
                ['Area', 'Device Name', 'Port Mode', 'Port Name', 'Virtual Port Mode', 'Virtual Port Name',
                 'Connected L2 Segment Name(Comma Separated)', 'L2 Name directly received by L3 Virtual Port (Comma Separated)'],
                [''],
                [''],
                ['<<END_MARK>>'],
            ]

            for row_idx, row_data in enumerate(empty_l2_data, start=1):
                for col_idx, cell_value in enumerate(row_data, start=1):
                    ws_l2.cell(row=row_idx, column=col_idx, value=cell_value)

            # ========== Create Master_Data_L3 sheet ==========
            ws_l3 = wb.create_sheet('Master_Data_L3')
            empty_l3_data = [
                ['<<L3_TABLE>>'],
                ['Area', 'Device Name', 'L3 IF Name', 'L3 Instance Name', 'IP Address / Subnet mask (Comma Separated)',
                 '[VPN] Target Device Name (Comma Separated)', '[VPN] Target L3 Port Name (Comma Separated)'],
                [''],
                [''],
                ['<<END_MARK>>'],
            ]

            for row_idx, row_data in enumerate(empty_l3_data, start=1):
                for col_idx, cell_value in enumerate(row_data, start=1):
                    ws_l3.cell(row=row_idx, column=col_idx, value=cell_value)

            # Save workbook
            wb.save(output_file_path)

            return {
                'status': 'success',
                'message': f'Empty master file created: {output_file_path}'
            }

        except Exception as e:
            import traceback
            traceback.print_exc()
            return {
                'status': 'error',
                'message': f'Failed to create empty master file: {str(e)}'
            }

    @staticmethod
    def add_l1_links_bulk(self, link_definitions, master_file_path):
        """
        ★★★ CORRECTED OPTIMIZED VERSION ★★★

        Fix: Use existing Excel write functions instead of custom implementation

        Performance: 15 links in ~20s, 256 links in ~3min
        """
        import ns_def
        import time

        try:
            start_time = time.time()
            total_links = len(link_definitions)
            print(f'\n{"=" * 70}')
            print(f'OPTIMIZED L1 LINK BULK ADDITION')
            print(f'{"=" * 70}')
            '''print(f'Total links to add: {total_links}')
            print(f'Master file: {os.path.basename(master_file_path)}')
            print(f'{"=" * 70}\n')'''

            if not link_definitions:
                return ([f'[ERROR] No link definitions provided'])

            # ==================== STEP 1: Load all arrays ====================
            print('[Step 1/8] Loading Excel data...')
            step_start = time.time()

            position_folder_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_FOLDER>>')
            position_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_SHAPE>>')
            style_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<STYLE_SHAPE>>')
            position_line_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_LINE>>')
            position_tag_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_TAG>>')
            ori_position_line_tuple = ns_def.convert_array_to_tuple(position_line_array)

            #print(f'  ✓ Data loaded in {time.time() - step_start:.2f}s\n')

            # Track used ports
            batch_used_ports = {}

            # Helper functions
            def check_duplicate_port(position_line_array, from_hostname, to_hostname, from_port, to_port):
                for item in position_line_array[2:]:
                    if len(item) < 2 or not isinstance(item[1], list) or len(item[1]) < 4:
                        continue

                    rd = item[1]
                    if (rd[0] == from_hostname and rd[2] == from_port) or \
                            (rd[1] == to_hostname and rd[3] == to_port) or \
                            (rd[0] == to_hostname and rd[2] == to_port) or \
                            (rd[1] == from_hostname and rd[3] == from_port):
                        return True
                return False

            def check_batch_duplicate(batch_used_ports, from_hostname, to_hostname, from_port, to_port, idx):
                for key in [(from_hostname, from_port), (to_hostname, to_port)]:
                    if key in batch_used_ports:
                        return (True, batch_used_ports[key])
                return (False, None)

            def add_link_to_array(position_line_array, from_h, to_h, from_p, to_p, direction, from_full, to_full):
                number = max([item[0] for item in position_line_array], default=2) + 1

                if direction == 'RIGHT_LEFT':
                    sides = ['RIGHT', 'LEFT']
                elif direction == 'LEFT_RIGHT':
                    sides = ['LEFT', 'RIGHT']
                else:
                    sides = ['', '']

                new_elem = [number, [
                    from_h, to_h, from_p, to_p,
                    sides[0], sides[1], '', '', '', '', '', '',
                    from_full, 'N/A', 'N/A', 'N/A',
                    to_full, 'N/A', 'N/A', 'N/A'
                ]]

                position_line_array.append(new_elem)
                return position_line_array

            # ==================== STEP 2: Add all links (IN MEMORY) ====================
            print(f'[Step 2/8] Processing {total_links} links...')
            step_start = time.time()

            added_links = []
            error_links = []
            all_affected_hostnames = set()

            for link_idx, link_def in enumerate(link_definitions):
                if (link_idx + 1) % 100 == 0:
                    elapsed = time.time() - step_start
                    rate = (link_idx + 1) / elapsed if elapsed > 0 else 0
                    remaining = (total_links - link_idx - 1) / rate if rate > 0 else 0
                    #print(f'  Progress: {link_idx + 1}/{total_links} - Est. {remaining:.0f}s remaining')

                # Validate
                if not isinstance(link_def, list) or len(link_def) != 4:
                    error_links.append({'idx': link_idx, 'err': 'Invalid format'})
                    continue

                from_h = str(link_def[0])
                to_h = str(link_def[1])
                from_pn = str(link_def[2])
                to_pn = str(link_def[3])

                if from_h > to_h:
                    from_h, to_h = to_h, from_h
                    from_pn, to_pn = to_pn, from_pn

                # Validate ports
                from_split = ns_def.split_portname(from_pn)
                to_split = ns_def.split_portname(to_pn)

                if not from_split[0] or not from_split[1]:
                    error_links.append({'idx': link_idx, 'err': 'Invalid from_port'})
                    continue

                if not to_split[0] or not to_split[1]:
                    error_links.append({'idx': link_idx, 'err': 'Invalid to_port'})
                    continue

                # Prepare
                from_adj = ns_def.adjust_portname(from_pn)
                to_adj = ns_def.adjust_portname(to_pn)

                from_p = f"{from_adj[0]} {from_split[1]}"
                to_p = f"{to_adj[0]} {to_split[1]}"
                from_full = from_split[0]
                to_full = to_split[0]

                # Check position
                result_dir = def_common.check_hostnames_in_same_element_static(
                    position_shape_array, from_h, to_h, position_folder_array
                )

                if isinstance(result_dir, str) and result_dir.startswith("Error:"):
                    error_links.append({'idx': link_idx, 'err': 'Invalid position'})
                    continue

                # Check duplicates
                if check_duplicate_port(position_line_array, from_h, to_h, from_p, to_p):
                    error_links.append({'idx': link_idx, 'err': 'Port used'})
                    continue

                is_dup, dup_idx = check_batch_duplicate(batch_used_ports, from_h, to_h, from_p, to_p, link_idx)
                if is_dup:
                    error_links.append({'idx': link_idx, 'err': f"Conflicts with Link {dup_idx + 1}"})
                    continue

                # Add
                position_line_array = add_link_to_array(
                    position_line_array, from_h, to_h, from_p, to_p, result_dir, from_full, to_full
                )

                batch_used_ports[(from_h, from_p)] = link_idx
                batch_used_ports[(to_h, to_p)] = link_idx
                all_affected_hostnames.add(from_h)
                all_affected_hostnames.add(to_h)
                added_links.append({'from': from_h, 'to': to_h})

            #print(f'  ✓ Added {len(added_links)} links in {time.time() - step_start:.2f}s\n')

            if error_links:
                error_msg = f'[ERROR] {len(error_links)} link(s) failed:\n'
                for err in error_links[:5]:
                    error_msg += f"  Link {err['idx'] + 1}: {err['err']}\n"
                if len(error_links) > 5:
                    error_msg += f"  ... and {len(error_links) - 5} more\n"
                return ([error_msg])

            # ==================== STEP 3: Sort ====================
            print('[Step 3/8] Sorting connections...')
            step_start = time.time()

            hostname_position = {}
            for row_idx, item in enumerate(position_shape_array):
                if len(item) >= 2 and isinstance(item[1], list):
                    for col_idx, hostname in enumerate(item[1]):
                        if hostname not in ['_AIR_', '<END>', '', '<<POSITION_SHAPE>>']:
                            hostname_position[hostname] = (row_idx, col_idx)

            header_rows = position_line_array[:2]
            data_rows = position_line_array[2:]

            def sort_key(row):
                if len(row) < 2 or not isinstance(row[1], list) or len(row[1]) < 2:
                    return (999, 999, 999, 999)
                fp = hostname_position.get(row[1][0], (999, 999))
                tp = hostname_position.get(row[1][1], (999, 999))
                return (fp[0], fp[1], tp[0], tp[1])

            data_rows.sort(key=sort_key)
            for idx, row in enumerate(data_rows, start=3):
                row[0] = idx

            position_line_array = header_rows + data_rows

            #print(f'  ✓ Sorted in {time.time() - step_start:.2f}s\n')

            # ==================== STEP 4: Calculate offsets ====================
            print(f'[Step 4/8] Calculating offsets for {len(all_affected_hostnames)} devices...')
            step_start = time.time()

            line_distance = 0.2

            def calc_offsets(count, dist):
                if count == 0: return []
                if count == 1: return [0]
                if count % 2 == 0:
                    half = dist / 2
                    result = []
                    for i in range(count // 2):
                        result.insert(0, -(half + i * dist))
                        result.append(half + i * dist)
                    return result
                else:
                    result = [0]
                    for i in range(1, (count + 1) // 2):
                        result.insert(0, -i * dist)
                        result.append(i * dist)
                    return result

            # Clear offsets
            for item in position_line_array[2:]:
                cd = item[1]
                if cd[0] in all_affected_hostnames or cd[1] in all_affected_hostnames:
                    cd[6] = cd[7] = cd[8] = cd[9] = ''

            # Calculate per device
            for proc_idx, hostname in enumerate(sorted(all_affected_hostnames)):
                #if (proc_idx + 1) % 50 == 0:
                    #print(f'  Progress: {proc_idx + 1}/{len(all_affected_hostnames)} devices')

                dir_groups = {'UP': [], 'DOWN': [], 'LEFT': [], 'RIGHT': []}

                for item in position_line_array[2:]:
                    cd = item[1]
                    if hostname not in [cd[0], cd[1]]:
                        continue

                    from_side = cd[4] if len(cd) > 4 else ''
                    to_side = cd[5] if len(cd) > 5 else ''

                    if from_side in ['RIGHT', 'LEFT'] and to_side in ['RIGHT', 'LEFT']:
                        conn_dir = 'RIGHT_LEFT' if from_side == 'RIGHT' else 'LEFT_RIGHT'
                    else:
                        conn_dir = def_common.check_hostnames_in_same_element_static(
                            position_shape_array, cd[0], cd[1], position_folder_array
                        )

                    other = cd[1] if hostname == cd[0] else cd[0]
                    other_pos = hostname_position.get(other, (999, 999))

                    if hostname == cd[0]:
                        if conn_dir == 'RIGHT_LEFT':
                            dir_groups['RIGHT'].append((other_pos[1], item))
                        elif conn_dir == 'LEFT_RIGHT':
                            dir_groups['LEFT'].append((other_pos[1], item))
                        elif conn_dir == 'UP_DOWN':
                            dir_groups['DOWN'].append((other_pos[1], item))
                        elif conn_dir == 'DOWN_UP':
                            dir_groups['UP'].append((other_pos[1], item))
                    else:
                        if conn_dir == 'RIGHT_LEFT':
                            dir_groups['LEFT'].append((other_pos[1], item))
                        elif conn_dir == 'LEFT_RIGHT':
                            dir_groups['RIGHT'].append((other_pos[1], item))
                        elif conn_dir == 'UP_DOWN':
                            dir_groups['UP'].append((other_pos[1], item))
                        elif conn_dir == 'DOWN_UP':
                            dir_groups['DOWN'].append((other_pos[1], item))

                for direction, items_pos in dir_groups.items():
                    if not items_pos:
                        continue

                    items_pos.sort(key=lambda x: x[0])
                    items = [x[1] for x in items_pos]
                    offsets = calc_offsets(len(items), line_distance)

                    for i, item in enumerate(items):
                        if i >= len(offsets):
                            continue

                        cd = item[1]
                        off = offsets[i]

                        if direction in ['LEFT', 'RIGHT']:
                            idx_set = 7 if hostname == cd[0] else 9
                            cd[idx_set] = off
                        else:
                            idx_set = 6 if hostname == cd[0] else 8
                            cd[idx_set] = off

            #print(f'  ✓ Offsets calculated in {time.time() - step_start:.2f}s\n')

            # ==================== STEP 5: Update shapes ====================
            print(f'[Step 5/8] Updating {len(all_affected_hostnames)} device shapes...')
            step_start = time.time()

            w_min, h_min = 0.4, 0.2
            wp_w_min = 0.5
            font_sz = 6.0

            for hostname in all_affected_hostnames:
                # Count connections
                up = down = left = right = 0

                for item in position_line_array[2:]:
                    cd = item[1]
                    if hostname not in [cd[0], cd[1]]:
                        continue

                    fs = cd[4] if len(cd) > 4 else ''
                    ts = cd[5] if len(cd) > 5 else ''

                    if fs in ['RIGHT', 'LEFT'] and ts in ['RIGHT', 'LEFT']:
                        d = 'RIGHT_LEFT' if fs == 'RIGHT' else 'LEFT_RIGHT'
                    else:
                        d = def_common.check_hostnames_in_same_element_static(
                            position_shape_array, cd[0], cd[1], position_folder_array
                        )

                    if hostname == cd[0]:
                        if d == 'RIGHT_LEFT':
                            right += 1
                        elif d == 'LEFT_RIGHT':
                            left += 1
                        elif d == 'UP_DOWN':
                            down += 1
                        elif d == 'DOWN_UP':
                            up += 1
                    else:
                        if d == 'RIGHT_LEFT':
                            left += 1
                        elif d == 'LEFT_RIGHT':
                            right += 1
                        elif d == 'UP_DOWN':
                            up += 1
                        elif d == 'DOWN_UP':
                            down += 1

                # Find device
                dev_idx = None
                is_wp = False

                for idx, item in enumerate(style_shape_array):
                    if len(item) >= 2 and isinstance(item[1], list) and len(item[1]) > 0:
                        if item[1][0] == hostname:
                            dev_idx = idx
                            if len(item[1]) > 4 and item[1][4] == 'BLUE':
                                is_wp = True
                            break

                if dev_idx is None:
                    continue

                dev_data = style_shape_array[dev_idx][1]

                # Calculate dimensions
                text_dims = ns_def.get_description_width_hight(font_sz, hostname)
                margin = 0.2 if is_wp else 0.1
                min_w = max(text_dims[0] + margin, wp_w_min if is_wp else w_min)
                min_h = max(text_dims[1] + margin, h_min)

                curr_w = float(dev_data[1]) if dev_data[1] != 'N/A' else min_w
                curr_h = float(dev_data[2]) if dev_data[2] != 'N/A' else min_h

                # Update HEIGHT
                max_lr = max(left, right)
                if max_lr > 0:
                    offs = calc_offsets(max_lr, line_distance)
                    if offs:
                        span = max(offs) - min(offs)
                        req_h = max(span + margin, min_h)
                        if curr_h < req_h:
                            dev_data[2] = req_h

                # Update WIDTH
                max_ud = max(up, down)
                if max_ud > 0:
                    offs = calc_offsets(max_ud, line_distance)
                    if offs:
                        span = max(offs) - min(offs)
                        req_w = max(span + margin, min_w)
                        if curr_w < req_w:
                            dev_data[1] = req_w

            # Update tags
            pl_tuple = ns_def.convert_array_to_tuple(position_line_array)
            tag_off_inch = 0.02

            for hostname in all_affected_hostnames:
                max_len = 1
                for key in pl_tuple:
                    if pl_tuple.get(key) == hostname and key[1] in [1, 2]:
                        tag_name = pl_tuple.get((key[0], 3 if key[1] == 1 else 4), '')
                        max_len = max(max_len, len(str(tag_name)))

                off_val = max_len * tag_off_inch

                for item in position_tag_array:
                    if len(item) >= 2 and isinstance(item[1], list) and len(item[1]) >= 5:
                        if item[1][0] == hostname:
                            curr = float(item[1][4]) if item[1][4] != '' else 0.0
                            if curr < off_val:
                                item[1][4] = off_val

            #print(f'  ✓ Shapes updated in {time.time() - step_start:.2f}s\n')

            # ==================== STEP 6: Write to Excel (using standard functions) ====================
            print('[Step 6/8] Writing to Excel...')
            step_start = time.time()

            # ★★★ FIX: Use existing Excel write functions ★★★

            # Convert to tuples
            position_line_tuple = ns_def.convert_array_to_tuple(position_line_array)
            style_shape_tuple = ns_def.convert_array_to_tuple(style_shape_array)
            position_tag_tuple = ns_def.convert_array_to_tuple(position_tag_array)

            # Write POSITION_LINE
            ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_position_line_tuple)
            ns_def.write_excel_meta(position_line_tuple, master_file_path, 'Master_Data', '<<POSITION_LINE>>', 0, 0)

            # Write STYLE_SHAPE
            ori_style_shape_tuple = ns_def.convert_array_to_tuple(
                ns_def.convert_master_to_array('Master_Data', master_file_path, '<<STYLE_SHAPE>>')
            )
            ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_style_shape_tuple)
            ns_def.write_excel_meta(style_shape_tuple, master_file_path, 'Master_Data', '<<STYLE_SHAPE>>', 0, 0)

            # Write POSITION_TAG
            ori_position_tag_tuple = ns_def.convert_array_to_tuple(
                ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_TAG>>')
            )
            ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_position_tag_tuple)
            ns_def.write_excel_meta(position_tag_tuple, master_file_path, 'Master_Data', '<<POSITION_TAG>>', 0, 0)

            #print(f'  ✓ Excel write completed in {time.time() - step_start:.2f}s\n')

            # ==================== STEP 7: Update POSITION_FOLDER ====================
            print('[Step 7/8] Updating POSITION_FOLDER...')
            step_start = time.time()

            pf_arr = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_FOLDER>>')
            ps_arr = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_SHAPE>>')
            ss_arr = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<STYLE_SHAPE>>')

            upd_pf_tuple = def_common.make_position_folder_tuple(
                ns_def.convert_array_to_tuple(pf_arr),
                ns_def.convert_array_to_tuple(ss_arr),
                ns_def.convert_array_to_tuple(ps_arr)
            )

            ori_pf = ns_def.convert_array_to_tuple(
                ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_FOLDER>>')
            )
            ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_pf)
            ns_def.write_excel_meta(upd_pf_tuple, master_file_path, 'Master_Data', '<<POSITION_FOLDER>>', 0, 0)

            res = def_common.recalculate_folder_sizes(master_file_path)
            if res['status'] != 'success':
                return ([f"[ERROR] Recalc failed: {res['message']}"])

            #print(f'  ✓ POSITION_FOLDER updated in {time.time() - step_start:.2f}s\n')

            # ==================== STEP 8: L2/L3 SYNC (ONCE) ====================
            print('[Step 8/8] Syncing with L2/L3 layers (may take 60s~)...')
            sync_start = time.time()

            import tkinter as tk
            dummy_tk = tk.Toplevel()
            dummy_tk.withdraw()

            self.full_filepath = master_file_path
            self.main1_1_entry_1 = tk.Entry(dummy_tk)
            self.main1_1_entry_1.insert(tk.END, master_file_path)

            self.inFileTxt_L3_1_1 = tk.Entry(dummy_tk)
            self.inFileTxt_L3_1_1.insert(tk.END, master_file_path)

            self.outFileTxt_11_2 = tk.Entry(dummy_tk)
            self.outFileTxt_11_2.insert(tk.END, master_file_path)

            self.inFileTxt_L2_1_1 = tk.Entry(dummy_tk)
            self.inFileTxt_L2_1_1.insert(tk.END, master_file_path)

            import ns_sync_between_layers
            ns_sync_between_layers.l1_master_device_and_line_sync_with_l2l3_master(self)

            tmp_del = master_file_path.replace('[MASTER]', '[L2_TABLE]')
            if os.path.isfile(tmp_del):
                os.remove(tmp_del)

            dummy_tk.destroy()

            sync_time = time.time() - sync_start
            #print(f'  ✓ L2/L3 sync completed in {sync_time:.1f}s\n')

            # ==================== FINAL REPORT ====================
            total_time = time.time() - start_time

            #print(f'{"=" * 70}')
            #print(f'BATCH ADDITION COMPLETE')
            #print(f'{"=" * 70}')

            return_msg = f'Links added: {len(added_links)}\n'
            return_msg += f'Total time: {total_time:.1f}s ({total_time / 60:.2f} min)\n'
            return_msg += f'  • Processing: {total_time - sync_time:.1f}s\n'
            return_msg += f'  • L2/L3 sync: {sync_time:.1f}s\n'
            return_msg += f'Average: {total_time / len(added_links):.3f}s/link\n'
            return_msg += f'Affected devices: {len(all_affected_hostnames)}\n\n'

            return_msg += 'Sample links:\n'
            for lnk in added_links[:3]:
                return_msg += f"  {lnk['from']} ↔ {lnk['to']}\n"
            if len(added_links) > 3:
                return_msg += f"  ... +{len(added_links) - 3} more\n"

            #print(return_msg)
            #print(f'{"=" * 70}\n')

            return ([return_msg])

        except Exception as e:
            import traceback
            traceback.print_exc()
            return ([f'[ERROR] Bulk addition failed: {str(e)}'])


    @staticmethod
    def update_device_location_batch(self, area_name, device_grid, master_file_path):
        """
        ★★★ OPTIMIZED BATCH VERSION ★★★
        Update device locations within a specific area (replace all devices)
        - Opens Excel ONCE at the beginning
        - Processes all devices in memory
        - Saves Excel ONCE at the end
        - Calls L2/L3 sync ONCE at the end

        Args:
            self: Reference to calling instance
            area_name: Name of the area to update
            device_grid: List of lists containing device names in desired grid layout
            master_file_path: Path to the master Excel file

        Returns:
            List with status message or error
        """
        import ns_def
        import openpyxl

        try:
            print(f'[Info] Starting batch device location update for area "{area_name}"...')

            # ★★★ STEP 0: Open Excel ONCE ★★★
            print('[Info] Opening Excel file...')
            wb = openpyxl.load_workbook(master_file_path)

            # Helper function to normalize and align device grid
            def normalize_device_grid(device_grid):
                """
                Normalize device grid:
                1. Center-align devices in each row (place devices in center, fill sides with _AIR_)
                2. Fill empty spaces with _AIR_
                3. Ensure all rows have the same width
                """
                if not device_grid:
                    return []

                AIR_MARK = "_AIR_"
                normalized_grid = []

                # Step 1: Process each row - clean and collect devices
                processed_rows = []
                for row in device_grid:
                    if not isinstance(row, list):
                        continue

                    # Filter out None and empty strings, keep actual devices and _AIR_ markers
                    cleaned_row = []
                    for device in row:
                        if device is None or device == '':
                            # Skip empty cells
                            continue
                        else:
                            cleaned_row.append(device)

                    # Only add non-empty rows
                    if cleaned_row:
                        processed_rows.append(cleaned_row)

                if not processed_rows:
                    return []

                # Step 2: Find maximum row width
                max_width = 0
                for row in processed_rows:
                    if len(row) > max_width:
                        max_width = len(row)

                # Step 3: Center-align each row and pad with _AIR_
                for row in processed_rows:
                    current_width = len(row)

                    if current_width < max_width:
                        # Calculate padding needed
                        total_padding = max_width - current_width
                        left_padding = total_padding // 2
                        right_padding = total_padding - left_padding

                        # Create centered row
                        centered_row = [AIR_MARK] * left_padding + row + [AIR_MARK] * right_padding
                        normalized_grid.append(centered_row)
                    else:
                        # Row is already max width
                        normalized_grid.append(row)

                return normalized_grid

            # Validate that area exists
            style_folder_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<STYLE_FOLDER>>')
            area_exists = False
            for item in style_folder_array:
                if item[0] not in [1, 2, 3] and len(item[1]) > 0:
                    if item[1][0] == area_name:
                        area_exists = True
                        break

            if not area_exists:
                wb.close()
                return ([f'[ERROR] Area "{area_name}" not found'])

            # Check if area is waypoint area
            is_waypoint_area = area_name.endswith('_wp_')

            # ★★★ Normalize device_grid before processing ★★★
            device_grid = normalize_device_grid(device_grid)

            if not device_grid:
                wb.close()
                return ([f'[ERROR] Device grid is empty after normalization'])

            # Get current devices in the area
            position_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_SHAPE>>')

            current_devices = []
            area_start_idx = None
            area_end_idx = None
            current_area = None

            for idx, item in enumerate(position_shape_array):
                if isinstance(item, list) and len(item) >= 2 and isinstance(item[1], list):
                    row = item[1]

                    # Check for area start
                    if len(row) > 0 and row[0] == area_name:
                        current_area = area_name
                        area_start_idx = idx
                        # Collect devices from first row
                        for val in row[1:]:
                            if val not in ['', '<END>', '_AIR_']:
                                current_devices.append(val)

                    # Collect devices from subsequent rows in the area
                    elif current_area == area_name and len(row) > 0 and row[0] == '':
                        for val in row[1:]:
                            if val not in ['', '<END>', '_AIR_']:
                                current_devices.append(val)

                    # Check for area end
                    elif current_area == area_name and len(row) == 1 and row[0] == '<END>':
                        area_end_idx = idx
                        break

            if area_start_idx is None:
                wb.close()
                return ([f'[ERROR] Area "{area_name}" not found in POSITION_SHAPE'])

            # Flatten device_grid to get new devices list (exclude _AIR_)
            new_devices = []
            for row in device_grid:
                if isinstance(row, list):
                    for device in row:
                        if device and device not in ['', '_AIR_']:
                            new_devices.append(device)

            # Check for duplicate device names in input
            if len(new_devices) != len(set(new_devices)):
                duplicates = [dev for dev in set(new_devices) if new_devices.count(dev) > 1]
                wb.close()
                return ([f'[ERROR] Duplicate device names found in input: {duplicates}'])

            # Determine devices to add and delete
            devices_to_add = [dev for dev in new_devices if dev not in current_devices]
            devices_to_delete = [dev for dev in current_devices if dev not in new_devices]

            print(f'[Info] Devices to add: {len(devices_to_add)}')
            print(f'[Info] Devices to delete: {len(devices_to_delete)}')

            # ★★★ Check NEW devices for conflicts in other areas ★★★
            for device_to_add in devices_to_add:
                # Check if device exists in OTHER areas (not current area)
                for item in position_shape_array:
                    if isinstance(item, list) and len(item) >= 2 and isinstance(item[1], list):
                        row = item[1]
                        folder = row[0] if len(row) > 0 else ''

                        # Skip current area
                        if folder == area_name:
                            continue

                        # Check if device exists in another area
                        if device_to_add in row:
                            wb.close()
                            other_is_waypoint = '_wp_' in str(folder)
                            if is_waypoint_area != other_is_waypoint:
                                return ([f'[ERROR] Device "{device_to_add}" already exists in a different area type (waypoint vs normal) in area "{folder}"'])
                            else:
                                return ([f'[ERROR] Device "{device_to_add}" already exists in area "{folder}"'])

            # Collect operations summary
            operations = []
            if devices_to_add:
                operations.append(f'Devices to add: {devices_to_add}')
            if devices_to_delete:
                operations.append(f'Devices to delete: {devices_to_delete}')

            # ========== BATCH STEP 1: Add new devices to STYLE_SHAPE, POSITION_TAG, ATTRIBUTE ==========
            print('[Info] Adding new devices (in memory)...')

            # ★★★ Prepare all device entries in memory FIRST ★★★
            shape_width_min = 0.4
            shape_hight_min = 0.2
            waypoint_width_min = 0.6
            waypoint_hight_min = 0.3
            shae_font_size = 6.0
            tag_offet_inche = 0.02

            # Load all arrays ONCE
            style_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<STYLE_SHAPE>>')
            position_tag_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_TAG>>')
            attribute_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<ATTRIBUTE>>')

            # Find max numbers ONCE
            max_style_number = max([item[0] for item in style_shape_array if isinstance(item[0], int)], default=3)
            max_tag_number = max([item[0] for item in position_tag_array if isinstance(item[0], int)], default=2)
            max_attr_number = max([item[0] for item in attribute_array if isinstance(item[0], int)], default=1)

            # Add all devices to arrays in memory
            for device_to_add in devices_to_add:
                # Determine device properties
                if is_waypoint_area:
                    device_color = 'BLUE'
                    device_roundness = 0.2  # Default for batch add
                    width_min = waypoint_width_min
                    height_min = waypoint_hight_min
                    attribute_default = '[\'WayPoint\', [220, 230, 242]]'
                else:
                    device_color = 'GREEN'
                    device_roundness = 0
                    width_min = shape_width_min
                    height_min = shape_hight_min
                    attribute_default = '[\'DEVICE\',[235, 241, 222]]'

                # Calculate dimensions
                text_dimensions = ns_def.get_description_width_hight(shae_font_size, device_to_add)
                num_char_width = text_dimensions[0]
                line_width = max(num_char_width, width_min)
                line_height = height_min

                # Add to STYLE_SHAPE array
                max_style_number += 1
                new_style_entry = [max_style_number, [
                    device_to_add,
                    line_width,
                    line_height,
                    device_roundness,
                    device_color
                ]]
                style_shape_array.append(new_style_entry)

                # Add to POSITION_TAG array
                max_tag_number += 1
                initial_tag_offset = 1 * tag_offet_inche
                new_tag_entry = [max_tag_number, [
                    device_to_add,
                    'LINE',
                    '',
                    '',
                    initial_tag_offset,
                    'YES'
                ]]
                position_tag_array.append(new_tag_entry)

                # Add to ATTRIBUTE array
                max_attr_number += 1
                new_attr_entry = [max_attr_number, [
                    device_to_add,
                    attribute_default,
                    '[\'<EMPTY>\', [255, 255, 255]]',
                    '[\'<EMPTY>\', [255, 255, 255]]',
                    '[\'<EMPTY>\', [255, 255, 255]]',
                    '[\'<EMPTY>\', [255, 255, 255]]',
                    '[\'<EMPTY>\', [255, 255, 255]]',
                    '[\'<EMPTY>\', [255, 255, 255]]',
                    '[\'<EMPTY>\', [255, 255, 255]]',
                    '[\'<EMPTY>\', [255, 255, 255]]'
                ]]
                attribute_array.append(new_attr_entry)

            # ========== BATCH STEP 2: Update POSITION_SHAPE with new grid ==========
            print('[Info] Updating POSITION_SHAPE (in memory)...')
            position_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_SHAPE>>')

            # Find area boundaries again
            area_start_idx = None
            area_end_idx = None
            for idx, item in enumerate(position_shape_array):
                if isinstance(item, list) and len(item) >= 2 and isinstance(item[1], list):
                    row = item[1]
                    if len(row) > 0 and row[0] == area_name:
                        area_start_idx = idx
                    elif area_start_idx is not None and len(row) == 1 and row[0] == '<END>':
                        area_end_idx = idx
                        break

            # Remove old area rows
            if area_start_idx is not None and area_end_idx is not None:
                del position_shape_array[area_start_idx:area_end_idx + 1]

            # Build new area rows from normalized device_grid
            new_area_rows = []
            for row_idx, row in enumerate(device_grid):
                if row_idx == 0:
                    new_row = [area_name] + row + ['<END>']
                else:
                    new_row = [''] + row + ['<END>']
                new_area_rows.append(new_row)
            new_area_rows.append(['<END>'])

            # Insert new rows
            for offset, new_row in enumerate(new_area_rows):
                position_shape_array.insert(area_start_idx + offset, [0, new_row])

            # Renumber
            for i, item in enumerate(position_shape_array):
                item[0] = i + 1

            # ========== BATCH STEP 3: Delete old devices (in memory) ==========
            print('[Info] Deleting old devices (in memory)...')
            if devices_to_delete:
                # STYLE_SHAPE
                style_shape_array = [item for item in style_shape_array
                                     if not (len(item) >= 2 and isinstance(item[1], list)
                                             and len(item[1]) > 0 and item[1][0] in devices_to_delete)]
                for i, item in enumerate(style_shape_array):
                    item[0] = i + 1

                # POSITION_TAG
                position_tag_array = [item for item in position_tag_array
                                      if not (len(item) >= 2 and isinstance(item[1], list)
                                              and len(item[1]) > 0 and item[1][0] in devices_to_delete)]
                for i, item in enumerate(position_tag_array):
                    item[0] = i + 1

                # ATTRIBUTE
                attribute_array = [item for item in attribute_array
                                   if not (len(item) >= 2 and isinstance(item[1], list)
                                           and len(item[1]) > 0 and item[1][0] in devices_to_delete)]
                for i, item in enumerate(attribute_array):
                    item[0] = i + 1

                # POSITION_LINE
                position_line_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_LINE>>')
                header_rows = position_line_array[:2]
                data_rows = [item for item in position_line_array[2:]
                             if not (len(item) >= 2 and isinstance(item[1], list)
                                     and len(item[1]) >= 2
                                     and (item[1][0] in devices_to_delete or item[1][1] in devices_to_delete))]
                position_line_array = header_rows + data_rows
                for i, item in enumerate(position_line_array):
                    item[0] = i + 1

            # ========== BATCH STEP 4: Convert to tuples and write ONCE ==========
            print('[Info] Writing all changes to Excel (ONCE)...')

            position_shape_tuple = ns_def.convert_array_to_tuple(position_shape_array)
            style_shape_tuple = ns_def.convert_array_to_tuple(style_shape_array)
            position_tag_tuple = ns_def.convert_array_to_tuple(position_tag_array)
            attribute_tuple = ns_def.convert_array_to_tuple(attribute_array)

            if devices_to_delete:
                position_line_tuple = ns_def.convert_array_to_tuple(position_line_array)

            # Write all sections
            ws_name = 'Master_Data'

            # POSITION_SHAPE
            ori_position_shape_tuple = ns_def.convert_array_to_tuple(
                ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_SHAPE>>')
            )
            ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_position_shape_tuple)
            ns_def.write_excel_meta(position_shape_tuple, master_file_path, ws_name, '<<POSITION_SHAPE>>', 0, 0)

            # STYLE_SHAPE
            ori_style_shape_tuple = ns_def.convert_array_to_tuple(
                ns_def.convert_master_to_array('Master_Data', master_file_path, '<<STYLE_SHAPE>>')
            )
            ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_style_shape_tuple)
            ns_def.write_excel_meta(style_shape_tuple, master_file_path, ws_name, '<<STYLE_SHAPE>>', 0, 0)

            # POSITION_TAG
            ori_position_tag_tuple = ns_def.convert_array_to_tuple(
                ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_TAG>>')
            )
            ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_position_tag_tuple)
            ns_def.write_excel_meta(position_tag_tuple, master_file_path, ws_name, '<<POSITION_TAG>>', 0, 0)

            # ATTRIBUTE
            ori_attribute_tuple = ns_def.convert_array_to_tuple(
                ns_def.convert_master_to_array('Master_Data', master_file_path, '<<ATTRIBUTE>>')
            )
            ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_attribute_tuple)
            ns_def.write_excel_meta(attribute_tuple, master_file_path, ws_name, '<<ATTRIBUTE>>', 0, 0)

            # POSITION_LINE (if needed)
            if devices_to_delete:
                ori_position_line_tuple = ns_def.convert_array_to_tuple(
                    ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_LINE>>')
                )
                ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_position_line_tuple)
                ns_def.write_excel_meta(position_line_tuple, master_file_path, ws_name, '<<POSITION_LINE>>', 0, 0)

            # ========== BATCH STEP 5: Recalculate POSITION_FOLDER ONCE ==========
            print('[Info] Recalculating POSITION_FOLDER...')

            position_folder_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_FOLDER>>')
            position_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_SHAPE>>')
            style_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<STYLE_SHAPE>>')

            position_folder_tuple = ns_def.convert_array_to_tuple(position_folder_array)
            position_shape_tuple = ns_def.convert_array_to_tuple(position_shape_array)
            style_shape_tuple = ns_def.convert_array_to_tuple(style_shape_array)

            ori_position_folder_tuple = position_folder_tuple

            update_position_folder_tuple = def_common.make_position_folder_tuple(
                position_folder_tuple,
                style_shape_tuple,
                position_shape_tuple
            )

            ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_position_folder_tuple)
            ns_def.write_excel_meta(update_position_folder_tuple, master_file_path,
                                    'Master_Data', '<<POSITION_FOLDER>>', 0, 0)

            # Recalculate folder sizes
            print('[Info] Recalculating folder sizes...')
            result = def_common.recalculate_folder_sizes(master_file_path)
            if result['status'] != 'success':
                wb.close()
                return ([f"[ERROR] Failed to recalculate folder sizes: {result['message']}"])

            # ========== BATCH STEP 6: Sync with L2/L3 ONCE at the end ==========
            print('[Info] Syncing with L2/L3 layers (ONCE)...')
            import tkinter as tk
            dummy_tk = tk.Toplevel()
            dummy_tk.withdraw()

            self.full_filepath = master_file_path
            self.main1_1_entry_1 = tk.Entry(dummy_tk)
            self.main1_1_entry_1.insert(tk.END, master_file_path)

            self.inFileTxt_L3_1_1 = tk.Entry(dummy_tk)
            self.inFileTxt_L3_1_1.delete(0, tkinter.END)
            self.inFileTxt_L3_1_1.insert(tk.END, master_file_path)

            self.outFileTxt_11_2 = tk.Entry(dummy_tk)
            self.outFileTxt_11_2.delete(0, tkinter.END)
            self.outFileTxt_11_2.insert(tk.END, master_file_path)

            self.inFileTxt_L2_1_1 = tk.Entry(dummy_tk)
            self.inFileTxt_L2_1_1.delete(0, tkinter.END)
            self.inFileTxt_L2_1_1.insert(tk.END, master_file_path)

            import ns_sync_between_layers
            ns_sync_between_layers.l1_master_device_and_line_sync_with_l2l3_master(self)

            tmp_delete_excel_name = master_file_path.replace('[MASTER]', '[L2_TABLE]')
            if os.path.isfile(tmp_delete_excel_name):
                os.remove(tmp_delete_excel_name)

            dummy_tk.destroy()

            # Close workbook
            wb.close()

            print('[Info] Batch device location update completed.')

            # Build return message

            return_msg = f'--- Device location updated for area "{area_name}" (BATCH MODE) ---\n'
            '''
            if devices_to_add:
                return_msg += f'Added devices: {devices_to_add}\n'
            if devices_to_delete:
                return_msg += f'Deleted devices: {devices_to_delete}\n'
            if not devices_to_add and not devices_to_delete:
                return_msg += 'Device layout rearranged (no additions or deletions)\n'
            return_msg += f'New device layout (normalized):\n'
            for row in device_grid:
                return_msg += f'  {row}\n'
            '''

            return ([return_msg.strip()])

        except Exception as e:
            import traceback
            traceback.print_exc()
            try:
                wb.close()
            except:
                pass
            return ([f'[ERROR] Failed to update device location: {str(e)}'])



    @staticmethod
    def validate_waypoint_layout(area_location_array):
        """
        Validate that waypoint areas (_wp_) are not placed horizontally adjacent to each other

        Args:
            area_location_array: List of lists containing area names in layout

        Returns:
            dict: {'status': 'success'/'error', 'message': str, 'invalid_rows': list}
        """
        invalid_rows = []

        for row_idx, row in enumerate(area_location_array):
            if not isinstance(row, list):
                continue

            # Check for consecutive waypoint areas in the row
            consecutive_waypoints = []
            current_waypoint_sequence = []

            for col_idx, area in enumerate(row):
                if isinstance(area, str) and area.endswith('_wp_'):
                    # This is a waypoint area
                    current_waypoint_sequence.append(area)
                else:
                    # Not a waypoint or empty cell
                    if len(current_waypoint_sequence) >= 2:
                        # Found 2 or more consecutive waypoints
                        consecutive_waypoints.append(current_waypoint_sequence.copy())
                    current_waypoint_sequence = []

            # Check the last sequence
            if len(current_waypoint_sequence) >= 2:
                consecutive_waypoints.append(current_waypoint_sequence.copy())

            # If consecutive waypoints found in this row, add to invalid rows
            if consecutive_waypoints:
                invalid_rows.append({
                    'row_index': row_idx,
                    'row': row,
                    'consecutive_waypoints': consecutive_waypoints
                })

        if invalid_rows:
            # Build error message
            error_msg = "[Error] Waypoint areas (_wp_) cannot be placed horizontally adjacent to each other.\n"
            for invalid_row in invalid_rows:
                row_num = invalid_row['row_index'] + 1
                error_msg += f"  Row {row_num}: {invalid_row['row']}\n"
                for wp_sequence in invalid_row['consecutive_waypoints']:
                    error_msg += f"    Consecutive waypoints detected: {wp_sequence}\n"
            error_msg += "  Please ensure waypoint areas are separated by non-waypoint areas or placed in different rows."

            return {
                'status': 'error',
                'message': error_msg,
                'invalid_rows': invalid_rows
            }

        return {
            'status': 'success',
            'message': 'Waypoint layout is valid',
            'invalid_rows': []
        }

    @staticmethod
    def add_devices_for_new_areas(new_areas, master_file_path):
        """
        Add devices for newly created areas

        Args:
            new_areas: List of new area names
            master_file_path: Path to the master Excel file

        Returns:
            dict: Status and message of the operation
        """
        import ns_def

        try:
            # Parameters (from ns_l1_master_create)
            shape_width_min = 0.4  # inches
            shape_hight_min = 0.2  # inches
            shae_font_size = 6.0  # pt
            tag_offet_inche = 0.02  # inches per character

            # Parameters for waypoints
            waypoint_width_min = 0.6  # inches
            waypoint_hight_min = 0.3  # inches

            # ★★★ NEW: Function to check if an area row contains only waypoint areas ★★★
            def is_waypoint_only_row(area_name, master_file_path):
                """
                Check if the area row containing this area has only waypoint areas (_wp_)
                Returns True if all areas in the same row end with _wp_
                """
                position_folder_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_FOLDER>>')

                # Find the row containing this area
                target_row_data = None
                for item in position_folder_array:
                    if len(item) >= 2 and isinstance(item[1], list):
                        row_data = item[1]
                        # Check if this is an area name row (starts with numeric value like 6.7 or 2.05)
                        if len(row_data) > 0 and isinstance(row_data[0], (int, float)):
                            # Check if our area is in this row
                            if area_name in row_data:
                                target_row_data = row_data
                                break

                if target_row_data is None:
                    return False

                # Check all non-empty area names in this row
                all_waypoint = True
                found_any_area = False
                for element in target_row_data[1:]:  # Skip the first element (row margin)
                    if isinstance(element, str) and element != '':
                        found_any_area = True
                        if not element.endswith('_wp_'):
                            all_waypoint = False
                            break

                # Return True only if we found at least one area and all are waypoints
                return found_any_area and all_waypoint

            for area_name in new_areas:
                # Generate device name based on area type
                if area_name.endswith('_wp_'):
                    # For waypoint areas: remove _wp_ suffix
                    device_name = area_name[:-4]
                    is_waypoint = True
                    device_color = 'BLUE'

                    # ★★★ NEW: Determine roundness based on whether row contains only waypoints ★★★
                    if is_waypoint_only_row(area_name, master_file_path):
                        device_roundness = 0.5  # All areas in row are waypoints
                    else:
                        device_roundness = 0.2  # Mixed: some areas are not waypoints

                    width_min = waypoint_width_min
                    height_min = waypoint_hight_min
                    attribute_default = '[\'WayPoint\', [220, 230, 242]]'
                else:
                    # For normal areas: add _device_ suffix
                    device_name = area_name + '_device_'
                    is_waypoint = False
                    device_color = 'GREEN'
                    device_roundness = 0
                    width_min = shape_width_min
                    height_min = shape_hight_min
                    attribute_default = '[\'DEVICE\',[235, 241, 222]]'

                # ========== Add to POSITION_SHAPE ==========
                position_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_SHAPE>>')
                ori_position_shape_tuple = ns_def.convert_array_to_tuple(position_shape_array)

                # Find the area in POSITION_SHAPE
                area_found = False
                area_index = -1
                current_area = None

                for i, item in enumerate(position_shape_array):
                    if isinstance(item, list) and len(item) >= 2 and isinstance(item[1], list):
                        # Check if this row starts a new area
                        if len(item[1]) > 0 and item[1][0] not in ['', '<END>', '<<POSITION_SHAPE>>', '_AIR_']:
                            current_area = item[1][0]

                        # Check if this is our target area
                        if current_area == area_name:
                            if not area_found:
                                area_found = True
                                area_index = i
                                # Add device to the first row of this area
                                if len(item[1]) > 0 and item[1][0] == area_name:
                                    # Find position to insert (before <END> if exists)
                                    if '<END>' in item[1]:
                                        end_index = item[1].index('<END>')
                                        item[1].insert(end_index, device_name)
                                    else:
                                        item[1].append(device_name)
                                    break

                # If area not found, add new entry at the end
                if not area_found:
                    # Find the last position (before the final <END> if it exists)
                    insert_position = len(position_shape_array)

                    # Create new area entry with device
                    new_shape_entry = [insert_position, [area_name, device_name, '<END>']]
                    position_shape_array.append(new_shape_entry)
                    # Add <END> marker for the area
                    position_shape_array.append([insert_position + 1, ['<END>']])

                    # Renumber all entries
                    for i, item in enumerate(position_shape_array):
                        item[0] = i + 1

                # Write updated POSITION_SHAPE
                position_shape_tuple = ns_def.convert_array_to_tuple(position_shape_array)
                ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_position_shape_tuple)
                ns_def.write_excel_meta(position_shape_tuple, master_file_path,
                                        'Master_Data', '<<POSITION_SHAPE>>', 0, 0)

                # ========== Add to STYLE_SHAPE ==========
                style_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<STYLE_SHAPE>>')
                ori_style_shape_tuple = ns_def.convert_array_to_tuple(style_shape_array)

                # Calculate device dimensions
                text_dimensions = ns_def.get_description_width_hight(shae_font_size, device_name)
                num_char_width = text_dimensions[0]
                line_width = max(num_char_width, width_min)
                line_height = height_min

                # Find max number
                max_number = 0
                for item in style_shape_array:
                    if isinstance(item[0], int) and item[0] > max_number:
                        max_number = item[0]

                if max_number < 3:
                    max_number = 3

                new_style_number = max_number + 1

                # Create new style entry with calculated roundness
                new_style_entry = [new_style_number, [
                    device_name,
                    line_width,
                    line_height,
                    device_roundness,  # ★★★ Uses calculated roundness value ★★★
                    device_color
                ]]

                style_shape_array.append(new_style_entry)

                # Write updated STYLE_SHAPE
                style_shape_tuple = ns_def.convert_array_to_tuple(style_shape_array)
                ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_style_shape_tuple)
                ns_def.write_excel_meta(style_shape_tuple, master_file_path,
                                        'Master_Data', '<<STYLE_SHAPE>>', 0, 0)

                # ========== Add to POSITION_TAG ==========
                position_tag_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_TAG>>')
                ori_position_tag_tuple = ns_def.convert_array_to_tuple(position_tag_array)

                # Find max number
                max_tag_number = 0
                for item in position_tag_array:
                    if isinstance(item[0], int) and item[0] > max_tag_number:
                        max_tag_number = item[0]

                if max_tag_number < 2:
                    max_tag_number = 2

                new_tag_number = max_tag_number + 1
                initial_tag_offset = 1 * tag_offet_inche

                # Create new tag entry
                new_tag_entry = [new_tag_number, [
                    device_name,
                    'LINE',
                    '',
                    '',
                    initial_tag_offset,
                    'YES'
                ]]

                position_tag_array.append(new_tag_entry)

                # Write updated POSITION_TAG
                position_tag_tuple = ns_def.convert_array_to_tuple(position_tag_array)
                ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_position_tag_tuple)
                ns_def.write_excel_meta(position_tag_tuple, master_file_path,
                                        'Master_Data', '<<POSITION_TAG>>', 0, 0)

                # ========== Add to ATTRIBUTE ==========
                attribute_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<ATTRIBUTE>>')
                ori_attribute_tuple = ns_def.convert_array_to_tuple(attribute_array)

                # Find max number
                max_attr_number = 0
                for item in attribute_array:
                    if isinstance(item[0], int) and item[0] > max_attr_number:
                        max_attr_number = item[0]

                if max_attr_number < 1:
                    max_attr_number = 1

                new_attr_number = max_attr_number + 1

                # Create new attribute entry
                new_attr_entry = [new_attr_number, [
                    device_name,
                    attribute_default,
                    '[\'<EMPTY>\', [255, 255, 255]]',
                    '[\'<EMPTY>\', [255, 255, 255]]',
                    '[\'<EMPTY>\', [255, 255, 255]]',
                    '[\'<EMPTY>\', [255, 255, 255]]',
                    '[\'<EMPTY>\', [255, 255, 255]]',
                    '[\'<EMPTY>\', [255, 255, 255]]',
                    '[\'<EMPTY>\', [255, 255, 255]]',
                    '[\'<EMPTY>\', [255, 255, 255]]'
                ]]

                attribute_array.append(new_attr_entry)

                # Write updated ATTRIBUTE
                attribute_tuple = ns_def.convert_array_to_tuple(attribute_array)
                ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_attribute_tuple)
                ns_def.write_excel_meta(attribute_tuple, master_file_path,
                                        'Master_Data', '<<ATTRIBUTE>>', 0, 0)

            return {'status': 'success', 'message': f'Added devices for {len(new_areas)} new areas'}

        except Exception as e:
            import traceback
            traceback.print_exc()
            return {'status': 'error', 'message': f'Error adding devices for new areas: {str(e)}'}



    @staticmethod
    def update_area_with_map(self, area_location_array, master_file_path):
        """
        Update area layout based on provided array structure

        Args:
            self: Reference to calling instance
            area_location_array: List of lists containing area names in desired layout
            master_file_path: Path to the master Excel file

        Returns:
            List with status message or error
        """
        import ns_def

        # Validate input format
        if not isinstance(area_location_array, list):
            return ([f'[ERROR] area_location array must be a list'])

        # ★★★ Add: Waypoint layout validation ★★★
        validation_result = def_common.validate_waypoint_layout(area_location_array)
        if validation_result['status'] == 'error':
            return ([validation_result['message']])
        # ★★★ End of waypoint validation ★★★

        # Get all existing areas from STYLE_FOLDER
        style_folder_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<STYLE_FOLDER>>')
        existing_areas = []
        for item in style_folder_array:
            if item[0] not in [1, 2, 3]:
                existing_areas.append(item[1][0])

        # Extract all areas from input array
        input_areas = []
        for row in area_location_array:
            if isinstance(row, list):
                for area in row:
                    if area and area not in input_areas:
                        input_areas.append(area)

        # Check if all existing areas are included in input
        missing_areas = []
        for existing_area in existing_areas:
            if existing_area not in input_areas:
                missing_areas.append(existing_area)

        if missing_areas:
            return ([f'[ERROR] All existing areas must be included. Missing areas: {missing_areas}'])

        # Identify new areas
        new_areas = []
        for area in input_areas:
            if area not in existing_areas:
                new_areas.append(area)

        # Check for duplicate area names and device name conflicts
        for new_area in new_areas:
            # Check if area name already exists
            if new_area in existing_areas:
                return ([f'[ERROR] Area name already exists: {new_area}'])

            # Generate device name for the new area
            if new_area.endswith('_wp_'):
                # Remove _wp_ suffix for waypoint areas
                device_name = new_area[:-4]  # Remove last 4 characters (_wp_)
            else:
                # Add _device_ suffix for normal areas
                device_name = new_area + '_device_'

            # Check if device name conflicts with existing devices
            position_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_SHAPE>>')
            for item in position_shape_array:
                if isinstance(item, list) and len(item) >= 2 and isinstance(item[1], list):
                    for val in item[1]:
                        if val == device_name:
                            return ([f'[ERROR] Device name already exists: {device_name}'])

        # Rebuild POSITION_FOLDER structure
        result = def_common.rebuild_position_folder(area_location_array, master_file_path)
        if result['status'] != 'success':
            return ([f"[ERROR] {result['message']}"])

        # Add new areas to STYLE_FOLDER if needed
        if new_areas:
            result = def_common.add_areas_to_style_folder(new_areas, master_file_path)
            if result['status'] != 'success':
                return ([f"[ERROR] {result['message']}"])

            # Add devices for new areas
            result = def_common.add_devices_for_new_areas(new_areas, master_file_path)
            if result['status'] != 'success':
                return ([f"[ERROR] {result['message']}"])

        # Recalculate folder sizes
        result = def_common.recalculate_folder_sizes(master_file_path)
        if result['status'] != 'success':
            return ([f"[ERROR] {result['message']}"])

        # Build return message
        return_msg = '--- Area layout updated ---\n'
        if new_areas:
            return_msg += f'New areas added: {new_areas}\n'
            # Show devices added
            for area in new_areas:
                if area.endswith('_wp_'):
                    device_name = area[:-4]
                else:
                    device_name = area + '_device_'
                return_msg += f'  Device added for {area}: {device_name}\n'

        # Show changes in layout
        return_msg += 'New area layout:\n'
        for row in area_location_array:
            return_msg += f'  {row}\n'

        return ([return_msg.strip()])

    @staticmethod
    def rebuild_position_folder(area_location_array, master_file_path):
        """
        Rebuild POSITION_FOLDER section based on new area layout
        Automatically inserts empty cells between adjacent non-waypoint areas
        Merges consecutive empty cells into one

        Args:
            area_location_array: List of lists containing area names in desired layout
            master_file_path: Path to the master Excel file

        Returns:
            dict: Status and message of the operation
        """
        import ns_def

        try:
            # Load current POSITION_FOLDER
            position_folder_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_FOLDER>>')

            # Store original for cleanup
            ori_position_folder_tuple = ns_def.convert_array_to_tuple(position_folder_array)

            # Get original area names and their corresponding widths
            original_area_widths = {}
            current_width_row = []
            for item in position_folder_array:
                if len(item) >= 2 and isinstance(item[1], list):
                    row_data = item[1]
                    # Check if this is a width row
                    if len(row_data) > 0 and row_data[0] in ['<<POSITION_FOLDER>>', '<SET_WIDTH>']:
                        current_width_row = row_data
                    # Check if this is an area name row
                    elif len(row_data) > 0 and isinstance(row_data[0], (int, float)) and len(current_width_row) > 0:
                        # Map area names to their widths
                        for col_idx in range(1, len(row_data)):
                            if col_idx < len(current_width_row):
                                area_name = row_data[col_idx]
                                width_value = current_width_row[col_idx]
                                if isinstance(area_name, str) and area_name != '' and isinstance(width_value, (int, float)):
                                    original_area_widths[area_name] = width_value

            # Function to process a row and insert empty cells between adjacent non-waypoint areas
            def process_row_with_empty_cells(row):
                """
                Insert empty cells between adjacent non-waypoint areas
                Rule: If two consecutive areas are both non-waypoint (don't end with _wp_), insert '' between them
                """
                if not row:
                    return row

                processed_row = []

                for i, current_area in enumerate(row):
                    # Add current area to processed row
                    processed_row.append(current_area)

                    # Check if we need to insert an empty cell after this area
                    # Only check if there's a next element
                    if i < len(row) - 1:
                        next_area = row[i + 1]

                        # Both current and next must be non-empty
                        if current_area and next_area:
                            current_is_waypoint = current_area.endswith('_wp_')
                            next_is_waypoint = next_area.endswith('_wp_')

                            # Insert empty cell if both are non-waypoint areas
                            if not current_is_waypoint and not next_is_waypoint:
                                processed_row.append('')  # Insert empty cell

                return processed_row

            # Function to merge consecutive empty cells
            def merge_consecutive_empty_cells(row):
                """
                Merge consecutive empty cells into one
                Policy: No more than one consecutive empty cell
                """
                if not row:
                    return row

                merged_row = []
                prev_was_empty = False

                for element in row:
                    # Check if current element is empty
                    is_empty = (element == '' or element is None)

                    if is_empty:
                        # Only add if previous was not empty
                        if not prev_was_empty:
                            merged_row.append('')
                            prev_was_empty = True
                        # Skip if previous was also empty (merge)
                    else:
                        # Non-empty element: add it
                        merged_row.append(element)
                        prev_was_empty = False

                return merged_row

            # ★★★ NEW: Function to remove trailing empty cells ★★★
            def remove_trailing_empty_cells(row):
                """
                Remove empty cells from the end of the row
                """
                if not row:
                    return row

                # Find the last non-empty element
                last_non_empty_idx = -1
                for i in range(len(row) - 1, -1, -1):
                    if row[i] != '' and row[i] is not None:
                        last_non_empty_idx = i
                        break

                # Return row up to (and including) the last non-empty element
                if last_non_empty_idx >= 0:
                    return row[:last_non_empty_idx + 1]
                else:
                    return []

            # Process area_location_array to insert empty cells
            processed_area_location_array = []
            for row in area_location_array:
                if isinstance(row, list):
                    # Step 1: Insert empty cells between adjacent non-waypoint areas
                    processed_row = process_row_with_empty_cells(row)
                    # Step 2: Merge consecutive empty cells
                    processed_row = merge_consecutive_empty_cells(processed_row)
                    # Step 3: Remove trailing empty cells
                    processed_row = remove_trailing_empty_cells(processed_row)
                    processed_area_location_array.append(processed_row)
                else:
                    processed_area_location_array.append(row)

            # Find the row with maximum number of elements (after processing)
            max_element_count = 0
            max_row_idx = -1
            for idx, row in enumerate(processed_area_location_array):
                if isinstance(row, list):
                    element_count = len(row)
                    if element_count > max_element_count:
                        max_element_count = element_count
                        max_row_idx = idx

            # Build new POSITION_FOLDER structure
            new_position_folder_array = []

            # Process each row in the processed array
            for row_idx, row in enumerate(processed_area_location_array):
                if not isinstance(row, list):
                    continue

                # Determine if this row needs empty cells on both sides
                needs_side_empty = (row_idx != max_row_idx)

                # Check if all non-empty areas in this row are waypoints
                non_empty_areas = [area for area in row if area]
                all_waypoints = all(area.endswith('_wp_') for area in non_empty_areas) if non_empty_areas else False

                if row_idx == 0:
                    # First row: <<POSITION_FOLDER>> serves as <SET_WIDTH>
                    width_row_num = 1
                    width_row = ['<<POSITION_FOLDER>>']

                    # Add empty cell at the beginning if needed
                    if needs_side_empty:
                        width_row.append(0.5)  # Default width for side empty area

                    # Process each element in the row (including empty strings)
                    for area in row:
                        if area and area != '':
                            # Non-empty area: use original width or default
                            if area in original_area_widths:
                                width_row.append(original_area_widths[area])
                            else:
                                # New area - use default based on type
                                if area.endswith('_wp_'):
                                    width_row.append(3.0)
                                else:
                                    width_row.append(2.5)
                        else:
                            # Empty string: add empty cell width
                            width_row.append(0.5)  # Default width for empty cell

                    # Add empty cell at the end if needed
                    if needs_side_empty:
                        width_row.append(0.5)  # Default width for side empty area

                    new_position_folder_array.append([width_row_num, width_row])

                    # Add area names row
                    area_row_num = 2
                    if all_waypoints:
                        area_row = [2.05]  # Waypoint row margin
                    else:
                        area_row = [6.7]  # Normal row margin

                    # Add empty cell at the beginning if needed
                    if needs_side_empty:
                        area_row.append('')

                    # Add area names (including empty strings)
                    for area in row:
                        area_row.append(area if area else '')

                    # Add empty cell at the end if needed
                    if needs_side_empty:
                        area_row.append('')

                    new_position_folder_array.append([area_row_num, area_row])

                else:
                    # Subsequent rows: use <SET_WIDTH>
                    width_row_num = (row_idx * 2) + 1
                    width_row = ['<SET_WIDTH>']

                    # Add empty cell at the beginning if needed
                    if needs_side_empty:
                        width_row.append(0.5)

                    # Process each element in the row (including empty strings)
                    for area in row:
                        if area and area != '':
                            # Non-empty area: use original width or default
                            if area in original_area_widths:
                                width_row.append(original_area_widths[area])
                            else:
                                # New area - use default based on type
                                if area.endswith('_wp_'):
                                    width_row.append(3.0)
                                else:
                                    width_row.append(2.5)
                        else:
                            # Empty string: add empty cell width
                            width_row.append(0.5)  # Default width for empty cell

                    # Add empty cell at the end if needed
                    if needs_side_empty:
                        width_row.append(0.5)

                    new_position_folder_array.append([width_row_num, width_row])

                    # Add area names row
                    area_row_num = (row_idx * 2) + 2
                    if all_waypoints:
                        area_row = [2.05]  # Waypoint row margin
                    else:
                        area_row = [6.7]  # Normal row margin

                    # Add empty cell at the beginning if needed
                    if needs_side_empty:
                        area_row.append('')

                    # Add area names (including empty strings)
                    for area in row:
                        area_row.append(area if area else '')

                    # Add empty cell at the end if needed
                    if needs_side_empty:
                        area_row.append('')

                    new_position_folder_array.append([area_row_num, area_row])

            # Convert to tuple for writing
            new_position_folder_tuple = ns_def.convert_array_to_tuple(new_position_folder_array)

            # Write new POSITION_FOLDER data
            ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_position_folder_tuple)
            ns_def.write_excel_meta(new_position_folder_tuple, master_file_path,
                                    'Master_Data', '<<POSITION_FOLDER>>', 0, 0)

            return {'status': 'success', 'message': 'POSITION_FOLDER rebuilt successfully'}

        except Exception as e:
            import traceback
            traceback.print_exc()
            return {'status': 'error', 'message': f'Error rebuilding POSITION_FOLDER: {str(e)}'}

    @staticmethod
    def recalculate_folder_sizes(master_file_path):
        """
        Recalculate and update folder sizes based on current layout
        Ensures max-width row has no leading/trailing empty cells
        Ensures non-max-width rows have leading/trailing empty cells

        Args:
            master_file_path: Path to the master Excel file

        Returns:
            dict: Status and message of the operation
        """
        import ns_def
        import openpyxl

        try:
            # Load necessary arrays
            position_folder_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_FOLDER>>')
            position_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_SHAPE>>')
            style_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<STYLE_SHAPE>>')

            # Convert to tuples
            position_folder_tuple = ns_def.convert_array_to_tuple(position_folder_array)
            position_shape_tuple = ns_def.convert_array_to_tuple(position_shape_array)
            style_shape_tuple = ns_def.convert_array_to_tuple(style_shape_array)

            # Parameters
            min_tag_inches = 0.3

            # Call get_folder_width_size
            master_folder_size_array = ns_def.get_folder_width_size(
                position_folder_tuple,
                style_shape_tuple,
                position_shape_tuple,
                min_tag_inches
            )

            # Use make_position_folder_tuple
            update_position_folder_tuple = def_common.make_position_folder_tuple(
                position_folder_tuple,
                style_shape_tuple,
                position_shape_tuple
            )

            # WRITE
            ori_position_folder_tuple = position_folder_tuple
            ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_position_folder_tuple)
            ns_def.write_excel_meta(update_position_folder_tuple, master_file_path,
                                    'Master_Data', '<<POSITION_FOLDER>>', 0, 0)

            # ★★★ POST-PROCESS: Array-based processing with proper empty cell detection ★★★
            position_folder_array_after = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_FOLDER>>')

            # Determine pattern
            has_widths_in_row_1 = False
            if len(position_folder_array_after) > 0 and len(position_folder_array_after[0]) >= 2:
                row_1_data = position_folder_array_after[0][1]
                if len(row_1_data) > 0 and row_1_data[0] == '<<POSITION_FOLDER>>':
                    for idx, val in enumerate(row_1_data):
                        if idx > 0 and isinstance(val, (int, float)) and val not in [10, 0.999, 1]:
                            has_widths_in_row_1 = True
                            break

            # Get width-area row pairs
            if has_widths_in_row_1:
                width_row_indices = [i for i in range(len(position_folder_array_after)) if i % 2 == 0]
            else:
                width_row_indices = [i for i in range(1, len(position_folder_array_after), 2)]

            # Collect row info WITH total width calculation (excluding empty cells)
            row_info_list = []
            for width_idx in width_row_indices:
                if width_idx >= len(position_folder_array_after):
                    continue

                area_idx = width_idx + 1
                if area_idx >= len(position_folder_array_after):
                    continue

                width_row_data = position_folder_array_after[width_idx][1]
                area_row_data = position_folder_array_after[area_idx][1]

                # ★★★ Extend area_row_data to match width_row_data length ★★★
                while len(area_row_data) < len(width_row_data):
                    area_row_data.append(None)

                # Count areas and calculate total width (EXCLUDING empty cells)
                area_count = 0
                first_area_idx = None
                last_area_idx = 0
                total_width = 0.0

                for col_idx in range(1, len(width_row_data)):
                    width_val = width_row_data[col_idx]
                    area_val = area_row_data[col_idx] if col_idx < len(area_row_data) else None

                    # ★★★ Check if this is an empty cell ★★★
                    is_empty_cell = (area_val is None or area_val == '')

                    # ★★★ Only add to total_width if NOT an empty cell ★★★
                    if isinstance(width_val, (int, float)) and width_val not in [10, 0.999, 1]:
                        if not is_empty_cell:
                            # This is an actual area or area separator (not empty cell)
                            total_width += width_val

                    # Count actual areas (not empty cells, not empty strings between areas)
                    if area_val is not None and isinstance(area_val, str) and area_val != '' and area_val not in ['<<POSITION_FOLDER>>', '<SET_WIDTH>']:
                        area_count += 1
                        if first_area_idx is None:
                            first_area_idx = col_idx
                        last_area_idx = col_idx

                row_info_list.append((width_idx, area_idx, area_count, first_area_idx, last_area_idx, total_width))

            # Find max total width (excluding empty cells)
            max_total_width = max([item[5] for item in row_info_list]) if row_info_list else 0

            # Process each row
            for width_idx, area_idx, area_count, first_area_idx, last_area_idx, total_width in row_info_list:
                # Compare by actual width
                is_max_width_row = abs(total_width - max_total_width) < 0.01

                width_row_data = position_folder_array_after[width_idx][1]
                area_row_data = position_folder_array_after[area_idx][1]

                if is_max_width_row:
                    # ★★★ Max-width row: Remove leading and trailing empty cells ★★★

                    # Remove leading empty cell if exists
                    if first_area_idx and first_area_idx > 1:
                        if len(width_row_data) > 1:
                            del width_row_data[1]
                        if len(area_row_data) > 1:
                            del area_row_data[1]

                        # Update indices
                        if first_area_idx:
                            first_area_idx -= 1
                        if last_area_idx:
                            last_area_idx -= 1

                    # ★★★ Remove ALL trailing elements (including empty cells) ★★★
                    if last_area_idx and last_area_idx > 0:
                        width_row_data[last_area_idx + 1:] = []
                        area_row_data[last_area_idx + 1:] = []
                else:
                    # ★★★ Non-max-width row: Ensure both ends have empties ★★★

                    # Ensure leading empty
                    if first_area_idx == 1:
                        area_row_data.insert(1, '')
                        width_row_data.insert(1, 0.5)
                        if first_area_idx:
                            first_area_idx += 1
                        if last_area_idx:
                            last_area_idx += 1

                    # ★★★ Ensure trailing empty exists ★★★
                    # Find actual last non-empty/non-None position
                    actual_last_idx = 0
                    for i in range(len(area_row_data) - 1, 0, -1):
                        val = area_row_data[i]
                        if val is not None and isinstance(val, str) and val != '':
                            actual_last_idx = i
                            break

                    if actual_last_idx > 0:
                        # Trim everything after actual last area
                        width_row_data[actual_last_idx + 1:] = []
                        area_row_data[actual_last_idx + 1:] = []

                        # Add trailing empty
                        area_row_data.append('')
                        left_empty_width = width_row_data[1] if len(width_row_data) > 1 and isinstance(width_row_data[1], (int, float)) else 0.5
                        width_row_data.append(left_empty_width)

            # Write back
            cleaned_tuple = ns_def.convert_array_to_tuple(position_folder_array_after)
            ns_def.remove_rows_under_section('Master_Data', master_file_path,
                                             ns_def.convert_array_to_tuple(ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_FOLDER>>')))
            ns_def.write_excel_meta(cleaned_tuple, master_file_path,
                                    'Master_Data', '<<POSITION_FOLDER>>', 0, 0)

            return {'status': 'success', 'message': 'Folder sizes recalculated successfully'}

        except Exception as e:
            import traceback
            traceback.print_exc()
            return {'status': 'error', 'message': f'Error recalculating folder sizes: {str(e)}'}


    @staticmethod
    def add_areas_to_style_folder(new_areas, master_file_path):
        """
        Add new areas to STYLE_FOLDER section

        Args:
            new_areas: List of new area names to add
            master_file_path: Path to the master Excel file

        Returns:
            dict: Status and message of the operation
        """
        import ns_def

        try:
            style_folder_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<STYLE_FOLDER>>')
            ori_style_folder_tuple = ns_def.convert_array_to_tuple(style_folder_array)

            # Find max number in style_folder_array
            max_number = 0
            for item in style_folder_array:
                if isinstance(item[0], int) and item[0] > max_number:
                    max_number = item[0]

            # Skip header rows
            if max_number < 3:
                max_number = 3

            # Add each new area to STYLE_FOLDER
            for new_area in new_areas:
                max_number += 1

                # Check if area name ends with '_wp_'
                if new_area.endswith('_wp_'):
                    # Waypoint areas: NO NO <AUTO> <AUTO>
                    new_entry = [max_number, [new_area, 'NO', 'NO', '<AUTO>', '<AUTO>']]
                else:
                    # Normal areas: YES UP <AUTO> <AUTO>
                    new_entry = [max_number, [new_area, 'YES', 'UP', '<AUTO>', '<AUTO>']]

                style_folder_array.append(new_entry)

            # Write updated STYLE_FOLDER
            new_style_folder_tuple = ns_def.convert_array_to_tuple(style_folder_array)
            ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_style_folder_tuple)
            ns_def.write_excel_meta(new_style_folder_tuple, master_file_path,
                                    'Master_Data', '<<STYLE_FOLDER>>', 0, 0)

            return {'status': 'success', 'message': f'Added {len(new_areas)} new areas to STYLE_FOLDER'}

        except Exception as e:
            import traceback
            traceback.print_exc()
            return {'status': 'error', 'message': f'Error adding areas to STYLE_FOLDER: {str(e)}'}

    @staticmethod
    def delete_device_or_waypoint_common(self, name, master_file_path, element_type='device'):
        """
        Common function to delete device or waypoint

        Args:
            self: Reference to calling instance
            name: Name of the device or waypoint to delete
            master_file_path: Path to the master Excel file
            element_type: 'device' or 'waypoint'

        Returns:
            List with status message or error
        """
        import ns_def

        # Load all necessary arrays
        position_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_SHAPE>>')
        style_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<STYLE_SHAPE>>')
        position_line_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_LINE>>')
        position_tag_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_TAG>>')
        attribute_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<ATTRIBUTE>>')
        position_folder_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_FOLDER>>')

        # Check if element exists and validate type
        element_found = False
        element_folder = None
        element_count_in_folder = 0
        is_wrong_type = False

        current_folder = None
        folder_element_count = {}

        for item in position_shape_array:
            if isinstance(item, list) and len(item) >= 2 and isinstance(item[1], list):
                row = item[1]

                # Skip header
                if len(row) > 0 and row[0] == '<<POSITION_SHAPE>>':
                    continue

                # Check for folder start
                if len(row) > 0 and row[0] and row[0] not in ['', '<END>', '_AIR_']:
                    current_folder = row[0]
                    if current_folder not in folder_element_count:
                        folder_element_count[current_folder] = 0

                # Count elements in current folder
                if current_folder and len(row) > 0 and row[0] != '<END>':
                    for val in row:
                        if val not in ['', '<END>', '_AIR_', current_folder]:
                            folder_element_count[current_folder] += 1
                            if val == name:
                                element_found = True
                                element_folder = current_folder
                                # Check type mismatch
                                if element_type == 'device' and '_wp_' in current_folder:
                                    is_wrong_type = True
                                elif element_type == 'waypoint' and '_wp_' not in current_folder:
                                    is_wrong_type = True

                # Reset when reaching folder end
                if len(row) == 1 and row[0] == '<END>':
                    current_folder = None

        # Error checks
        if not element_found:
            element_label = 'Device' if element_type == 'device' else 'Waypoint'
            return ([f"[Error] {element_label} '{name}' not found"])

        if is_wrong_type:
            if element_type == 'device':
                return ([f"[Error] '{name}' is a waypoint (in folder '{element_folder}'). Please use 'delete waypoint' command instead."])
            else:
                return ([f"[Error] '{name}' is a device (in folder '{element_folder}'). Please use 'delete device' command instead."])

        # Check if it's the last element in the folder
        if element_folder and folder_element_count.get(element_folder, 0) == 1:
            element_label = 'device' if element_type == 'device' else 'waypoint'
            return ([f"[Error] Cannot delete '{name}' as it is the last {element_label} in folder '{element_folder}'. Please use 'delete area' command to remove the entire area."])

        # Store originals for cleanup
        ori_position_shape_tuple = ns_def.convert_array_to_tuple(position_shape_array)
        ori_style_shape_tuple = ns_def.convert_array_to_tuple(style_shape_array)
        ori_position_line_tuple = ns_def.convert_array_to_tuple(position_line_array)
        ori_position_tag_tuple = ns_def.convert_array_to_tuple(position_tag_array)
        ori_attribute_tuple = ns_def.convert_array_to_tuple(attribute_array)
        ori_position_folder_tuple = ns_def.convert_array_to_tuple(position_folder_array)

        # 1. Delete from POSITION_SHAPE and replace with _AIR_
        grid = []
        for item in position_shape_array:
            if isinstance(item, list) and len(item) >= 2 and isinstance(item[1], list):
                row = item[1]
                # Replace element name with _AIR_
                for i, val in enumerate(row):
                    if val == name:
                        row[i] = '_AIR_'
                grid.append(row)
            else:
                grid.append([''])

        # Apply the remove_air_only_rows_and_cols function
        grid = def_common.remove_air_only_rows_and_cols(grid)

        # Convert grid back to position_shape_array format
        position_shape_array = []
        for i, row in enumerate(grid, start=1):
            position_shape_array.append([i, row])

        # 2. Delete from STYLE_SHAPE
        style_shape_array = [item for item in style_shape_array
                             if not (len(item) >= 2 and isinstance(item[1], list)
                                     and len(item[1]) > 0 and item[1][0] == name)]

        # Renumber STYLE_SHAPE
        for i, item in enumerate(style_shape_array):
            item[0] = i + 1

        # 3. Delete from POSITION_LINE
        header_rows = position_line_array[:2]
        data_rows = [item for item in position_line_array[2:]
                     if not (len(item) >= 2 and isinstance(item[1], list)
                             and len(item[1]) >= 2
                             and (item[1][0] == name or item[1][1] == name))]

        position_line_array = header_rows + data_rows

        # Renumber POSITION_LINE
        for i, item in enumerate(position_line_array):
            item[0] = i + 1

        # 4. Delete from POSITION_TAG
        position_tag_array = [item for item in position_tag_array
                              if not (len(item) >= 2 and isinstance(item[1], list)
                                      and len(item[1]) > 0 and item[1][0] == name)]

        # Renumber POSITION_TAG
        for i, item in enumerate(position_tag_array):
            item[0] = i + 1

        # 5. Delete from ATTRIBUTE
        attribute_array = [item for item in attribute_array
                           if not (len(item) >= 2 and isinstance(item[1], list)
                                   and len(item[1]) > 0 and item[1][0] == name)]

        # Renumber ATTRIBUTE
        for i, item in enumerate(attribute_array):
            item[0] = i + 1

        # Convert to tuples
        position_shape_tuple = ns_def.convert_array_to_tuple(position_shape_array)
        style_shape_tuple = ns_def.convert_array_to_tuple(style_shape_array)
        position_line_tuple = ns_def.convert_array_to_tuple(position_line_array)
        position_tag_tuple = ns_def.convert_array_to_tuple(position_tag_array)
        attribute_tuple = ns_def.convert_array_to_tuple(attribute_array)

        # Write updates to Excel
        ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_position_shape_tuple)
        ns_def.write_excel_meta(position_shape_tuple, master_file_path,
                                'Master_Data', '<<POSITION_SHAPE>>', 0, 0)

        ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_style_shape_tuple)
        ns_def.write_excel_meta(style_shape_tuple, master_file_path,
                                'Master_Data', '<<STYLE_SHAPE>>', 0, 0)

        ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_position_line_tuple)
        ns_def.write_excel_meta(position_line_tuple, master_file_path,
                                'Master_Data', '<<POSITION_LINE>>', 0, 0)

        ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_position_tag_tuple)
        ns_def.write_excel_meta(position_tag_tuple, master_file_path,
                                'Master_Data', '<<POSITION_TAG>>', 0, 0)

        ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_attribute_tuple)
        ns_def.write_excel_meta(attribute_tuple, master_file_path,
                                'Master_Data', '<<ATTRIBUTE>>', 0, 0)

        # 6. Recalculate POSITION_FOLDER
        position_folder_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_FOLDER>>')
        position_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_SHAPE>>')
        style_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<STYLE_SHAPE>>')

        position_folder_tuple = ns_def.convert_array_to_tuple(position_folder_array)
        position_shape_tuple = ns_def.convert_array_to_tuple(position_shape_array)
        style_shape_tuple = ns_def.convert_array_to_tuple(style_shape_array)

        update_position_folder_tuple = def_common.make_position_folder_tuple(
            position_folder_tuple,
            style_shape_tuple,
            position_shape_tuple
        )

        ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_position_folder_tuple)
        ns_def.write_excel_meta(update_position_folder_tuple, master_file_path,
                                'Master_Data', '<<POSITION_FOLDER>>', 0, 0)

        # ★★★ ADD: Call recalculate_folder_sizes ★★★
        result = def_common.recalculate_folder_sizes(master_file_path)
        if result['status'] != 'success':
            return ([f"[ERROR] Failed to recalculate folder sizes: {result['message']}"])
        # ★★★ End of addition ★★★

        # Sync with L2/L3 layers
        import tkinter as tk
        dummy_tk = tk.Toplevel()
        dummy_tk.withdraw()
        self.full_filepath = master_file_path
        self.main1_1_entry_1 = tk.Entry(dummy_tk)
        self.main1_1_entry_1.insert(tk.END, master_file_path)

        import ns_sync_between_layers
        ns_sync_between_layers.l1_master_device_and_line_sync_with_l2l3_master(self)

        dummy_tk.destroy()

        element_label = 'Device' if element_type == 'device' else 'Waypoint'
        return_text = f'--- {element_label} deleted --- {name}'
        return ([return_text])



    @staticmethod
    def remove_air_only_rows_and_cols(grid):
        """
        Remove rows and columns that contain only _AIR_ within each group
        Also handles folder name movement when first row contains only _AIR_ devices
        """
        if not grid:
            return grid

        AIR_MARK = "_AIR_"
        END_MARK = "<END>"

        new_grid = []
        current_group = []

        for r, row in enumerate(grid):
            # Always keep <<POSITION_SHAPE>> row as is
            if r == 0:
                new_grid.append(row)
            # Process group when we hit <END> row
            elif len(row) == 1 and row[0] == END_MARK:
                # Process current group
                if current_group:
                    # Check if first row (with folder name) has only _AIR_ in device positions
                    if len(current_group) > 0:
                        first_row = current_group[0]
                        folder_name = first_row[0] if first_row[0] and first_row[0] not in [AIR_MARK, ''] else ''

                        if folder_name:
                            # Check if all elements after folder name are _AIR_ or empty or <END>
                            device_elements = [v for v in first_row[1:] if v not in [AIR_MARK, '', END_MARK]]

                            if not device_elements and len(current_group) > 1:
                                # First row has only _AIR_ after folder name, move folder name to next row
                                # Find the next row with actual content
                                for next_idx in range(1, len(current_group)):
                                    next_row = current_group[next_idx]
                                    next_device_elements = [v for v in next_row if v not in [AIR_MARK, '', END_MARK]]
                                    if next_device_elements:
                                        # Move folder name to this row
                                        current_group[next_idx][0] = folder_name
                                        # Remove the first row from group
                                        current_group.pop(0)
                                        break

                    # Now process the group to remove _AIR_-only rows
                    filtered_group = []
                    for idx, group_row in enumerate(current_group):
                        # Check if row has any non-AIR content (excluding <END>)
                        # Keep row if it has folder name or any actual device
                        has_folder_name = group_row[0] and group_row[0] not in [AIR_MARK, '']
                        has_device = any(v not in [AIR_MARK, '', END_MARK] for v in group_row)

                        if has_folder_name or has_device:
                            filtered_group.append(group_row)
                        # Skip rows that are entirely _AIR_ or empty

                    # Remove _AIR_-only columns from filtered group
                    if filtered_group:
                        max_cols = max(len(row) for row in filtered_group)
                        cols_to_remove = []

                        for c in range(max_cols):
                            has_content = False
                            for row in filtered_group:
                                if c < len(row) and row[c] not in [AIR_MARK, '', END_MARK]:
                                    has_content = True
                                    break
                            if not has_content:
                                cols_to_remove.append(c)

                        # Remove columns (but keep at least the folder name column if it exists)
                        for row in filtered_group:
                            # Track if this row has folder name before removing columns
                            has_folder = row[0] and row[0] not in [AIR_MARK, '']

                            for c in sorted(cols_to_remove, reverse=True):
                                # Don't remove column 0 if it contains folder name
                                if c == 0 and has_folder:
                                    continue
                                if c < len(row) and row[c] != END_MARK:
                                    row.pop(c)

                        # Add filtered group to new grid
                        new_grid.extend(filtered_group)

                # Add the <END> row - ALWAYS just [<END>]
                new_grid.append([END_MARK])
                current_group = []
            else:
                # Accumulate group rows
                current_group.append(row)

        # Handle last group if no ending <END>
        if current_group:
            # Same logic as above for the last group
            if len(current_group) > 0:
                first_row = current_group[0]
                folder_name = first_row[0] if first_row[0] and first_row[0] not in [AIR_MARK, ''] else ''

                if folder_name:
                    device_elements = [v for v in first_row[1:] if v not in [AIR_MARK, '', END_MARK]]

                    if not device_elements and len(current_group) > 1:
                        for next_idx in range(1, len(current_group)):
                            next_row = current_group[next_idx]
                            next_device_elements = [v for v in next_row if v not in [AIR_MARK, '', END_MARK]]
                            if next_device_elements:
                                current_group[next_idx][0] = folder_name
                                current_group.pop(0)
                                break

            filtered_group = []
            for idx, group_row in enumerate(current_group):
                has_folder_name = group_row[0] and group_row[0] not in [AIR_MARK, '']
                has_device = any(v not in [AIR_MARK, '', END_MARK] for v in group_row)

                if has_folder_name or has_device:
                    filtered_group.append(group_row)

            new_grid.extend(filtered_group)

        # Ensure last row is <END> if not already
        if not new_grid or new_grid[-1] != [END_MARK]:
            new_grid.append([END_MARK])

        return new_grid

    @staticmethod
    def add_device_or_waypoint_common(self, name, reference_device, direction, master_file_path, device_type='device'):
        """
        Common function to add device or waypoint

        Args:
            self: Reference to calling instance
            name: Name of the device or waypoint to add
            reference_device: Reference device for positioning
            direction: Direction relative to reference device
            master_file_path: Path to the master Excel file
            device_type: 'device' or 'waypoint'

        Returns:
            List with status message
        """
        import ns_def

        # Validate direction
        valid_directions = [
            'UP', 'DOWN', 'LEFT', 'RIGHT',
            'UP_WITH_GRID', 'DOWN_WITH_GRID',
            'LEFT_WITH_GRID', 'RIGHT_WITH_GRID'
        ]

        if direction not in valid_directions:
            return ([f"[Error] Invalid direction '{direction}'. Valid directions: {', '.join(valid_directions)}"])

        # Load and prepare position_shape array
        position_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_SHAPE>>')

        # Add to POSITION_SHAPE
        result_position = def_common.add_device_position_shape(
            device_name=name,
            reference_device=reference_device,
            direction=direction,
            position_shape_array=position_shape_array,
            master_file_path=master_file_path
        )

        # Process result
        if result_position['status'] == 'success':
            # Add to STYLE_SHAPE, POSITION_TAG, and ATTRIBUTE with appropriate type
            if device_type == 'waypoint':
                # Pass direction to add_waypoint_style_shape_tag for roundness determination
                result_style_tag = def_common.add_waypoint_style_shape_tag(
                    waypoint_name=name,
                    master_file_path=master_file_path,
                    direction=direction  # Pass direction parameter
                )
                return_text = f"--- Waypoint added --- {name} relative to {reference_device} with direction {direction}"
            else:
                result_style_tag = def_common.add_device_style_shape_tag(
                    device_name=name,
                    master_file_path=master_file_path
                )
                return_text = f"--- Device added --- {name} relative to {reference_device} with direction {direction}"

            if result_style_tag['status'] == 'success':
                # Sync with L2/L3 layers (for both device and waypoint)
                import tkinter as tk
                dummy_tk = tk.Toplevel()
                dummy_tk.withdraw()

                self.full_filepath = master_file_path
                self.main1_1_entry_1 = tk.Entry(dummy_tk)
                self.main1_1_entry_1.insert(tk.END, master_file_path)

                # ★★★ Add missing attributes ★★★
                self.inFileTxt_L3_1_1 = tk.Entry(dummy_tk)
                self.inFileTxt_L3_1_1.delete(0, tkinter.END)
                self.inFileTxt_L3_1_1.insert(tk.END, master_file_path)

                self.outFileTxt_11_2 = tk.Entry(dummy_tk)
                self.outFileTxt_11_2.delete(0, tkinter.END)
                self.outFileTxt_11_2.insert(tk.END, master_file_path)

                self.inFileTxt_L2_1_1 = tk.Entry(dummy_tk)
                self.inFileTxt_L2_1_1.delete(0, tkinter.END)
                self.inFileTxt_L2_1_1.insert(tk.END, master_file_path)

                import ns_sync_between_layers
                ns_sync_between_layers.l1_master_device_and_line_sync_with_l2l3_master(self)

                # ★★★ Delete L2_TABLE file if exists ★★★
                tmp_delete_excel_name = master_file_path.replace('[MASTER]', '[L2_TABLE]')
                if os.path.isfile(tmp_delete_excel_name):
                    os.remove(tmp_delete_excel_name)

                dummy_tk.destroy()

                return ([return_text])
            else:
                return ([f"[Error] {result_style_tag['message']}"])
        else:
            return ([f"[Error] {result_position['message']}"])


    @staticmethod
    def add_device_style_shape_tag(device_name, master_file_path):
        """
        Add a new device to <<STYLE_SHAPE>> and <<POSITION_TAG>> sections

        Args:
            device_name: Name of the device to add
            master_file_path: Path to the master Excel file

        Returns:
            dict: Status and message of the operation
        """
        import ns_def

        try:
            # Parameters (from ns_l1_master_create)
            shape_width_min = 0.4  # inches
            shape_hight_min = 0.2  # inches
            shae_font_size = 6.0  # pt
            tag_offet_inche = 0.02  # inches per character
            default_color = 'GREEN'  # Default color for devices
            default_roundness = 0  # Default roundness (0.0-1.0)

            # ========== Add to <<STYLE_SHAPE>> ==========
            style_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<STYLE_SHAPE>>')

            # Store original for cleanup
            ori_style_shape_tuple = ns_def.convert_array_to_tuple(style_shape_array)

            # Check if device already exists in STYLE_SHAPE
            device_exists_in_style = False
            for item in style_shape_array:
                if len(item) >= 2 and isinstance(item[1], list) and len(item[1]) > 0:
                    if item[1][0] == device_name:
                        device_exists_in_style = True
                        break

            if not device_exists_in_style:
                # Calculate device name width and height
                text_dimensions = ns_def.get_description_width_hight(shae_font_size, device_name)
                num_char_width = text_dimensions[0]

                # Determine dimensions (ensure minimum values)
                line_width = num_char_width if num_char_width > shape_width_min else shape_width_min
                line_hight = shape_hight_min

                # Find the highest number in style_shape_array
                max_number = 0
                for item in style_shape_array:
                    if isinstance(item[0], int) and item[0] > max_number:
                        max_number = item[0]

                # Skip header rows (usually rows 1-3 contain headers and defaults)
                if max_number < 3:
                    max_number = 3

                new_style_number = max_number + 1

                # Create new style entry
                new_style_entry = [new_style_number, [
                    device_name,  # Device name
                    line_width,  # Width (inches)
                    line_hight,  # Height (inches)
                    default_roundness,  # Roundness (0.0-1.0)
                    default_color  # Color
                ]]

                # Add to array
                style_shape_array.append(new_style_entry)

                # Convert to tuple for writing
                style_shape_tuple = ns_def.convert_array_to_tuple(style_shape_array)

                # Remove existing rows under section before writing new data
                ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_style_shape_tuple)

                # Write new data
                ns_def.write_excel_meta(style_shape_tuple, master_file_path,
                                        'Master_Data', '<<STYLE_SHAPE>>', 0, 0)

            # ========== Add to <<POSITION_TAG>> ==========
            position_tag_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_TAG>>')

            # Store original for cleanup
            ori_position_tag_tuple = ns_def.convert_array_to_tuple(position_tag_array)

            # Check if device already exists in POSITION_TAG
            device_exists_in_tag = False
            for item in position_tag_array:
                if len(item) >= 2 and isinstance(item[1], list) and len(item[1]) > 0:
                    if item[1][0] == device_name:
                        device_exists_in_tag = True
                        break

            if not device_exists_in_tag:
                # Find the highest number in position_tag_array
                max_tag_number = 0
                for item in position_tag_array:
                    if isinstance(item[0], int) and item[0] > max_tag_number:
                        max_tag_number = item[0]

                # Skip header rows
                if max_tag_number < 2:
                    max_tag_number = 2

                new_tag_number = max_tag_number + 1

                # Calculate initial tag offset (can be updated later based on actual tag names)
                initial_tag_offset = 1 * tag_offet_inche  # Default 1 character

                # Create new tag entry
                new_tag_entry = [new_tag_number, [
                    device_name,  # Device name
                    'LINE',  # Type (LINE or SHAPE)
                    '',  # Offset_SHAPE_X (empty for LINE type)
                    '',  # Offset_SHAPE_Y (empty for LINE type)
                    initial_tag_offset,  # Offset_LINE (inches)
                    'YES'  # Adjust_LINE_Angle
                ]]

                # Add to array
                position_tag_array.append(new_tag_entry)

                # Convert to tuple for writing
                position_tag_tuple = ns_def.convert_array_to_tuple(position_tag_array)

                # Remove existing rows under section before writing new data
                ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_position_tag_tuple)

                # Write new data
                ns_def.write_excel_meta(position_tag_tuple, master_file_path,
                                        'Master_Data', '<<POSITION_TAG>>', 0, 0)

            # ========== Add to <<ATTRIBUTE>> ==========
            attribute_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<ATTRIBUTE>>')

            # Store original for cleanup
            ori_attribute_tuple = ns_def.convert_array_to_tuple(attribute_array)

            # Check if device already exists in ATTRIBUTE
            device_exists_in_attribute = False
            for item in attribute_array:
                if len(item) >= 2 and isinstance(item[1], list) and len(item[1]) > 0:
                    if item[1][0] == device_name:
                        device_exists_in_attribute = True
                        break

            if not device_exists_in_attribute:
                # Find the highest number in attribute_array
                max_attr_number = 0
                for item in attribute_array:
                    if isinstance(item[0], int) and item[0] > max_attr_number:
                        max_attr_number = item[0]

                # Skip header row
                if max_attr_number < 1:
                    max_attr_number = 1

                new_attr_number = max_attr_number + 1

                # Create new attribute entry
                new_attr_entry = [new_attr_number, [
                    device_name,  # Device name
                    '[\'DEVICE\',[235, 241, 222]]',  # Default (GREEN color for DEVICE)
                    '[\'<EMPTY>\', [255, 255, 255]]',  # Attribute-A
                    '[\'<EMPTY>\', [255, 255, 255]]',  # Attribute-B
                    '[\'<EMPTY>\', [255, 255, 255]]',  # Attribute-C
                    '[\'<EMPTY>\', [255, 255, 255]]',  # Attribute-D
                    '[\'<EMPTY>\', [255, 255, 255]]',  # Attribute-E
                    '[\'<EMPTY>\', [255, 255, 255]]',  # Attribute-F
                    '[\'<EMPTY>\', [255, 255, 255]]',  # Attribute-G
                    '[\'<EMPTY>\', [255, 255, 255]]'  # Attribute-H
                ]]

                # Add to array
                attribute_array.append(new_attr_entry)

                # Convert to tuple for writing
                attribute_tuple = ns_def.convert_array_to_tuple(attribute_array)

                # Remove existing rows under section before writing new data
                ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_attribute_tuple)

                # Write new data
                ns_def.write_excel_meta(attribute_tuple, master_file_path,
                                        'Master_Data', '<<ATTRIBUTE>>', 0, 0)

            # ========== Update <<POSITION_FOLDER>> using make_position_folder_tuple ==========
            # Read the latest data after all updates
            position_folder_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_FOLDER>>')
            position_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_SHAPE>>')
            style_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<STYLE_SHAPE>>')

            # Convert to tuples
            position_folder_tuple = ns_def.convert_array_to_tuple(position_folder_array)
            position_shape_tuple = ns_def.convert_array_to_tuple(position_shape_array)
            style_shape_tuple = ns_def.convert_array_to_tuple(style_shape_array)

            # Store original for cleanup
            ori_position_folder_tuple = position_folder_tuple

            # Call make_position_folder_tuple to update folder sizes
            update_position_folder_tuple = def_common.make_position_folder_tuple(
                position_folder_tuple,
                style_shape_tuple,
                position_shape_tuple
            )

            # Remove existing rows under section before writing new data
            ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_position_folder_tuple)

            # Write updated position folder data
            ns_def.write_excel_meta(update_position_folder_tuple, master_file_path,
                                    'Master_Data', '<<POSITION_FOLDER>>', 0, 0)

            # ★★★ ADD: Call recalculate_folder_sizes ★★★
            result = def_common.recalculate_folder_sizes(master_file_path)
            if result['status'] != 'success':
                return {
                    'status': 'error',
                    'message': f"Failed to recalculate folder sizes: {result['message']}"
                }
            # ★★★ End of addition ★★★

            return {
                'status': 'success',
                'message': f"Device '{device_name}' successfully added to STYLE_SHAPE, POSITION_TAG, ATTRIBUTE, and POSITION_FOLDER updated"
            }

        except Exception as e:
            import traceback
            traceback.print_exc()
            return {
                'status': 'error',
                'message': f"Error adding device to STYLE_SHAPE/POSITION_TAG: {str(e)}"
            }

    @staticmethod
    def add_waypoint_style_shape_tag(waypoint_name, master_file_path, direction=None):
        """
        Add a new waypoint to <<STYLE_SHAPE>> and <<POSITION_TAG>> sections
        Waypoints have specific styling: BLUE color, roundness depends on direction, min width 0.6, min height 0.3

        Args:
            waypoint_name: Name of the waypoint to add
            master_file_path: Path to the master Excel file
            direction: Direction of the waypoint (UP/DOWN/LEFT/RIGHT) - optional

        Returns:
            dict: Status and message of the operation
        """
        import ns_def

        try:
            # Parameters specific for waypoints
            shape_width_min = 0.6  # inches (larger than device)
            shape_hight_min = 0.3  # inches (larger than device)
            shae_font_size = 6.0  # pt
            tag_offet_inche = 0.02  # inches per character
            waypoint_color = 'BLUE'  # Waypoint color

            # Determine roundness based on direction
            if direction in ['RIGHT', 'LEFT']:
                waypoint_roundness = 0.5  # Higher roundness for RIGHT/LEFT
            else:
                waypoint_roundness = 0.2  # Default roundness for UP/DOWN

            # ========== Add to <<STYLE_SHAPE>> ==========
            style_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<STYLE_SHAPE>>')

            # Store original for cleanup
            ori_style_shape_tuple = ns_def.convert_array_to_tuple(style_shape_array)

            # Check if waypoint already exists in STYLE_SHAPE
            waypoint_exists_in_style = False
            for item in style_shape_array:
                if len(item) >= 2 and isinstance(item[1], list) and len(item[1]) > 0:
                    if item[1][0] == waypoint_name:
                        waypoint_exists_in_style = True
                        break

            if not waypoint_exists_in_style:
                # Calculate waypoint name width and height
                text_dimensions = ns_def.get_description_width_hight(shae_font_size, waypoint_name)
                num_char_width = text_dimensions[0]

                # Determine dimensions (ensure minimum values for waypoint)
                line_width = num_char_width if num_char_width > shape_width_min else shape_width_min
                line_hight = shape_hight_min

                # Find the highest number in style_shape_array
                max_number = 0
                for item in style_shape_array:
                    if isinstance(item[0], int) and item[0] > max_number:
                        max_number = item[0]

                # Skip header rows
                if max_number < 3:
                    max_number = 3

                new_style_number = max_number + 1

                # Create new style entry for waypoint
                new_style_entry = [new_style_number, [
                    waypoint_name,  # Waypoint name
                    line_width,  # Width (inches)
                    line_hight,  # Height (inches)
                    waypoint_roundness,  # Roundness (0.5 for RIGHT/LEFT, 0.2 for UP/DOWN)
                    waypoint_color  # Color (BLUE for waypoint)
                ]]

                # Add to array
                style_shape_array.append(new_style_entry)

                # Convert to tuple for writing
                style_shape_tuple = ns_def.convert_array_to_tuple(style_shape_array)

                # Remove existing rows under section before writing new data
                ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_style_shape_tuple)

                # Write new data
                ns_def.write_excel_meta(style_shape_tuple, master_file_path,
                                        'Master_Data', '<<STYLE_SHAPE>>', 0, 0)

            # ========== Add to <<POSITION_TAG>> ==========
            position_tag_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_TAG>>')

            # Store original for cleanup
            ori_position_tag_tuple = ns_def.convert_array_to_tuple(position_tag_array)

            # Check if waypoint already exists in POSITION_TAG
            waypoint_exists_in_tag = False
            for item in position_tag_array:
                if len(item) >= 2 and isinstance(item[1], list) and len(item[1]) > 0:
                    if item[1][0] == waypoint_name:
                        waypoint_exists_in_tag = True
                        break

            if not waypoint_exists_in_tag:
                # Find the highest number in position_tag_array
                max_tag_number = 0
                for item in position_tag_array:
                    if isinstance(item[0], int) and item[0] > max_tag_number:
                        max_tag_number = item[0]

                # Skip header rows
                if max_tag_number < 2:
                    max_tag_number = 2

                new_tag_number = max_tag_number + 1

                # Calculate initial tag offset
                initial_tag_offset = 1 * tag_offet_inche

                # Create new tag entry
                new_tag_entry = [new_tag_number, [
                    waypoint_name,  # Waypoint name
                    'LINE',  # Type
                    '',  # Offset_SHAPE_X
                    '',  # Offset_SHAPE_Y
                    initial_tag_offset,  # Offset_LINE
                    'YES'  # Adjust_LINE_Angle
                ]]

                # Add to array
                position_tag_array.append(new_tag_entry)

                # Convert to tuple for writing
                position_tag_tuple = ns_def.convert_array_to_tuple(position_tag_array)

                # Remove existing rows under section before writing new data
                ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_position_tag_tuple)

                # Write new data
                ns_def.write_excel_meta(position_tag_tuple, master_file_path,
                                        'Master_Data', '<<POSITION_TAG>>', 0, 0)

            # ========== Add to <<ATTRIBUTE>> ==========
            attribute_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<ATTRIBUTE>>')

            # Store original for cleanup
            ori_attribute_tuple = ns_def.convert_array_to_tuple(attribute_array)

            # Check if waypoint already exists in ATTRIBUTE
            waypoint_exists_in_attribute = False
            for item in attribute_array:
                if len(item) >= 2 and isinstance(item[1], list) and len(item[1]) > 0:
                    if item[1][0] == waypoint_name:
                        waypoint_exists_in_attribute = True
                        break

            if not waypoint_exists_in_attribute:
                # Find the highest number in attribute_array
                max_attr_number = 0
                for item in attribute_array:
                    if isinstance(item[0], int) and item[0] > max_attr_number:
                        max_attr_number = item[0]

                # Skip header row
                if max_attr_number < 1:
                    max_attr_number = 1

                new_attr_number = max_attr_number + 1

                # Create new attribute entry for waypoint
                new_attr_entry = [new_attr_number, [
                    waypoint_name,  # Waypoint name
                    '[\'WayPoint\', [220, 230, 242]]',  # Default (BLUE color for WayPoint)
                    '[\'<EMPTY>\', [255, 255, 255]]',  # Attribute-A
                    '[\'<EMPTY>\', [255, 255, 255]]',  # Attribute-B
                    '[\'<EMPTY>\', [255, 255, 255]]',  # Attribute-C
                    '[\'<EMPTY>\', [255, 255, 255]]',  # Attribute-D
                    '[\'<EMPTY>\', [255, 255, 255]]',  # Attribute-E
                    '[\'<EMPTY>\', [255, 255, 255]]',  # Attribute-F
                    '[\'<EMPTY>\', [255, 255, 255]]',  # Attribute-G
                    '[\'<EMPTY>\', [255, 255, 255]]'  # Attribute-H
                ]]

                # Add to array
                attribute_array.append(new_attr_entry)

                # Convert to tuple for writing
                attribute_tuple = ns_def.convert_array_to_tuple(attribute_array)

                # Remove existing rows under section before writing new data
                ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_attribute_tuple)

                # Write new data
                ns_def.write_excel_meta(attribute_tuple, master_file_path,
                                        'Master_Data', '<<ATTRIBUTE>>', 0, 0)

            # ========== Update <<POSITION_FOLDER>> using make_position_folder_tuple ==========
            # Read the latest data after all updates
            position_folder_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_FOLDER>>')
            position_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<POSITION_SHAPE>>')
            style_shape_array = ns_def.convert_master_to_array('Master_Data', master_file_path, '<<STYLE_SHAPE>>')

            # Convert to tuples
            position_folder_tuple = ns_def.convert_array_to_tuple(position_folder_array)
            position_shape_tuple = ns_def.convert_array_to_tuple(position_shape_array)
            style_shape_tuple = ns_def.convert_array_to_tuple(style_shape_array)

            # Store original for cleanup
            ori_position_folder_tuple = position_folder_tuple

            # Call make_position_folder_tuple to update folder sizes
            update_position_folder_tuple = def_common.make_position_folder_tuple(
                position_folder_tuple,
                style_shape_tuple,
                position_shape_tuple
            )

            # Remove existing rows under section before writing new data
            ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_position_folder_tuple)

            # Write updated position folder data
            ns_def.write_excel_meta(update_position_folder_tuple, master_file_path,
                                    'Master_Data', '<<POSITION_FOLDER>>', 0, 0)

            # ★★★ ADD: Call recalculate_folder_sizes ★★★
            result = def_common.recalculate_folder_sizes(master_file_path)
            if result['status'] != 'success':
                return {
                    'status': 'error',
                    'message': f"Failed to recalculate folder sizes: {result['message']}"
                }
            # ★★★ End of addition ★★★

            return {
                'status': 'success',
                'message': f"Waypoint '{waypoint_name}' successfully added to STYLE_SHAPE, POSITION_TAG, ATTRIBUTE, and POSITION_FOLDER updated"
            }

        except Exception as e:
            import traceback
            traceback.print_exc()
            return {
                'status': 'error',
                'message': f"Error adding waypoint to STYLE_SHAPE/POSITION_TAG: {str(e)}"
            }



    def add_device_position_shape(device_name, reference_device, direction, position_shape_array, master_file_path):
        """
        Add a new device at a specific position relative to a reference device in position_shape array
        """
        import copy
        import ns_def

        try:
            # Store original position_shape for cleanup
            ori_position_shape_tuple = ns_def.convert_array_to_tuple(position_shape_array)

            # Convert position_shape_array to 2D grid (skip the first element which is row number)
            grid = []
            for item in position_shape_array:
                if isinstance(item, list) and len(item) >= 2 and isinstance(item[1], list):
                    grid.append(item[1])
                else:
                    grid.append([''])

            # Set target device and new device name
            TARGET = reference_device
            NEWDEV = device_name
            END_MARK = "<END>"
            AIR_MARK = "_AIR_"
            HEADER_MARK = "<<POSITION_SHAPE>>"

            # Helper functions
            def find_target_with_group(grid, target):
                """Return the (r, c, group_start, group_end) position where target appears in grid."""
                current_group_start = 1  # Skip <<POSITION_SHAPE>> row

                for r, row in enumerate(grid):
                    # Skip header row
                    if r == 0:
                        continue

                    # Check if this is a group header (first element is non-empty and not a special marker)
                    if len(row) > 0 and row[0] and row[0] not in ['', END_MARK, HEADER_MARK]:
                        current_group_start = r

                    # Check if this is the end of a group
                    if len(row) == 1 and row[0] == END_MARK:
                        # Search for target in current group
                        for gr in range(current_group_start, r):
                            for c, val in enumerate(grid[gr]):
                                if val == target:  # Exact match
                                    return (gr, c, current_group_start, r - 1)
                        current_group_start = r + 1

                return None

            def clone_grid(grid):
                """Create a deep copy of the grid"""
                return copy.deepcopy(grid)

            def normalize_grid(grid):
                """Ensure all rows end with <END> at the same column position and fill with _AIR_"""
                if not grid:
                    return grid

                # Process each group separately
                current_group_start = 1  # Skip <<POSITION_SHAPE>> row

                for r in range(len(grid)):
                    row = grid[r]

                    # Skip header row - don't modify it
                    if r == 0:
                        continue

                    # Keep <END>-only rows as is
                    if len(row) == 1 and row[0] == END_MARK:
                        current_group_start = r + 1
                        continue

                    # Process group when we reach the last row or next <END> row
                    next_is_end = (r + 1 < len(grid) and len(grid[r + 1]) == 1 and grid[r + 1][0] == END_MARK)
                    is_last = (r == len(grid) - 1)

                    if next_is_end or is_last:
                        group_end = r

                        if group_end >= current_group_start:
                            # Find max END position in this group
                            max_end_col = 0
                            for gr in range(current_group_start, group_end + 1):
                                for c, val in enumerate(grid[gr]):
                                    if val == END_MARK:
                                        max_end_col = max(max_end_col, c)
                                        break

                            # Normalize rows in this group
                            for gr in range(current_group_start, group_end + 1):
                                row_to_norm = grid[gr]

                                # Remove existing <END> markers
                                while END_MARK in row_to_norm:
                                    row_to_norm.remove(END_MARK)

                                # Fill with _AIR_ to reach max_end_col if needed
                                while len(row_to_norm) < max_end_col:
                                    row_to_norm.append(AIR_MARK)

                                # Add <END> at the end
                                row_to_norm.append(END_MARK)

                return grid



            # Find target position with group information
            target_info = find_target_with_group(grid, TARGET)
            if target_info is None:
                return {
                    'status': 'error',
                    'message': f"Reference device '{reference_device}' not found in position_shape"
                }

            r, c, group_start, group_end = target_info

            # Check if new device already exists
            if find_target_with_group(grid, NEWDEV) is not None:
                return {
                    'status': 'error',
                    'message': f"Device '{device_name}' already exists in position_shape"
                }

            # Apply transformation based on direction
            new_grid = clone_grid(grid)

            # Determine if we need to add a row/column (edge case handling)
            need_with_grid = False

            if direction == 'UP':
                # Check if we're at the top of the group
                if r == group_start:
                    need_with_grid = True
                elif r > group_start:
                    # Check if position exists and is not AIR/empty
                    if c >= len(new_grid[r - 1]):
                        need_with_grid = True
                    elif new_grid[r - 1][c] not in [AIR_MARK, '']:
                        need_with_grid = True

                if need_with_grid:
                    direction = 'UP_WITH_GRID'

            elif direction == 'DOWN':
                # Check if we're at the bottom of the group
                if r == group_end:
                    need_with_grid = True
                elif r < group_end:
                    # Check if position exists and is not AIR/empty
                    if c >= len(new_grid[r + 1]):
                        need_with_grid = True
                    elif new_grid[r + 1][c] not in [AIR_MARK, '']:
                        need_with_grid = True

                if need_with_grid:
                    direction = 'DOWN_WITH_GRID'

            elif direction == 'LEFT':
                if c == 0 or (c == 1 and r == group_start and new_grid[r][0] not in ['', AIR_MARK]):
                    need_with_grid = True
                elif c > 0 and new_grid[r][c - 1] not in [AIR_MARK, '']:
                    need_with_grid = True

                if need_with_grid:
                    direction = 'LEFT_WITH_GRID'

            elif direction == 'RIGHT':
                # Check if we're at the right edge (before <END>)
                if c >= len(new_grid[r]) - 2:  # -2 because of <END>
                    need_with_grid = True
                elif c < len(new_grid[r]) - 1 and new_grid[r][c + 1] == END_MARK:
                    need_with_grid = True
                elif c < len(new_grid[r]) - 1 and new_grid[r][c + 1] not in [AIR_MARK, '']:
                    need_with_grid = True

                if need_with_grid:
                    direction = 'RIGHT_WITH_GRID'

            # Execute the transformation
            if direction == 'UP':
                # Ensure the row above has enough columns
                while len(new_grid[r - 1]) <= c:
                    # Insert before <END> if it exists
                    insert_pos = len(new_grid[r - 1]) - 1 if END_MARK in new_grid[r - 1] else len(new_grid[r - 1])
                    new_grid[r - 1].insert(insert_pos, AIR_MARK)

                if new_grid[r - 1][c] == AIR_MARK:
                    new_grid[r - 1][c] = NEWDEV
                else:
                    new_grid[r - 1].insert(c, NEWDEV)

            elif direction == 'DOWN':
                # Ensure the row below has enough columns
                while len(new_grid[r + 1]) <= c:
                    # Insert before <END> if it exists
                    insert_pos = len(new_grid[r + 1]) - 1 if END_MARK in new_grid[r + 1] else len(new_grid[r + 1])
                    new_grid[r + 1].insert(insert_pos, AIR_MARK)

                if new_grid[r + 1][c] == AIR_MARK:
                    new_grid[r + 1][c] = NEWDEV
                else:
                    new_grid[r + 1].insert(c, NEWDEV)

            elif direction == 'LEFT':
                if new_grid[r][c - 1] == AIR_MARK:
                    # Simply replace the AIR with the new device
                    new_grid[r][c - 1] = NEWDEV
                else:
                    # Need to shift: use LEFT_WITH_GRID logic
                    # Check all rows in the group to maintain column alignment
                    for i in range(group_start, group_end + 1):
                        # Find <END> and remove it temporarily if it exists
                        had_end = END_MARK in new_grid[i]
                        if had_end:
                            new_grid[i].remove(END_MARK)

                        if i == r:
                            # Insert new device at position c
                            new_grid[i].insert(c, NEWDEV)
                        else:
                            # Check what's at position c in this row
                            if c < len(new_grid[i]):
                                # Check the content at position c and c-1
                                content_at_c = new_grid[i][c] if c < len(new_grid[i]) else AIR_MARK
                                content_at_c_minus_1 = new_grid[i][c - 1] if c > 0 and c - 1 < len(new_grid[i]) else AIR_MARK

                                # Only insert if we need to maintain column alignment
                                if content_at_c not in [AIR_MARK, ''] or content_at_c_minus_1 not in [AIR_MARK, '']:
                                    new_grid[i].insert(c, AIR_MARK)

                        # Add <END> back at the end if it was there
                        if had_end:
                            new_grid[i].append(END_MARK)

            elif direction == 'RIGHT':
                if new_grid[r][c + 1] == AIR_MARK:
                    # Simply replace the AIR with the new device
                    new_grid[r][c + 1] = NEWDEV
                else:
                    # Need to shift: use RIGHT_WITH_GRID logic
                    # Check all rows in the group to maintain column alignment
                    for i in range(group_start, group_end + 1):
                        # Find <END> and remove it temporarily if it exists
                        had_end = END_MARK in new_grid[i]
                        if had_end:
                            new_grid[i].remove(END_MARK)

                        if i == r:
                            # Insert new device at position c+1
                            new_grid[i].insert(c + 1, NEWDEV)
                        else:
                            # Check what's at position c+1 in this row
                            if c + 1 < len(new_grid[i]):
                                # Check the content at position c+1
                                content_at_c_plus_1 = new_grid[i][c + 1] if c + 1 < len(new_grid[i]) else AIR_MARK

                                # Only insert if we need to maintain column alignment
                                if content_at_c_plus_1 not in [AIR_MARK, '']:
                                    new_grid[i].insert(c + 1, AIR_MARK)

                        # Add <END> back at the end if it was there
                        if had_end:
                            new_grid[i].append(END_MARK)

            elif direction == 'UP_WITH_GRID':
                # Add new row at the top of the group
                # Get the maximum columns in the group
                max_cols = 0
                for gr in range(group_start, group_end + 1):
                    cols = len([v for v in new_grid[gr] if v != END_MARK])
                    max_cols = max(max_cols, cols)

                # If adding to the first row of the group, need to handle group name
                if r == group_start:
                    group_name = new_grid[r][0] if new_grid[r][0] not in ['', AIR_MARK] else ''
                    # Create new row with group name
                    new_row = []
                    if group_name:
                        new_row.append(group_name)
                        # Clear group name from original row
                        new_grid[r][0] = ''
                    else:
                        new_row.append('')

                    # Fill the rest of the row
                    for i in range(1, max(max_cols, c + 1)):
                        new_row.append(NEWDEV if i == c else AIR_MARK)

                    new_grid.insert(r, new_row)
                else:
                    # Normal row addition
                    new_row = ['' if i == 0 else (NEWDEV if i == c else AIR_MARK) for i in range(max(max_cols, c + 1))]
                    new_grid.insert(r, new_row)

            elif direction == 'DOWN_WITH_GRID':
                # Add new row at the bottom of the group
                max_cols = 0
                for gr in range(group_start, group_end + 1):
                    cols = len([v for v in new_grid[gr] if v != END_MARK])
                    max_cols = max(max_cols, cols)

                new_row = ['' if i == 0 else (NEWDEV if i == c else AIR_MARK) for i in range(max(max_cols, c + 1))]
                new_grid.insert(r + 1, new_row)

            elif direction == 'LEFT_WITH_GRID':
                # Add new column to the left of target device within the group only
                for i in range(group_start, group_end + 1):
                    # For group header row, don't insert before the group name
                    insert_pos = c
                    if i == group_start and c == 0 and new_grid[i][0] not in ['', AIR_MARK]:
                        insert_pos = 1  # Insert after group name

                    # Find <END> and remove it temporarily if it exists
                    had_end = END_MARK in new_grid[i]
                    if had_end:
                        new_grid[i].remove(END_MARK)

                    # Insert at the calculated position
                    if i == r:
                        new_grid[i].insert(insert_pos, NEWDEV)
                    else:
                        new_grid[i].insert(insert_pos, AIR_MARK)

                    # Add <END> back at the end if it was there
                    if had_end:
                        new_grid[i].append(END_MARK)

            elif direction == 'RIGHT_WITH_GRID':
                # Add new column to the right of target device within the group only
                # Insert position should be c+1 (right after the target)
                for i in range(group_start, group_end + 1):
                    # Find <END> and remove it temporarily if it exists
                    had_end = END_MARK in new_grid[i]
                    if had_end:
                        new_grid[i].remove(END_MARK)

                    # Insert at position c+1 (right after the target)
                    if i == r:
                        new_grid[i].insert(c + 1, NEWDEV)
                    else:
                        new_grid[i].insert(c + 1, AIR_MARK)

                    # Add <END> back at the end if it was there
                    if had_end:
                        new_grid[i].append(END_MARK)

            else:
                return {
                    'status': 'error',
                    'message': f"Invalid direction: {direction}"
                }

            # Normalize grid (align <END> markers and fill with _AIR_)
            new_grid = normalize_grid(new_grid)

            # Remove _AIR_-only rows and columns
            new_grid = def_common.remove_air_only_rows_and_cols(new_grid)

            # Convert grid back to position_shape_array format
            new_position_shape_array = []
            for i, row in enumerate(new_grid, start=1):
                new_position_shape_array.append([i, row])

            # Convert to tuple format for writing
            position_shape_tuple = ns_def.convert_array_to_tuple(new_position_shape_array)

            # Remove existing rows under section before writing new data
            ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_position_shape_tuple)

            # Write new data
            ns_def.write_excel_meta(position_shape_tuple, master_file_path,'Master_Data', '<<POSITION_SHAPE>>', 0, 0)

            return {
                'status': 'success',
                'message': f"Device '{device_name}' added successfully with direction '{direction}'"
            }

        except Exception as e:
            import traceback
            traceback.print_exc()
            return {
                'status': 'error',
                'message': f"Error adding device: {str(e)}"
            }




    def make_position_folder_tuple(master_folder_tuple, master_style_shape_tuple, master_shape_tuple):
        # Minimum tag inches between side of folder and edge shape (left and right)
        min_tag_inches = 0.3

        # Get each folder size
        import ns_def
        master_folder_size_array = ns_def.get_folder_width_size(
            master_folder_tuple, master_style_shape_tuple, master_shape_tuple, min_tag_inches
        )

        update_master_folder_tuple = {}

        for tmp_master_width_size_y_grid in master_folder_size_array[1]:
            for tmp_master_folder_size in master_folder_size_array[2]:
                if tmp_master_width_size_y_grid[0] == tmp_master_folder_size[0]:
                    if master_folder_size_array[0] == tmp_master_width_size_y_grid[1]:  # check max width in the slide
                        for tmp_master_folder_tuple in master_folder_tuple:
                            if (
                                    tmp_master_folder_tuple[0] == tmp_master_folder_size[0]
                                    and master_folder_tuple[tmp_master_folder_tuple] == tmp_master_folder_size[1][0][0]
                            ):
                                update_master_folder_tuple[tmp_master_folder_tuple[0] - 1, tmp_master_folder_tuple[1]] = tmp_master_folder_size[1][0][1]
                                update_master_folder_tuple[tmp_master_folder_tuple[0], tmp_master_folder_tuple[1]] = master_folder_tuple[tmp_master_folder_tuple]
                            elif (
                                    tmp_master_folder_tuple[0] == tmp_master_folder_size[0]
                                    and master_folder_tuple[tmp_master_folder_tuple] == ''
                            ):
                                update_master_folder_tuple[tmp_master_folder_tuple[0] - 1, tmp_master_folder_tuple[1]] = tmp_master_width_size_y_grid[2]
                                update_master_folder_tuple[tmp_master_folder_tuple[0], tmp_master_folder_tuple[1]] = ''
                            elif tmp_master_folder_tuple[1] == 1:  # write ALL column =1
                                update_master_folder_tuple[tmp_master_folder_tuple] = master_folder_tuple[tmp_master_folder_tuple]

                    else:  # insert empty folder to left and right side
                        tmp_max_row = 0
                        tmp_max_column = 0

                        tmp_bothside_empty = (master_folder_size_array[0] - tmp_master_width_size_y_grid[1]) * 0.25

                        for tmp_master_folder_tuple in master_folder_tuple:
                            # set first column
                            if tmp_master_folder_tuple[0] == tmp_master_folder_size[0]:

                                # check existence of tuples for bug fix 001
                                flag_exist_POSITION_FOLDER = False
                                flag_exist_SET_WIDTH = False
                                for tmp_bug_fix_tuple in master_folder_tuple:
                                    if (
                                            tmp_bug_fix_tuple[0] == tmp_master_folder_tuple[0] - 1
                                            and tmp_bug_fix_tuple[1] == tmp_master_folder_tuple[1] - 1
                                    ):
                                        if master_folder_tuple[tmp_master_folder_tuple[0] - 1, tmp_master_folder_tuple[1] - 1] == '<<POSITION_FOLDER>>':
                                            flag_exist_POSITION_FOLDER = True
                                        if master_folder_tuple[tmp_master_folder_tuple[0] - 1, tmp_master_folder_tuple[1] - 1] == '<SET_WIDTH>':
                                            flag_exist_SET_WIDTH = True

                                if (
                                        tmp_master_folder_tuple[0] != 1
                                        and tmp_master_folder_tuple[1] == 2
                                        and (flag_exist_SET_WIDTH or flag_exist_POSITION_FOLDER)
                                ):
                                    update_master_folder_tuple[tmp_master_folder_tuple[0] - 1, tmp_master_folder_tuple[1]] = tmp_bothside_empty
                                elif (
                                        tmp_master_folder_tuple[0] != 1
                                        and tmp_master_folder_tuple[1] == 2
                                        and flag_exist_SET_WIDTH
                                        and flag_exist_POSITION_FOLDER
                                ):
                                    update_master_folder_tuple[tmp_master_folder_tuple[0], tmp_master_folder_tuple[1]] = ''

                            # set body
                            if tmp_master_folder_tuple[0] == tmp_master_folder_size[0] and tmp_master_folder_tuple[1] != 1:
                                if (
                                        tmp_master_folder_tuple[0] == tmp_master_folder_size[0]
                                        and master_folder_tuple[tmp_master_folder_tuple] == tmp_master_folder_size[1][0][0]
                                ):
                                    update_master_folder_tuple[tmp_master_folder_tuple[0] - 1, tmp_master_folder_tuple[1] + 1] = tmp_master_folder_size[1][0][1]
                                    update_master_folder_tuple[tmp_master_folder_tuple[0], tmp_master_folder_tuple[1] + 1] = master_folder_tuple[tmp_master_folder_tuple]
                                elif (
                                        tmp_master_folder_tuple[0] == tmp_master_folder_size[0]
                                        and master_folder_tuple[tmp_master_folder_tuple] == ''
                                ):
                                    update_master_folder_tuple[tmp_master_folder_tuple[0] - 1, tmp_master_folder_tuple[1] + 1] = tmp_master_width_size_y_grid[2]
                                    update_master_folder_tuple[tmp_master_folder_tuple[0], tmp_master_folder_tuple[1] + 1] = ''

                            # set last column
                            if tmp_master_folder_tuple[0] == tmp_master_folder_size[0] and tmp_master_folder_tuple[0] > tmp_max_row:
                                tmp_max_row = tmp_master_folder_tuple[0]
                            if tmp_master_folder_tuple[0] == tmp_master_folder_size[0] and tmp_master_folder_tuple[1] > tmp_max_column:
                                tmp_max_column = tmp_master_folder_tuple[1]

                        if (tmp_max_column + 2) == 3:
                            # kyusai only empty row
                            update_master_folder_tuple[tmp_max_row - 1, tmp_max_column + 1] = 10
                        else:
                            update_master_folder_tuple[tmp_max_row - 1, tmp_max_column + 2] = tmp_bothside_empty
                            update_master_folder_tuple[tmp_max_row, tmp_max_column + 2] = ''

            flag_wp_only = True
            for tmp_master_folder_size in master_folder_size_array[2]:
                if tmp_master_width_size_y_grid[0] == tmp_master_folder_size[0]:
                    if '_wp_' not in str(tmp_master_folder_size[1][0][0]):
                        flag_wp_only = False
                        break

            # only way point
            if flag_wp_only:
                for tmp_tmp_master_folder_tuple in master_folder_tuple:
                    if tmp_master_width_size_y_grid[0] == tmp_tmp_master_folder_tuple[0] and tmp_tmp_master_folder_tuple[1] != 1:
                        if ns_def.check_tuple_num_exist(update_master_folder_tuple, tmp_tmp_master_folder_tuple[0] - 1, tmp_tmp_master_folder_tuple[1] + 2):
                            update_master_folder_tuple[tmp_tmp_master_folder_tuple[0] - 1, tmp_tmp_master_folder_tuple[1]] = update_master_folder_tuple[tmp_tmp_master_folder_tuple[0] - 1, tmp_tmp_master_folder_tuple[1] + 2]
                        else:
                            update_master_folder_tuple[tmp_tmp_master_folder_tuple[0] - 1, tmp_tmp_master_folder_tuple[1]] = 0.999
                    elif master_folder_tuple[tmp_tmp_master_folder_tuple] == 0.999:  # bug fix 001
                        update_master_folder_tuple[tmp_tmp_master_folder_tuple] = master_folder_tuple[tmp_tmp_master_folder_tuple]

        # update best height size
        pre_update_master_folder_tuple = update_master_folder_tuple
        update_master_folder_tuple = {}
        for tmp_pre_update_master_folder_tuple in pre_update_master_folder_tuple:
            if tmp_pre_update_master_folder_tuple[1] != 1 or pre_update_master_folder_tuple[tmp_pre_update_master_folder_tuple] == '<SET_WIDTH>':
                update_master_folder_tuple[tmp_pre_update_master_folder_tuple] = pre_update_master_folder_tuple[tmp_pre_update_master_folder_tuple]
            else:
                for tmp_master_folder_size_array in master_folder_size_array[4]:
                    if tmp_master_folder_size_array[0] == tmp_pre_update_master_folder_tuple[0]:
                        update_master_folder_tuple[tmp_pre_update_master_folder_tuple] = tmp_master_folder_size_array[1]

        # Additional fix related to bug 1~3 at ver 2.3.4
        if (1, 2) not in update_master_folder_tuple and (2, 3) in update_master_folder_tuple and (1, 4) in update_master_folder_tuple:
            if isinstance(update_master_folder_tuple[(2, 3)], str) and '_wp_' in update_master_folder_tuple[(2, 3)]:
                update_master_folder_tuple[1, 2] = update_master_folder_tuple[1, 4]

        # Remove entries where the first coordinate is 0
        update_master_folder_tuple = {k: v for k, v in update_master_folder_tuple.items() if k[0] != 0}

        # Remove duplicate empty columns
        def remove_duplicate_empty_columns(folder_tuple):
            changes_made = True

            while changes_made:
                changes_made = False

                # Get min and max columns for each row to check full range
                rows = {}
                for key in folder_tuple.keys():
                    row = key[0]
                    col = key[1]
                    if row not in rows:
                        rows[row] = {'min': col, 'max': col}
                    else:
                        rows[row]['min'] = min(rows[row]['min'], col)
                        rows[row]['max'] = max(rows[row]['max'], col)

                # Process each row
                for row in sorted(rows.keys()):
                    if changes_made:
                        break

                    # Check all columns in range (including non-existent keys)
                    min_col = rows[row]['min']
                    max_col = rows[row]['max']

                    # Check consecutive columns for empty pattern
                    for current_col in range(min_col, max_col):
                        next_col = current_col + 1

                        current_key = (row, current_col)
                        next_key = (row, next_col)

                        # Get values, treating non-existent keys as empty ''
                        current_value = folder_tuple.get(current_key, '')
                        next_value = folder_tuple.get(next_key, '')

                        # Check if both are empty (including non-existent)
                        if current_value == '' and next_value == '':
                            current_above_key = (row - 1, current_col)
                            next_above_key = (row - 1, next_col)

                            # Get the values above
                            current_above_value = folder_tuple.get(current_above_key, float('inf'))
                            next_above_value = folder_tuple.get(next_above_key, float('inf'))

                            # Store original above values before any modification
                            original_current_above = current_above_value
                            original_next_above = next_above_value

                            # Treat non-numeric values as infinity
                            if not isinstance(current_above_value, (int, float)):
                                current_above_value = float('inf')
                            if not isinstance(next_above_value, (int, float)):
                                next_above_value = float('inf')

                            # Skip if both above values are infinity (both non-existent)
                            if current_above_value == float('inf') and next_above_value == float('inf'):
                                continue

                            # Determine which column to remove (the one with smaller value above)
                            # and preserve the larger value
                            if current_above_value <= next_above_value:
                                col_to_remove = current_col
                                col_to_keep = next_col
                                value_to_preserve = original_next_above
                            else:
                                col_to_remove = next_col
                                col_to_keep = current_col
                                value_to_preserve = original_current_above

                            # Calculate the final position after shift
                            if col_to_keep > col_to_remove:
                                final_col_position = col_to_keep - 1
                            else:
                                final_col_position = col_to_keep

                            # Store the value to preserve before deletion
                            value_above_to_preserve = value_to_preserve

                            # Store existing values that should be preserved (columns before col_to_remove)
                            affected_rows = [row - 1, row]
                            preserved_values = {}
                            for k, v in list(folder_tuple.items()):
                                if k[0] in affected_rows and k[1] < col_to_remove:
                                    preserved_values[k] = v

                            # Remove cells in the column to remove ONLY for the current row and row above
                            keys_to_remove = [k for k in list(folder_tuple.keys())
                                              if k[1] == col_to_remove and k[0] in affected_rows]
                            for k in keys_to_remove:
                                del folder_tuple[k]

                            # Shift columns to the right of deleted column to the left
                            # ONLY for the current row and row above
                            keys_to_shift = sorted([k for k in list(folder_tuple.keys())
                                                    if k[1] > col_to_remove and k[0] in affected_rows])

                            # Create temporary storage for shifted values
                            temp_shifted = {}
                            for k in keys_to_shift:
                                new_key = (k[0], k[1] - 1)
                                temp_shifted[new_key] = folder_tuple[k]
                                del folder_tuple[k]

                            # Apply shifted values
                            for k, v in temp_shifted.items():
                                folder_tuple[k] = v

                            # Restore preserved values (to prevent overwriting)
                            for k, v in preserved_values.items():
                                folder_tuple[k] = v

                            # Set the larger value in the correct position (row - 1, final_col_position)
                            # Only if it's a valid numeric value
                            if value_above_to_preserve != float('inf') and isinstance(value_above_to_preserve, (int, float)):
                                folder_tuple[(row - 1, final_col_position)] = value_above_to_preserve

                            changes_made = True
                            break

                    if changes_made:
                        break

            return folder_tuple

        # Execute removal of consecutive empty columns
        update_master_folder_tuple = remove_duplicate_empty_columns(update_master_folder_tuple)

        # NEW: Add +1.0 buffer to width values for rows containing '_wp_' folders
        wp_buffer = 2.0  # Buffer value to add

        # Find all rows that contain '_wp_' folders
        rows_with_wp = set()
        for key, value in update_master_folder_tuple.items():
            if isinstance(value, str) and '_wp_' in value:
                rows_with_wp.add(key[0])

        # Add buffer to width values (one row above) for rows containing '_wp_'
        for wp_row in rows_with_wp:
            width_row = wp_row - 1  # Width values are in the row above

            # Add buffer to all numeric width values in this row
            for key in list(update_master_folder_tuple.keys()):
                if key[0] == width_row and key[1] >= 2:  # Width values start from column 2
                    if isinstance(update_master_folder_tuple[key], (int, float)) and update_master_folder_tuple[key] not in [10, 0.999, 1]:
                        # Add buffer
                        update_master_folder_tuple[key] = update_master_folder_tuple[key] + wp_buffer

        # Change (1, 1) to '<<POSITION_FOLDER>>'
        if (1, 1) in update_master_folder_tuple:
            update_master_folder_tuple[(1, 1)] = '<<POSITION_FOLDER>>'

        return update_master_folder_tuple




    def check_hostnames_in_same_element_static(position_shape_array, from_hostname, to_hostname, position_folder_array=None):
        """
        Static version of check_hostnames_in_same_element for external use
        Check if both hostnames exist in the same element (horizontal) or different elements (vertical)
        If devices are in different groups, use position_folder_array to determine direction
        """
        excluded_values = ['_AIR_', '<END>', '', '<<POSITION_SHAPE>>']

        # Find which group each hostname belongs to
        from_group = None
        to_group = None
        from_element_number = None
        to_element_number = None
        from_index = None
        to_index = None

        for item in position_shape_array:
            element_number = item[0]
            hostname_list = item[1]

            # First element is the group name
            group_name = hostname_list[0] if len(hostname_list) > 0 else ''
            filtered_list = [host for host in hostname_list[1:] if host not in excluded_values]

            if from_hostname in filtered_list:
                from_group = group_name
                from_element_number = element_number
                from_index = filtered_list.index(from_hostname)

            if to_hostname in filtered_list:
                to_group = group_name
                to_element_number = element_number
                to_index = filtered_list.index(to_hostname)

        # If not found, return error
        if from_group is None:
            return f"Error: from_hostname '{from_hostname}' not found in position_shape_array"
        if to_group is None:
            return f"Error: to_hostname '{to_hostname}' not found in position_shape_array"

        # If both in the same group and same row → horizontal (LEFT/RIGHT)
        if from_group == to_group and from_element_number == to_element_number:
            return 'LEFT_RIGHT' if to_index < from_index else 'RIGHT_LEFT'

        # If both in the same group but different rows → vertical (UP/DOWN)
        if from_group == to_group:
            return 'UP_DOWN' if from_element_number < to_element_number else 'DOWN_UP'

        # ★★★ If in different groups, use position_folder_array ★★★
        if position_folder_array is not None:
            # Find group positions in position_folder_array
            from_group_row = None
            from_group_col = None
            to_group_row = None
            to_group_col = None

            for folder_item in position_folder_array:
                if len(folder_item) < 2 or not isinstance(folder_item[1], list):
                    continue

                folder_row = folder_item[0]
                folder_list = folder_item[1]

                for col_idx, folder_name in enumerate(folder_list):
                    # ★★★ Skip non-string values AND empty strings ★★★
                    if not isinstance(folder_name, str) or folder_name == '':
                        continue

                    # Check with and without _wp_ suffix for both directions
                    folder_name_base = folder_name.replace('_wp_', '') if '_wp_' in folder_name else folder_name
                    from_group_base = from_group.replace('_wp_', '') if '_wp_' in from_group else from_group
                    to_group_base = to_group.replace('_wp_', '') if '_wp_' in to_group else to_group

                    if folder_name == from_group or folder_name == from_group + '_wp_' or folder_name_base == from_group_base:
                        from_group_row = folder_row
                        from_group_col = col_idx
                    if folder_name == to_group or folder_name == to_group + '_wp_' or folder_name_base == to_group_base:
                        to_group_row = folder_row
                        to_group_col = col_idx

            # Determine direction based on folder positions
            if from_group_row is not None and to_group_row is not None:
                # ★★★ Different rows → always vertical ★★★
                if from_group_row != to_group_row:
                    return 'UP_DOWN' if from_group_row < to_group_row else 'DOWN_UP'
                # Same row in folder_array → horizontal
                else:
                    return 'LEFT_RIGHT' if to_group_col < from_group_col else 'RIGHT_LEFT'

        # Fallback: use position_shape_array row numbers
        return 'UP_DOWN' if from_element_number < to_element_number else 'DOWN_UP'





    def process_l1_link_common(self, position_shape_array, style_shape_array, position_line_array,
                               position_tag_array, master_file_path, ori_position_line_tuple,
                               allow_shrink=False, affected_hostnames_only=None, position_folder_array=None):
        """
        Common processing for L1 link operations (add/delete)
        Recalculates offsets and updates shapes after connection changes

        Args:
            self: Reference to calling instance
            position_shape_array: Array containing position and shape information
            style_shape_array: Array containing style and shape information
            position_line_array: Array containing connection line information
            position_tag_array: Array containing position tag information
            master_file_path: Path to the master Excel file
            ori_position_line_tuple: Original position line tuple for cleanup
            allow_shrink: If True, allows shape dimensions to shrink (for delete operation)
            affected_hostnames_only: If set, only recalculate offsets for these hostnames (for add operation)
            position_folder_array: Array containing folder/group layout information

        Returns:
            Tuple of (updated position_line_array, updated style_shape_array)
        """
        import ns_def

        # Define all helper functions
        def count_interface_directions(hostname, position_shape_array, position_line_array):
            """
            Count the number of interfaces in each direction (UP, DOWN, LEFT, RIGHT) for a given hostname
            """
            up_count = down_count = left_count = right_count = 0

            for item in position_line_array[2:]:
                connection_data = item[1]
                from_hostname = connection_data[0]
                to_hostname = connection_data[1]

                # ★★★ Use stored direction from data (same as main offset loop) ★★★
                from_side = connection_data[4] if len(connection_data) > 4 else ''
                to_side = connection_data[5] if len(connection_data) > 5 else ''

                # Determine connection direction from stored side values
                if from_side in ['RIGHT', 'LEFT'] and to_side in ['RIGHT', 'LEFT']:
                    # Horizontal connection (RIGHT/LEFT)
                    if from_side == 'RIGHT':
                        direction = 'RIGHT_LEFT'
                    else:
                        direction = 'LEFT_RIGHT'
                else:
                    # Vertical connection or not explicitly set - calculate
                    direction = def_common.check_hostnames_in_same_element_static(position_shape_array, from_hostname, to_hostname, position_folder_array)

                if hostname == from_hostname:
                    if direction == 'RIGHT_LEFT':
                        right_count += 1
                    elif direction == 'LEFT_RIGHT':
                        left_count += 1
                    elif direction == 'UP_DOWN':
                        down_count += 1
                    elif direction == 'DOWN_UP':
                        up_count += 1

                elif hostname == to_hostname:
                    if direction == 'RIGHT_LEFT':
                        left_count += 1
                    elif direction == 'LEFT_RIGHT':
                        right_count += 1
                    elif direction == 'UP_DOWN':
                        up_count += 1
                    elif direction == 'DOWN_UP':
                        down_count += 1

            return [up_count, down_count, left_count, right_count]

        def calculate_offset_values(line_count, line_distance):
            """
            Calculate offset values for lines based on count
            """
            if line_count == 0:
                return []
            elif line_count == 1:
                return [0]
            elif line_count % 2 == 0:
                half_distance = line_distance / 2
                offsets = []
                for i in range(line_count // 2):
                    offsets.insert(0, -(half_distance + i * line_distance))
                for i in range(line_count // 2):
                    offsets.append(half_distance + i * line_distance)
                return offsets
            else:
                offsets = [0]
                for i in range(1, (line_count + 1) // 2):
                    offsets.insert(0, -i * line_distance)
                    offsets.append(i * line_distance)
                return offsets

        def update_shape_dimensions(style_shape_array, hostname, position_shape_array, position_line_array, line_distance=0.2, margin=0.1, allow_shrink=False):
            """
            Update Width or Height in style_shape_array for a specific hostname

            Args:
                allow_shrink: If True, allows shrinking to minimum size based on text length and connections
            """
            # Parameters from ns_l1_master_create
            shape_width_min = 0.4
            shape_hight_min = 0.2
            shae_font_size = 6.0

            # Width margin settings
            normal_shape_width_margin = 0.1  # Margin for normal devices
            waypoint_shape_width_margin = 0.2  # Margin for WayPoints (BLUE color)

            # Height margin settings
            normal_shape_height_margin = 0.1  # Margin for normal devices
            waypoint_shape_height_margin = 0.2  # Margin for WayPoints (BLUE color)

            # Minimum dimensions for waypoints
            waypoint_width_min = 0.5  # Minimum width specifically for waypoints

            # Get interface direction counts for the hostname
            direction_counts = count_interface_directions(hostname, position_shape_array, position_line_array)
            up_count, down_count, left_count, right_count = direction_counts

            # Find the hostname in style_shape_array
            hostname_row_index = None
            is_waypoint = False  # Track if hostname is a waypoint

            for idx, item in enumerate(style_shape_array):
                if len(item) >= 2 and isinstance(item[1], list) and len(item[1]) > 0:
                    if item[1][0] == hostname:
                        hostname_row_index = idx
                        # Check if it's a waypoint by color (BLUE indicates waypoint)
                        if len(item[1]) > 4 and item[1][4] == 'BLUE':
                            is_waypoint = True
                        break

            # If hostname not found, try to create from <DEFAULT>
            if hostname_row_index is None:
                for idx, item in enumerate(style_shape_array):
                    if len(item) >= 2 and isinstance(item[1], list) and len(item[1]) > 0:
                        if item[1][0] == '<DEFAULT>':
                            default_data = item[1]
                            new_number = max(row[0] for row in style_shape_array) + 1
                            new_entry = [new_number, [hostname, default_data[1], default_data[2], default_data[3], default_data[4]]]
                            style_shape_array.append(new_entry)
                            hostname_row_index = len(style_shape_array) - 1
                            break

            if hostname_row_index is None:
                return style_shape_array

            # Get current dimensions
            hostname_data = style_shape_array[hostname_row_index][1]

            try:
                current_width = float(hostname_data[1]) if hostname_data[1] != 'N/A' else shape_width_min
                current_height = float(hostname_data[2]) if hostname_data[2] != 'N/A' else shape_hight_min
            except (ValueError, IndexError):
                current_width = shape_width_min
                current_height = shape_hight_min

            # Calculate minimum dimensions based on hostname text length
            text_dimensions = ns_def.get_description_width_hight(shae_font_size, hostname)
            num_char_width = text_dimensions[0]
            num_char_height = text_dimensions[1]

            # Apply appropriate margins based on device type
            if is_waypoint:
                text_width_with_margin = num_char_width + waypoint_shape_width_margin
                text_height_with_margin = num_char_height + waypoint_shape_height_margin
            else:
                text_width_with_margin = num_char_width + normal_shape_width_margin
                text_height_with_margin = num_char_height + normal_shape_height_margin

            # Determine absolute minimum dimensions (text or default, whichever is larger)
            # For waypoints, use waypoint_width_min if larger than calculated value
            if is_waypoint:
                absolute_min_width = max(text_width_with_margin, waypoint_width_min)
            else:
                absolute_min_width = max(text_width_with_margin, shape_width_min)

            absolute_min_height = max(text_height_with_margin, shape_hight_min)

            # For LEFT/RIGHT connections (horizontal), lines go up/down, so update HEIGHT
            max_lr_count = max(left_count, right_count)
            if max_lr_count > 0:
                offset_values = calculate_offset_values(max_lr_count, line_distance)
                if len(offset_values) > 0:
                    span = max(offset_values) - min(offset_values)
                    # Apply appropriate height margin based on device type
                    if is_waypoint:
                        required_height = span + waypoint_shape_height_margin
                    else:
                        required_height = span + normal_shape_height_margin
                else:
                    # Use appropriate margin as minimum when no span
                    required_height = waypoint_shape_height_margin if is_waypoint else normal_shape_height_margin

                # Never go below absolute minimum
                final_height = max(required_height, absolute_min_height)

                if allow_shrink:
                    hostname_data[2] = final_height
                else:
                    if current_height < final_height:
                        hostname_data[2] = final_height
            else:
                if allow_shrink:
                    hostname_data[2] = absolute_min_height

            # For UP/DOWN connections (vertical), lines go left/right, so update WIDTH
            max_ud_count = max(up_count, down_count)
            if max_ud_count > 0:
                offset_values = calculate_offset_values(max_ud_count, line_distance)
                if len(offset_values) > 0:
                    span = max(offset_values) - min(offset_values)
                    # Apply appropriate width margin based on device type
                    if is_waypoint:
                        required_width = span + waypoint_shape_width_margin
                    else:
                        required_width = span + normal_shape_width_margin
                else:
                    # Use appropriate margin as minimum when no span
                    required_width = waypoint_shape_width_margin if is_waypoint else normal_shape_width_margin

                # Never go below absolute minimum (includes waypoint_width_min for waypoints)
                final_width = max(required_width, absolute_min_width)

                if allow_shrink:
                    hostname_data[1] = final_width
                else:
                    if current_width < final_width:
                        hostname_data[1] = final_width
            else:
                # No UP/DOWN connections - set to absolute minimum
                if allow_shrink:
                    hostname_data[1] = absolute_min_width

            return style_shape_array




        def determine_line_order(position_line_array, position_shape_array):
            """
            Determine the order of lines based on device positions
            """
            hostname_position = {}
            for row_idx, item in enumerate(position_shape_array):
                if len(item) >= 2 and isinstance(item[1], list):
                    for col_idx, hostname in enumerate(item[1]):
                        if hostname not in ['_AIR_', '<END>', '', '<<POSITION_SHAPE>>']:
                            if hostname not in hostname_position:
                                hostname_position[hostname] = (row_idx, col_idx)

            header_rows = position_line_array[:2]
            data_rows = position_line_array[2:]

            def sort_key(row):
                if len(row) < 2 or not isinstance(row[1], list):
                    return (999, 999, 999, 999, 999, 999)

                data = row[1]
                from_hostname = data[0] if len(data) > 0 else ''
                to_hostname = data[1] if len(data) > 1 else ''

                from_pos = hostname_position.get(from_hostname, (999, 999))
                to_pos = hostname_position.get(to_hostname, (999, 999))

                direction = 'UNKNOWN'
                if from_hostname and to_hostname:
                    direction = def_common.check_hostnames_in_same_element_static(position_shape_array, from_hostname, to_hostname, position_folder_array)

                if direction in ['UP_DOWN', 'DOWN_UP']:
                    if from_pos[0] < to_pos[0]:
                        key = (to_pos[0], to_pos[1], 0, from_pos[1], row[0])
                    else:
                        key = (from_pos[0], from_pos[1], 1, to_pos[1], to_pos[0], row[0])
                elif direction in ['LEFT_RIGHT', 'RIGHT_LEFT']:
                    if from_pos[1] < to_pos[1]:
                        key = (from_pos[0], from_pos[1], 2, to_pos[1], row[0])
                    else:
                        key = (from_pos[0], to_pos[1], 3, from_pos[1], row[0])
                else:
                    key = (from_pos[0], from_pos[1], 4, to_pos[0], to_pos[1], row[0])

                return key

            sorted_data_rows = sorted(data_rows, key=sort_key)
            for idx, row in enumerate(sorted_data_rows, start=3):
                row[0] = idx

            return header_rows + sorted_data_rows

        def get_tag_offset(hostname, master_line_tuple, tag_offet_inche):
            """
            Get tag offset value based on maximum tag name length
            """
            max_len_num = 1
            for tmp in master_line_tuple:
                if master_line_tuple.get(tmp) == hostname and tmp[1] in [1, 2]:
                    tmp_len_char = len(master_line_tuple[tmp[0], 3 if tmp[1] == 1 else 4])
                    if max_len_num < tmp_len_char:
                        max_len_num = tmp_len_char
            return max_len_num * tag_offet_inche

        def update_position_tag_array_single(position_tag_array, hostname, offset_value):
            """
            Update Offset_LINE value for a single hostname
            """
            for item in position_tag_array:
                if len(item) >= 2 and isinstance(item[1], list) and len(item[1]) >= 5:
                    row_data = item[1]
                    if row_data[0] == hostname:
                        try:
                            current_offset = float(row_data[4]) if row_data[4] != '' else 0.0
                            if current_offset < offset_value:
                                row_data[4] = offset_value
                        except (ValueError, TypeError):
                            row_data[4] = offset_value
            return position_tag_array

        # ========== MAIN PROCESSING ==========

        # Sort connections
        position_line_array = determine_line_order(position_line_array, position_shape_array)

        # Get unique hostnames
        unique_hostnames = set()
        for item in position_line_array[2:]:
            connection_data = item[1]
            unique_hostnames.add(connection_data[0])
            unique_hostnames.add(connection_data[1])

        # Determine which hostnames to process
        if affected_hostnames_only is not None:
            # For add operation: only process affected devices
            hostnames_to_process = affected_hostnames_only & unique_hostnames

            # Clear offsets only for connections involving affected hostnames
            for item in position_line_array[2:]:
                connection_data = item[1]
                if connection_data[0] in hostnames_to_process or connection_data[1] in hostnames_to_process:
                    connection_data[6] = ''
                    connection_data[7] = ''
                    connection_data[8] = ''
                    connection_data[9] = ''
        else:
            # For delete operation: process all devices
            hostnames_to_process = unique_hostnames

            # Clear all offsets
            for item in position_line_array[2:]:
                connection_data = item[1]
                connection_data[6] = ''
                connection_data[7] = ''
                connection_data[8] = ''
                connection_data[9] = ''

        # Recalculate offsets
        line_distance = 0.2

        for current_hostname in sorted(hostnames_to_process):

            # Build hostname position mapping
            hostname_position = {}
            for row_idx, item in enumerate(position_shape_array):
                if len(item) >= 2 and isinstance(item[1], list):
                    for col_idx, hostname_iter in enumerate(item[1]):
                        if hostname_iter not in ['_AIR_', '<END>', '', '<<POSITION_SHAPE>>']:
                            if hostname_iter not in hostname_position:
                                hostname_position[hostname_iter] = (row_idx, col_idx)

            # Group connections by direction
            direction_groups = {'UP': [], 'DOWN': [], 'LEFT': [], 'RIGHT': []}

            for sort_index, item in enumerate(position_line_array[2:]):
                connection_data = item[1]
                from_h = connection_data[0]
                to_h = connection_data[1]

                if current_hostname == from_h or current_hostname == to_h:
                    # Use stored direction from data
                    from_side = connection_data[4] if len(connection_data) > 4 else ''
                    to_side = connection_data[5] if len(connection_data) > 5 else ''


                    # Determine connection direction from stored side values
                    if from_side in ['RIGHT', 'LEFT'] and to_side in ['RIGHT', 'LEFT']:
                        # Horizontal connection (RIGHT/LEFT)
                        conn_direction = 'RIGHT_LEFT' if from_side == 'RIGHT' else 'LEFT_RIGHT'
                    else:
                        # Vertical connection (UP/DOWN)
                        conn_direction = def_common.check_hostnames_in_same_element_static(
                            position_shape_array, from_h, to_h, position_folder_array
                        )

                    other_hostname = to_h if current_hostname == from_h else from_h
                    other_pos = hostname_position.get(other_hostname, (999, 999))

                    # Classify into direction group
                    if current_hostname == from_h:
                        if conn_direction == 'RIGHT_LEFT':
                            direction_groups['RIGHT'].append((other_pos[0], other_pos[1], sort_index, item))
                        elif conn_direction == 'LEFT_RIGHT':
                            direction_groups['LEFT'].append((other_pos[0], other_pos[1], sort_index, item))
                        elif conn_direction == 'UP_DOWN':
                            # current_hostname is FROM and going DOWN (to device below)
                            direction_groups['DOWN'].append((other_pos[0], other_pos[1], sort_index, item))
                        elif conn_direction == 'DOWN_UP':
                            # current_hostname is FROM and going UP (to device above)
                            direction_groups['UP'].append((other_pos[0], other_pos[1], sort_index, item))
                    elif current_hostname == to_h:
                        if conn_direction == 'RIGHT_LEFT':
                            direction_groups['LEFT'].append((other_pos[0], other_pos[1], sort_index, item))
                        elif conn_direction == 'LEFT_RIGHT':
                            direction_groups['RIGHT'].append((other_pos[0], other_pos[1], sort_index, item))
                        elif conn_direction == 'UP_DOWN':
                            # current_hostname is TO and connection is coming from UP (from device above)
                            direction_groups['UP'].append((other_pos[0], other_pos[1], sort_index, item))
                        elif conn_direction == 'DOWN_UP':
                            # current_hostname is TO and connection is coming from DOWN (from device below)
                            direction_groups['DOWN'].append((other_pos[0], other_pos[1], sort_index, item))

            # Calculate and assign offsets for EACH direction independently
            for direction, indexed_items in direction_groups.items():
                if len(indexed_items) == 0:
                    continue

                # Sort by column for spatial ordering
                indexed_items.sort(key=lambda x: x[1])

                items = [x[3] for x in indexed_items]
                offset_values = calculate_offset_values(len(items), line_distance)

                for idx, item in enumerate(items):
                    if idx < len(offset_values):
                        connection_data = item[1]
                        from_h = connection_data[0]
                        to_h = connection_data[1]
                        offset_val = offset_values[idx]

                        if direction in ['LEFT', 'RIGHT']:
                            # Horizontal connections: use Y-axis offset
                            index_to_set = 7 if current_hostname == from_h else 9
                            connection_data[index_to_set] = offset_val
                        else:
                            # Vertical connections: use X-axis offset
                            index_to_set = 6 if current_hostname == from_h else 8
                            connection_data[index_to_set] = offset_val

        # Update shape dimensions with allow_shrink flag
        shape_margin = 0.1
        for affected_host in hostnames_to_process:
            style_shape_array = update_shape_dimensions(style_shape_array, affected_host, position_shape_array, position_line_array, line_distance, shape_margin, allow_shrink)

        # ========== SAVE TO MASTER FILE ==========

        excel_file_path = master_file_path
        worksheet_name = 'Master_Data'
        offset_row = 0
        offset_column = 0

        # Write position_line_array
        self.position_line_tuple = ns_def.convert_array_to_tuple(position_line_array)
        section_write_to = '<<POSITION_LINE>>'
        ns_def.remove_rows_under_section('Master_Data', master_file_path, ori_position_line_tuple)
        ns_def.write_excel_meta(self.position_line_tuple, master_file_path, 'Master_Data', section_write_to, offset_row, offset_column)

        # Write style_shape_array
        self.style_shape_array_tuple = ns_def.convert_array_to_tuple(style_shape_array)
        section_write_to = '<<STYLE_SHAPE>>'
        ns_def.overwrite_excel_meta(self.style_shape_array_tuple, excel_file_path, worksheet_name, section_write_to, offset_row, offset_column)

        # Update position_tag_array
        tag_offet_inche = 0.02
        for target_hostname in hostnames_to_process:
            offset_value = get_tag_offset(target_hostname, self.position_line_tuple, tag_offet_inche)
            update_position_tag_array_single(position_tag_array, target_hostname, offset_value)

        self.position_tag_array_tuple = ns_def.convert_array_to_tuple(position_tag_array)
        section_write_to = '<<POSITION_TAG>>'
        ns_def.overwrite_excel_meta(self.position_tag_array_tuple, excel_file_path, worksheet_name, section_write_to, offset_row, offset_column)

        # Sync with Master_Data_L2 and Master_Data_L3
        self.full_filepath = master_file_path
        try:
            self.main1_1_entry_1.delete(0, tkinter.END)
            self.main1_1_entry_1.insert(tk.END, master_file_path)
        except AttributeError:
            dummy_root = tk.Tk()
            dummy_root.withdraw()
            self.main1_1_entry_1 = tk.Entry(dummy_root)
            self.main1_1_entry_1.insert(tk.END, master_file_path)

        import ns_sync_between_layers
        ns_sync_between_layers.l1_master_device_and_line_sync_with_l2l3_master(self)

        return position_line_array, style_shape_array




def get_next_arg(argv_array, target):
    try:
        index = argv_array.index(target)
        return argv_array[index + 1]
    except (ValueError, IndexError):
        return None

def print_type(self, argv_array,source):
    if  '--one_msg' in argv_array:
        print(source)
    else:
        for item in source:
            print(item)
    return

def get_device_waypoint_array(master_file_path):
    import ns_def
    l2_table_array = ns_def.convert_master_to_array('Master_Data_L2', master_file_path, '<<L2_TABLE>>')

    new_l2_table_array = []
    for tmp_l2_table_array in l2_table_array:
        if tmp_l2_table_array[0] != 1 and tmp_l2_table_array[0] != 2:
            tmp_l2_table_array[1].extend(['', '', '', '', '', '', '', ''])
            del tmp_l2_table_array[1][8:]
            new_l2_table_array.append(tmp_l2_table_array)

    device_list_array = []
    area_device_list_array = []
    wp_list_array = []
    for tmp_new_l2_table_array in new_l2_table_array:
        if tmp_new_l2_table_array[1][1] not in device_list_array and tmp_new_l2_table_array[1][1] not in wp_list_array:
            if tmp_new_l2_table_array[1][0] == 'N/A':
                wp_list_array.append(tmp_new_l2_table_array[1][1])
            else:
                device_list_array.append(tmp_new_l2_table_array[1][1])
                area_device_list_array.append([tmp_new_l2_table_array[1][0], tmp_new_l2_table_array[1][1]])

    return ([device_list_array,wp_list_array])

