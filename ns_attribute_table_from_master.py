'''
SPDX-License-Identifier: Apache-2.0

Copyright 2023 Cisco Systems, Inc. and its affiliates

Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.
'''

import ns_def, ns_egt_maker, openpyxl
from openpyxl.styles import Font, Color
import ast

class  ns_attribute_table_from_master():
    def __init__(self):
        excel_maseter_file = self.inFileTxt_L3_1_1.get()
        write_excel_file = self.outFileTxt_11_2.get().replace('[MASTER]', '')
        ws_name = 'Master_Data'
        tmp_ws_name = '_tmp_'

        '''
        Create L3 Table excel file
        '''
        self.position_shape_array = ns_def.convert_master_to_array(ws_name, excel_maseter_file, '<<POSITION_SHAPE>>')
        self.position_shape_tuple = ns_def.convert_array_to_tuple(self.position_shape_array)
        self.attribute_array = ns_def.convert_master_to_array('Master_Data', excel_maseter_file, '<<ATTRIBUTE>>')
        #print(self.attribute_array )

        remake_attribute_table_array = []

        '''make meta data for egt_maker '''
        #make remake_attribute_table_array
        header_array = self.attribute_array[1]
        index = header_array[1].index('Device Name')
        header_array[1].insert(index, '<HEADER>')
        header_array[1].insert(index + 1, 'Area')

        data = ['<RANGE>', '<END>']
        number_of_ones = len(self.attribute_array[1][1]) - 2
        ones_to_insert = ['1'] * number_of_ones
        data[1:1] = ones_to_insert

        egt_maker_width_array = ['25', '25', '15', '15', '15', '15', '15', '15', '15', '15', '15']
        remake_attribute_table_array.append([1, data])
        remake_attribute_table_array.append(header_array)

        # make pre_remake_attribute_table_array
        last_num = 3
        pre_remake_attribute_table_array = []
        get_shape_folder_tuple = ns_def.get_shape_folder_tuple(self.position_shape_tuple)
        color_grid_rgb_array = []

        for tmp_attribute_array in self.attribute_array:
            if tmp_attribute_array[0] != 1 and tmp_attribute_array[0] != 2:
                tmp_array = []
                tmp_color_grid_rgb_array = []
                for kari_tmp_attribute_array in tmp_attribute_array[1]:
                    if kari_tmp_attribute_array.startswith('[') and kari_tmp_attribute_array.endswith(']'):
                        tmp_array.append(ast.literal_eval(kari_tmp_attribute_array)[0])
                        tmp_color_grid_rgb_array.append(kari_tmp_attribute_array)
                    else:
                        tmp_array.append(kari_tmp_attribute_array)
                        tmp_color_grid_rgb_array.append(kari_tmp_attribute_array)

                if '_wp_' in str(get_shape_folder_tuple.get(tmp_attribute_array[1][0])):
                    area_value = '_N/A_'
                else:
                    area_value = str(get_shape_folder_tuple.get(tmp_attribute_array[1][0]))
                tmp_array.insert(0, '')
                tmp_array.insert(1, area_value)
                tmp_array.append('<END>')
                tmp_color_grid_rgb_array.insert(0, '')
                tmp_color_grid_rgb_array.insert(1, area_value)
                tmp_color_grid_rgb_array.append('<END>')
                pre_remake_attribute_table_array.append([tmp_attribute_array[0],tmp_array])
                color_grid_rgb_array.append([tmp_attribute_array[0], tmp_color_grid_rgb_array])
                last_num = tmp_attribute_array[0]

        pre_remake_attribute_table_array = sorted(pre_remake_attribute_table_array, key=lambda x: (x[1][1], x[1][2]))
        color_grid_rgb_array = sorted(color_grid_rgb_array, key=lambda x: (x[1][1], x[1][2]))

        for i, item in enumerate(pre_remake_attribute_table_array, start=3):
            item[0] = i
        for i, item in enumerate(color_grid_rgb_array, start=3):
            item[0] = i

        remake_attribute_table_array = remake_attribute_table_array + pre_remake_attribute_table_array
        remake_attribute_table_array.append([last_num + 1, ['<END>']])
        #print(remake_attribute_table_array)

        '''Create Attribute Table'''
        attribute_master_data_tuple = {}
        attribute_master_data_tuple = ns_def.convert_array_to_tuple(remake_attribute_table_array)

        # Create _tmp_ sheet
        ns_def.create_excel_sheet(excel_maseter_file, tmp_ws_name)

        # Write normal tuple to excel
        master_excel_meta = attribute_master_data_tuple
        excel_file_path = excel_maseter_file
        worksheet_name = tmp_ws_name
        section_write_to = '<<N/A>>'
        offset_row = 0
        offset_column = 0
        ns_def.write_excel_meta(master_excel_meta, excel_file_path, worksheet_name, section_write_to, offset_row, offset_column)

        #create dummy Attribute Table file
        tmp_tmp_ws_name = '_tmp_tmp_'
        ns_def.create_excel_sheet(excel_maseter_file, tmp_tmp_ws_name)
        input_tree_tuple = {}
        input_tree_tuple[1,1] = 'Dummy' # first excel sheet has bug. That's why the first sheet is dummy sheet. this sheet will be deleted soon.
        input_tree_tuple[2,1] = 'Attribute'
        master_excel_meta = input_tree_tuple
        excel_file_path = excel_maseter_file
        worksheet_name = tmp_tmp_ws_name
        section_write_to = '<<N/A>>'
        offset_row = 0
        offset_column = 0
        ns_def.write_excel_meta(master_excel_meta, excel_file_path, worksheet_name, section_write_to, offset_row, offset_column)

        #Add Attribute Table sheet
        #ns_def.create_excel_sheet(write_excel_file, 'Attribute')
        input_excel_name = excel_maseter_file
        output_excel_name = write_excel_file

        NEW_OR_ADD = 'ADD'
        ns_egt_maker.create_excel_gui_tree(input_excel_name,output_excel_name,NEW_OR_ADD, egt_maker_width_array)
        ns_def.remove_excel_sheet(output_excel_name, 'Dummy')

        #Add attribute Table from meta
        self.input_tree_excel = openpyxl.load_workbook(output_excel_name)
        worksheet_name = 'Attribute'
        start_row = 1
        start_column = 0
        custom_table_name = excel_maseter_file
        self.input_tree_excel = ns_egt_maker.insert_custom_excel_table(self.input_tree_excel, worksheet_name, start_row, start_column, custom_table_name)

        # Set the specified sheet as active
        self.input_tree_excel.active = self.input_tree_excel.sheetnames.index('L1 Table')

        # Remove _tmp_ sheet from excel master
        ns_def.remove_excel_sheet(excel_maseter_file, tmp_ws_name)
        ns_def.remove_excel_sheet(excel_maseter_file, tmp_tmp_ws_name)

        ''' Apply color to each cell '''
        sheet_name = 'Attribute'
        for tmp_color_grid_rgb_array in color_grid_rgb_array:
            for i, tmp_tmp_color_grid_rgb_array in enumerate(tmp_color_grid_rgb_array[1],start=0):
                if i >= 3 and tmp_tmp_color_grid_rgb_array != '<END>':
                    if i >= 3:
                        row = tmp_color_grid_rgb_array[0]
                        column = i
                        rgb_color = ast.literal_eval(tmp_tmp_color_grid_rgb_array)[1]
                        ns_def.change_cell_color(self.input_tree_excel, sheet_name, row, column, rgb_color)

        ''' Apply color to title cells that can be changed '''
        row = 2
        for column in range(3, 12):
            ns_def.change_cell_color(self.input_tree_excel, sheet_name, row, column, [246,252,252])
            sheet = self.input_tree_excel[sheet_name]
            cell = sheet.cell(row=row, column=column)
            cell.font = Font(name=cell.font.name, size=cell.font.size, bold=cell.font.bold, color='000000')

        #save
        self.input_tree_excel.save(output_excel_name)
        self.input_tree_excel.close()
