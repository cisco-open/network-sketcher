'''
SPDX-License-Identifier: Apache-2.0

Copyright 2023 Cisco Systems, Inc. and its affiliates

Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.
'''

import openpyxl
import tkinter.filedialog, tkinter.messagebox
import ns_def

class  ns_attribute_table_sync_master():
    def __init__(self):
        #print('--- attribute table sync to master')
        '''
        update Excel master data file from attribute table excel file
        '''
        # parameter
        attribute_table_ws_name = 'Attribute'
        attribute_table_file = self.main1_1_entry_1.get()
        excel_master_ws_name_attribute = 'Master_Data'
        excel_maseter_file = self.sub3_1_entry_1.get()

        #get attribute Table Excel file
        attribute_table_array = []
        start_row = 2
        end_colum = 11
        attribute_table_array = ns_attribute_table_sync_master.convert_excel_attribute_to_array(attribute_table_ws_name, attribute_table_file, start_row, end_colum)
        #print('--- attribute_table_array ---')
        #print(attribute_table_array)

        '''overwrite_excel_meta <<ATTRIBUTE>>'''
        #print('--- last_attribute_table_array ---')
        #print(last_attribute_table_array)

        last_attribute_table_tuple = {}
        last_attribute_table_tuple = ns_def.convert_array_to_tuple(attribute_table_array)
        #print(last_attribute_table_tuple)

        ### overwerite overwrite_last_attribute_table_tuple to excel master data
        master_excel_meta = last_attribute_table_tuple
        excel_file_path = excel_maseter_file
        worksheet_name = excel_master_ws_name_attribute
        section_write_to = '<<ATTRIBUTE>>'
        offset_row = 1
        offset_column = 0
        ns_def.overwrite_excel_meta(master_excel_meta, excel_file_path, worksheet_name, section_write_to, offset_row, offset_column)
        return (True)

    def convert_excel_attribute_to_array(ws_name, excel_file, start_row, end_colum):
        input_ppt_mata_excel = openpyxl.load_workbook(excel_file)
        input_ppt_mata_excel.active = input_ppt_mata_excel[ws_name]

        # GET Folder names
        flag_finish = False
        current_row = 1
        empty_count = 0

        while flag_finish == False:
            if str(input_ppt_mata_excel.active.cell(current_row, 1).value) == 'None' and str(input_ppt_mata_excel.active.cell(current_row, 2).value) == 'None':
                empty_count += 1
            else:
                empty_count = 0

            if empty_count >= 100:
                flag_finish = True
                end_row = current_row - 100
            current_row += 1

        return_array = []
        for tmp_row in range(start_row, end_row + 1):
            tmp_array = []
            current_row_array = []
            for tmp_column in range(1, end_colum + 1):
                # get cell color
                cell = input_ppt_mata_excel.active.cell(row=tmp_row, column=tmp_column)
                fill = cell.fill

                if fill.start_color.type == "rgb":
                    argb_color = fill.start_color.rgb
                    rgb_color = argb_color[2:]
                    r = int(rgb_color[0:2], 16)
                    g = int(rgb_color[2:4], 16)
                    b = int(rgb_color[4:6], 16)

                elif fill.start_color.type == "theme":
                    rgb_color = fill.start_color.theme
                    tmp_value = rgb_color * 8
                    tmp_value += 128
                    if tmp_value >= 256:
                        tmp_value = 200
                    r = int(tmp_value)
                    g = int(tmp_value)
                    b = int(tmp_value)
                    print('[Warning]Theme color was used in the Attribute sheet and it could not be read properly, please use Standard color.')

                if str(input_ppt_mata_excel.active.cell(tmp_row, tmp_column).value) != 'None':
                    if tmp_column == 2 and tmp_row == 2:
                        current_row_array.append('Device Name')
                    elif tmp_column <= 2 or tmp_row == 2:
                        current_row_array.append(input_ppt_mata_excel.active.cell(tmp_row, tmp_column).value)
                    else:
                        current_row_array.append(str([input_ppt_mata_excel.active.cell(tmp_row, tmp_column).value, [r, g, b]]))
                else:
                    current_row_array.append('[\'<EMPTY>\', [255, 255, 255]]')
            current_row_array.append('<END>')
            current_row_array.pop(0)

            if len(current_row_array) != 0:
                return_array.append([tmp_row - start_row + 1, current_row_array])

        input_ppt_mata_excel.close()
        return (return_array)