'''
SPDX-License-Identifier: Apache-2.0

Copyright 2023 Cisco Systems, Inc. and its affiliates

Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.
'''

import tkinter as tk ,tkinter.ttk as ttk,tkinter.filedialog, tkinter.messagebox
import sys, os
import ns_def
import openpyxl

class ns_front_run():
    def __init__(self):
        # create the root frame
        root = tk.Tk()
        root.title("Network Sketcher Ver 2.xx(development mode) < Cisco Internal Use Only >  ")
        root.geometry("860x320")
        # Notebook
        nb = ttk.Notebook(width=200, height=200)

        # create Tabs
        tab0 = tk.Frame(nb)
        tab1 = tk.Frame(nb)
        tab2 = tk.Frame(nb)
        tabx = tk.Frame(nb)
        tab_L2_1 = tk.Frame(nb)
        tab_L2_2 = tk.Frame(nb)
        tab_L3_1 = tk.Frame(nb)
        tab_L3_2 = tk.Frame(nb)

        nb.add(tab0, text='   Master Data   ', padding=5)
        nb.add(tab2, text='   L1 Table   ', padding=5)
        nb.add(tab1, text='   L1 Diagram   ', padding=5)
        nb.add(tab_L2_1, text='   L2 Table   ', padding=5)
        nb.add(tab_L2_2, text='   L2 Diagram   ', padding=5)
        nb.add(tab_L3_1, text='   L3 Table   ', padding=5)
        nb.add(tab_L3_2, text='   L3 Diagram   ', padding=5)
        nb.add(tabx, text='   For Cisco   ', padding=5)

        nb.pack(expand=1, fill='both')

        self.click_value_dummy = '' # for click action 12-3

        '''
        tab0  <<Master Data>>
        '''
        ### [INPUT] Sketch file or Layer1 Diagram file(per area / no tag):
        stepZero = tk.LabelFrame(tab0, text="[INPUT] Sketch file or Diagram(Per Area)     [OUTPUT] Excel Master Data", font=("", 14), height=1, background="#DCE6f1")
        stepZero.grid(row=0, columnspan=7, sticky='W', padx=5, pady=5, ipadx=5, ipady=5)
        inEncLbl_1_1 = tk.Label(stepZero, text="                                  ")
        inEncLbl_1_1.grid(row=0, column=1, sticky='E', padx=5, pady=2)
        inEncLbl_1_2 = tk.Label(stepZero, text="                                  ")
        inEncLbl_1_2.grid(row=0, column=2, sticky='E', padx=5, pady=2)
        inEncLbl_1_3 = tk.Label(stepZero, text="                                  ")
        inEncLbl_1_3.grid(row=0, column=3, sticky='E', padx=5, pady=2)
        inEncLbl_1_4 = tk.Label(stepZero, text="                                  ")
        inEncLbl_1_4.grid(row=0, column=4, sticky='E', padx=5, pady=2)

        inFileLbl_1_1 = tk.Label(stepZero, text="   1. Select the Sketch or Per Area file", font=("", 10), background="#DCE6f1")
        inFileLbl_1_1.grid(row=0, column=0, sticky='W', padx=5, pady=2)
        self.inFileTxt_1_1 = tk.Entry(stepZero)
        self.inFileTxt_1_1.grid(row=0, column=1, columnspan=7, sticky="WE", pady=3)
        self.outFileTxt_1_4_1 = tk.Entry(stepZero)
        inFileBtn_1_1 = tk.Button(stepZero, text="Browse ...", command=lambda: self.click_action('1-1'))
        inFileBtn_1_1.grid(row=0, column=8, sticky='W', padx=5, pady=2)
        self.outFileTxt_1_1 = tk.Entry(stepZero)
        self.outFileTxt_1_2 = tk.Entry(stepZero)
        inFileLbl_1_x = tk.Label(stepZero, text="   2.Click the Create button", font=("", 10), background="#DCE6f1")
        inFileLbl_1_x.grid(row=3, column=0, sticky='W', padx=5, pady=2)
        Create_ND_1_1 = tk.Button(stepZero, text="Convert to [Excel Master Data]", font=("", 12), command=(lambda: self.click_action('1-4')))
        Create_ND_1_1.grid(row=3, column=1, columnspan=7, sticky='WE', padx=5, pady=2)

        '''
        tab_L3_1  <<L3 Table>>
        '''
        ### [INPUT] Excel Master Data [OUTPUT] Excel L3 Table:
        stepZero_L3 = tk.LabelFrame(tab_L3_1, text="[INPUT] Excel Master Data     [OUTPUT] Excel L3 Table", font=("", 14), height=1, background="#DAFBFE")
        stepZero_L3.grid(row=0, columnspan=7, sticky='W', padx=5, pady=5, ipadx=5, ipady=5)
        inEncLb_L3_1_1 = tk.Label(stepZero_L3, text="                                  ")
        inEncLb_L3_1_1.grid(row=0, column=1, sticky='E', padx=5, pady=2)
        inEncLb_L3_1_2 = tk.Label(stepZero_L3, text="                                  ")
        inEncLb_L3_1_2.grid(row=0, column=2, sticky='E', padx=5, pady=2)
        inEncLb_L3_1_3 = tk.Label(stepZero_L3, text="                                  ")
        inEncLb_L3_1_3.grid(row=0, column=3, sticky='E', padx=5, pady=2)
        inEncLb_L3_1_4 = tk.Label(stepZero_L3, text="                                  ")
        inEncLb_L3_1_4.grid(row=0, column=4, sticky='E', padx=5, pady=2)

        inFileLb_L3_1_1 = tk.Label(stepZero_L3, text="   1. Select the Excel Master Data file", font=("", 10), background="#DAFBFE")
        inFileLb_L3_1_1.grid(row=0, column=0, sticky='W', padx=5, pady=2)
        self.inFileTxt_L3_1_1 = tk.Entry(stepZero_L3)
        self.inFileTxt_L3_1_1.grid(row=0, column=1, columnspan=7, sticky="WE", pady=3)
        self.outFileTxt_L3_1_4_1 = tk.Entry(stepZero_L3)
        inFileBtn_L3_1_1 = tk.Button(stepZero_L3, text="Browse ...", command=lambda: self.click_action('L3-1-1'))
        inFileBtn_L3_1_1.grid(row=0, column=8, sticky='W', padx=5, pady=2)
        self.outFileTxt_L3_1_1 = tk.Entry(stepZero_L3)
        self.outFileTxt_L3_1_2 = tk.Entry(stepZero_L3)
        inFileLb_L3_1_x = tk.Label(stepZero_L3, text="   2.Click the Create button", font=("", 10), background="#DAFBFE")
        inFileLb_L3_1_x.grid(row=3, column=0, sticky='W', padx=5, pady=2)
        Create_ND_L3_1_1 = tk.Button(stepZero_L3, text="Create [Excel L3 Table]", font=("", 12), command=(lambda: self.click_action('L3-1-2')))
        Create_ND_L3_1_1.grid(row=3, column=1, columnspan=7, sticky='WE', padx=5, pady=2)

        # [INPUT] L3 Table file:
        StepTwo_L3 = tk.LabelFrame(tab_L3_1, text=" [INPUT] Updated L3 Table     [Sync with] Excel Master Data ", font=("", 14), height=1, background="#FCEEFC")
        StepTwo_L3.grid(row=6, columnspan=7, sticky='W', padx=5, pady=5, ipadx=5, ipady=5)
        inEncLbl_L3_2_1 = tk.Label(StepTwo_L3, text="                                  ")
        inEncLbl_L3_2_1.grid(row=0, column=1, sticky='E', padx=5, pady=2)
        inEncLbl_L3_2_2 = tk.Label(StepTwo_L3, text="                                  ")
        inEncLbl_L3_2_2.grid(row=0, column=2, sticky='E', padx=5, pady=2)
        inEncLbl_L3_2_3 = tk.Label(StepTwo_L3, text="                                  ")
        inEncLbl_L3_2_3.grid(row=0, column=3, sticky='E', padx=5, pady=2)
        inEncLbl_L3_2_4 = tk.Label(StepTwo_L3, text="                                  ")
        inEncLbl_L3_2_4.grid(row=0, column=4, sticky='E', padx=5, pady=2)

        inFileLbl_L3_2_1 = tk.Label(StepTwo_L3, text="   1.Select the Updated L3 Table file", font=("", 10), background="#FCEEFC")
        inFileLbl_L3_2_1.grid(row=0, column=0, sticky='W', padx=5, pady=2)
        self.inFileTxt_L3_2_1 = tk.Entry(StepTwo_L3)
        self.inFileTxt_L3_2_1_2 = tk.Entry(StepTwo_L3)
        self.outFileTxt_L3_2_3_1 = tk.Entry(StepTwo_L3)
        self.inFileTxt_L3_2_1.grid(row=0, column=1, columnspan=7, sticky="WE", pady=3)
        inFileBtn_L3_2_1 = tk.Button(StepTwo_L3, text="Browse ...", command=lambda: self.click_action('L3-2-1'))
        inFileBtn_L3_2_1.grid(row=0, column=8, sticky='W', padx=5, pady=2)

        inFileLbl_L3_2_2 = tk.Label(StepTwo_L3, text="   2.Select the Target Excel Master file", font=("", 10), background="#FCEEFC")
        inFileLbl_L3_2_2.grid(row=1, column=0, sticky='W', padx=5, pady=2)
        self.inFileTxt_L3_2_2 = tk.Entry(StepTwo_L3)
        self.inFileTxt_L3_2_2_2 = tk.Entry(StepTwo_L3)
        self.inFileTxt_L3_2_2_3 = tk.Entry(StepTwo_L3)
        self.inFileTxt_L3_2_2.grid(row=1, column=1, columnspan=7, sticky="WE", pady=3)
        inFileBtn_L3_2_2 = tk.Button(StepTwo_L3, text="Browse ...", command=lambda: self.click_action('L3-2-2'))
        inFileBtn_L3_2_2.grid(row=1, column=8, sticky='W', padx=5, pady=2)

        inFileLbl_L3_2_x = tk.Label(StepTwo_L3, text="   3.Click the Sync button", font=("", 10), background="#FCEEFC")
        inFileLbl_L3_2_x.grid(row=3, column=0, sticky='W', padx=5, pady=2)
        Create_ND_L3_2_3 = tk.Button(StepTwo_L3, text="Sync", font=("", 12), command=lambda: self.click_action('L3-2-3'))
        Create_ND_L3_2_3.grid(row=3, column=1, columnspan=7, sticky='WE', padx=5, pady=2)

        '''
        tab_L3_2  <<L3 Diagram>>
        '''
        ### [INPUT] Excel Master Data [OUTPUT] Excel L3 Diagram:
        stepZero_L3_2 = tk.LabelFrame(tab_L3_2, text="[INPUT] Excel Master Data     [OUTPUT] PPT L3 Diagram", font=("", 14), height=1, background="#FFFFEB")
        stepZero_L3_2.grid(row=0, columnspan=7, sticky='W', padx=5, pady=5, ipadx=5, ipady=5)
        inEncLb_L3_3_1 = tk.Label(stepZero_L3_2, text="                                  ")
        inEncLb_L3_3_1.grid(row=0, column=1, sticky='E', padx=5, pady=2)
        inEncLb_L3_3_2 = tk.Label(stepZero_L3_2, text="                                  ")
        inEncLb_L3_3_2.grid(row=0, column=2, sticky='E', padx=5, pady=2)
        inEncLb_L3_3_3 = tk.Label(stepZero_L3_2, text="                                  ")
        inEncLb_L3_3_3.grid(row=0, column=3, sticky='E', padx=5, pady=2)
        inEncLb_L3_3_4 = tk.Label(stepZero_L3_2, text="                                  ")
        inEncLb_L3_3_4.grid(row=0, column=4, sticky='E', padx=5, pady=2)

        inFileLb_L3_3_1 = tk.Label(stepZero_L3_2, text="   1. Select the Excel Master Data file", font=("", 10), background="#FFFFEB")
        inFileLb_L3_3_1.grid(row=0, column=0, sticky='W', padx=5, pady=2)
        self.inFileTxt_L3_3_1 = tk.Entry(stepZero_L3_2)
        self.inFileTxt_L3_3_1.grid(row=0, column=1, columnspan=7, sticky="WE", pady=3)
        self.outFileTxt_L3_3_4_1 = tk.Entry(stepZero_L3_2)
        inFileBtn_L3_3_1 = tk.Button(stepZero_L3_2, text="Browse ...", command=lambda: self.click_action('L3-3-1'))
        inFileBtn_L3_3_1.grid(row=0, column=8, sticky='W', padx=5, pady=2)
        self.outFileTxt_L3_3_1 = tk.Entry(stepZero_L3_2)
        self.outFileTxt_L3_3_2 = tk.Entry(stepZero_L3_2)
        inFileLb_L3_3_x = tk.Label(stepZero_L3_2, text="   2.Click the Create button", font=("", 10), background="#FFFFEB")
        inFileLb_L3_3_x.grid(row=3, column=0, sticky='W', padx=5, pady=2)
        Create_ND_L3_3_2 = tk.Button(stepZero_L3_2, text=" Per Area ", font=("", 12), command=(lambda: self.click_action('L3-3-2')))
        Create_ND_L3_3_2.grid(row=3, column=2, columnspan=1, sticky='WE', padx=5, pady=2)


        '''
        tab_L2_1  <<L2 Table>>
        '''
        ### [INPUT] Excel Master Data [OUTPUT] Excel L2 Table:
        stepZero_L2 = tk.LabelFrame(tab_L2_1, text="[INPUT] Excel Master Data     [OUTPUT] Excel L2 Table", font=("", 14), height=1, background="#DCE6f1")
        stepZero_L2.grid(row=0, columnspan=7, sticky='W', padx=5, pady=5, ipadx=5, ipady=5)
        inEncLb_L2_1_1 = tk.Label(stepZero_L2, text="                                  ")
        inEncLb_L2_1_1.grid(row=0, column=1, sticky='E', padx=5, pady=2)
        inEncLb_L2_1_2 = tk.Label(stepZero_L2, text="                                  ")
        inEncLb_L2_1_2.grid(row=0, column=2, sticky='E', padx=5, pady=2)
        inEncLb_L2_1_3 = tk.Label(stepZero_L2, text="                                  ")
        inEncLb_L2_1_3.grid(row=0, column=3, sticky='E', padx=5, pady=2)
        inEncLb_L2_1_4 = tk.Label(stepZero_L2, text="                                  ")
        inEncLb_L2_1_4.grid(row=0, column=4, sticky='E', padx=5, pady=2)

        inFileLb_L2_1_1 = tk.Label(stepZero_L2, text="   1. Select the Excel Master Data file", font=("", 10), background="#DCE6f1")
        inFileLb_L2_1_1.grid(row=0, column=0, sticky='W', padx=5, pady=2)
        self.inFileTxt_L2_1_1 = tk.Entry(stepZero_L2)
        self.inFileTxt_L2_1_1.grid(row=0, column=1, columnspan=7, sticky="WE", pady=3)
        self.outFileTxt_L2_1_4_1 = tk.Entry(stepZero_L2)
        inFileBtn_L2_1_1 = tk.Button(stepZero_L2, text="Browse ...", command=lambda: self.click_action('L2-1-1'))
        inFileBtn_L2_1_1.grid(row=0, column=8, sticky='W', padx=5, pady=2)
        self.outFileTxt_L2_1_1 = tk.Entry(stepZero_L2)
        self.outFileTxt_L2_1_2 = tk.Entry(stepZero_L2)
        inFileLb_L2_1_x = tk.Label(stepZero_L2, text="   2.Click the Create button", font=("", 10), background="#DCE6f1")
        inFileLb_L2_1_x.grid(row=3, column=0, sticky='W', padx=5, pady=2)
        Create_ND_L2_1_1 = tk.Button(stepZero_L2, text="Create [Excel L2 Table]", font=("", 12), command=(lambda: self.click_action('L2-1-2')))
        Create_ND_L2_1_1.grid(row=3, column=1, columnspan=7, sticky='WE', padx=5, pady=2)

        # [INPUT] L2 Table file:
        StepTwo_L2 = tk.LabelFrame(tab_L2_1, text=" [INPUT] Updated L2 Table     [Sync with] Excel Master Data ", font=("", 14), height=1, background="#E6E0EC")
        StepTwo_L2.grid(row=6, columnspan=7, sticky='W', padx=5, pady=5, ipadx=5, ipady=5)
        inEncLbl_L2_2_1 = tk.Label(StepTwo_L2, text="                                  ")
        inEncLbl_L2_2_1.grid(row=0, column=1, sticky='E', padx=5, pady=2)
        inEncLbl_L2_2_2 = tk.Label(StepTwo_L2, text="                                  ")
        inEncLbl_L2_2_2.grid(row=0, column=2, sticky='E', padx=5, pady=2)
        inEncLbl_L2_2_3 = tk.Label(StepTwo_L2, text="                                  ")
        inEncLbl_L2_2_3.grid(row=0, column=3, sticky='E', padx=5, pady=2)
        inEncLbl_L2_2_4 = tk.Label(StepTwo_L2, text="                                  ")
        inEncLbl_L2_2_4.grid(row=0, column=4, sticky='E', padx=5, pady=2)

        inFileLbl_L2_2_1 = tk.Label(StepTwo_L2, text="   1.Select the Updated L2 Table file", font=("", 10), background="#E6E0EC")
        inFileLbl_L2_2_1.grid(row=0, column=0, sticky='W', padx=5, pady=2)
        self.inFileTxt_L2_2_1 = tk.Entry(StepTwo_L2)
        self.inFileTxt_L2_2_1_2 = tk.Entry(StepTwo_L2)
        self.outFileTxt_L2_2_3_1 = tk.Entry(StepTwo_L2)
        self.inFileTxt_L2_2_1.grid(row=0, column=1, columnspan=7, sticky="WE", pady=3)
        inFileBtn_L2_2_1 = tk.Button(StepTwo_L2, text="Browse ...", command=lambda: self.click_action('L2-2-1'))
        inFileBtn_L2_2_1.grid(row=0, column=8, sticky='W', padx=5, pady=2)

        inFileLbl_L2_2_2 = tk.Label(StepTwo_L2, text="   2.Select the Target Excel Master file", font=("", 10), background="#E6E0EC")
        inFileLbl_L2_2_2.grid(row=1, column=0, sticky='W', padx=5, pady=2)
        self.inFileTxt_L2_2_2 = tk.Entry(StepTwo_L2)
        self.inFileTxt_L2_2_2_2 = tk.Entry(StepTwo_L2)
        self.inFileTxt_L2_2_2_3 = tk.Entry(StepTwo_L2)
        self.inFileTxt_L2_2_2.grid(row=1, column=1, columnspan=7, sticky="WE", pady=3)
        inFileBtn_L2_2_2 = tk.Button(StepTwo_L2, text="Browse ...", command=lambda: self.click_action('L2-2-2'))
        inFileBtn_L2_2_2.grid(row=1, column=8, sticky='W', padx=5, pady=2)

        inFileLbl_L2_2_x = tk.Label(StepTwo_L2, text="   3.Click the Sync button", font=("", 10), background="#E6E0EC")
        inFileLbl_L2_2_x.grid(row=3, column=0, sticky='W', padx=5, pady=2)
        Create_ND_L2_2_3 = tk.Button(StepTwo_L2, text="Sync", font=("", 12), command=lambda: self.click_action('L2-2-3'))
        Create_ND_L2_2_3.grid(row=3, column=1, columnspan=7, sticky='WE', padx=5, pady=2)

        '''
        tab_L2_2  <<L2 Diagram>>
        '''
        ### [INPUT] Excel Master Data [OUTPUT] Excel L2 Diagram:
        stepZero_L2_2 = tk.LabelFrame(tab_L2_2, text="[INPUT] Excel Master Data     [OUTPUT] Excel L2 Diagram", font=("", 14), height=1, background="#fff2cc")
        stepZero_L2_2.grid(row=0, columnspan=7, sticky='W', padx=5, pady=5, ipadx=5, ipady=5)
        inEncLb_L2_3_1 = tk.Label(stepZero_L2_2, text="                                  ")
        inEncLb_L2_3_1.grid(row=0, column=1, sticky='E', padx=5, pady=2)
        inEncLb_L2_3_2 = tk.Label(stepZero_L2_2, text="                                  ")
        inEncLb_L2_3_2.grid(row=0, column=2, sticky='E', padx=5, pady=2)
        inEncLb_L2_3_3 = tk.Label(stepZero_L2_2, text="                                  ")
        inEncLb_L2_3_3.grid(row=0, column=3, sticky='E', padx=5, pady=2)
        inEncLb_L2_3_4 = tk.Label(stepZero_L2_2, text="                                  ")
        inEncLb_L2_3_4.grid(row=0, column=4, sticky='E', padx=5, pady=2)

        inFileLb_L2_3_1 = tk.Label(stepZero_L2_2, text="   1. Select the Excel Master Data file", font=("", 10), background="#fff2cc")
        inFileLb_L2_3_1.grid(row=0, column=0, sticky='W', padx=5, pady=2)
        self.inFileTxt_L2_3_1 = tk.Entry(stepZero_L2_2)
        self.inFileTxt_L2_3_1.grid(row=0, column=1, columnspan=7, sticky="WE", pady=3)
        self.outFileTxt_L2_3_4_1 = tk.Entry(stepZero_L2_2)
        inFileBtn_L2_3_1 = tk.Button(stepZero_L2_2, text="Browse ...", command=lambda: self.click_action('L2-3-1'))
        inFileBtn_L2_3_1.grid(row=0, column=8, sticky='W', padx=5, pady=2)
        self.outFileTxt_L2_3_1 = tk.Entry(stepZero_L2_2)
        self.outFileTxt_L2_3_2 = tk.Entry(stepZero_L2_2)
        inFileLb_L2_3_x = tk.Label(stepZero_L2_2, text="   2.Click the Create button", font=("", 10), background="#fff2cc")
        inFileLb_L2_3_x.grid(row=3, column=0, sticky='W', padx=5, pady=2)
        Create_ND_L2_3_2 = tk.Button(stepZero_L2_2, text=" Per Area ", font=("", 12), command=(lambda: self.click_action('L2-3-2')))
        Create_ND_L2_3_2.grid(row=3, column=2, columnspan=1, sticky='WE', padx=5, pady=2)
        Create_ND_L2_3_3 = tk.Button(stepZero_L2_2, text=" Per Device ", font=("", 12), command=(lambda: self.click_action('L2-3-3')))
        Create_ND_L2_3_3.grid(row=3, column=3, columnspan=1, sticky='WE', padx=5, pady=2)

        '''
        tabX  <<For Cisco>>
        '''
        ### SVG file from DNAC  convert to Master file
        stepTwo = tk.LabelFrame(tabx, text="[INPUT] SVG file from DNAC    [OUTPUT] Excel Master Data", font=("", 14), height=1, background="#ffE6f1")
        stepTwo.grid(row=0, columnspan=7, sticky='W', padx=5, pady=5, ipadx=5, ipady=5)
        inEncLbl_1a_1 = tk.Label(stepTwo, text="                                  ")
        inEncLbl_1a_1.grid(row=0, column=1, sticky='E', padx=5, pady=2)
        inEncLbl_1a_2 = tk.Label(stepTwo, text="                                  ")
        inEncLbl_1a_2.grid(row=0, column=2, sticky='E', padx=5, pady=2)
        inEncLbl_1a_3 = tk.Label(stepTwo, text="                                  ")
        inEncLbl_1a_3.grid(row=0, column=3, sticky='E', padx=5, pady=2)
        inEncLbl_1a_4 = tk.Label(stepTwo, text="                                  ")
        inEncLbl_1a_4.grid(row=0, column=4, sticky='E', padx=5, pady=2)

        inFileLbl_1a_1 = tk.Label(stepTwo, text="   1. Select the SVG file from DNAC", font=("", 10), background="#ffE6f1")
        inFileLbl_1a_1.grid(row=0, column=0, sticky='W', padx=5, pady=2)
        self.inFileTxt_1a_1 = tk.Entry(stepTwo)
        self.inFileTxt_1a_1.grid(row=0, column=1, columnspan=7, sticky="WE", pady=3)
        self.outFileTxt_1a_4_1 = tk.Entry(stepTwo)
        inFileBtn_1a_1 = tk.Button(stepTwo, text="Browse ...", command=lambda: self.click_action('1-1a'))
        inFileBtn_1a_1.grid(row=0, column=8, sticky='W', padx=5, pady=2)
        self.outFileTxt_1a_1 = tk.Entry(stepTwo)
        self.outFileTxt_1a_2 = tk.Entry(stepTwo)
        inFileLbl_1a_x = tk.Label(stepTwo, text="   2.Click the Create button", font=("", 10), background="#ffE6f1")
        inFileLbl_1a_x.grid(row=3, column=0, sticky='W', padx=5, pady=2)
        Create_ND_1a_1 = tk.Button(stepTwo, text="Convert to [Excel Master Data]", font=("", 12), command=(lambda: self.click_action('1-4a')))
        Create_ND_1a_1.grid(row=3, column=1, columnspan=7, sticky='WE', padx=5, pady=2)

        ### YAML file from CML  convert to Master file
        stepTwo = tk.LabelFrame(tabx, text="[INPUT] YAML file from CML    [OUTPUT] Excel Master Data", font=("", 14), height=1, background="#FFFFCC")
        stepTwo.grid(row=1, columnspan=7, sticky='W', padx=5, pady=5, ipadx=5, ipady=5)
        inEncLbl_1b_1 = tk.Label(stepTwo, text="                                  ")
        inEncLbl_1b_1.grid(row=0, column=1, sticky='E', padx=5, pady=2)
        inEncLbl_1b_2 = tk.Label(stepTwo, text="                                  ")
        inEncLbl_1b_2.grid(row=0, column=2, sticky='E', padx=5, pady=2)
        inEncLbl_1b_3 = tk.Label(stepTwo, text="                                  ")
        inEncLbl_1b_3.grid(row=0, column=3, sticky='E', padx=5, pady=2)
        inEncLbl_1b_4 = tk.Label(stepTwo, text="                                  ")
        inEncLbl_1b_4.grid(row=0, column=4, sticky='E', padx=5, pady=2)

        inFileLbl_1b_1 = tk.Label(stepTwo, text="   1. Select the YAML file from CML", font=("", 10), background="#FFFFCC")
        inFileLbl_1b_1.grid(row=0, column=0, sticky='W', padx=5, pady=2)
        self.inFileTxt_1b_1 = tk.Entry(stepTwo)
        self.inFileTxt_1b_1.grid(row=0, column=1, columnspan=7, sticky="WE", pady=3)
        self.outFileTxt_1b_4_1 = tk.Entry(stepTwo)
        inFileBtn_1b_1 = tk.Button(stepTwo, text="Browse ...", command=lambda: self.click_action('1-1b'))
        inFileBtn_1b_1.grid(row=0, column=8, sticky='W', padx=5, pady=2)
        self.outFileTxt_1b_1 = tk.Entry(stepTwo)
        self.outFileTxt_1b_2 = tk.Entry(stepTwo)
        inFileLbl_1b_x = tk.Label(stepTwo, text="   2.Click the Create button", font=("", 10), background="#FFFFCC")
        inFileLbl_1b_x.grid(row=3, column=0, sticky='W', padx=5, pady=2)
        Create_ND_1b_1 = tk.Button(stepTwo, text="Convert to [Excel Master Data]", font=("", 12), command=(lambda: self.click_action('1-4b')))
        Create_ND_1b_1.grid(row=3, column=1, columnspan=7, sticky='WE', padx=5, pady=2)

        '''
        tab1  <<NW Diagram>>
        '''
        # [INPUT] Excel Master file:
        StepTwo = tk.LabelFrame(tab1, text=" [INPUT] Excel Master Data     [OUTPUT] Various types of Diagram ", font=("", 14), height=1, background="#fff2cc")
        StepTwo.grid(row=6, columnspan=7, sticky='W', padx=5, pady=5, ipadx=5, ipady=5)
        inEncLbl_2_1 = tk.Label(StepTwo, text="                                  ")
        inEncLbl_2_1.grid(row=0, column=1, sticky='E', padx=5, pady=2)
        inEncLbl_2_2 = tk.Label(StepTwo, text="                                  ")
        inEncLbl_2_2.grid(row=0, column=2, sticky='E', padx=5, pady=2)
        inEncLbl_2_3 = tk.Label(StepTwo, text="                                  ")
        inEncLbl_2_3.grid(row=0, column=3, sticky='E', padx=5, pady=2)
        inEncLbl_2_4 = tk.Label(StepTwo, text="                                  ")
        inEncLbl_2_4.grid(row=0, column=4, sticky='E', padx=5, pady=2)

        inFileLbl_2_1 = tk.Label(StepTwo, text="   1. Select the Excel Master Data file", font=("", 10), background="#fff2cc", justify='left')
        inFileLbl_2_1.grid(row=0, column=0, sticky='W', padx=5, pady=2)
        self.inFileTxt_2_1 = tk.Entry(StepTwo)
        self.inFileTxt_2_1.grid(row=0, column=1, columnspan=7, sticky="WE", pady=3)
        inFileBtn_2_1 = tk.Button(StepTwo, text="Browse ...", command=lambda: self.click_action('2-1'))
        inFileBtn_2_1.grid(row=0, column=8, sticky='W', padx=5, pady=2)

        self.outFileTxt_2_1 = tk.Entry(StepTwo)
        self.outFileTxt_2_2 = tk.Entry(StepTwo)
        self.outFileTxt_2_3 = tk.Entry(StepTwo)
        self.outFileTxt_2_4 = tk.Entry(StepTwo)

        inFileLbl_2_x = tk.Label(StepTwo, text="   2. Click the Diagram type button", font=("", 10), background="#fff2cc")
        inFileLbl_2_x.grid(row=3, column=0, sticky='W', padx=5, pady=2)
        Create_ND_2_1_1 = tk.Button(StepTwo, text="Per Area", font=("", 12), command=lambda: self.click_action('2-4-1'))
        Create_ND_2_1_1.grid(row=3, column=1, columnspan=2, sticky='WE', padx=5, pady=2)
        Create_ND_2_1_2 = tk.Button(StepTwo, text="Per Area with IF Tag", font=("", 12), command=lambda: self.click_action('2-4-2'))
        Create_ND_2_1_2.grid(row=3, column=3, columnspan=2, sticky='WE', padx=5, pady=2)
        Create_ND_2_1_3 = tk.Button(StepTwo, text="All Areas", font=("", 12), command=lambda: self.click_action('2-4-3'))
        Create_ND_2_1_3.grid(row=4, column=1, columnspan=2, sticky='WE', padx=5, pady=2)
        Create_ND_2_1_4 = tk.Button(StepTwo, text="All Areas with IF Tag", font=("", 12), command=lambda: self.click_action('2-4-4'))
        Create_ND_2_1_4.grid(row=4, column=3, columnspan=2, sticky='WE', padx=5, pady=2)

        # [Input] diagram per area [Sync with] Excel Master file:
        StepThree = tk.LabelFrame(tab1, text=" [INPUT] Updated Diagram(Per Area)   [Sync with] Excel Master Data ", font=("", 14), height=1, background="#E6E0EC")
        StepThree.grid(row=7, columnspan=7, sticky='W', padx=5, pady=5, ipadx=5, ipady=5)
        inEncLbl_92_1 = tk.Label(StepThree, text="                                  ")
        inEncLbl_92_1.grid(row=0, column=1, sticky='E', padx=5, pady=2)
        inEncLbl_92_2 = tk.Label(StepThree, text="                                  ")
        inEncLbl_92_2.grid(row=0, column=2, sticky='E', padx=5, pady=2)
        inEncLbl_92_3 = tk.Label(StepThree, text="                                  ")
        inEncLbl_92_3.grid(row=0, column=3, sticky='E', padx=5, pady=2)
        inEncLbl_92_4 = tk.Label(StepThree, text="                                  ")
        inEncLbl_92_4.grid(row=0, column=4, sticky='E', padx=5, pady=2)

        inFileLbl_92_1 = tk.Label(StepThree, text="   1.Select the Updated Diagram(Per Area) file", font=("", 10), background="#E6E0EC")
        inFileLbl_92_1.grid(row=0, column=0, sticky='W', padx=5, pady=2)
        self.inFileTxt_92_1 = tk.Entry(StepThree)
        self.inFileTxt_92_1_2 = tk.Entry(StepThree)
        self.outFileTxt_92_3_1 = tk.Entry(StepThree)
        self.inFileTxt_92_1.grid(row=0, column=1, columnspan=7, sticky="WE", pady=3)
        inFileBtn_92_1 = tk.Button(StepThree, text="Browse ...", command=lambda: self.click_action('92-1'))
        inFileBtn_92_1.grid(row=0, column=8, sticky='W', padx=5, pady=2)

        inFileLbl_92_2 = tk.Label(StepThree, text="   2.Select the Target Excel Master file", font=("", 10), background="#E6E0EC")
        inFileLbl_92_2.grid(row=1, column=0, sticky='W', padx=5, pady=2)
        self.inFileTxt_92_2 = tk.Entry(StepThree)
        self.inFileTxt_92_2_2 = tk.Entry(StepThree)
        self.inFileTxt_92_2_3 = tk.Entry(StepThree)
        self.inFileTxt_92_2.grid(row=1, column=1, columnspan=7, sticky="WE", pady=3)
        inFileBtn_92_2 = tk.Button(StepThree, text="Browse ...", command=lambda: self.click_action('92-2'))
        inFileBtn_92_2.grid(row=1, column=8, sticky='W', padx=5, pady=2)

        inFileLbl_92_x = tk.Label(StepThree, text="   3.Click the Sync button", font=("", 10), background="#E6E0EC")
        inFileLbl_92_x.grid(row=3, column=0, sticky='W', padx=5, pady=2)
        Create_ND_92_3 = tk.Button(StepThree, text="Sync", font=("", 12), command=lambda: self.click_action('92-3'))
        Create_ND_92_3.grid(row=3, column=1, columnspan=4, sticky='WE', padx=5, pady=2)

        #credit
        frame3_4_5_1 = tk.Frame(tab0, pady=10)
        frame3_4_5_1.grid(column=0, row=23, sticky='W')
        label3_4_5_1 = tk.Label(frame3_4_5_1, font=("", 10), text="\n\n\n\n\n")
        label3_4_5_1.grid(column=0, row=23)
        frame3_4_5 = tk.Frame(tab0, pady=10)
        frame3_4_5.grid(column=0, row=25, sticky='W')
        label3_4_5 = tk.Label(frame3_4_5, font=("", 10), text="Author :  Yusuke Ogawa ( yuogawa@cisco.com )  \n  © 2022  Cisco and/or its affiliates. All rights reserved.", background='#e4eee8')
        label3_4_5.grid(column=0, row=25)

        '''
        tab2  <<Device Table>>
        '''
        # [INPUT] Excel Master Data file
        stepOne_2 = tk.LabelFrame(tab2, text="[INPUT] Excel Master Data     [OUTPUT] Device(L1) Table ", font=("", 14), height=1, background="#FDEADA")
        stepOne_2.grid(row=0, columnspan=7, sticky='W', padx=5, pady=5, ipadx=5, ipady=5)
        inEncLbl_11_1 = tk.Label(stepOne_2, text="                                  ")
        inEncLbl_11_1.grid(row=0, column=1, sticky='E', padx=5, pady=2)
        inEncLbl_11_2 = tk.Label(stepOne_2, text="                                  ")
        inEncLbl_11_2.grid(row=0, column=2, sticky='E', padx=5, pady=2)
        inEncLbl_11_3 = tk.Label(stepOne_2, text="                                  ")
        inEncLbl_11_3.grid(row=0, column=3, sticky='E', padx=5, pady=2)
        inEncLbl_11_4 = tk.Label(stepOne_2, text="                                  ")
        inEncLbl_11_4.grid(row=0, column=4, sticky='E', padx=5, pady=2)

        inFileLbl_11_1 = tk.Label(stepOne_2, text="   1. Select the Excel Master file", font=("", 10), background="#FDEADA")
        inFileLbl_11_1.grid(row=0, column=0, sticky='W', padx=5, pady=2)
        self.inFileTxt_11_1 = tk.Entry(stepOne_2)
        self.inFileTxt_11_1.grid(row=0, column=1, columnspan=7, sticky="WE", pady=3)
        inFileBtn_11_1 = tk.Button(stepOne_2, text="Browse ...", command=lambda: self.click_action('11-1'))
        inFileBtn_11_1.grid(row=0, column=8, sticky='W', padx=5, pady=2)

        inFileLbl_11_x = tk.Label(stepOne_2, text="   2. Click the Create button", font=("", 10), background="#FDEADA")
        inFileLbl_11_x.grid(row=3, column=0, sticky='W', padx=5, pady=2)
        self.outFileTxt_11_2 = tk.Entry(stepOne_2)
        Create_ND_11_1 = tk.Button(stepOne_2, text="Create  [Device Table]", font=("", 12), command=lambda: self.click_action('11-4'))
        Create_ND_11_1.grid(row=3, column=1, columnspan=7, sticky='WE', padx=5, pady=2)

        # [INPUT] Device Table file:
        #inEncLbl_12_1 = tk.Label(tab2, text="                                  ")
        #inEncLbl_12_1.grid(row=5, column=1, sticky='E', padx=5, pady=2)
        StepTwo_2 = tk.LabelFrame(tab2, text=" [INPUT] Updated L1 Table     [Sync with] Excel Master Data ", font=("", 14), height=1, background="#E6E0EC")
        StepTwo_2.grid(row=6, columnspan=7, sticky='W', padx=5, pady=5, ipadx=5, ipady=5)
        inEncLbl_12_1 = tk.Label(StepTwo_2, text="                                  ")
        inEncLbl_12_1.grid(row=0, column=1, sticky='E', padx=5, pady=2)
        inEncLbl_12_2 = tk.Label(StepTwo_2, text="                                  ")
        inEncLbl_12_2.grid(row=0, column=2, sticky='E', padx=5, pady=2)
        inEncLbl_12_3 = tk.Label(StepTwo_2, text="                                  ")
        inEncLbl_12_3.grid(row=0, column=3, sticky='E', padx=5, pady=2)
        inEncLbl_12_4 = tk.Label(StepTwo_2, text="                                  ")
        inEncLbl_12_4.grid(row=0, column=4, sticky='E', padx=5, pady=2)

        inFileLbl_12_1 = tk.Label(StepTwo_2, text="   1.Select the Updated L1 Table file", font=("", 10), background="#E6E0EC")
        inFileLbl_12_1.grid(row=0, column=0, sticky='W', padx=5, pady=2)
        self.inFileTxt_12_1 = tk.Entry(StepTwo_2)
        self.inFileTxt_12_1_2 = tk.Entry(StepTwo_2)
        self.outFileTxt_12_3_1 = tk.Entry(StepTwo_2)
        self.inFileTxt_12_1.grid(row=0, column=1, columnspan=7, sticky="WE", pady=3)
        inFileBtn_12_1 = tk.Button(StepTwo_2, text="Browse ...", command=lambda: self.click_action('12-1'))
        inFileBtn_12_1.grid(row=0, column=8, sticky='W', padx=5, pady=2)

        inFileLbl_12_2 = tk.Label(StepTwo_2, text="   2.Select the Target Excel Master file", font=("", 10), background="#E6E0EC")
        inFileLbl_12_2.grid(row=1, column=0, sticky='W', padx=5, pady=2)
        self.inFileTxt_12_2 = tk.Entry(StepTwo_2)
        self.inFileTxt_12_2_2 = tk.Entry(StepTwo_2)
        self.inFileTxt_12_2_3 = tk.Entry(StepTwo_2)
        self.inFileTxt_12_2.grid(row=1, column=1, columnspan=7, sticky="WE", pady=3)
        inFileBtn_12_2 = tk.Button(StepTwo_2, text="Browse ...", command=lambda: self.click_action('12-2'))
        inFileBtn_12_2.grid(row=1, column=8, sticky='W', padx=5, pady=2)

        inFileLbl_12_x = tk.Label(StepTwo_2, text="   3.Click the Sync button", font=("", 10), background="#E6E0EC")
        inFileLbl_12_x.grid(row=3, column=0, sticky='W', padx=5, pady=2)
        Create_ND_12_3 = tk.Button(StepTwo_2, text="Sync", font=("", 12), command=lambda: self.click_action('12-3'))
        Create_ND_12_3.grid(row=3, column=1, columnspan=7, sticky='WE', padx=5, pady=2)

        # main loop
        root.mainloop()

    def click_action(self,click_value):
        if click_value == '1-1': # select browse
            fTyp = [("", ".pptx")]
            iDir = os.path.abspath(os.path.dirname(sys.argv[0]))
            full_filepath = tk.filedialog.askopenfilename(filetypes=fTyp, initialdir=iDir)
            self.inFileTxt_1_1.delete(0, tkinter.END)
            self.inFileTxt_1_1.insert(tk.END, full_filepath)

            basename_without_ext = os.path.splitext(os.path.basename(full_filepath))[0]
            self.outFileTxt_1_1.delete(0, tkinter.END)
            self.outFileTxt_1_1.insert(tk.END, iDir + ns_def.return_os_slash() + '[L1_DIAGRAM]AllAreasTag_' + basename_without_ext  + '.pptx')
            #print(self.outFileTxt_1_1.get())
            self.outFileTxt_1_2.delete(0, tkinter.END)
            self.outFileTxt_1_2.insert(tk.END, iDir + ns_def.return_os_slash() + '[MASTER]' + basename_without_ext  + '.xlsx')
            #print(self.outFileTxt_1_2.get())

            # create device table
            basename_without_ext = os.path.splitext(os.path.basename(self.outFileTxt_1_1.get()))[0]
            self.outFileTxt_1_4_1.delete(0, tkinter.END)
            self.outFileTxt_1_4_1.insert(tk.END, iDir + ns_def.return_os_slash() + '[DEVICE]' + basename_without_ext + '.xlsx')
            #print(self.outFileTxt_1_4_1.get())

            self.output_ppt_file = self.outFileTxt_1_1.get()  # default pptx file path

        if click_value == '1-1a': # select browse
            fTyp = [("", ".svg")]
            iDir = os.path.abspath(os.path.dirname(sys.argv[0]))
            full_filepath = tk.filedialog.askopenfilename(filetypes=fTyp, initialdir=iDir)
            self.inFileTxt_1a_1.delete(0, tkinter.END)
            self.inFileTxt_1a_1.insert(tk.END, full_filepath)

            basename_without_ext = os.path.splitext(os.path.basename(full_filepath))[0]
            self.outFileTxt_1a_1.delete(0, tkinter.END)
            self.outFileTxt_1a_1.insert(tk.END, iDir + ns_def.return_os_slash() + '[L1_DIAGRAM]AllAreasTag_' + basename_without_ext  + '.pptx')
            #print(self.outFileTxt_1_1.get())
            self.outFileTxt_1a_2.delete(0, tkinter.END)
            self.outFileTxt_1a_2.insert(tk.END, iDir + ns_def.return_os_slash() + '[MASTER]' + basename_without_ext  + '.xlsx')
            #print(self.outFileTxt_1_2.get())

            # create device table
            basename_without_ext = os.path.splitext(os.path.basename(self.outFileTxt_1a_1.get()))[0]
            self.outFileTxt_1a_4_1.delete(0, tkinter.END)
            self.outFileTxt_1a_4_1.insert(tk.END, iDir + ns_def.return_os_slash() + '[DEVICE]' + basename_without_ext + '.xlsx')
            #print(self.outFileTxt_1_4_1.get())

            self.output_ppt_file = self.outFileTxt_1a_1.get()  # default svg file path

        if click_value == '1-1b': # select browse
            fTyp = [("", ".yaml")]
            iDir = os.path.abspath(os.path.dirname(sys.argv[0]))
            full_filepath = tk.filedialog.askopenfilename(filetypes=fTyp, initialdir=iDir)
            self.inFileTxt_1b_1.delete(0, tkinter.END)
            self.inFileTxt_1b_1.insert(tk.END, full_filepath)

            basename_without_ext = os.path.splitext(os.path.basename(full_filepath))[0]
            self.outFileTxt_1b_1.delete(0, tkinter.END)
            self.outFileTxt_1b_1.insert(tk.END, iDir + ns_def.return_os_slash() + '[L1_DIAGRAM]AllAreasTag_' + basename_without_ext  + '.pptx')
            #print(self.outFileTxt_1_1.get())
            self.outFileTxt_1b_2.delete(0, tkinter.END)
            self.outFileTxt_1b_2.insert(tk.END, iDir + ns_def.return_os_slash() + '[MASTER]' + basename_without_ext  + '.xlsx')
            #print(self.outFileTxt_1_2.get())

            # create device table
            basename_without_ext = os.path.splitext(os.path.basename(self.outFileTxt_1a_1.get()))[0]
            self.outFileTxt_1a_4_1.delete(0, tkinter.END)
            self.outFileTxt_1a_4_1.insert(tk.END, iDir + ns_def.return_os_slash() + '[DEVICE]' + basename_without_ext + '.xlsx')
            #print(self.outFileTxt_1_4_1.get())

            self.output_ppt_file = self.outFileTxt_1b_1.get()  # default yaml file path

        if click_value == '1-4':  # select create
            # check file open
            if ns_def.check_file_open(self.outFileTxt_1_1.get()) == True:
                return ()
            if ns_def.check_file_open(self.outFileTxt_1_2.get()) == True:
                return ()

            # remove exist ppt file
            if os.path.isfile(self.outFileTxt_1_1.get()) == True:
                os.remove(self.outFileTxt_1_1.get())

            self.click_value = '1-4'
            # import module
            import ns_l1_master_create
            ns_l1_master_create.ns_l1_master_create.__init__(self)

            # input master excel file path to 2-1  , ver 1.1
            iDir = os.path.abspath(os.path.dirname(self.outFileTxt_1_2.get()))
            self.inFileTxt_2_1.delete(0, tkinter.END)
            self.inFileTxt_2_1.insert(tk.END, self.outFileTxt_1_2.get())

            basename_without_ext = os.path.splitext(os.path.basename(self.outFileTxt_1_2.get()))[0]
            self.outFileTxt_2_1.delete(0, tkinter.END)
            self.outFileTxt_2_1.insert(tk.END, iDir + ns_def.return_os_slash() + '[L1_DIAGRAM]PerArea_' + basename_without_ext.replace('[MASTER]', '') + '.pptx')
            self.outFileTxt_2_2.delete(0, tkinter.END)
            self.outFileTxt_2_2.insert(tk.END, iDir + ns_def.return_os_slash() + '[L1_DIAGRAM]PerAreaTag_' + basename_without_ext.replace('[MASTER]', '') + '.pptx')
            self.outFileTxt_2_3.delete(0, tkinter.END)
            self.outFileTxt_2_3.insert(tk.END, iDir + ns_def.return_os_slash() + '[L1_DIAGRAM]AllAreas_' + basename_without_ext.replace('[MASTER]', '') + '.pptx')
            self.outFileTxt_2_4.delete(0, tkinter.END)
            self.outFileTxt_2_4.insert(tk.END, iDir + ns_def.return_os_slash() + '[L1_DIAGRAM]AllAreasTag_' + basename_without_ext.replace('[MASTER]', '') + '.pptx')

            ### input master excel file path to 11-1  , ver 1.11
            iDir = os.path.abspath(os.path.dirname(self.outFileTxt_1_2.get()))
            self.inFileTxt_11_1.delete(0, tkinter.END)
            self.inFileTxt_11_1.insert(tk.END, self.outFileTxt_1_2.get())

            # SET Device file patch
            basename_without_ext = os.path.splitext(os.path.basename(self.outFileTxt_1_2.get()))[0]
            self.outFileTxt_11_2.delete(0, tkinter.END)
            self.outFileTxt_11_2.insert(tk.END, iDir + ns_def.return_os_slash() + '[DEVICE]' + basename_without_ext + '.xlsx')

            #view complete
            #tkinter.messagebox.showinfo('info', 'successfully completed')

        if click_value == '1-4a':  # select convert to master data file from SVG
            # check file open
            if ns_def.check_file_open(self.outFileTxt_1a_1.get()) == True:
                return ()
            if ns_def.check_file_open(self.outFileTxt_1a_2.get()) == True:
                return ()

            # remove exist ppt file
            if os.path.isfile(self.outFileTxt_1a_1.get()) == True:
                os.remove(self.outFileTxt_1a_1.get())

            self.click_value = '1-4a'
            # import module
            import ns_option_convert_to_master
            ns_option_convert_to_master.ns_option_convert_to_master_svg.__init__(self)

            # import module
            import ns_l1_master_create
            ns_l1_master_create.ns_l1_master_create.__init__(self)

            # remove exist ppt file
            if os.path.isfile("./_tmp_tmp_tmp_.pptx") == True:
                os.remove("./_tmp_tmp_tmp_.pptx")

            # input master excel file path to 2-1  , ver 1.1
            iDir = os.path.abspath(os.path.dirname(self.outFileTxt_1a_2.get()))
            self.inFileTxt_2_1.delete(0, tkinter.END)
            self.inFileTxt_2_1.insert(tk.END, self.outFileTxt_1a_2.get())

            basename_without_ext = os.path.splitext(os.path.basename(self.outFileTxt_1a_2.get()))[0]
            self.outFileTxt_2_1.delete(0, tkinter.END)
            self.outFileTxt_2_1.insert(tk.END, iDir + ns_def.return_os_slash() + '[L1_DIAGRAM]PerArea_' + basename_without_ext.replace('[MASTER]', '') + '.pptx')
            self.outFileTxt_2_2.delete(0, tkinter.END)
            self.outFileTxt_2_2.insert(tk.END, iDir + ns_def.return_os_slash() + '[L1_DIAGRAM]PerAreaTag_' + basename_without_ext.replace('[MASTER]', '') + '.pptx')
            self.outFileTxt_2_3.delete(0, tkinter.END)
            self.outFileTxt_2_3.insert(tk.END, iDir + ns_def.return_os_slash() + '[L1_DIAGRAM]AllAreas_' + basename_without_ext.replace('[MASTER]', '') + '.pptx')
            self.outFileTxt_2_4.delete(0, tkinter.END)
            self.outFileTxt_2_4.insert(tk.END, iDir + ns_def.return_os_slash() + '[L1_DIAGRAM]AllAreasTag_' + basename_without_ext.replace('[MASTER]', '') + '.pptx')

            #view complete
            tkinter.messagebox.showinfo('info', 'successfully completed')

        if click_value == '1-4b':  # select convert to master data file from YAML
            # check file open
            print(self.full_filepath)
            self.yaml_full_filepath = self.full_filepath

            self.click_value = '1-4b'

            # input master excel file path  , ver 2.1.0
            iDir = os.path.abspath(os.path.dirname(self.full_filepath))
            basename_without_ext = os.path.splitext(os.path.basename(self.full_filepath))[0]
            self.excel_file_path = iDir + ns_def.return_os_slash() + '[MASTER]' + basename_without_ext.replace('.yaml', '') + '.xlsx'
            print(self.excel_file_path )


            # import module and create a tmp powerpoint figure
            import ns_option_convert_to_master
            ns_option_convert_to_master.ns_option_convert_to_master_yaml.__init__(self)

            # import module and create a Master data excel file
            import ns_l1_master_create
            ns_l1_master_create.ns_l1_master_create.__init__(self)

            # Update the Master data excel file. Overwrite lines values.
            ns_option_convert_to_master.ns_overwrite_line_to_master_yaml.__init__(self)

            # remove exist ppt file
            if os.path.isfile("./_tmp_tmp_tmp_.pptx") == True:
                os.remove("./_tmp_tmp_tmp_.pptx")

            ### run L2-1-2 in network_sketcher_dev ,  add l2 master sheet
            self.inFileTxt_L2_1_1 = tk.Entry(self.main1_1)
            self.inFileTxt_L3_1_1 = tk.Entry(self.main1_1)
            self.inFileTxt_L2_1_1.delete(0, tkinter.END)
            self.inFileTxt_L2_1_1.insert(tk.END, iDir + ns_def.return_os_slash() + '[MASTER]' + basename_without_ext + '.xlsx')
            self.inFileTxt_L3_1_1.delete(0, tkinter.END)
            self.inFileTxt_L3_1_1.insert(tk.END, iDir + ns_def.return_os_slash() + '[MASTER]' + basename_without_ext + '.xlsx')

            self.click_value = 'L2-1-2'
            ns_front_run.click_action(self,'L2-1-2')

            # remove exist L2/ file
            if os.path.isfile(self.inFileTxt_L2_1_1.get().replace('[MASTER]', '[L2_TABLE]')) == True:
                os.remove(self.inFileTxt_L2_1_1.get().replace('[MASTER]', '[L2_TABLE]'))

            ### run L3-1-2 in network_sketcher_dev ,  add l3 master sheet
            self.click_value = 'L3-1-2'
            ns_front_run.click_action(self,'L3-1-2')

            # remove exist L3/ file
            if os.path.isfile(self.inFileTxt_L2_1_1.get().replace('[MASTER]', '[L3_TABLE]')) == True:
                os.remove(self.inFileTxt_L2_1_1.get().replace('[MASTER]', '[L3_TABLE]'))

            ##view complete
            #tkinter.messagebox.showinfo('info', 'successfully completed')

            ### open master panel
            file_type_array = ['EXCEL_MASTER', 'EXCEL_MASTER']
            self.full_filepath = iDir + ns_def.return_os_slash() + '[MASTER]' + basename_without_ext + '.xlsx'
            self.filename = os.path.basename(self.full_filepath)
            import network_sketcher
            network_sketcher.ns_front_run.sub_excel_master_1(self, file_type_array)

            '''add L3 config in yaml to Master file'''
            ns_option_convert_to_master.ns_l3_config_to_master_yaml.__init__(self)


        if click_value == '2-1': # select browse
            fTyp = [("", ".xlsx")]
            iDir = os.path.abspath(os.path.dirname(sys.argv[0]))
            full_filepath = tk.filedialog.askopenfilename(filetypes=fTyp, initialdir=iDir)
            self.inFileTxt_2_1.delete(0, tkinter.END)
            self.inFileTxt_2_1.insert(tk.END, full_filepath)

            basename_without_ext = os.path.splitext(os.path.basename(full_filepath))[0]
            self.outFileTxt_2_1.delete(0, tkinter.END)
            self.outFileTxt_2_1.insert(tk.END, iDir + ns_def.return_os_slash() + '[L1_DIAGRAM]PerArea_' + basename_without_ext.replace('[MASTER]', '')  + '.pptx')
            self.outFileTxt_2_2.delete(0, tkinter.END)
            self.outFileTxt_2_2.insert(tk.END, iDir + ns_def.return_os_slash() + '[L1_DIAGRAM]PerAreaTag_' + basename_without_ext.replace('[MASTER]', '')  + '.pptx')
            self.outFileTxt_2_3.delete(0, tkinter.END)
            self.outFileTxt_2_3.insert(tk.END, iDir + ns_def.return_os_slash() + '[L1_DIAGRAM]AllAreas_' + basename_without_ext.replace('[MASTER]', '')  + '.pptx')
            self.outFileTxt_2_4.delete(0, tkinter.END)
            self.outFileTxt_2_4.insert(tk.END, iDir + ns_def.return_os_slash() + '[L1_DIAGRAM]AllAreasTag_' + basename_without_ext.replace('[MASTER]', '')  + '.pptx')

        if click_value == '2-4-1':  # select create from master
            self.click_value = '2-4-1'
            self.output_ppt_file = self.outFileTxt_2_1.get()

            # check : file is being opened
            if ns_def.check_file_open(self.outFileTxt_2_1.get()) == True:
                return ()

            # remove exist ppt file
            if os.path.isfile(self.outFileTxt_2_1.get()) == True:
                os.remove(self.outFileTxt_2_1.get())

            # check : file is being opened
            if ns_def.check_file_open(self.inFileTxt_2_1.get()) == True:
                return ()

            # import module
            import ns_l1_diagram_create
            ns_l1_diagram_create.ns_l1_diagram_create.__init__(self)

            # view complete
            ns_def.messagebox_file_open(self.output_ppt_file)

        if click_value == '2-4-2':  # select create from master
            self.click_value = '2-4-2'
            self.output_ppt_file = self.outFileTxt_2_2.get()

            # check : file is being opened
            if ns_def.check_file_open(self.outFileTxt_2_2.get()) == True:
                return ()

            # remove exist ppt file
            if os.path.isfile(self.outFileTxt_2_2.get()) == True:
                os.remove(self.outFileTxt_2_2.get())

            # check : file is being opened
            if ns_def.check_file_open(self.inFileTxt_2_1.get()) == True:
                return ()

            # import module
            import ns_l1_diagram_create
            ns_l1_diagram_create.ns_l1_diagram_create.__init__(self)

            # view complete
            ns_def.messagebox_file_open(self.output_ppt_file)

        if click_value == '2-4-3':  # select create from master
            if self.click_value == 'VPN-1-1':
                self.output_ppt_file = self.outFileTxt_2_3.get().replace('[L1_DIAGRAM]AllAreas_','[VPNs_on_L1]')
            else:
                self.click_value = '2-4-3'
                self.output_ppt_file = self.outFileTxt_2_3.get()

            # check : file is being opened
            if ns_def.check_file_open(self.output_ppt_file) == True:
                return ()

            # remove exist ppt file
            if os.path.isfile(self.output_ppt_file) == True:
                os.remove(self.output_ppt_file)

            # check : file is being opened
            if ns_def.check_file_open(self.inFileTxt_2_1.get()) == True:
                return ()

            # import module
            import ns_l1_diagram_create
            ns_l1_diagram_create.ns_l1_diagram_create.__init__(self)

            # view complete
            ns_def.messagebox_file_open(self.output_ppt_file)

            # return value of self.click_value for re-run 2-4-3 after VPN-1-1'
            self.click_value = ''


        if click_value == '2-4-4':  # select create from master
            self.click_value = '2-4-4'
            self.output_ppt_file = self.outFileTxt_2_4.get()

            # check : file is being opened
            if ns_def.check_file_open(self.outFileTxt_2_4.get()) == True:
                return ()

            # remove exist ppt file
            if os.path.isfile(self.outFileTxt_2_4.get()) == True:
                os.remove(self.outFileTxt_2_4.get())

            # check : file is being opened
            if ns_def.check_file_open(self.inFileTxt_2_1.get()) == True:
                return ()

            # import module
            import ns_l1_diagram_create
            ns_l1_diagram_create.ns_l1_diagram_create.__init__(self)

            # view complete
            ns_def.messagebox_file_open(self.output_ppt_file)

        if click_value == '11-1': # select browse
            fTyp = [("", ".xlsx")]
            iDir = os.path.abspath(os.path.dirname(sys.argv[0]))
            full_filepath = tk.filedialog.askopenfilename(filetypes=fTyp, initialdir=iDir)
            self.inFileTxt_11_1.delete(0, tkinter.END)
            self.inFileTxt_11_1.insert(tk.END, full_filepath)

            # SET Device file patch
            basename_without_ext = os.path.splitext(os.path.basename(full_filepath))[0]
            self.outFileTxt_11_2.delete(0, tkinter.END)
            self.outFileTxt_11_2.insert(tk.END, iDir + ns_def.return_os_slash() + '[DEVICE]' + basename_without_ext  + '.xlsx')
            #print(self.outFileTxt_11_2.get())


        if click_value == '11-4':  # select create from master

            self.click_value = '11-4'
            ## check file open
            if ns_def.check_file_open(str(self.outFileTxt_11_2.get()).replace('[MASTER]','')) == True:
                return ()

            # remove exist device file
            if os.path.isfile(self.outFileTxt_11_2.get()) == True:
                os.remove(self.outFileTxt_11_2.get())

            # check : file is being opened
            if ns_def.check_file_open(self.inFileTxt_11_1.get()) == True:
                return ()

            # import module
            import ns_l1_table_from_master
            ns_l1_table_from_master.ns_l1_table_from_master.__init__(self)

            # view complete
            #tkinter.messagebox.showinfo('info', 'successfully completed')

        if click_value == '12-1':  # select browse
            fTyp = [("", ".xlsx")]
            iDir = os.path.abspath(os.path.dirname(sys.argv[0]))
            full_filepath = tk.filedialog.askopenfilename(filetypes=fTyp, initialdir=iDir)
            self.inFileTxt_12_1.delete(0, tkinter.END)
            self.inFileTxt_12_1.insert(tk.END, full_filepath)

            # SET other file path
            basename_without_ext = os.path.splitext(os.path.basename(full_filepath))[0]
            self.inFileTxt_12_1_2.delete(0, tkinter.END)
            self.inFileTxt_12_1_2.insert(tk.END, iDir + ns_def.return_os_slash() + basename_without_ext+ '_backup'  + '.xlsx')

        if click_value == '12-2':  # select browse
            fTyp = [("", ".xlsx")]
            iDir = os.path.abspath(os.path.dirname(sys.argv[0]))
            full_filepath = tk.filedialog.askopenfilename(filetypes=fTyp, initialdir=iDir)
            self.inFileTxt_12_2.delete(0, tkinter.END)
            self.inFileTxt_12_2.insert(tk.END, full_filepath)

            basename_without_ext = os.path.splitext(os.path.basename(full_filepath))[0]
            self.inFileTxt_12_2_2.delete(0, tkinter.END)
            self.inFileTxt_12_2_2.insert(tk.END, iDir + ns_def.return_os_slash() + basename_without_ext+ '_backup'  + '.xlsx')

            #make ppt diagram backup patch
            self.inFileTxt_12_2_3.delete(0, tkinter.END)
            self.inFileTxt_12_2_3.insert(tk.END, iDir + ns_def.return_os_slash() + '[L1_DIAGRAM]AllAreasTag_' + str(basename_without_ext).replace('[MASTER]','') + '_backup'  + '.pptx')

            # SET ppt diagram file patch
            basename_without_ext = os.path.splitext(os.path.basename(full_filepath))[0]
            self.outFileTxt_12_3_1.delete(0, tkinter.END)
            self.outFileTxt_12_3_1.insert(tk.END, iDir + ns_def.return_os_slash() + '[L1_DIAGRAM]AllAreasTag_' + str(basename_without_ext).replace('[MASTER]','')  + '.pptx')
            #print(self.outFileTxt_12_3_1.get())

        if click_value == '12-3':  # select create from master

            self.click_value = '12-3'
            # check : file is being opened
            '''if ns_def.check_file_open(self.inFileTxt_12_1.get()) == True:
                return ()
            if ns_def.check_file_open(self.inFileTxt_12_2.get()) == True:
                return ()
            if ns_def.check_file_open(self.inFileTxt_12_2_2.get()) == True:
                return ()
            if ns_def.check_file_open(self.inFileTxt_12_1_2.get()) == True:
                return ()
            if ns_def.check_file_open(self.inFileTxt_12_2_3.get()) == True:
                return ()
            if ns_def.check_file_open(self.outFileTxt_12_3_1.get()) == True:
                return ()'''

            # confirm to exist device table and master data file
            if os.path.isfile(self.inFileTxt_12_1.get()) == False:
                tkinter.messagebox.showerror('Error', 'Could not find the Device Table file')
            else:
                if os.path.isfile(self.inFileTxt_12_2.get()) == False:
                    tkinter.messagebox.showerror('Error', 'Could not find the Excel Data file')
                else:
                    # import module
                    import ns_l1_table_sync_master
                    flag_return = ns_l1_table_sync_master.ns_l1_table_sync_master.__init__(self)

                    print(self.inFileTxt_12_2.get())
                    full_filepath = self.inFileTxt_12_2.get()

                    self.inFileTxt_11_1.delete(0, tkinter.END)
                    self.inFileTxt_11_1.insert(tk.END,full_filepath)

                    # SET Device file patch
                    basename_without_ext = os.path.splitext(os.path.basename(full_filepath))[0]
                    iDir = os.path.dirname(full_filepath)
                    self.outFileTxt_11_2.delete(0, tkinter.END)
                    self.outFileTxt_11_2.insert(tk.END, iDir + ns_def.return_os_slash() + '[DEVICE]' + basename_without_ext + '.xlsx')

                    # view complete
                    #tkinter.messagebox.showinfo('info', 'successfully completed')

        if click_value == '92-1':  # select browse
            fTyp = [("", ".pptx")]
            iDir = os.path.abspath(os.path.dirname(sys.argv[0]))
            full_filepath = tk.filedialog.askopenfilename(filetypes=fTyp, initialdir=iDir)
            self.inFileTxt_92_1.delete(0, tkinter.END)
            self.inFileTxt_92_1.insert(tk.END, full_filepath)

        if click_value == '92-2':  # select browse
            fTyp = [("", ".xlsx")]
            iDir = os.path.abspath(os.path.dirname(sys.argv[0]))
            full_filepath = tk.filedialog.askopenfilename(filetypes=fTyp, initialdir=iDir)
            self.inFileTxt_92_2.delete(0, tkinter.END)
            self.inFileTxt_92_2.insert(tk.END, full_filepath)

            # make MASTER backup patch
            basename_without_ext = os.path.splitext(os.path.basename(full_filepath))[0]
            self.inFileTxt_92_2_2.delete(0, tkinter.END)
            self.inFileTxt_92_2_2.insert(tk.END, iDir + ns_def.return_os_slash() + basename_without_ext + '_backup' + '.xlsx')

        if click_value == '92-3':  # select browse
            self.click_value = '92-3'
            self.updated_name_array = []

            # check : file is being opened
            if ns_def.check_file_open(self.inFileTxt_92_2_2.get()) == True:
                return ()

            import ns_l1_master_create
            ns_l1_master_create.ns_l1_master_create.__init__(self)

            import ns_l1_diagram_sync_master
            ns_l1_diagram_sync_master.ns_l1_diagram_sync_master.__init__(self)

            if self.click_value_2nd == 'self.sub1_1_button_3':
                ns_def.remove_excel_sheet(self.excel_file_path, 'Master_Data')
                wb = openpyxl.load_workbook(self.excel_file_path)
                ws = wb['Master_Data_tmp_']
                ws.title = 'Master_Data'
                wb.save(self.excel_file_path)
                wb.close()

            # view complete
            #tkinter.messagebox.showinfo('info', 'successfully completed')

            #SET filepath to [OUTPUT] Diagrams
            self.inFileTxt_2_1.delete(0, tkinter.END)
            self.inFileTxt_2_1.insert(tk.END, self.inFileTxt_92_2.get())

            full_filepath = self.inFileTxt_92_2.get()
            iDir = os.path.dirname(full_filepath)
            basename_without_ext = os.path.splitext(os.path.basename(full_filepath))[0]
            self.outFileTxt_2_1.delete(0, tkinter.END)
            self.outFileTxt_2_1.insert(tk.END, iDir + ns_def.return_os_slash() + '[L1_DIAGRAM]PerArea_' + basename_without_ext.replace('[MASTER]', '') + '.pptx')
            self.outFileTxt_2_2.delete(0, tkinter.END)
            self.outFileTxt_2_2.insert(tk.END, iDir + ns_def.return_os_slash() + '[L1_DIAGRAM]PerAreaTag_' + basename_without_ext.replace('[MASTER]', '') + '.pptx')
            self.outFileTxt_2_3.delete(0, tkinter.END)
            self.outFileTxt_2_3.insert(tk.END, iDir + ns_def.return_os_slash() + '[L1_DIAGRAM]AllAreas_' + basename_without_ext.replace('[MASTER]', '') + '.pptx')
            self.outFileTxt_2_4.delete(0, tkinter.END)
            self.outFileTxt_2_4.insert(tk.END, iDir + ns_def.return_os_slash() + '[L1_DIAGRAM]AllAreasTag_' + basename_without_ext.replace('[MASTER]', '') + '.pptx')

        if click_value == 'L2-1-1': # select browse
            fTyp = [("", ".xlsx")]
            iDir = os.path.abspath(os.path.dirname(sys.argv[0]))
            full_filepath = tk.filedialog.askopenfilename(filetypes=fTyp, initialdir=iDir)
            self.inFileTxt_L2_1_1.delete(0, tkinter.END)
            self.inFileTxt_L2_1_1.insert(tk.END, full_filepath)

        if click_value == 'L2-1-2':  # select create from master
            self.click_value = 'L2-1-2'

            # check : file is being opened
            if ns_def.check_file_open(self.inFileTxt_L2_1_1.get().replace('[MASTER]', '[L2_TABLE]')) == True:
                return ()

            # remove exist L2 file
            if os.path.isfile(self.inFileTxt_L2_1_1.get().replace('[MASTER]', '[L2_TABLE]')) == True:
                os.remove(self.inFileTxt_L2_1_1.get().replace('[MASTER]', '[L2_TABLE]'))

            # check Master_Data_L2 sheet already exits in Excel Master data file
            import ns_l2_table_from_master
            input_excel_master_data = openpyxl.load_workbook(self.inFileTxt_L2_1_1.get())
            ws_list = input_excel_master_data.sheetnames

            if 'Master_Data_L2' not in ws_list:
                #### create L2 table excel file and add L2 Data sheet to Master file from master L1 only
                ns_l2_table_from_master.ns_l2_table_from_master.__init__(self)

            else:
                #### create L2 table from master and l2 data sheet
                ns_l2_table_from_master.ns_l2_table_from_master_l2_sheet.__init__(self)

            # view complete
            #tkinter.messagebox.showinfo('info', 'successfully completed')

        if click_value == 'L2-2-1': # select browse
            fTyp = [("", ".xlsx")]
            iDir = os.path.abspath(os.path.dirname(sys.argv[0]))
            full_filepath = tk.filedialog.askopenfilename(filetypes=fTyp, initialdir=iDir)
            self.inFileTxt_L2_2_1.delete(0, tkinter.END)
            self.inFileTxt_L2_2_1.insert(tk.END, full_filepath)

        if click_value == 'L2-2-2': # select browse
            fTyp = [("", ".xlsx")]
            iDir = os.path.abspath(os.path.dirname(sys.argv[0]))
            full_filepath = tk.filedialog.askopenfilename(filetypes=fTyp, initialdir=iDir)
            self.inFileTxt_L2_2_2.delete(0, tkinter.END)
            self.inFileTxt_L2_2_2.insert(tk.END, full_filepath)

        if click_value == 'L2-2-3':  #  L2 Table sync with the master data
            self.click_value = 'L2-2-3'

            iDir = os.path.abspath(os.path.dirname(sys.argv[0]))
            self.inFileTxt_L2_2_2_backup= iDir + ns_def.return_os_slash() + os.path.splitext(os.path.basename(self.inFileTxt_L2_2_2.get()))[0] + '_backup' + '.xlsx'

            # check : file is being opened
            '''if ns_def.check_file_open(self.inFileTxt_L2_2_2.get()) == True:
                return ()'''
            # confirm to exist device table and master data file
            if os.path.isfile(self.inFileTxt_L2_2_1.get()) == False:
                tkinter.messagebox.showerror('Error', 'Could not find the L2 Table file')
            else:
                if os.path.isfile(self.inFileTxt_L2_2_2.get()) == False:
                    tkinter.messagebox.showerror('Error', 'Could not find the Excel Data file')
                else:
                    # import module
                    import ns_l2_table_sync_master
                    ns_l2_table_sync_master.ns_l2_table_sync_master.__init__(self)

                    # view complete
                    #tkinter.messagebox.showinfo('info', 'successfully completed')

        if click_value == 'L2-3-1': # select browse
            fTyp = [("", ".xlsx")]
            iDir = os.path.abspath(os.path.dirname(sys.argv[0]))
            full_filepath = tk.filedialog.askopenfilename(filetypes=fTyp, initialdir=iDir)
            self.inFileTxt_L2_3_1.delete(0, tkinter.END)
            self.inFileTxt_L2_3_1.insert(tk.END, full_filepath)

        if click_value == 'L2-3-2' or click_value == 'L2-3-3':  # select create from master
            full_filepath = self.inFileTxt_L2_3_1.get()
            iDir = os.path.dirname(full_filepath)
            basename_without_ext = os.path.splitext(os.path.basename(full_filepath))[0]

            if click_value == 'L2-3-2':
                self.click_value = 'L2-3-2'
                self.outFileTxt_L2_3_4_1.delete(0, tkinter.END)
                self.outFileTxt_L2_3_4_1.insert(tk.END, iDir + ns_def.return_os_slash() + '[L2_DIAGRAM]PerArea_' + basename_without_ext.replace('[MASTER]', '') + '.pptx')
                self.output_ppt_file = self.outFileTxt_L2_3_4_1.get()

                # check : file is being opened
                if ns_def.check_file_open(self.outFileTxt_L2_3_4_1.get()) == True:
                    return ()

                # remove exist L2 file
                if os.path.isfile(self.outFileTxt_L2_3_4_1.get()) == True:
                    os.remove(self.outFileTxt_L2_3_4_1.get())

            elif click_value == 'L2-3-3':
                self.click_value = 'L2-3-3'
                self.outFileTxt_L2_3_4_1.delete(0, tkinter.END)
                self.outFileTxt_L2_3_4_1.insert(tk.END, iDir + ns_def.return_os_slash() + '[L2_DIAGRAM]PerDevice_' + basename_without_ext.replace('[MASTER]', '') + '.pptx')
                self.output_ppt_file = self.outFileTxt_L2_3_4_1.get()

                # check : file is being opened
                if ns_def.check_file_open(self.outFileTxt_L2_3_4_1.get()) == True:
                    return ()

                # remove exist L2 file
                if os.path.isfile(self.outFileTxt_L2_3_4_1.get()) == True:
                    os.remove(self.outFileTxt_L2_3_4_1.get())

            # check Master_Data_L2 sheet already exits in Excel Master data file
            import ns_l2_diagram_create
            ns_l2_diagram_create.ns_l2_diagram_create.__init__(self)

            # view complete
            ns_def.messagebox_file_open(self.output_ppt_file)


        if click_value == 'L3-1-1': # select browse
            fTyp = [("", ".xlsx")]
            iDir = os.path.abspath(os.path.dirname(sys.argv[0]))
            full_filepath = tk.filedialog.askopenfilename(filetypes=fTyp, initialdir=iDir)
            self.inFileTxt_L3_1_1.delete(0, tkinter.END)
            self.inFileTxt_L3_1_1.insert(tk.END, full_filepath)

        if click_value == 'L3-1-2':  # select create from master
            self.click_value = 'L3-1-2'

            # check : file is being opened
            if ns_def.check_file_open(self.inFileTxt_L3_1_1.get().replace('[MASTER]', '[L3_TABLE]')) == True:
                return ()

            # remove exist L3 file
            if os.path.isfile(self.inFileTxt_L3_1_1.get().replace('[MASTER]', '[L3_TABLE]')) == True:
                os.remove(self.inFileTxt_L3_1_1.get().replace('[MASTER]', '[L3_TABLE]'))

            # check Master_Data_L3 sheet already exits in Excel Master data file
            import ns_l3_table_from_master

            input_excel_master_data = openpyxl.load_workbook(self.inFileTxt_L3_1_1.get())
            ws_list = input_excel_master_data.sheetnames

            if 'Master_Data_L3' not in ws_list:
                #### create L3 table excel file and add L3 Data sheet to Master file from master L1 only
                ns_l3_table_from_master.ns_l3_table_from_master.__init__(self)

            else:
                #### create L3 table from master and L3 data sheet
                ns_l3_table_from_master.ns_l3_table_from_master_l3_sheet.__init__(self)

            # view complete
            #tkinter.messagebox.showinfo('info', 'successfully completed')

        if click_value == 'L3-2-1': # select browse
            fTyp = [("", ".xlsx")]
            iDir = os.path.abspath(os.path.dirname(sys.argv[0]))
            full_filepath = tk.filedialog.askopenfilename(filetypes=fTyp, initialdir=iDir)
            self.inFileTxt_L3_2_1.delete(0, tkinter.END)
            self.inFileTxt_L3_2_1.insert(tk.END, full_filepath)

        if click_value == 'L3-2-2': # select browse
            fTyp = [("", ".xlsx")]
            iDir = os.path.abspath(os.path.dirname(sys.argv[0]))
            full_filepath = tk.filedialog.askopenfilename(filetypes=fTyp, initialdir=iDir)
            self.inFileTxt_L3_2_2.delete(0, tkinter.END)
            self.inFileTxt_L3_2_2.insert(tk.END, full_filepath)

        if click_value == 'L3-2-3':  #  L3 Table sync with the master data
            self.click_value = 'L3-2-3'

            iDir = os.path.abspath(os.path.dirname(sys.argv[0]))
            self.inFileTxt_L3_2_2_backup= iDir + ns_def.return_os_slash() + os.path.splitext(os.path.basename(self.inFileTxt_L3_2_2.get()))[0] + '_backup' + '.xlsx'

            # check : file is being opened
            '''if ns_def.check_file_open(self.inFileTxt_L3_2_2.get()) == True:
                return ()'''
            # confirm to exist device table and master data file
            if os.path.isfile(self.inFileTxt_L3_2_1.get()) == False:
                tkinter.messagebox.showerror('Error', 'Could not find the L3 Table file')
            else:
                if os.path.isfile(self.inFileTxt_L3_2_2.get()) == False:
                    tkinter.messagebox.showerror('Error', 'Could not find the Excel Data file')
                else:
                    # import module
                    import ns_l3_table_sync_master
                    ns_l3_table_sync_master.ns_l3_table_sync_master.__init__(self)

                    # view complete
                    #tkinter.messagebox.showinfo('info', 'successfully completed')

        if click_value == 'L3-3-1': # select browse
            fTyp = [("", ".xlsx")]
            iDir = os.path.abspath(os.path.dirname(sys.argv[0]))
            full_filepath = tk.filedialog.askopenfilename(filetypes=fTyp, initialdir=iDir)
            self.inFileTxt_L3_3_1.delete(0, tkinter.END)
            self.inFileTxt_L3_3_1.insert(tk.END, full_filepath)

        if click_value == 'L3-3-2' or click_value == 'L3-3-3':  # select create from master
            ### TEST MODE ###
            #self.inFileTxt_L3_3_1.delete(0, tkinter.END)
            #self.inFileTxt_L3_3_1.insert(tk.END, 'C:/work/Network Sketcher/Network Skecher Ver2.0/[MASTER]Sample figure5.xlsx')
            #################

            full_filepath = self.inFileTxt_L3_3_1.get()
            iDir = os.path.dirname(full_filepath)
            basename_without_ext = os.path.splitext(os.path.basename(full_filepath))[0]

            if click_value == 'L3-3-2':
                self.click_value = 'L3-3-2'
                self.outFileTxt_L3_3_4_1.delete(0, tkinter.END)
                self.outFileTxt_L3_3_4_1.insert(tk.END, iDir + ns_def.return_os_slash() + '[L3_DIAGRAM]PerArea_' + basename_without_ext.replace('[MASTER]', '') + '.pptx')
                self.output_ppt_file = self.outFileTxt_L3_3_4_1.get()

                # check : file is being opened
                if ns_def.check_file_open(self.outFileTxt_L3_3_4_1.get()) == True:
                    return ()

                # remove exist L3 file
                if os.path.isfile(self.outFileTxt_L3_3_4_1.get()) == True:
                    os.remove(self.outFileTxt_L3_3_4_1.get())

            # check Master_Data_L3 sheet already exits in Excel Master data file
            import ns_l3_diagram_create
            ns_l3_diagram_create.ns_l3_diagram_create.__init__(self)

            # view complete
            ns_def.messagebox_file_open(self.output_ppt_file)

if __name__ == '__main__':
    ns_front_run()
