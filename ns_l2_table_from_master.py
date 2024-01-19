'''
SPDX-License-Identifier: Apache-2.0

Copyright 2023 Cisco Systems, Inc. and its affiliates

Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.
'''

from pptx import *
import sys, os, re
import numpy as np
import math
import openpyxl
import ns_def ,ns_ddx_figure

class  ns_l2_table_from_master():
    def __init__(self):
        '''
        make L2 Table excel file
        '''
        print('### ns_l2_table_from_master() ###')
        #parameter
        ws_name = 'Master_Data'
        tmp_ws_name = '_tmp_'
        excel_maseter_file = self.inFileTxt_L2_1_1.get()
        #write_excel_file = self.outFileTxt_11_2.get().replace('[MASTER]', '')

        #convert from master to array and convert to tuple
        self.position_folder_array = ns_def.convert_master_to_array(ws_name, excel_maseter_file,'<<POSITION_FOLDER>>')
        self.position_shape_array = ns_def.convert_master_to_array(ws_name, excel_maseter_file, '<<POSITION_SHAPE>>')
        self.position_line_array = ns_def.convert_master_to_array(ws_name, excel_maseter_file, '<<POSITION_LINE>>')
        self.position_style_shape_array = ns_def.convert_master_to_array(ws_name, excel_maseter_file, '<<STYLE_SHAPE>>')
        self.position_tag_array = ns_def.convert_master_to_array(ws_name, excel_maseter_file, '<<POSITION_TAG>>')
        self.root_folder_array = ns_def.convert_master_to_array(ws_name, excel_maseter_file, '<<ROOT_FOLDER>>')
        self.position_folder_tuple = ns_def.convert_array_to_tuple(self.position_folder_array)
        self.position_shape_tuple = ns_def.convert_array_to_tuple(self.position_shape_array)
        self.position_line_tuple = ns_def.convert_array_to_tuple(self.position_line_array)
        self.position_style_shape_tuple = ns_def.convert_array_to_tuple(self.position_style_shape_array)
        self.position_tag_tuple = ns_def.convert_array_to_tuple(self.position_tag_array)
        self.root_folder_tuple = ns_def.convert_array_to_tuple(self.root_folder_array)

        print('---- self.position_folder_tuple ----')
        #print(self.position_folder_tuple)
        print('---- self.position_shape_tuple ----')
        #print(self.position_shape_tuple)
        print('---- self.position_line_tuple ----')
        #print(self.position_line_tuple)

        # GET Folder and wp name List
        folder_wp_name_array = ns_def.get_folder_wp_array_from_master(ws_name, excel_maseter_file)
        print('---- folder_wp_name_array ----')
        #print(folder_wp_name_array)

        '''
        Create L2 Table excel file
        '''

        master_l2_table_tuple = {}
        l2_table_array = []

        egt_maker_width_array = ['25', '25', '20', '25', '20', '30', '60', '60']
        l2_table_array.append([1, ['<RANGE>', '1', '1', '1', '1', '1', '1', '1', '1', '<END>']])
        l2_table_array.append([2, ['<HEADER>', 'Area', 'Device Name', 'Port Mode','Port Name', 'Virtual Port Mode', 'Virtual Port Name',  'Connected L2 Segment Name(Comma Separated)',  'L2 Name directly received by L3 Virtual Port (Comma Separated)', '<END>']])

        port_mode_excel_function = '=IF(@INDIRECT(ADDRESS(ROW(),COLUMN()+3))= "",IF(@INDIRECT(ADDRESS(ROW(),COLUMN()+4))= "",IF(@INDIRECT(ADDRESS(ROW(),COLUMN()+1))= "","","Routed (L3)"),IF(@INDIRECT(ADDRESS(ROW(),COLUMN()+1))= "","","Switch (L2)")),IF(@INDIRECT(ADDRESS(ROW(),COLUMN()+1))= "","","Switch (L2)"))'
        virtual_port_mode_excel_function = '=IF(@INDIRECT(ADDRESS(ROW(),COLUMN()+1))= "",IF(@INDIRECT(ADDRESS(ROW(),COLUMN()+1))= "","","Routed (L3)"),IF(@INDIRECT(ADDRESS(ROW(),COLUMN()+2))= "",IF(@INDIRECT(ADDRESS(ROW(),COLUMN()-1))= "","Loopback (L3)","Routed (L3)"),IF(@INDIRECT(ADDRESS(ROW(),COLUMN()-1))= "","Routed (L3)","Switch (L2)")))'

        start_row = 3
        start_column = 2
        egt_prefix = '>>' # for egt maker value.

        #sort folder name
        print('---- new_folder_wp_name_array ----')
        new_folder_wp_name_array = sorted(folder_wp_name_array[0], key=str.lower)
        new_folder_wp_name_array.extend(sorted(folder_wp_name_array[1], key=str.lower))
        #print(new_folder_wp_name_array)

        #Run each folder name
        for tmp_new_folder_wp_name_array in new_folder_wp_name_array:
            tmp_current_array = []
            Flag_first_folder = True
            tmp_sort_array = []

            # extend folder name
            tmp_current_array.extend([''])
            if '_wp_' not in tmp_new_folder_wp_name_array:
                tmp_current_array.extend([tmp_new_folder_wp_name_array])
            else:
                tmp_current_array.extend(['N/A'])
            #get shape name in the folder and sort
            shape_folder_tuple = ns_def.get_shape_folder_tuple(self.position_shape_tuple)
            print('---- shape_folder_tuple ----')
            #print(shape_folder_tuple)
            for tmp_shape_folder_tuple in shape_folder_tuple:
                #print(shape_folder_tuple[tmp_shape_folder_tuple] ,tmp_new_folder_wp_name_array )
                if shape_folder_tuple[tmp_shape_folder_tuple] == tmp_new_folder_wp_name_array:
                    tmp_sort_array.append(tmp_shape_folder_tuple)
            new_tmp_sort_array = sorted(tmp_sort_array, key=str.lower)
            print('---- new_tmp_sort_array ----')
            #print(new_tmp_sort_array)

            '''get tag of the shape'''
            flag_shape_first = True
            for tmp_new_tmp_sort_array in new_tmp_sort_array:
                for tmp_position_line_tuple in self.position_line_tuple:
                    if tmp_position_line_tuple[0] != 1 and tmp_position_line_tuple[0] != 2:
                        if tmp_position_line_tuple[1] == 1 and self.position_line_tuple[tmp_position_line_tuple] == tmp_new_tmp_sort_array:
                            if flag_shape_first == True and Flag_first_folder == True:
                                tmp_current_array.extend([self.position_line_tuple[tmp_position_line_tuple[0], 1]])
                                flag_shape_first = False
                                Flag_first_folder = False
                            elif flag_shape_first == True and Flag_first_folder == False:
                                tmp_current_array.extend(['', ''])
                                tmp_current_array.extend([self.position_line_tuple[tmp_position_line_tuple[0], 1]])
                                flag_shape_first = False
                            else:
                                tmp_current_array.extend(['','',''])

                            # change part of port number between port name and abbr
                            from_split_port_name = str(self.position_line_tuple[tmp_position_line_tuple[0], 3]).split(' ')
                            if ' ' in str(self.position_line_tuple[tmp_position_line_tuple[0], 3]):
                                from_port_full = self.position_line_tuple[tmp_position_line_tuple[0], 13] + ' ' + from_split_port_name[-1]
                                from_port_abbr = from_split_port_name[0]
                            else:
                                from_port_full = self.position_line_tuple[tmp_position_line_tuple[0],13]
                                from_port_abbr = self.position_line_tuple[tmp_position_line_tuple[0], 3]

                            tmp_current_array.extend([port_mode_excel_function,from_port_full ,virtual_port_mode_excel_function,'<EMPTY>','<EMPTY>','<EMPTY>','<END>'])
                            l2_table_array.append([start_row, tmp_current_array])
                            tmp_current_array = []
                            start_row += 1

                        if tmp_position_line_tuple[1] == 2 and self.position_line_tuple[tmp_position_line_tuple] == tmp_new_tmp_sort_array:
                            if flag_shape_first == True and Flag_first_folder == True:
                                tmp_current_array.extend([self.position_line_tuple[tmp_position_line_tuple[0], 2]])
                                flag_shape_first = False
                                Flag_first_folder = False
                            elif flag_shape_first == True and Flag_first_folder == False:
                                tmp_current_array.extend(['', ''])
                                tmp_current_array.extend([self.position_line_tuple[tmp_position_line_tuple[0], 2]])
                                flag_shape_first = False
                            else:
                                tmp_current_array.extend(['', '', ''])

                            # split part of port number between port name and abbr
                            from_split_port_name = str(self.position_line_tuple[tmp_position_line_tuple[0], 4]).split(' ')
                            if ' ' in str(self.position_line_tuple[tmp_position_line_tuple[0], 4]):
                                from_port_full = self.position_line_tuple[tmp_position_line_tuple[0], 17] + ' ' + from_split_port_name[-1]
                                from_port_abbr = from_split_port_name[0]
                            else:
                                from_port_full = self.position_line_tuple[tmp_position_line_tuple[0], 17]
                                from_port_abbr = self.position_line_tuple[tmp_position_line_tuple[0], 4]

                            tmp_current_array.extend([port_mode_excel_function,from_port_full ,virtual_port_mode_excel_function,'<EMPTY>','<EMPTY>','<EMPTY>','<END>'])
                            l2_table_array.append([start_row, tmp_current_array])
                            #print(tmp_current_array)
                            tmp_current_array = []
                            start_row += 1
                flag_shape_first = True

        l2_table_array.append([start_row, ['<END>']])

        print('---- l2_table_array ----')
        #print(l2_table_array)

        ''' sort Interafce number'''
        current_if_array = []
        tuple_num = 1
        overwrite_if_array = []
        for tmp_l2_table_array in l2_table_array:
            #print(tmp_l2_table_array[1])
            if str(tmp_l2_table_array[1][0]) == '':
                if str(tmp_l2_table_array[1][2]) != '':
                    if len(current_if_array) != 0:
                        # sort interface number
                        current_if_array = sorted(current_if_array, reverse=False, key=lambda x:(x[0][7],x[0][6]))  # sort for tmp_end_top

                        for tmp_current_if_array in current_if_array:
                            overwrite_if_array.append([tuple_num,tmp_current_if_array[:9]])
                            #print(tuple_num, tmp_current_if_array[:7])
                            tuple_num += 1

                    current_if_array = []
                    current_shape_name = tmp_l2_table_array[1][3]
                    if_value = ns_def.get_if_value(tmp_l2_table_array[1][4])
                    tmp_l2_table_array[1].extend([if_value])
                    tmp_l2_table_array[1].extend([str(ns_def.split_portname(tmp_l2_table_array[1][4])[0])])
                    current_if_array.append(tmp_l2_table_array[1][3:])
                else:
                    if_value = ns_def.get_if_value(tmp_l2_table_array[1][4])
                    tmp_l2_table_array[1].extend([if_value])
                    tmp_l2_table_array[1].extend([str(ns_def.split_portname(tmp_l2_table_array[1][4])[0])])
                    current_if_array.append(tmp_l2_table_array[1][3:])
                    last_append = current_if_array

        ### kyuusai last shape interfaces
        if len(current_if_array) >= 2:
            for tmp_current_if_array in current_if_array:
                last_append = sorted(last_append, reverse=False, key=lambda x:(x[0][7],x[0][6]))  # sort for tmp_end_top
                overwrite_if_array.append([tuple_num, tmp_current_if_array[:9]])
                tuple_num += 1

        print('---- overwrite_if_array ----')
        #print(overwrite_if_array)
        overwrite_if_tuple = ns_def.convert_array_to_tuple(overwrite_if_array)
        print('---- overwrite_if_tuple ----')
        #print(overwrite_if_tuple)

        ### Convert to tuple
        master_l2_table_tuple = ns_def.convert_array_to_tuple(l2_table_array)
        print('---- master_l2_table_tuple ----')
        #print(master_l2_table_tuple)

        # Create _tmp_ sheet
        ns_def.create_excel_sheet(excel_maseter_file, tmp_ws_name)

        # Write normal tuple to excel
        master_excel_meta = master_l2_table_tuple
        excel_file_path = excel_maseter_file
        worksheet_name = tmp_ws_name
        section_write_to = '<<N/A>>'
        offset_row = 0
        offset_column = 0
        ns_def.write_excel_meta(master_excel_meta, excel_file_path, worksheet_name, section_write_to, offset_row, offset_column)

        # Write overwrite tuple to excel
        master_excel_meta = overwrite_if_tuple
        excel_file_path = excel_maseter_file
        worksheet_name = tmp_ws_name
        section_write_to = '<<N/A>>'
        offset_row = 2
        offset_column = 3
        ns_def.overwrite_excel_meta(master_excel_meta, excel_file_path, worksheet_name, section_write_to, offset_row, offset_column)

        '''Create L2 Table file'''
        import ns_egt_maker #module import

        #create L2 Table file
        tmp_tmp_ws_name = '_tmp_tmp_'
        ns_def.create_excel_sheet(excel_maseter_file, tmp_tmp_ws_name)
        input_tree_tuple = {}
        input_tree_tuple[1,1] = 'Dummy' # first excel sheet has bug. That's why the first sheet is dummy sheet. this sheet will be deleted soon.
        input_tree_tuple[2, 1] = 'L2 Table'
        master_excel_meta = input_tree_tuple
        excel_file_path = excel_maseter_file
        worksheet_name = tmp_tmp_ws_name
        section_write_to = '<<N/A>>'
        offset_row = 0
        offset_column = 0
        ns_def.write_excel_meta(master_excel_meta, excel_file_path, worksheet_name, section_write_to, offset_row, offset_column)
        #Add L2 Table sheet
        input_excel_name = excel_maseter_file
        output_excel_name = self.inFileTxt_L2_1_1.get().replace('[MASTER]', '[L2_TABLE]')

        NEW_OR_ADD = 'NEW'
        ns_egt_maker.create_excel_gui_tree(input_excel_name,output_excel_name,NEW_OR_ADD, egt_maker_width_array)
        ns_def.remove_excel_sheet(output_excel_name, 'Dummy')

        #Add L2 Table from meta
        #print(output_excel_name)
        self.input_tree_excel = openpyxl.load_workbook(output_excel_name)
        worksheet_name = 'L2 Table'
        start_row = 1
        start_column = 0
        custom_table_name = excel_maseter_file
        self.input_tree_excel = ns_egt_maker.insert_custom_excel_table(self.input_tree_excel, worksheet_name, start_row, start_column, custom_table_name)
        self.input_tree_excel.save(output_excel_name)

        # Remove _tmp_ sheet from excel master
        ns_def.remove_excel_sheet(excel_maseter_file, tmp_ws_name)
        ns_def.remove_excel_sheet(excel_maseter_file, tmp_tmp_ws_name)

        '''
        ADD L2 MASTER DATA Sheet to Excel Master file
        '''
        l2_table_ws_name = 'L2 Table'
        l2_table_file = output_excel_name
        excel_master_ws_name = 'Master_Data'
        excel_master_ws_name_l2 = 'Master_Data_L2'

        # get L2 Table Excel file
        l2_table_array = []
        l2_table_array = ns_def.convert_excel_to_array(l2_table_ws_name, l2_table_file, 3)
        #print(l2_table_array)

        # make L2 Line array
        l2_master_data_array = []
        l2_master_data_array.append([1, ['<<L2_TABLE>>']])
        l2_master_data_array.append([2, ['Area', 'Device Name', 'Port Mode','Port Name', 'Virtual Port Mode', 'Virtual Port Name',  'Connected L2 Segment Name(Comma Separated)','L2 Name directly received by L3 Virtual Port (Comma Separated)']])
        tmp_row = 3
        current_device_name = 'dummy'
        current_area_name = 'dummy'
        for tmp_l2_table_array in l2_table_array:
            # set number of array = 6
            tmp_l2_table_array[1].extend(['', '', '', '', ''])
            del tmp_l2_table_array[1][7:]

            if tmp_l2_table_array[1][1] != '' and tmp_l2_table_array[1][1] != current_device_name:
                current_device_name = tmp_l2_table_array[1][1]

            if tmp_l2_table_array[1][0] != '' and tmp_l2_table_array[1][0] != current_area_name:
                current_area_name = tmp_l2_table_array[1][0]
            #l2_master_data_array.append([tmp_row , [current_area_name, current_device_name, tmp_l2_table_array[1][2], tmp_l2_table_array[1][3], tmp_l2_table_array[1][4], tmp_l2_table_array[1][5]]])
            l2_master_data_array.append([tmp_row , [current_area_name, current_device_name, '', tmp_l2_table_array[1][3], '', tmp_l2_table_array[1][5]]])
            tmp_row += 1
        print('--- l2_master_data_array ---')
        #print(l2_master_data_array)

        l2_master_data_tuple = {}
        l2_master_data_tuple = ns_def.convert_array_to_tuple(l2_master_data_array)

        # create Master_Data_L2 sheet
        ns_def.create_excel_sheet(excel_maseter_file, excel_master_ws_name_l2)
        print('create Master_Data_L2 sheet')

        # write l2_master_data_array tupple to Master data L2 sheet
        offset_row = 0
        offset_column = 0
        write_to_section = '_template_'
        ns_def.write_excel_meta(l2_master_data_tuple, excel_maseter_file, excel_master_ws_name_l2, write_to_section, offset_row, offset_column)


class ns_l2_table_from_master_l2_sheet():
    def __init__(self):
        '''
        make L2 Table sheet from master l2 sheet
        '''
        print('---ns_l2_table_from_master_l2()---')

        tmp_ws_name = '_tmp_'
        excel_maseter_file = self.inFileTxt_L2_1_1.get()
        write_excel_file = self.outFileTxt_11_2.get().replace('[MASTER]','')
        excel_master_ws_name_l2 = 'Master_Data_L2'

        master_l2_array = []
        master_l2_array = ns_def.convert_master_to_array('Master_Data_L2', excel_maseter_file, '<<L2_TABLE>>')
        #print(master_l2_array)

        l2_table_array = []

        egt_maker_width_array = ['25', '25', '20', '25', '20', '30', '60', '60']
        l2_table_array.append([1, ['<RANGE>', '1', '1', '1', '1', '1', '1', '1', '1', '<END>']])
        l2_table_array.append([2, ['<HEADER>', 'Area', 'Device Name', 'Port Mode','Port Name', 'Virtual Port Mode', 'Virtual Port Name',  'Connected L2 Segment Name(Comma Separated)',  'L2 Name directly received by L3 Virtual Port (Comma Separated)', '<END>']])

        port_mode_excel_function = '=IF(@INDIRECT(ADDRESS(ROW(),COLUMN()+3))= "",IF(@INDIRECT(ADDRESS(ROW(),COLUMN()+4))= "",IF(@INDIRECT(ADDRESS(ROW(),COLUMN()+1))= "","","Routed (L3)"),IF(@INDIRECT(ADDRESS(ROW(),COLUMN()+1))= "","","Switch (L2)")),IF(@INDIRECT(ADDRESS(ROW(),COLUMN()+1))= "","","Switch (L2)"))'
        virtual_port_mode_excel_function = '=IF(@INDIRECT(ADDRESS(ROW(),COLUMN()+1))= "",IF(@INDIRECT(ADDRESS(ROW(),COLUMN()+1))= "","","Routed (L3)"),IF(@INDIRECT(ADDRESS(ROW(),COLUMN()+2))= "",IF(@INDIRECT(ADDRESS(ROW(),COLUMN()-1))= "","Loopback (L3)","Routed (L3)"),IF(@INDIRECT(ADDRESS(ROW(),COLUMN()-1))= "","Routed (L3)","Switch (L2)")))'

        max_row_num = 0
        for tmp_master_l2_array in master_l2_array:
            if tmp_master_l2_array[0] != 1 and tmp_master_l2_array[0] != 2:

                tmp_master_l2_array[1].extend(['<EMPTY>','<EMPTY>','<EMPTY>','<EMPTY>','<EMPTY>'])
                tmp_master_l2_array[1].insert(0, '')
                del tmp_master_l2_array[1][9:]
                tmp_master_l2_array[1].append('<END>')
                max_row_num = tmp_master_l2_array[0]

                #print(tmp_master_l2_array)
                tmp_master_l2_array[1][3] = port_mode_excel_function
                tmp_master_l2_array[1][5] = virtual_port_mode_excel_function

                #insert '>>' or <EMPTY> to column when greater than 7
                for i in range(6,9):
                    if tmp_master_l2_array[1][i] != '<EMPTY>' and tmp_master_l2_array[1][i] !=  '':
                        tmp_master_l2_array[1][i] = '>>' + str(tmp_master_l2_array[1][i])
                    elif tmp_master_l2_array[1][i] == '':
                        tmp_master_l2_array[1][i] = '<EMPTY>'

                l2_table_array.append(tmp_master_l2_array)

        l2_table_array.append([max_row_num + 1 ,['<END>']])

        print('--- l2_table_array ---')
        #print(l2_table_array)

        # remove duplicated value in row 1 or 2
        new_l2_table_array = []
        tmp_current_value = 'dummy'
        for tmp_l2_table_array in l2_table_array:
            if tmp_l2_table_array[0] != 1 and tmp_l2_table_array[0] != 2 and tmp_l2_table_array[1][0] != '<END>':
                if tmp_l2_table_array[1][2] == tmp_current_value:
                    tmp_l2_table_array[1][2] = ''
                else:
                    tmp_current_value = tmp_l2_table_array[1][2]
                new_l2_table_array.append(tmp_l2_table_array)
            else:
                new_l2_table_array.append(tmp_l2_table_array)
                if tmp_l2_table_array[1][0] != '<END>':
                    tmp_current_value = tmp_l2_table_array[1][2]

        tmp_current_value = '_dummy_'
        new_new_l2_table_array = []
        for tmp_l2_table_array in new_l2_table_array:
            if tmp_l2_table_array[0] != 1 and tmp_l2_table_array[0] != 2 and tmp_l2_table_array[1][0] != '<END>':
                if tmp_l2_table_array[1][1] == tmp_current_value and tmp_l2_table_array[1][1] != 'N/A':
                    tmp_l2_table_array[1][1] = ''
                elif tmp_l2_table_array[1][1] == 'N/A' and tmp_l2_table_array[1][2] == '':
                    tmp_l2_table_array[1][1] = ''
                else:
                    tmp_current_value = tmp_l2_table_array[1][1]
                new_new_l2_table_array.append(tmp_l2_table_array)
            else:
                new_new_l2_table_array.append(tmp_l2_table_array)
                if tmp_l2_table_array[1][0] != '<END>':
                    tmp_current_value = tmp_l2_table_array[1][1]

        # Add 'N/A' empty value in row 4
        new_new_new_l2_table_array = []
        for tmp_l2_table_array in new_new_l2_table_array:
            if tmp_l2_table_array[0] != 1 and tmp_l2_table_array[0] != 2 and tmp_l2_table_array[1][0] != '<END>':
                if tmp_l2_table_array[1][4] == '':
                    tmp_l2_table_array[1][4] = '_EMPTY_'

            new_new_new_l2_table_array.append(tmp_l2_table_array)

        print('--- new_new_new_l2_table_array ---')
        #print(new_new_new_l2_table_array)

        l2_table_tuple = {}
        l2_table_tuple = ns_def.convert_array_to_tuple(new_new_new_l2_table_array)

        '''Create _tmp_ sheet for l2_table_tuple'''
        ns_def.create_excel_sheet(write_excel_file, tmp_ws_name)

        # Write normal tuple to excel
        master_excel_meta = l2_table_tuple
        excel_file_path = write_excel_file
        worksheet_name = tmp_ws_name
        section_write_to = '<<N/A>>'
        offset_row = 0
        offset_column = 0
        print('---master_excel_meta---')
        #print(master_excel_meta)
        ns_def.write_excel_meta(master_excel_meta, excel_file_path, worksheet_name, section_write_to, offset_row, offset_column)

        '''Create L2 Table file'''
        import ns_egt_maker  # module import

        # create L2 Table file
        tmp_tmp_ws_name = '_tmp_tmp_'
        ns_def.create_excel_sheet(write_excel_file, tmp_tmp_ws_name)
        input_tree_tuple = {}
        input_tree_tuple[1, 1] = 'Dummy'  # first excel sheet has bug. That's why the first sheet is dummy sheet. this sheet will be deleted soon.
        input_tree_tuple[2, 1] = 'L2 Table'
        master_excel_meta = input_tree_tuple
        excel_file_path = write_excel_file
        worksheet_name = tmp_tmp_ws_name
        section_write_to = '<<N/A>>'
        offset_row = 0
        offset_column = 0
        ns_def.write_excel_meta(master_excel_meta, excel_file_path, worksheet_name, section_write_to, offset_row, offset_column)
        # Add L2 Table sheet
        input_excel_name = write_excel_file
        output_excel_name = write_excel_file

        NEW_OR_ADD = 'ADD'
        ns_egt_maker.create_excel_gui_tree(input_excel_name, output_excel_name, NEW_OR_ADD, egt_maker_width_array)
        ns_def.remove_excel_sheet(output_excel_name, 'Dummy')

        # Add L2 Table from meta
        # print(output_excel_name)
        self.input_tree_excel = openpyxl.load_workbook(output_excel_name)
        worksheet_name = 'L2 Table'
        start_row = 1
        start_column = 0
        custom_table_name = write_excel_file
        self.input_tree_excel = ns_egt_maker.insert_custom_excel_table(self.input_tree_excel, worksheet_name, start_row, start_column, custom_table_name)
        self.input_tree_excel.save(output_excel_name)

        # Remove _tmp_ sheet from excel master
        ns_def.remove_excel_sheet(write_excel_file, tmp_ws_name)
        ns_def.remove_excel_sheet(write_excel_file, tmp_tmp_ws_name)

        '''
        additional change_string_color
        '''
        self.input_tree_excel = openpyxl.load_workbook(output_excel_name)
        ws_name = worksheet_name
        current_device_name = ''
        pre_if_name = ''
        for tmp_master_l2_array in master_l2_array:
            if tmp_master_l2_array[0] != 1 and tmp_master_l2_array[0] != 2 and tmp_master_l2_array[1][4] != '' and tmp_master_l2_array[1][4] != '_EMPTY_':
                if current_device_name == tmp_master_l2_array[1][2]:
                    if pre_if_name == tmp_master_l2_array[1][4]:
                        #print(current_device_name,pre_if_name,tmp_master_l2_array[1][3])
                        ns_egt_maker.change_string_color(self.input_tree_excel, ws_name, tmp_master_l2_array[0], 3, 'GRAY')
                        ns_egt_maker.change_string_color(self.input_tree_excel, ws_name, tmp_master_l2_array[0], 4, 'GRAY')
                        ns_egt_maker.change_string_color(self.input_tree_excel, ws_name, tmp_master_l2_array[0], 5, 'GRAY')
                        pre_if_name = tmp_master_l2_array[1][4]
                    else:
                        pre_if_name = tmp_master_l2_array[1][4]
                else:
                    current_device_name = tmp_master_l2_array[1][2]
                    pre_if_name = ''

        self.input_tree_excel.save(output_excel_name)

