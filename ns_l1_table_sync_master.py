'''
SPDX-License-Identifier: Apache-2.0

Copyright 2023 Cisco Systems, Inc. and its affiliates

Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.
'''

from pptx import *
import sys, os, re
import numpy as np
import math
import openpyxl
import tkinter as tk ,tkinter.ttk as ttk,tkinter.filedialog, tkinter.messagebox
import ns_def ,ns_ddx_figure
import shutil

class  ns_l1_table_sync_master():
    def __init__(self):
        '''
        update Excel master data file from device table excel file
        '''
        # parameter
        ws_name = 'Master_Data'
        device_table_ws_name = 'L1 Table'
        device_table_file = self.inFileTxt_12_1.get()
        device_table_file_backup = self.inFileTxt_12_1_2.get()
        excel_master_ws_name = 'Master_Data'
        excel_maseter_file = self.inFileTxt_12_2.get()
        excel_maseter_file_backup = self.inFileTxt_12_2_2.get()

        # convert from master to array and convert to tuple
        self.position_line_array = ns_def.convert_master_to_array(excel_master_ws_name, excel_maseter_file, '<<POSITION_LINE>>')
        self.position_line_tuple = ns_def.convert_array_to_tuple(self.position_line_array)

        print('---- self.position_line_tuple ----')
        #print(self.position_line_tuple)


        # GET Folder and wp name List
        folder_wp_name_array = ns_def.get_folder_wp_array_from_master(ws_name, excel_maseter_file)
        print('---- folder_wp_name_array ----')
        #print(folder_wp_name_array)

        '''GET start row and column in Device table'''
        start_row = 1
        start_column = 1

        excel_device_table = openpyxl.load_workbook(device_table_file)
        excel_device_table.active = excel_device_table.sheetnames.index(device_table_ws_name)

        flag_get_start = False
        for tmp_column in range(1, 100):
            for tmp_row in range(1, 100):
                #print(str(excel_device_table.active.cell(tmp_row, tmp_column).fill.bgColor.value).lower())
                if '4f81bd' in str(excel_device_table.active.cell(tmp_row, tmp_column).fill.bgColor.value).lower() and 'dce6f1' in str(excel_device_table.active.cell(tmp_row +1, tmp_column).fill.bgColor.value).lower() :
                    start_folder_array = tmp_row +1,tmp_column,excel_device_table.active.cell(tmp_row + 1, tmp_column).value
                    for tmp_folder_wp_name_array in folder_wp_name_array[0]:
                        if tmp_folder_wp_name_array == start_folder_array[2]:
                            flag_get_start = True
                            break

            if flag_get_start == True:
                break
        #print('start_folder_array')
        #print(start_folder_array)

        # convert from device table to array and convert to tuple
        excel_ws_name = 'L1 Table'
        excel_file = device_table_file
        start_row = start_folder_array[0]
        self.device_table_array = ns_def.convert_excel_to_array(excel_ws_name, excel_file, start_row)
        self.device_table_tuple = ns_def.convert_array_to_tuple(self.position_line_array)

        print('---- self.device_table_array ----')
        #print(self.device_table_array)

        ### convert to shape,interface array
        current_if_array = []
        tuple_num = 1
        new_devicetable_array = []
        for tmp_device_table_array in self.device_table_array:
            if str(tmp_device_table_array[1][1]) != '':
                if len(current_if_array) != 0:
                    for tmp_current_if_array in current_if_array:
                        new_devicetable_array.append([current_shape_name,tmp_current_if_array[:12]])
                        tuple_num += 1
                    #print(current_shape_name, current_if_array)

                current_if_array = []
                current_shape_name = tmp_device_table_array[1][1]
                if_value = ns_def.get_if_value(tmp_device_table_array[1][3])
                tmp_device_table_array[1].extend([if_value])
                current_if_array.append(tmp_device_table_array[1][2:12])
                last_append = current_if_array
            else:
                if_value = ns_def.get_if_value(tmp_device_table_array[1][3])
                tmp_device_table_array[1].extend([if_value])
                current_if_array.append(tmp_device_table_array[1][2:12])
                last_append = current_if_array

        for tmp_last_append in last_append:
            new_devicetable_array.append([current_shape_name,tmp_last_append ])
            #print('last_append  ' + str(current_shape_name) + '   ' + str(tmp_last_append) )


        print('---- new_devicetable_array ----')
        #print(new_devicetable_array)

        '''
        create overwrite port number
        '''
        ### update excel master file ###
        self.update_port_num_array = []
        overwrite_line_tuple = {}
        done_line_from_to_array = []
        port_no_value_num = 10001
        for tmp_new_devicetable_array in new_devicetable_array:
            from_shape_name = tmp_new_devicetable_array[0]
            to_shape_name = tmp_new_devicetable_array[1][7]

            # get device portname with excel abbr format
            return_array            = ns_def.split_portname(tmp_new_devicetable_array[1][0])
            device_from_portname = return_array[0]
            return_array = ns_def.split_portname(tmp_new_devicetable_array[1][6])
            device_from_portnum_abbr = return_array[1]
            device_from_portfullname_abbr = tmp_new_devicetable_array[1][6]
            return_array            = ns_def.split_portname(tmp_new_devicetable_array[1][8])
            device_to_portnum_abbr = return_array[1]
            device_to_portfullname_abbr = tmp_new_devicetable_array[1][8]


            #get device_to_portname_abbr from device table
            flag_match_to_port = False
            for tmp_tmp_new_devicetable_array in new_devicetable_array:
                if tmp_tmp_new_devicetable_array[1][5] == from_shape_name and tmp_tmp_new_devicetable_array[1][6]  == device_from_portfullname_abbr and \
                    tmp_tmp_new_devicetable_array[1][7] == to_shape_name and tmp_tmp_new_devicetable_array[1][8]  == device_to_portfullname_abbr:
                        device_to_portfullname_abbr =tmp_new_devicetable_array[1][8]
                        flag_match_to_port = True
                        break

            #print(device_from_portname_abbr,device_to_portname_abbr)
            #print(from_shape_name,to_shape_name)

            for tmp_position_line_tuple in self.position_line_tuple:
                if tmp_position_line_tuple[0] != 1 and tmp_position_line_tuple[0] != 2 and [tmp_position_line_tuple[0],tmp_position_line_tuple[1]] not in done_line_from_to_array \
                        and (tmp_position_line_tuple[1] == 1 or tmp_position_line_tuple[1] == 2 ):

                    #append updated port name array for ver 2.0
                    original_port_name = device_from_portfullname_abbr
                    updated_port_name = tmp_new_devicetable_array[1][1] +' '+ str(ns_def.split_portname(tmp_new_devicetable_array[1][0])[1])
                    if original_port_name != updated_port_name:
                        target_device_name = from_shape_name
                        tag_if_name = original_port_name
                        position_line_tuple = self.position_line_tuple
                        full_original_port_name =  ns_def.get_full_name_from_tag_name(target_device_name, tag_if_name, position_line_tuple)
                        self.update_port_num_array.append([from_shape_name,full_original_port_name,tmp_new_devicetable_array[1][0]])


                    #### port number update ####
                    if flag_match_to_port == True:
                        # if device_to_portname_abbr match to others
                        # excel masetr line, from side
                        if self.position_line_tuple[tmp_position_line_tuple] == from_shape_name and self.position_line_tuple[tmp_position_line_tuple[0],2] == to_shape_name \
                                and self.position_line_tuple[tmp_position_line_tuple[0],3] == device_from_portfullname_abbr and self.position_line_tuple[tmp_position_line_tuple[0],4] == device_to_portfullname_abbr:

                            ### overwrite port settings to tuple
                            overwrite_line_tuple[tmp_position_line_tuple[0], 13] = device_from_portname.replace(' ','')
                            overwrite_line_tuple[tmp_position_line_tuple[0], 14] = tmp_new_devicetable_array[1][2]
                            overwrite_line_tuple[tmp_position_line_tuple[0], 15] = tmp_new_devicetable_array[1][3]
                            overwrite_line_tuple[tmp_position_line_tuple[0], 16] = tmp_new_devicetable_array[1][4]

                            ### overwrite port num to tuple
                            overwrite_line_tuple[tmp_position_line_tuple[0],3] = tmp_new_devicetable_array[1][1] +' '+ str(ns_def.split_portname(tmp_new_devicetable_array[1][0])[1])
                            done_line_from_to_array.append([tmp_position_line_tuple[0],tmp_position_line_tuple[1]])
                            break


                        # excel masetr line, to side
                        elif self.position_line_tuple[tmp_position_line_tuple] == from_shape_name and self.position_line_tuple[tmp_position_line_tuple[0],1] == to_shape_name \
                                and self.position_line_tuple[tmp_position_line_tuple[0],4] == device_from_portfullname_abbr and self.position_line_tuple[tmp_position_line_tuple[0],3] == device_to_portfullname_abbr:

                            ### overwrite port settings to tuple
                            overwrite_line_tuple[tmp_position_line_tuple[0], 17] = device_from_portname.replace(' ','')
                            overwrite_line_tuple[tmp_position_line_tuple[0], 18] = tmp_new_devicetable_array[1][2]
                            overwrite_line_tuple[tmp_position_line_tuple[0], 19] = tmp_new_devicetable_array[1][3]
                            overwrite_line_tuple[tmp_position_line_tuple[0], 20] = tmp_new_devicetable_array[1][4]

                            ### overwrite port num to tuple
                            overwrite_line_tuple[tmp_position_line_tuple[0], 4] = tmp_new_devicetable_array[1][1] + ' ' + str(ns_def.split_portname(tmp_new_devicetable_array[1][0])[1])
                            done_line_from_to_array.append([tmp_position_line_tuple[0], tmp_position_line_tuple[1]])
                            break

                    elif flag_match_to_port == False:
                        #if device_to_portname_abbr no match to others
                        # excel masetr line, from side
                        if self.position_line_tuple[tmp_position_line_tuple] == from_shape_name and self.position_line_tuple[tmp_position_line_tuple[0], 2] == to_shape_name:

                            ### overwrite port settings to tuple
                            overwrite_line_tuple[tmp_position_line_tuple[0], 13] = device_from_portname.replace(' ','')
                            overwrite_line_tuple[tmp_position_line_tuple[0], 14] = tmp_new_devicetable_array[1][2]
                            overwrite_line_tuple[tmp_position_line_tuple[0], 15] = tmp_new_devicetable_array[1][3]
                            overwrite_line_tuple[tmp_position_line_tuple[0], 16] = tmp_new_devicetable_array[1][4]

                            ### overwrite port num to tuple
                            if device_from_portfullname_abbr != str(self.position_line_tuple[tmp_position_line_tuple[0], 3]):
                                overwrite_line_tuple[tmp_position_line_tuple[0], 3] = tmp_new_devicetable_array[1][1] + ' ' + str(port_no_value_num)
                                done_line_from_to_array.append([tmp_position_line_tuple[0], tmp_position_line_tuple[1]])
                                port_no_value_num += 1
                                break

                        # excel masetr line, to side
                        elif self.position_line_tuple[tmp_position_line_tuple] == from_shape_name and self.position_line_tuple[tmp_position_line_tuple[0], 1] == to_shape_name:

                            ### overwrite port settings to tuple
                            overwrite_line_tuple[tmp_position_line_tuple[0], 17] = device_from_portname.replace(' ','')
                            overwrite_line_tuple[tmp_position_line_tuple[0], 18] = tmp_new_devicetable_array[1][2]
                            overwrite_line_tuple[tmp_position_line_tuple[0], 19] = tmp_new_devicetable_array[1][3]
                            overwrite_line_tuple[tmp_position_line_tuple[0], 20] = tmp_new_devicetable_array[1][4]

                            ### overwrite port num to tuple
                            if device_from_portfullname_abbr != str(self.position_line_tuple[tmp_position_line_tuple[0], 4]):
                                # print(str(tmp_new_devicetable_array[1][1] + ' ' + device_from_portnum_abbr), str(self.position_line_tuple[tmp_position_line_tuple[0], 4]))
                                overwrite_line_tuple[tmp_position_line_tuple[0], 4] = tmp_new_devicetable_array[1][1] + ' ' + device_from_portnum_abbr
                                done_line_from_to_array.append([tmp_position_line_tuple[0], tmp_position_line_tuple[1]])
                                break

        #print(done_line_from_to_array)
        print('---- overwrite_line_tuple ----')
        #print(overwrite_line_tuple)

        ### Sync file copy ###
        '''if os.path.isfile(excel_maseter_file_backup) == True:
            os.remove(excel_maseter_file_backup)
            shutil.copy(excel_maseter_file, excel_maseter_file_backup)
        else:
            shutil.copy(excel_maseter_file, excel_maseter_file_backup)'''
        ### overwerite overwrite_line_tuple to excel master data
        master_excel_meta = overwrite_line_tuple
        excel_file_path = excel_maseter_file
        worksheet_name = excel_master_ws_name
        section_write_to = '<<POSITION_LINE>>'
        offset_row = 0
        offset_column = 0
        ns_def.overwrite_excel_meta(master_excel_meta, excel_file_path, worksheet_name, section_write_to, offset_row, offset_column)


        #print(done_line_from_to_array)
        print('---- overwrite_line_tuple ----')
        #print(overwrite_line_tuple)

        ### overwerite overwrite_line_tuple to excel master data
        master_excel_meta = overwrite_line_tuple
        excel_file_path = excel_maseter_file
        worksheet_name = excel_master_ws_name
        section_write_to = '<<POSITION_LINE>>'
        offset_row = 0
        offset_column = 0
        ns_def.overwrite_excel_meta(master_excel_meta, excel_file_path, worksheet_name, section_write_to, offset_row, offset_column)

        self.update_port_num_array = ns_def.get_l2_broadcast_domains.get_unique_list(self.update_port_num_array)  # for ver 2.0

        if flag_get_start == False:
            tkinter.messagebox.showerror('Error', 'Could not find the L1 Table')
            return(False)

        return (True)
