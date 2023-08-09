'''
SPDX-License-Identifier: Apache-2.0

Copyright 2023 Cisco Systems, Inc. and its affiliates

Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.
'''

import sys, os
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill
import tkinter as tk ,tkinter.ttk as ttk, tkinter.filedialog, tkinter.messagebox
from tkinter import *
import subprocess

'''
 common variables
'''
##### Default Font #####
defalut_font = Font(name="Futura", size=11, color='FF000000')  # default font
table_header_font = Font(name="Futura", bold=True, size=11, color='FFFFFFFF')

##### Font and style Settings for each bullet #####
first_paragraph_font = Font(name="Futura", bold=True, size=16, color='FF1f497d', )
first_paragraph_height = 20.25
first_paragraph_bullet_front = '['
first_paragraph_bullet_back = ']'

second_paragraph_font = Font(name="Futura", bold=True, underline='single', size=16, color='FF1f497d', )
second_paragraph_height = 20.25
second_paragraph_bullet_front = '⁃'
second_paragraph_bullet_back = ''

third_paragraph_font = Font(name="Futura", bold=True, underline='single', size=16, color='FF000000', )
third_paragraph_height = 20.25
third_paragraph_bullet_front = '•'
third_paragraph_bullet_back = ''

fourth_paragraph_font = Font(name="Futura", underline='single', size=16, color='FF000000', )
fourth_paragraph_height = 20.25
fourth_paragraph_bullet_front = '◦'
fourth_paragraph_bullet_back = ''

fifth_paragraph_font = Font(name="Futura", underline='single', size=16, color='FF000000', )
fifth_paragraph_height = 20.25
fifth_paragraph_bullet_front = '‣'
fifth_paragraph_bullet_back = ''

num_to_font = {1: first_paragraph_font, 2: second_paragraph_font, 3: third_paragraph_font, 4: fourth_paragraph_font, 5: fifth_paragraph_font}
num_to_height = {1: first_paragraph_height, 2: second_paragraph_height, 3: third_paragraph_height, 4: fourth_paragraph_height, 5: fifth_paragraph_height}
num_to_bullet_front = {1: first_paragraph_bullet_front, 2: second_paragraph_bullet_front, 3: third_paragraph_bullet_front, 4: fourth_paragraph_bullet_front, 5: fifth_paragraph_bullet_front}
num_to_bullet_back = {1: first_paragraph_bullet_back, 2: second_paragraph_bullet_back, 3: third_paragraph_bullet_back, 4: fourth_paragraph_bullet_back, 5: fifth_paragraph_bullet_back}

##### Enable or Disable mark #####
Enable_mark = '🗹'
Disable_mark = '☐'

##### type of bullet depend on usage type #####
bullet_direct_front = '>>'
bullet_direct_back = ''
bullet_select_front = '>'
bullet_select_back = ''

##### Excel Border Settings #####
white_border = Border(left=Side(border_style="thin", color="FFFFFF"),
                      right=Side(border_style="thin", color="FFFFFF"),
                      top=Side(border_style="thin", color="FFFFFF"),
                      bottom=Side(border_style="thin", color="FFFFFF"))
blue_border_start = Border(left=Side(border_style="thin", color="000000"),
                           right=Side(border_style="thin", color="000000"),   # update for Network Sketcher Ver 2.0
                           top=Side(border_style="thin", color="000000"),
                           bottom=Side(border_style="thin", color="000000"))
blue_border_medium = Border(left=Side(border_style="thin", color="4f81bd"),
                            right=Side(border_style="thin", color="4f81bd"),
                            top=Side(border_style="thin", color="000000"),
                            bottom=Side(border_style="thin", color="000000"))
blue_border_end = Border(left=Side(border_style="thin", color="4f81bd"),
                         right=Side(border_style="thin", color="000000"),
                         top=Side(border_style="thin", color="000000"),
                         bottom=Side(border_style="thin", color="000000"))
gray_blue_border_start = Border(left=Side(border_style="thin", color="000000"),
                           right=Side(border_style="thin", color="000000"),         #Change Value for Network Sketcher ver 2.0
                           top=Side(border_style="thin", color="000000"),
                           bottom=Side(border_style="thin", color="000000"))
gray_blue_border_medium = Border(left=Side(border_style="thin", color="dce6f1"),
                            right=Side(border_style="thin", color="dce6f1"),
                            top=Side(border_style="thin", color="000000"),
                            bottom=Side(border_style="thin", color="000000"))
gray_blue_border_end = Border(left=Side(border_style="thin", color="dce6f1"),
                         right=Side(border_style="thin", color="000000"),
                         top=Side(border_style="thin", color="000000"),
                         bottom=Side(border_style="thin", color="000000"))
gray_blue_border_start_updown_same = Border(left=Side(border_style="thin", color="000000"),
                           right=Side(border_style="thin", color="dce6f1"),
                           top=Side(border_style="thin", color="dce6f1"),
                           bottom=Side(border_style="thin", color="dce6f1"))
gray_blue_border_medium_updown_same = Border(left=Side(border_style="thin", color="dce6f1"),
                            right=Side(border_style="thin", color="dce6f1"),
                            top=Side(border_style="thin", color="dce6f1"),
                            bottom=Side(border_style="thin", color="dce6f1"))
gray_blue_border_end_updown_same = Border(left=Side(border_style="thin", color="dce6f1"),
                         right=Side(border_style="thin", color="000000"),
                         top=Side(border_style="thin", color="dce6f1"),
                         bottom=Side(border_style="thin", color="dce6f1"))
gray_blue_border_start_bottom_same = Border(left=Side(border_style="thin", color="000000"),
                           right=Side(border_style="thin", color="dce6f1"),
                           top=Side(border_style="thin", color="000000"),
                           bottom=Side(border_style="thin", color="dce6f1"))
gray_blue_border_medium_bottom_same = Border(left=Side(border_style="thin", color="dce6f1"),
                            right=Side(border_style="thin", color="dce6f1"),
                            top=Side(border_style="thin", color="000000"),
                            bottom=Side(border_style="thin", color="dce6f1"))
gray_blue_border_end_bottom_same = Border(left=Side(border_style="thin", color="dce6f1"),
                         right=Side(border_style="thin", color="000000"),
                         top=Side(border_style="thin", color="000000"),
                         bottom=Side(border_style="thin", color="dce6f1"))
gray_blue_border_start_top_same = Border(left=Side(border_style="thin", color="000000"),
                           right=Side(border_style="thin", color="dce6f1"),
                           top=Side(border_style="thin", color="dce6f1"),
                           bottom=Side(border_style="thin", color="000000"))
gray_blue_border_medium_top_same = Border(left=Side(border_style="thin", color="dce6f1"),
                            right=Side(border_style="thin", color="dce6f1"),
                            top=Side(border_style="thin", color="dce6f1"),
                            bottom=Side(border_style="thin", color="000000"))
gray_blue_border_end_top_same = Border(left=Side(border_style="thin", color="dce6f1"),
                         right=Side(border_style="thin", color="000000"),
                         top=Side(border_style="thin", color="dce6f1"),
                         bottom=Side(border_style="thin", color="000000"))
black_border_start = Border(left=Side(border_style="thin", color="000000"),
                            right=Side(border_style="thin", color="000000"),  # change valule from FFFFFF for Network Sketcher Ver2.0
                            top=Side(border_style="thin", color="000000"),
                            bottom=Side(border_style="thin", color="000000"))
black_border_medium = Border(left=Side(border_style="thin", color="FFFFFF"),
                             right=Side(border_style="thin", color="FFFFFF"),
                             top=Side(border_style="thin", color="000000"),
                             bottom=Side(border_style="thin", color="000000"))
black_border_end = Border(left=Side(border_style="thin", color="FFFFFF"),
                          right=Side(border_style="thin", color="000000"),
                          top=Side(border_style="thin", color="000000"),
                          bottom=Side(border_style="thin", color="000000"))
black_border_start_updown_same = Border(left=Side(border_style="thin", color="000000"),
                            right=Side(border_style="thin", color="FFFFFF"),
                            top=Side(border_style="thin", color="FFFFFF"),
                            bottom=Side(border_style="thin", color="FFFFFF"))
black_border_medium_updown_same = Border(left=Side(border_style="thin", color="FFFFFF"),
                             right=Side(border_style="thin", color="FFFFFF"),
                             top=Side(border_style="thin", color="FFFFFF"),
                             bottom=Side(border_style="thin", color="FFFFFF"))
black_border_end_updown_same = Border(left=Side(border_style="thin", color="FFFFFF"),
                          right=Side(border_style="thin", color="000000"),
                          top=Side(border_style="thin", color="FFFFFF"),
                          bottom=Side(border_style="thin", color="FFFFFF"))
black_border_start_bottom_same = Border(left=Side(border_style="thin", color="000000"),
                            right=Side(border_style="thin", color="FFFFFF"),
                            top=Side(border_style="thin", color="000000"),
                            bottom=Side(border_style="thin", color="FFFFFF"))
black_border_medium_bottom_same = Border(left=Side(border_style="thin", color="FFFFFF"),
                             right=Side(border_style="thin", color="FFFFFF"),
                             top=Side(border_style="thin", color="000000"),
                             bottom=Side(border_style="thin", color="FFFFFF"))
black_border_end_bottom_same = Border(left=Side(border_style="thin", color="FFFFFF"),
                          right=Side(border_style="thin", color="000000"),
                          top=Side(border_style="thin", color="000000"),
                          bottom=Side(border_style="thin", color="FFFFFF"))
black_border_start_top_same = Border(left=Side(border_style="thin", color="000000"),
                            right=Side(border_style="thin", color="000000"),            # change valule from FFFFFF for Network Sketcher Ver2.0
                            top=Side(border_style="thin", color="FFFFFF"),
                            bottom=Side(border_style="thin", color="000000"))
black_border_medium_top_same = Border(left=Side(border_style="thin", color="FFFFFF"),
                             right=Side(border_style="thin", color="FFFFFF"),
                             top=Side(border_style="thin", color="FFFFFF"),
                             bottom=Side(border_style="thin", color="000000"))
black_border_end_top_same = Border(left=Side(border_style="thin", color="FFFFFF"),
                          right=Side(border_style="thin", color="000000"),
                          top=Side(border_style="thin", color="FFFFFF"),
                          bottom=Side(border_style="thin", color="000000"))
TOP_WHITE_border_start = Border(left=Side(border_style="thin", color="000000"),
                                right=Side(border_style="thin", color="FFFFFF"),
                                top=Side(border_style="thin", color="FFFFFF"),
                                # bottom=Side(border_style="thin", color="000000") #please keep '#'
                                )
TOP_WHITE_border_medium = Border(left=Side(border_style="thin", color="FFFFFF"),
                                 right=Side(border_style="thin", color="FFFFFF"),
                                 top=Side(border_style="thin", color="FFFFFF"),
                                 # bottom=Side(border_style="thin", color="000000") #please keep '#'
                                 )
TOP_WHITE_border_end = Border(left=Side(border_style="thin", color="FFFFFF"),
                              right=Side(border_style="thin", color="000000"),
                              top=Side(border_style="thin", color="FFFFFF"),
                              # bottom=Side(border_style="thin", color="000000") #please keep '#'
                              )
BOTTOM_WHITE_border_start = Border(left=Side(border_style="thin", color="000000"),
                                   right=Side(border_style="thin", color="FFFFFF"),
                                   # top=Side(border_style="thin", color="000000"),
                                   bottom=Side(border_style="thin", color="FFFFFF"))
BOTTOM_WHITE_border_medium = Border(left=Side(border_style="thin", color="FFFFFF"),
                                    right=Side(border_style="thin", color="FFFFFF"),
                                    # top=Side(border_style="thin", color="000000"), #please keep '#'
                                    bottom=Side(border_style="thin", color="FFFFFF"))
BOTTOM_WHITE_border_end = Border(left=Side(border_style="thin", color="FFFFFF"),
                                 right=Side(border_style="thin", color="000000"),
                                 # top=Side(border_style="thin", color="000000"),  #please keep '#'
                                 bottom=Side(border_style="thin", color="FFFFFF"))
BOTTOM_LAST_border_start = Border(left=Side(border_style="thin", color="000000"),
                                  right=Side(border_style="thin", color="FFFFFF"),
                                  top=Side(border_style="thin", color="FFFFFF"),
                                  bottom=Side(border_style="thin", color="000000"))
BOTTOM_LAST_border_medium = Border(left=Side(border_style="thin", color="FFFFFF"),
                                   right=Side(border_style="thin", color="FFFFFF"),
                                   top=Side(border_style="thin", color="FFFFFF"),
                                   bottom=Side(border_style="thin", color="000000"))
BOTTOM_LAST_border_end = Border(left=Side(border_style="thin", color="FFFFFF"),
                                right=Side(border_style="thin", color="000000"),
                                top=Side(border_style="thin", color="FFFFFF"),
                                bottom=Side(border_style="thin", color="000000"))

##### cell color #####
blue_cell = PatternFill(patternType='solid', fgColor='4f81bd', bgColor='4f81bd')
white_cell = PatternFill(patternType='solid', fgColor='FFFFFF', bgColor='FFFFFF')

gray_blue_cell = PatternFill(patternType='solid', fgColor='dce6f1', bgColor='dce6f1')

'''
 DEF get common variables
'''
def get_excel_cell_format(num_paragraph, type): # type--> font,height,bullet_front,bullet_back,Enable_mark,Disable_mark,direct_front,direct_back,select_front,select_back
    num_to_font = {1: first_paragraph_font, 2: second_paragraph_font, 3: third_paragraph_font, 4: fourth_paragraph_font, 5: fifth_paragraph_font}
    num_to_height = {1: first_paragraph_height, 2: second_paragraph_height, 3: third_paragraph_height, 4: fourth_paragraph_height, 5: fifth_paragraph_height}
    num_to_bullet_front = {1: first_paragraph_bullet_front, 2: second_paragraph_bullet_front, 3: third_paragraph_bullet_front, 4: fourth_paragraph_bullet_front, 5: fifth_paragraph_bullet_front}
    num_to_bullet_back = {1: first_paragraph_bullet_back, 2: second_paragraph_bullet_back, 3: third_paragraph_bullet_back, 4: fourth_paragraph_bullet_back, 5: fifth_paragraph_bullet_back}

    if type == 'font':
        return_value = num_to_font[num_paragraph]
    elif type == 'height':
        return_value = num_to_height[num_paragraph]
    elif type =='bullet_front':
        return_value = num_to_bullet_front[num_paragraph]
    elif type == 'bullet_back':
        return_value = num_to_bullet_back[num_paragraph]
    elif type == 'Enable_mark':
        return_value = Enable_mark
    elif type == 'Disable_mark':
        return_value = Disable_mark
    elif type == 'direct_front':
        return_value = bullet_direct_front
    elif type == 'direct_back':
        return_value = bullet_direct_back
    elif type == 'select_front':
        return_value = bullet_select_front
    elif type == 'select_back':
        return_value = bullet_select_back

    return(return_value)

'''
 CREATE EXCEL FILE with GUI TREE format
'''
def create_excel_gui_tree(input_excel_name,output_excel_name,NEW_OR_ADD,egt_maker_width_array):  #Add egt_maker_width_array for Network Sketcher ver 2.0
    ##### User Settings #####
    excel_file_end_count = 3  # Set the number of consecutive empty input_row in input excel file for detecting end.
    limit_max_column = 5  # Set the max number of column each input_row for detecting column end.
    size_of_column_width = 22  # Set the width value of column
    size_of_row_height = 14.25  # Set the height value of row
    format_number_of_column = 15  # format number of column. format border/width/height For Example, Column:A = 1  Column:AA = 27 Column:CZ = 104.
    number_of_add_modify_rows = 10  # add number of coordinated empty rows under ended row.
    add_empty_row_paragraph = True  # If true, empty row is added under the paragraph

    ##### Default font and message #####
    defalut_message_to_paragraph = 'N/A or Omitted because there is little necessity.'  # add the message to under paragraph. '' is not used.

    ##### System Settings #####
    empty_count = 0
    empty_count_sub = 0
    column = 1
    input_row = 1
    output_row = 1
    number_of_sheets = 1

    ##### Main script prepared #####
    input_tree_excel = openpyxl.load_workbook(input_excel_name)
    print('Read the input excel file -> ' + input_excel_name)

    ##### NEW Sheet , Insert paragraph##
    if NEW_OR_ADD == 'ADD':
        output_tree_excel = openpyxl.load_workbook(output_excel_name)
        number_of_sheets += 1

    elif NEW_OR_ADD == 'PARA':
        number_of_add_modify_rows = -1
        output_tree_excel = openpyxl.load_workbook(output_excel_name)
        number_of_sheets += 1
        output_row = int(input_tree_excel.active.cell(1, 1).value)+3
        input_tree_excel.active.cell(1, 1).value = None
        sheetNames = input_tree_excel.get_sheet_names()
        output_tree_excel.active = output_tree_excel.sheetnames.index(sheetNames[0])
        insert_excel_empty_row(output_tree_excel, sheetNames[0], output_row-1)
        insert_excel_empty_row(output_tree_excel, sheetNames[0], output_row-1)
        insert_excel_empty_row(output_tree_excel, sheetNames[0], output_row-1)
    else:
        output_tree_excel = openpyxl.Workbook()


    ##### Main script #####
    while empty_count <= excel_file_end_count:
        for column in range(1, limit_max_column + 1):
            if column == 1 and input_tree_excel.active.cell(input_row, column).value != None:
                ### make a new sheet and set value###
                if number_of_sheets == 1:
                    output_tree_excel.active.title = input_tree_excel.active.cell(input_row, column).value
                    output_tree_excel.active.cell(output_row, column).value = num_to_bullet_front[column] + input_tree_excel.active.cell(input_row, column).value + num_to_bullet_back[column]
                    number_of_sheets += 1
                    first_sheet_name = str(output_tree_excel.active.title)
                    print('Create a new sheet -> ' + output_tree_excel.active.cell(output_row, column).value)
                else:
                    output_row = 1
                    output_tree_excel.create_sheet(title=str(input_tree_excel.active.cell(input_row, column).value))
                    output_tree_excel.active = output_tree_excel[str(input_tree_excel.active.cell(input_row, column).value)]
                    output_tree_excel.active.cell(output_row, column).value = num_to_bullet_front[column] + input_tree_excel.active.cell(input_row, column).value + num_to_bullet_back[column]
                    number_of_sheets += 1
                    print('Create a new sheet -> ' + output_tree_excel.active.cell(output_row, column).value)

                output_tree_excel.active.row_dimensions[output_row].height = first_paragraph_height
                output_tree_excel.active.cell(output_row, column).font = first_paragraph_font
                empty_count = 0

            ### set value to more than second column###
            elif input_tree_excel.active.cell(input_row, column).value != None:
                output_tree_excel.active.cell(output_row, column).value = num_to_bullet_front[column] + input_tree_excel.active.cell(input_row, column).value + num_to_bullet_back[column]
                print('     Add -> ' +input_tree_excel.active.cell(input_row, column).value)
                empty_count = 0
                ###  set font and height to more than second colum ###
                output_tree_excel.active.row_dimensions[output_row].height = num_to_height[column]
                output_tree_excel.active.cell(output_row, column).font = num_to_font[column]

                ### set defalut message to under the row ###
                if defalut_message_to_paragraph != '':
                    output_row += 1
                    output_tree_excel.active.cell(output_row, column).value = defalut_message_to_paragraph
                    output_tree_excel.active.cell(output_row, column).font = defalut_font
                    for i in range(1, format_number_of_column + 1):
                        output_tree_excel.active.cell(output_row, output_tree_excel.active.cell(output_row, i).column).border = white_border

                ### set empty row to under the row ###
                if add_empty_row_paragraph == True:
                    output_row += 1
                    for i in range(1, format_number_of_column + 1):
                        output_tree_excel.active.cell(output_row, output_tree_excel.active.cell(output_row, i).column).border = white_border

            ### If all of columns are None###
            else:
                empty_count_sub += 1

        ### check end of current sheet###
        if empty_count_sub >= limit_max_column:
            ### add empty row with decided height###
            for i in range(output_row + 1, output_row + number_of_add_modify_rows):
                output_tree_excel.active.row_dimensions[i].height = size_of_row_height
            ### add empty row with white border###
            for i in range(1, format_number_of_column + 1):
                for k in range(output_row, output_row + number_of_add_modify_rows):
                    output_tree_excel.active.cell(k, output_tree_excel.active.cell(k, i).column).border = white_border

            empty_count += 1
            empty_count_sub = 0

        ### increment the input_row number until empty_count reatch the value###
        input_row += 1
        output_row += 1

        ### STEP1->change size of column and row. STEP2->make white border in row###
        if NEW_OR_ADD != 'PARA': ### Safety function. If PARA, do not following scriput.
            for i in range(1, format_number_of_column + 1):
                output_tree_excel.active.column_dimensions[str(openpyxl.utils.get_column_letter(i))].width = size_of_column_width

                #Add Customize width value function for Network Sketcher ver 2.0
                if len(egt_maker_width_array) >= i:
                    output_tree_excel.active.column_dimensions[str(openpyxl.utils.get_column_letter(i))].width = egt_maker_width_array[i - 1]

                for k in range(1, output_row):
                    output_tree_excel.active.cell(k, output_tree_excel.active.cell(k, i).column).border = white_border

    ### move to first sheet and save the file###
    if NEW_OR_ADD == 'NEW':
        output_tree_excel.active = output_tree_excel[first_sheet_name]
    output_tree_excel.save(output_excel_name)
    print('Complete to Save the output excel file -> ' + output_excel_name)
    return(output_tree_excel)

'''
 EXCEL DEF
'''
def write_excel_paragraph(input_tree_excel, worksheet_name ,start_row, start_column, input_msg, paragraph_level):
    input_tree_excel.active = input_tree_excel[worksheet_name]
    input_tree_excel.active.cell(start_row, start_column-1).value = None
    input_tree_excel.active.cell(start_row, start_column).value = num_to_bullet_front[paragraph_level]+input_msg +num_to_bullet_back[paragraph_level]
    input_tree_excel.active.row_dimensions[start_row].height = num_to_height[paragraph_level]
    input_tree_excel.active.cell(start_row, start_column).font = num_to_font[paragraph_level]
    return(input_tree_excel)

def write_excel_cell(input_excel, worksheet_name ,start_row, start_column, input_msg, input_num_column, cell_style, write_style):
    input_excel.active = input_excel[worksheet_name]

    #####Insert row, if write_style == INSERT.  ######
    if write_style == 'INSERT':
        input_excel  = insert_excel_empty_row(input_excel, worksheet_name, start_row)

    #####write columns belong row ######
    write_column = start_column

    if input_msg == '<EMPTY>' or input_msg == '_EMPTY_':                                                  # Add <EMPTY> function for Network Sketcher ver 2.0
        input_excel.active.cell(start_row, write_column).value = ''
    else:
        input_excel.active.cell(start_row, write_column).value = input_msg

    if cell_style == 'TABLE_HEADER':
        input_excel.active.cell(start_row, write_column).font = table_header_font
        for c in range(write_column, write_column + input_num_column):
            input_excel.active.cell(start_row, c).fill = blue_cell
            if c == start_column:
                input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = blue_border_start
            elif c == (write_column+input_num_column-1):
                input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = blue_border_end
            else:
                input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = blue_border_medium
        write_column = write_column + input_num_column

    elif cell_style == 'SUB_TITLE':
        input_excel.active.cell(start_row, write_column).font = defalut_font
        for c in range(write_column, write_column + input_num_column):
            input_excel.active.cell(start_row, c).fill = gray_blue_cell
            if c == start_column:
                input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = gray_blue_border_start
            elif c == (write_column + input_num_column - 1):
                input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = gray_blue_border_end
            else:
                input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = gray_blue_border_medium

            if input_msg == None or input_msg == '':
                if c == start_column:
                    input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = gray_blue_border_start_updown_same
                elif c == (write_column + input_num_column - 1):
                    input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = gray_blue_border_end_updown_same
                else:
                    input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = gray_blue_border_medium_updown_same

                ### First or last normal table. Border is changed
            if (input_msg != None or input_msg != '') and write_style == 'BOTTOM_SAME':
                if c == start_column:
                    input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = gray_blue_border_start_bottom_same
                elif c == (write_column + input_num_column - 1):
                    input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = gray_blue_border_end_bottom_same
                else:
                    input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = gray_blue_border_medium_bottom_same

            if (input_msg != None or input_msg != '') and write_style == 'TOP_SAME':
                if c == start_column:
                    input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = gray_blue_border_start_top_same
                elif c == (write_column + input_num_column - 1):
                    input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = gray_blue_border_end_top_same
                else:
                    input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = gray_blue_border_medium_top_same

        write_column = write_column + input_num_column

    elif cell_style == 'TABLE_NORMAL':
        input_excel.active.cell(start_row, write_column).font = defalut_font
        for c in range(write_column, write_column + input_num_column):
            #### Add TABLE_NORMAL cell color###
            input_excel.active.cell(start_row, c).fill = white_cell

            #### Add TABLE_NORMAL script###
            if c == start_column:
                input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = black_border_start
            elif c == (write_column+input_num_column-1):
                input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = black_border_end
            else:
                input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = black_border_medium

            if input_msg == None or input_msg == '':
                if c == start_column:
                    input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = black_border_start_updown_same
                elif c == (write_column + input_num_column - 1):
                    input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = black_border_end_updown_same
                else:
                    input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = black_border_medium_updown_same

                ### First or last normal table. Border is changed
            if (input_msg != None or input_msg != '')  and write_style == 'BOTTOM_SAME':
                if c == start_column:
                    input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = black_border_start_bottom_same
                elif c == (write_column + input_num_column - 1):
                    input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = black_border_end_bottom_same
                else:
                    input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = black_border_medium_bottom_same

            if write_style == 'TOP_SAME':
                if c == start_column:
                    input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = black_border_start_top_same
                elif c == (write_column + input_num_column - 1):
                    input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = black_border_end_top_same
                else:
                    input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = black_border_medium_top_same

            if write_style == 'EDGE':
                if c == start_column:
                    input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = black_border_medium
                elif c == (write_column + input_num_column - 1):
                    input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = black_border_end
                else:
                    input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = black_border_medium

        write_column = write_column + input_num_column

    elif cell_style == 'TABLE_TOP_WHITE':
        input_excel.active.cell(start_row, write_column).font = defalut_font
        for c in range(write_column, write_column + input_num_column):
            input_excel.active.cell(start_row, c).fill = white_cell
            if c == start_column:
                input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = TOP_WHITE_border_start
                input_excel.active.cell(start_row-1, input_excel.active.cell(start_row, c).column).border = BOTTOM_WHITE_border_start
            elif c == (write_column+input_num_column-1):
                input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = TOP_WHITE_border_end
                input_excel.active.cell(start_row-1, input_excel.active.cell(start_row, c).column).border = BOTTOM_WHITE_border_end
            else:
                input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = TOP_WHITE_border_medium
                input_excel.active.cell(start_row-1, input_excel.active.cell(start_row, c).column).border = BOTTOM_WHITE_border_medium
        write_column = write_column + input_num_column

    elif cell_style == 'TOP_LAST':
        input_excel.active.cell(start_row, write_column).font = defalut_font
        for c in range(write_column, write_column + input_num_column):
            input_excel.active.cell(start_row, c).fill = white_cell
            if c == start_column:
                #input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = TOP_BLACK_border_medium
                input_excel.active.cell(start_row - 1, input_excel.active.cell(start_row, c).column).border = BOTTOM_LAST_border_start
            elif c == (write_column+input_num_column-1):
                #input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = TOP_BLACK_border_medium
                input_excel.active.cell(start_row - 1, input_excel.active.cell(start_row, c).column).border = BOTTOM_LAST_border_end
            else:
                #input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = TOP_BLACK_border_medium
                input_excel.active.cell(start_row - 1, input_excel.active.cell(start_row, c).column).border = BOTTOM_LAST_border_medium
        write_column = write_column + input_num_column

    return (input_excel)

def write_excel_table(input_excel, worksheet_name ,start_row, start_column, dict_input_msg, dict_input_num_column, cell_style, write_style):
    input_excel.active = input_excel[worksheet_name]

    #####Insert row, if write_style == INSERT.  For non vertical######
    if write_style == 'INSERT' and '_Vertical' not in cell_style:  #exclude vertical pattern
        input_excel  = insert_excel_empty_row(input_excel, worksheet_name, start_row)

    #####write columns belong row ######
    write_column = start_column
    write_row = start_row
    temp_flag_insert = False
    d=0
    while d < len(dict_input_msg):
        if cell_style == 'TABLE_HEADER':
            input_excel.active.cell(start_row, write_column).value = dict_input_msg[d]
            input_excel.active.cell(start_row, write_column).font = table_header_font
            for c in range(write_column, write_column + dict_input_num_column[d]):
                input_excel.active.cell(start_row, c).fill = blue_cell
                if c == start_column:
                    input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = blue_border_start
                elif c == (write_column+dict_input_num_column[d]-1):
                    input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = blue_border_end
                else:
                    input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = blue_border_medium
            write_column = write_column + dict_input_num_column[d]

        elif cell_style == 'TABLE_NORMAL':
            input_excel.active.cell(start_row, write_column).value = dict_input_msg[d]
            input_excel.active.cell(start_row, write_column).font = defalut_font
            for c in range(write_column, write_column + dict_input_num_column[d]):
                input_excel.active.cell(start_row, c).fill = white_cell
                if c == start_column:
                    input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = black_border_start
                elif c == (write_column+dict_input_num_column[d]-1):
                    input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = black_border_end
                else:
                    input_excel.active.cell(start_row, input_excel.active.cell(start_row, c).column).border = black_border_medium
            write_column = write_column + dict_input_num_column[d]

        elif cell_style == 'TABLE_HEADER_Vertical':
            if write_style == 'INSERT' and temp_flag_insert != True:
                for o in range(0,len(dict_input_msg)):
                    input_excel = insert_excel_empty_row(input_excel, worksheet_name, start_row)
                temp_flag_insert = True

            input_excel.active.cell(write_row, start_column).value = dict_input_msg[d]
            input_excel.active.cell(write_row, start_column).font = defalut_font
            for c in range(start_column , start_column + dict_input_num_column[0]):
                input_excel.active.cell(write_row, c).fill = gray_blue_cell
                if c == start_column:
                    input_excel.active.cell(write_row, input_excel.active.cell(write_row, c).column).border = gray_blue_border_start
                elif c == (write_column+dict_input_num_column[0]-1):
                    input_excel.active.cell(write_row, input_excel.active.cell(write_row, c).column).border = gray_blue_border_end
                else:
                    input_excel.active.cell(write_row, input_excel.active.cell(write_row, c).column).border = gray_blue_border_medium
            write_row += 1

        elif cell_style == 'TABLE_NORMAL_Vertical':
            input_excel.active.cell(write_row, start_column).value = dict_input_msg[d]
            input_excel.active.cell(write_row, start_column).font = defalut_font
            for c in range(start_column, start_column + dict_input_num_column[0]):
                input_excel.active.cell(write_row, c).fill = white_cell
                if c == start_column:
                    input_excel.active.cell(write_row, input_excel.active.cell(write_row, c).column).border = black_border_start
                elif c == (write_column + dict_input_num_column[0] - 1):
                    input_excel.active.cell(write_row, input_excel.active.cell(write_row, c).column).border = black_border_end
                else:
                    input_excel.active.cell(write_row, input_excel.active.cell(write_row, c).column).border = black_border_medium
            write_row += 1

        d += 1

    return (input_excel)

def insert_excel_empty_row(input_excel, worksheet_name, insert_row):
    ##### User Settings #####
    format_number_of_column = 15  # format number of column. format border/width/height For Example, Column:A = 1  Column:AA = 27 Column:CZ = 104.
    size_of_row_height = 14.25                  #Set the height value of row

    ##### Excel Border Settings #####
    white_border = Border(left=Side(border_style="thin",color="FFFFFF"),
                            right=Side(border_style="thin",color="FFFFFF"),
                            top=Side(border_style="thin",color="FFFFFF"),
                            bottom=Side(border_style="thin",color="FFFFFF"))

    ##### insert row and modify fond and height #####
    input_excel.active = input_excel[worksheet_name]
    get_under_row_height = input_excel.active.row_dimensions[insert_row+1].height
    input_excel.active.insert_rows(insert_row)
    input_excel.active.row_dimensions[insert_row+1].height = size_of_row_height
    input_excel.active.row_dimensions[insert_row+2].height = get_under_row_height
    for i in range(1, format_number_of_column+1):
            input_excel.active.cell(insert_row, input_excel.active.cell(insert_row, i).column).border = white_border
    return(input_excel)

def get_start_row_or_column(input_tree_excel_file,worksheet_name,paragraph_name, row_or_column):
    ##### user settings  #####
    max_number_of_serch_colmun = 5000

    ##### system settings  #####
    temp_found_flag = False
    temp_row = 1
    temp_column = 1

    ##### Open the input excel file #####
    input_tree_excel = openpyxl.load_workbook(input_tree_excel_file)
    input_tree_excel.active = input_tree_excel[worksheet_name]

    ##### Search locaton of the paragraph name #####
    while temp_row <= max_number_of_serch_colmun and temp_found_flag == False:
        for temp_column in range(1, 6):
            if input_tree_excel.active.cell(temp_row, temp_column).value == paragraph_name:
                num_row = temp_row + 1
                num_colmun = temp_column + 1
                temp_found_flag = True
                print(' ->Found the number of '+ row_or_column +'  ' + paragraph_name)
        if temp_row >= max_number_of_serch_colmun:
            temp_found_flag = True
            print('<<<ERROR the search paragraph name can not be find in the Worksheet>>>')
            sys.exit()
        temp_row +=1

    if row_or_column == 'row':
        num_row_or_colmun = num_row
    elif row_or_column == 'column':
        num_row_or_colmun = num_colmun

    return(num_row_or_colmun)

def insert_custom_excel_table(input_excel, worksheet_name ,start_row, start_column,custom_table_name):
    ##### init setting #####
    max_num_row = 0
    custom_table_row = 0
    temp_size_column = 0
    dict_input_num_column = ['']
    Row_OF_Column_Range = 1

    ##### Main script prepared #####
    input_tree_excel = input_excel
    input_custom_table = openpyxl.load_workbook(custom_table_name)
    input_custom_table.active = input_custom_table.sheetnames.index('_tmp_')  #### Add for Network Sketcher

    ##### check number of table's row #####
    for n in range(1, 10000):
        if input_custom_table.active.cell(n, 1).value == '<END>':
            max_num_row = n-1
            break
        elif n == 9999:
            tkinter.messagebox.showerror('error', 'Uneble to find <END> on row 1 in the meta_excel_table_file or number of row over 9999')
            return ('Uneble to find <END> on row 1')

    ##### Insert empty rows #####
    for s in range(0, max_num_row ):
        input_tree_excel = insert_excel_empty_row(input_tree_excel, worksheet_name, start_row+1+s)

    ##### Insert table with each column #####
    for temp_row in range(1, max_num_row):
        current_row = start_row + temp_row
        custom_table_row = temp_row+1
        temp_current_column = start_column + 1


        #### Get number of column end #####
        temp_flag_end = False
        temp_column_end = 0
        while (temp_flag_end == False):
            temp_column_end += 1
            if input_custom_table.active.cell(custom_table_row, temp_column_end + 1).value == '<END>':
                temp_flag_end = True
            elif temp_column_end == 9999:
                tkinter.messagebox.showerror('error', 'Uneble to find <END> on any coulmn in the meta_excel_table_file or number of column over 9999')
                break

        #### Change to rows to new Column range#####
        if input_custom_table.active.cell(custom_table_row, 1).value == '<RANGE>':
            Row_OF_Column_Range = temp_row + 1
            start_row -= 1

            '''
            ROW is <HEADER>
            '''
        elif input_custom_table.active.cell(custom_table_row , 1).value == '<HEADER>':
            input_msg = input_custom_table.active.cell(custom_table_row, 2).value
            temp_size_column = 0

            for c in range(2, temp_column_end+1):
                if input_custom_table.active.cell(custom_table_row, c+1).value != None or input_custom_table.active.cell(custom_table_row, c+1).value == '<END>':
                    input_num_column = int(input_custom_table.active.cell(Row_OF_Column_Range, c).value) + temp_size_column
                    write_excel_cell(input_excel, worksheet_name, current_row, temp_current_column, input_msg, input_num_column, 'TABLE_HEADER', 'NONE')
                    temp_current_column = temp_current_column + input_num_column
                    input_msg = input_custom_table.active.cell(custom_table_row, c + 1).value
                elif input_custom_table.active.cell(custom_table_row, c+1).value == None:
                    temp_size_column = temp_size_column + int(input_custom_table.active.cell(Row_OF_Column_Range, c).value)
            input_num_column = 0
            temp_current_column = start_column + 1

        ### ROW is <END> ###
        elif input_custom_table.active.cell(custom_table_row, 1).value == '<END>':
            print('ROW is <END>')

            '''
            ROW is Normal 
            '''
        else:
            input_msg = input_custom_table.active.cell(custom_table_row, 2).value
            for c in range(2, temp_column_end + 1):
                #### Check cell_style     - >buttom border line is covert to White ####
                cell_style = 'TABLE_NORMAL'
                write_style = 'NONE'

                ### set border first row and last row for normal###
                if input_custom_table.active.cell(custom_table_row, c).value != None and input_custom_table.active.cell(custom_table_row + 1, c).value == '<WHITE>' and input_custom_table.active.cell(custom_table_row, c).value != '<WHITE>':
                    write_style = 'BOTTOM_SAME'
                ### set border first row and last row for sub_title###
                if input_custom_table.active.cell(custom_table_row, c).value != None and input_custom_table.active.cell(custom_table_row + 1, c).value == None and input_custom_table.active.cell(custom_table_row + 1, c).value != '<END>' and input_custom_table.active.cell(custom_table_row + 1, c-1).value != '<END>':
                    write_style = 'BOTTOM_SAME'
                if input_custom_table.active.cell(custom_table_row, c).value != None and input_custom_table.active.cell(custom_table_row + 1, 1).value != None and input_custom_table.active.cell(custom_table_row - 1, c).value != None and '>' not in input_custom_table.active.cell(custom_table_row - 1, c).value:
                    write_style = 'TOP_SAME'
                if  input_custom_table.active.cell(custom_table_row + 1, 1).value == '<END>':
                    write_style = 'NONE'
                if input_custom_table.active.cell(custom_table_row, c).value == None and input_custom_table.active.cell(custom_table_row + 1, 1).value == '<END>':
                    write_style = 'TOP_SAME'
                if input_custom_table.active.cell(custom_table_row, c).value == '<WHITE>' and input_custom_table.active.cell(custom_table_row + 1, 1).value == '<END>':
                    write_style = 'TOP_SAME'


                ### check sub_title###
                if input_custom_table.active.cell(custom_table_row, c).value != None:
                    temp_first_character = input_custom_table.active.cell(custom_table_row, c).value[:1]
                    if temp_first_character != '>' and temp_first_character != '<':
                        cell_style = 'SUB_TITLE'
                else:
                    cell_style = 'SUB_TITLE'

                ### check white###
                if input_custom_table.active.cell(custom_table_row, c).value == '<WHITE>':
                    input_msg = ''

                ### check white###
                if input_custom_table.active.cell(custom_table_row, c).value == '<EDGE>':
                    input_msg = ''
                    write_style = 'EDGE'

                    ###Enable Disable Mark###
                if input_custom_table.active.cell(custom_table_row, c + 1).value != None or input_custom_table.active.cell(custom_table_row, c + 1).value == '<END>':
                    input_num_column = int(input_custom_table.active.cell(Row_OF_Column_Range, c).value)
                    #### Specific mark are converted ####
                    input_msg = str(input_msg)
                    if input_msg == '<ENABLE>':
                        input_msg = ''
                        for t in range(1,int(input_num_column*2.4)):
                            input_msg = input_msg + ' '
                        input_msg = input_msg + Enable_mark
                    elif input_msg == '<DISABLE>':
                        input_msg = ''
                        for t in range(1,int(input_num_column*2.4)):
                            input_msg = input_msg + ' '
                        input_msg = input_msg + Disable_mark
                    elif input_msg == 'None':
                        input_msg = ''

                    #### sub main ####
                    write_excel_cell(input_excel, worksheet_name, current_row, temp_current_column, input_msg, input_num_column, cell_style, write_style)
                    temp_current_column = temp_current_column + input_num_column
                    input_msg = input_custom_table.active.cell(custom_table_row, c + 1).value.replace('>>', '') # add replace >> for Network Sketcher

                elif input_custom_table.active.cell(custom_table_row, c+1).value == None:
                    input_num_column = int(input_custom_table.active.cell(Row_OF_Column_Range, c).value)
                    write_excel_cell(input_excel, worksheet_name, current_row, temp_current_column, input_msg, input_num_column, cell_style, write_style)
                    temp_current_column = temp_current_column + input_num_column
                    input_msg = input_custom_table.active.cell(custom_table_row, c + 1).value
                    input_msg = ''

            input_num_column = 0
            temp_current_column = start_column + 1
            temp_size_column=0


    return(input_tree_excel)

def diff_worksheets(COM_input_tree_excel, MAS_input_tree_excel, OPTION1, OPTION2):
    format_number_of_column = 104  # format number of column. format border/width/height For Example, Column:A = 1  Column:AA = 27 Column:CZ = 104.
    number_of_add_modify_rows = 10  # add number of coordinated empty rows under ended row.
    empty_count = 0
    COM_current_row_num = 2
    MAS_current_row_num = 2
    COM_current_bullet = None
    MAS_current_bullet = None
    empty_flag = False

    OPTION1.insert(tkinter.END, '"Row number","Column number","Master sheet","Compared sheet"')
    ###Main###
    while empty_count <= number_of_add_modify_rows:
        for temp_column_num in range(2, format_number_of_column + 1):
            ###input bullet###
            temp_MAS_str = str(MAS_input_tree_excel.active.cell(MAS_current_row_num,temp_column_num).value)
            temp_COM_str = str(COM_input_tree_excel.active.cell(COM_current_row_num, temp_column_num).value)
            if temp_MAS_str.startswith(second_paragraph_bullet_front) or temp_MAS_str.startswith(third_paragraph_bullet_front) or temp_MAS_str.startswith(fourth_paragraph_bullet_front) or temp_MAS_str.startswith(fifth_paragraph_bullet_front):
                MAS_current_bullet = MAS_input_tree_excel.active.cell(MAS_current_row_num,temp_column_num).value
            if temp_COM_str.startswith(second_paragraph_bullet_front) or temp_COM_str.startswith(third_paragraph_bullet_front) or temp_COM_str.startswith(fourth_paragraph_bullet_front) or temp_COM_str.startswith(fifth_paragraph_bullet_front):
                COM_current_bullet = COM_input_tree_excel.active.cell(COM_current_row_num,temp_column_num).value

            ###check diff###
            if str(MAS_input_tree_excel.active.cell(MAS_current_row_num,temp_column_num).value) != str(COM_input_tree_excel.active.cell(COM_current_row_num,temp_column_num).value):
                if temp_COM_str.startswith('>') or  Enable_mark in temp_COM_str or Disable_mark in temp_COM_str :
                    ### Change Enable and Disable Mark to <Enable> and <Disable> because of tk can not use the specific mark.
                    temptemp_MAS_str = temp_MAS_str
                    temptemp_COM_str = temp_COM_str
                    if Enable_mark in temp_MAS_str:
                        temptemp_MAS_str = '<Enable>'
                    elif Disable_mark in temp_MAS_str:
                        temptemp_MAS_str = '<Disable>'
                    if Enable_mark in temp_COM_str:
                        temptemp_COM_str = '<Enable>'
                    elif Disable_mark in temp_COM_str:
                        temptemp_COM_str = '<Disable>'

                    #### write color####
                    OPTION1.insert(tkinter.END, '\n\"'+str(COM_current_row_num)+'\",\"'+str(temp_column_num)+'\",\"'+temptemp_MAS_str+'\",\"'+temptemp_COM_str+'\"')
                    if OPTION2 != 'ONLY_LOG':
                        COM_input_tree_excel.active.cell(COM_current_row_num,temp_column_num).font = openpyxl.styles.fonts.Font(name="Futura", size=11,color='FF0000')
                    else:
                        COM_input_tree_excel.active.cell(COM_current_row_num,temp_column_num).font = openpyxl.styles.fonts.Font(name="Futura", size=11,color='000000')

            ### flag check end of sheet###
            if MAS_input_tree_excel.active.cell(MAS_current_row_num,temp_column_num).value != None:
                empty_flag = False
                empty_count = 0

        if MAS_current_bullet != COM_current_bullet:
            COM_current_row_num += 1
        else:
            MAS_current_row_num  += 1
            COM_current_row_num += 1

        ### check end of sheet###
        if empty_flag == True:
            empty_count += 1
        empty_flag = True

    return(COM_input_tree_excel)


'''
 RUN GUI TREE MAKER
'''
class qui_tree_run():
    def __init__(self):
        # crate the root frame
        root = tk.Tk()
        root.title("EXCEL GUI TREE MAKER Ver 1.3 < Cisco Internal Use Only >  ")
        root.geometry("1024x800")
        # Notebook
        nb = ttk.Notebook(width=200, height=200)

        # create Tabs
        tab1 = tk.Frame(nb)
        tab2 = tk.Frame(nb)
        tab3 = tk.Frame(nb)
        tab4 = tk.Frame(nb)
        tab5 = tk.Frame(nb)
        nb.add(tab2, text='   Excel Frame Editer   ', padding=15)
        nb.add(tab4, text='   Difference check   ', padding=15)
        nb.add(tab5, text='   DIFF to PowerPoint   ', padding=15)
        nb.add(tab1, text='   Create Excel Frame   ', padding=15 )
        nb.add(tab3, text='   Advanced Module   ', padding=15)

        nb.pack(expand=1, fill='both')

        # TAB5 initial settings
        self.flag_set_diff = False

        '''
        TAB1  <<CREATE EXCEL>>
        '''
        # label setting
        label1_0 = tk.Label(tab1,text="Create Excel Frame file from META file",font=("",16),height=2,background="#ffffff")
        label1_0.pack(fill="x")

        # input file
        frame1_1 = tk.Frame(tab1,pady=10)
        frame1_1.pack()

        label1_1 = tk.Label(frame1_1,font=("",14),text="INPUT File Path(Excel meta)")
        label1_1.pack(side="left")
        self.entry1_1 = tk.Entry(frame1_1,font=("",10),justify="left",width=60)
        self.entry1_1.pack(side="left")
        button1_1 = tk.Button(frame1_1,text="Browse...",font=("",10),width=8,command=lambda:self.click_action('1-1'))
        button1_1.pack(side="left")

        # output file
        frame1_2 = tk.Frame(tab1,pady=10)
        frame1_2.pack()
        label1_2 = tk.Label(frame1_2,font=("",14),text="OUTPUT File Path(Excel frame)")
        label1_2.pack(side="left")
        self.entry1_2 = tk.Entry(frame1_2,font=("",10),justify="left",width=60)
        self.entry1_2.pack(side="left")
        button1_2 = tk.Button(frame1_2,text="Browse...",font=("",10),width=8,command=lambda:self.click_action('1-2'))
        button1_2.pack(side="left")
        # Run
        button1_3 = tk.Button(tab1,text="CREATE",font=("",14),width=15,command=lambda:self.click_action('1-3'))
        button1_3.pack()

        '''
        TAB2  <<EDIT EXCEL>>
        '''
        # self  setting
        self.input_tree_excel = None
        self.selected_1 = None
        self.selected_2 = None
        self.selected_3 = None
        self.selected_4 = None
        self.selected_5 = None
        self.selected_4_1 = None
        self.selected_4_2 = None
        self.selected_row_1 = None
        self.selected_row_2 = None
        self.selected_row_3 = None
        self.selected_row_4 = None
        self.selected_row_5 = None
        self.temp_max_row = None

        # label setting
        label2_0 = tk.Label(tab2,text="Excel Frame Editer",font=("",16),height=1,background="#ffffff")
        label2_0.grid(column=0, row=0)

        # input file
        frame2_1 = tk.Frame(tab2,pady=10)
        frame2_1.grid(column=0, row=1, sticky = W)

        label2_1 = tk.Label(frame2_1,font=("",14),text="<STEP1> Input File Path(Excel frame)",background="#fff2cc")
        label2_1.grid(column=0, row=2, sticky = W)
        self.entry2_1 = tk.Entry(frame2_1,font=("",10),justify="left",width=100)
        self.entry2_1.grid(column=0, row=3)
        button2_1 = tk.Button(frame2_1,text="Browse...",font=("",10),width=15,command=lambda:self.click_action('2-1'))
        button2_1.grid(column=1, row=3)
        label2_1_open = tk.Label(frame2_1,font=("",14),text=" ")
        label2_1_open.grid(column=2, row=3 , sticky = W)
        button2_1_open = tk.Button(frame2_1,text=" Open The Selected File",font=("",10),width=17,command=lambda:self.click_action('2-1-open'))
        button2_1_open.grid(column=3, row=3)

        def select_listbox2_1(event):
            self.click_action('2-2')
        def select_listbox2_2(event):
            self.click_action('2-3')
        def select_listbox2_3(event):
            self.click_action('2-4')
        def select_listbox2_4(event):
            self.click_action('2-5')
        def select_listbox2_5(event):
            self.click_action('2-6')

        # Listbox
        frame2_2 = tk.Frame(tab2, pady=10)
        frame2_2.grid(column=0, row=4, sticky = W)
        label2_2 = tk.Label(frame2_2,font=("",14),text="<STEP2> Select the Worksheet or Paragraph ",background="#fff2cc")
        label2_2.grid(column=0, row=5 ,columnspan = 5, sticky = W)
        self.listbox2_1 = tk.Listbox(frame2_2, listvariable=StringVar(value='<Worksheet_name>'))
        self.listbox2_1.grid(column=0, row=7)
        self.listbox2_2 = tk.Listbox(frame2_2, listvariable=StringVar(value='<2nd>'))
        self.listbox2_2.grid(column=1, row=7)
        self.listbox2_3 = tk.Listbox(frame2_2, listvariable=StringVar(value='<3rd>'))
        self.listbox2_3.grid(column=2, row=7)
        self.listbox2_4 = tk.Listbox(frame2_2, listvariable=StringVar(value='<4th>'))
        self.listbox2_4.grid(column=3, row=7)
        self.listbox2_5 = tk.Listbox(frame2_2, listvariable=StringVar(value='<5th>'))
        self.listbox2_5.grid(column=4, row=7)

        self.listbox2_1.bind('<<ListboxSelect>>', select_listbox2_1)
        self.listbox2_2.bind('<<ListboxSelect>>', select_listbox2_2)
        self.listbox2_3.bind('<<ListboxSelect>>', select_listbox2_3)
        self.listbox2_4.bind('<<ListboxSelect>>', select_listbox2_4)
        self.listbox2_5.bind('<<ListboxSelect>>', select_listbox2_5)
        '''  At first, Listbox was action button style. In ver 1.1a, changed to active click style.
        button2_2 = tk.Button(frame2_2, text="Select", font=("", 12), width=6,  command=lambda: self.click_action('2-2')).grid(column=0, row=8)
        button2_3 = tk.Button(frame2_2, text="Select", font=("", 12), width=6,  command=lambda: self.click_action('2-3')).grid(column=1, row=8)
        button2_4 = tk.Button(frame2_2, text="Select", font=("", 12), width=6,  command=lambda: self.click_action('2-4')).grid(column=2, row=8)
        button2_5 = tk.Button(frame2_2, text="Select", font=("", 12), width=6,  command=lambda: self.click_action('2-5')).grid(column=3, row=8)
        button2_6 = tk.Button(frame2_2, text="Select", font=("", 12), width=6,  command=lambda: self.click_action('2-6')).grid(column=4, row=8)
        '''

        # Set Edit point
        frame2_3 = tk.Frame(tab2, pady=10)
        frame2_3.grid(column=0, row=8, sticky = W)
        label2_3 = tk.Label(frame2_3,font=("",14),text="Selected Edit point. You can customise each value.")
        label2_3.grid(column=0, row=9 ,columnspan = 5, sticky = W)

        label2_4 = tk.Label(frame2_3, font=("", 14), text="   Selected name")
        label2_4.grid(column=1, row=10)
        self.entry2_2 = tk.Entry(frame2_3, font=("", 12), justify="left", width=20)
        self.entry2_2.grid(column=2, row=10)
        label2_5 = tk.Label(frame2_3, font=("", 14), text="  Column")
        label2_5.grid(column=3, row=10)
        self.entry2_3 = tk.Entry(frame2_3, font=("", 12), justify="center", width=20)
        self.entry2_3.grid(column=4, row=10)
        label2_6 = tk.Label(frame2_3, font=("", 14), text="  Row")
        label2_6.grid(column=5, row=10)
        self.entry2_4 = tk.Entry(frame2_3, font=("", 12), justify="center", width=20)
        self.entry2_4.grid(column=6, row=10)

        ###STEP3###
        frame2_4 = tk.Frame(tab2, pady=10)
        frame2_4.grid(column=0, row=11, sticky = W)

        label2_7 = tk.Label(frame2_4,font=("",14),text="<STEP3> Click Action " ,background="#fff2cc")
        label2_7.grid(column=0, row=12 ,columnspan = 5, sticky = W)

        ###STEP3 Insert Basic table###
        button2_31 = tk.Button(frame2_4, text="Insert Basic Table ", font=("", 14), width=16,  command=lambda: self.click_action('2-31')).grid(column=1, row=14)
        label2_30 = tk.Label(frame2_4,font=("",14),text="   columns X rows >> ")
        label2_30.grid(column=2, row=14 , sticky = W)
        self.entry2_31 = tk.Entry(frame2_4, font=("", 12), justify="center", width=10)
        self.entry2_31.grid(column=3, row=14)
        self.entry2_31.delete(0, 'end')
        self.entry2_31.insert(tk.END, 2)
        label2_31 = tk.Label(frame2_4,font=("",14),text=" X ")
        label2_31.grid(column=4, row=14 , sticky = W)
        self.entry2_32 = tk.Entry(frame2_4, font=("", 12), justify="center", width=10)
        self.entry2_32.grid(column=5, row=14)
        self.entry2_32.delete(0, 'end')
        self.entry2_32.insert(tk.END, 2)
        label2_32 = tk.Label(frame2_4,font=("",14),text="   Number of columns per Cell >> ")
        label2_32.grid(column=6, row=14 , sticky = W)
        self.entry2_33 = tk.Entry(frame2_4, font=("", 12), justify="center", width=10)
        self.entry2_33.grid(column=7, row=14)
        self.entry2_33.delete(0, 'end')
        self.entry2_33.insert(tk.END, 10)

        ###STEP3###
        frame2_4_1 = tk.Frame(tab2, pady=10)
        frame2_4_1.grid(column=0, row=15, sticky = W)

        ###STEP3 Insert Custom table###
        button2_32 = tk.Button(frame2_4_1, text="Insert Custom Table ", font=("", 14), width=16,  command=lambda: self.click_action('2-32')).grid(column=1, row=15)
        label2_32_2 = tk.Label(frame2_4_1,font=("",14),text="   Input Meta Table File(Excel)")
        label2_32_2.grid(column=2, row=15, sticky = W)
        self.entry2_32_2 = tk.Entry(frame2_4_1,font=("",10),justify="left",width=60)
        self.entry2_32_2.grid(column=3, row=15,columnspan = 3)
        button2_32_2 = tk.Button(frame2_4_1,text="Browse...",font=("",10),width=8,command=lambda:self.click_action('2-32-2'))
        button2_32_2.grid(column=7, row=15)

        ###STEP3 Add New Worksheet###
        frame2_5 = tk.Frame(tab2, pady=10)
        frame2_5.grid(column=0, row=16, sticky = W)

        button2_33 = tk.Button(frame2_5, text="Add New Worksheet", font=("", 14), width=16,  command=lambda: self.click_action('2-33')).grid(column=1, row=17)
        label2_35 = tk.Label(frame2_5, font=("", 14), text="   New Worksheet name")
        label2_35.grid(column=2, row=17)
        self.entry2_36 = tk.Entry(frame2_5, font=("", 12), justify="left", width=20)
        self.entry2_36.insert(tk.END,"New Worksheet name")
        self.entry2_36.grid(column=3, row=17)

        ###STEP3 Add New paragraph###
        button2_34 = tk.Button(frame2_5, text="Insert New paragraph", font=("", 14), width=16,  command=lambda: self.click_action('2-34')).grid(column=1, row=16)
        label2_36 = tk.Label(frame2_5, font=("", 14), text="   Insert Paragraph name")
        label2_36.grid(column=2, row=16)
        self.entry2_37 = tk.Entry(frame2_5, font=("", 12), justify="left", width=20)
        self.entry2_37.insert(tk.END,"New Paragraph name")
        self.entry2_37.grid(column=3, row=16)
        label2_36 = tk.Label(frame2_5, font=("", 14), text="   Paragraph level(2~5)")
        label2_36.grid(column=4, row=16)
        self.entry2_38 = tk.Entry(frame2_5, font=("", 12), justify="center", width=10)
        self.entry2_38.insert(tk.END,"2")
        self.entry2_38.grid(column=5, row=16)

        ###STEP3 Insert empyt row###
        button2_7 = tk.Button(frame2_5, text="Insert Empty row ", font=("", 14), width=16,  command=lambda: self.click_action('2-7')).grid(column=1, row=18)

        frame2_9 = tk.Frame(tab2, pady=10)
        frame2_9.grid(column=0, row=18, sticky = W)
        label2_9 = tk.Label(frame2_9,font=("",10),text="© 2019  Cisco and/or its affiliates. All rights reserved. \nCisco Confidential \n( Contact Name :  Yusuke Ogawa , yuogawa@cisco.com )",background='#e4eee8')
        label2_9.grid(column=0, row=18)

        '''
        TAB3  <<Advanced Module>>
        '''
        self.input_ise_policy_file = None

        # label setting
        label3_10 = tk.Label(tab3,text="Module for specific products",font=("",16),height=2,background="#ffffff")
        label3_10.grid(column=0, row=0)

        # input file
        frame3_11 = tk.Frame(tab3,pady=10)
        frame3_11.grid(column=0, row=1, sticky = W)

        label3_11 = tk.Label(frame3_11,font=("",14),text="<STEP1> Input File Path(Excel frame)")
        label3_11.grid(column=0, row=2, sticky = W)
        self.entry3_11 = tk.Entry(frame3_11,font=("",10),justify="left",width=100)
        self.entry3_11.grid(column=0, row=3)
        button3_11 = tk.Button(frame3_11,text="Browse...",font=("",10),width=8,command=lambda:self.click_action('3-11'))
        button3_11.grid(column=1, row=3)

        ###STEP2###
        frame3_21 = tk.Frame(tab3, pady=10)
        frame3_21.grid(column=0, row=11, sticky = W)

        ###STEP2 Insert empyt row###
        label3_21 = tk.Label(frame3_21,font=("",14),text="<STEP2> --- Cisco ISE(Beta) ---")
        label3_21.grid(column=0, row=12 ,columnspan = 5, sticky = W)
        label3_22 = tk.Label(frame3_21,font=("",14),text="Input File Path(Policy file)")
        label3_22.grid(column=0, row=13, sticky = W)
        self.entry3_21 = tk.Entry(frame3_21,font=("",10),justify="left",width=100)
        self.entry3_21.grid(column=0, row=14)
        button3_21 = tk.Button(frame3_21,text="Browse...",font=("",10),width=8,command=lambda:self.click_action('3-21'))
        button3_21.grid(column=1, row=14)
        button3_22 = tk.Button(frame3_21, text="Run the ISE Module ", font=("", 14), width=20,  command=lambda: self.click_action('3-22')).grid(column=0, row=15)

        '''
        TAB4  <<Difference Check>>
        '''
        # label setting
        label4_0 = tk.Label(tab4, text="Difference Check", font=("", 16), height=1, background="#ffffff")
        label4_0.grid(column=0, row=0)

        # input file
        frame4_1 = tk.Frame(tab4, pady=10)
        frame4_1.grid(column=0, row=1, sticky=W)

        label4_1 = tk.Label(frame4_1, font=("", 14), text="<STEP1> Input File Path(Excel frame)", background="#fff2cc")
        label4_1.grid(column=0, row=2, sticky=W)
        self.entry4_1 = tk.Entry(frame4_1, font=("", 10), justify="left", width=100)
        self.entry4_1.grid(column=0, row=3)
        button4_1 = tk.Button(frame4_1, text="Browse...", font=("", 10), width=15,  command=lambda: self.click_action('4-1'))
        button4_1.grid(column=1, row=3)

        frame4_2 = tk.Frame(tab4, pady=10)
        frame4_2.grid(column=0, row=4, sticky=W)
        label4_2 = tk.Label(frame4_2, font=("", 14), text="<STEP2> Select the Compared Worksheet and Master Worksheet.", background="#fff2cc")
        label4_2.grid(column=0, row=5,columnspan = 5, sticky=W)
        label4_3 = tk.Label(frame4_2, font=("", 14), text="[Compared]")
        label4_3.grid(column=0, row=7, columnspan = 3,sticky=W)
        label4_3 = tk.Label(frame4_2, font=("", 14), text="      ")
        label4_3.grid(column=1, row=7, sticky=W)
        label4_4 = tk.Label(frame4_2, font=("", 14), text="[Master]")
        label4_4.grid(column=2, row=7, columnspan = 3,sticky=W)
        self.listbox4_1 = tk.Listbox(frame4_2, width=30,listvariable=StringVar(value='<Worksheet>'))
        self.listbox4_1.grid(column=0, row=8)
        self.listbox4_2 = tk.Listbox(frame4_2, width=30,listvariable=StringVar(value='<Worksheet>'))
        self.listbox4_2.grid(column=2, row=8)

        label4_5 = tk.Label(frame4_2, font=("", 14), text="Selected name")
        label4_5.grid(column=0, row=9, columnspan = 3,sticky=W)
        label4_6 = tk.Label(frame4_2, font=("", 14), text="      ")
        label4_6.grid(column=1, row=9, sticky=W)
        label4_7 = tk.Label(frame4_2, font=("", 14), text="Selected name")
        label4_7.grid(column=2, row=9, columnspan = 3,sticky=W)
        self.entry4_2 = tk.Entry(frame4_2, font=("", 14), justify="left", width=25)
        self.entry4_2.grid(column=0, row=10)
        self.entry4_3 = tk.Entry(frame4_2, font=("", 14), justify="left", width=25)
        self.entry4_3.grid(column=2, row=10)

        def select_listbox4_1(event):
            self.click_action('4-2')
        def select_listbox4_2(event):
            self.click_action('4-3')

        self.listbox4_1.bind('<<ListboxSelect>>', select_listbox4_1)
        self.listbox4_2.bind('<<ListboxSelect>>', select_listbox4_2)

        frame4_3 = tk.Frame(tab4, pady=10)
        frame4_3.grid(column=0, row=11, sticky=W)
        label4_3_1 = tk.Label(frame4_3, font=("", 14), text="<STEP3> Click Action", background="#fff2cc")
        label4_3_1.grid(column=0, row=12,columnspan = 5, sticky=W)
        button4_3_1 = tk.Button(frame4_3, text=" Run DIFF and the Difference points are changed to RED ", font=("", 14), width=50,  command=lambda: self.click_action('4-3-1')).grid(column=0, row=13)
        label4_3_3 = tk.Label(frame4_3, font=("",5), text=" ")
        label4_3_3.grid(column=0, row=14, sticky=W)

        ### no color mode ###
        button4_3_2 = tk.Button(frame4_3, text=" Run DIFF and the Difference points are changed to Black ", font=("", 14), width=50,  command=lambda: self.click_action('4-3-2')).grid(column=0, row=15)
        label4_3_5 = tk.Label(frame4_3, font=("",5), text=" ")
        label4_3_5.grid(column=0, row=16, sticky=W)

        ### log text box###
        self.text4_1 = Text(frame4_3, height=12, width=120)
        self.text4_1.grid(column=0, row=17,columnspan = 5)
        self.text4_1.insert(tkinter.END, 'Diff log with CSV')

        '''
        TAB5  <<DIFF to PowerPoint>>
        '''
        # label setting
        label5_0 = tk.Label(tab5, text="Create The Change procedure with PowerPoint using [Difference Check] Tab data", font=("", 16), height=1, background="#ffffff")
        label5_0.grid(column=0, row=0)

        # input file
        frame5_1 = tk.Frame(tab5, pady=10)
        frame5_1.grid(column=0, row=1, sticky=W)
        label5_1 = tk.Label(frame5_1, font=("", 14), text="<STEP1> Set [Difference Check] Result", background="#fff2cc")
        label5_1.grid(column=0, row=2, sticky=W,columnspan = 5)
        label5_1_2 = tk.Label(frame5_1, font=("", 6), text="   ")
        label5_1_2.grid(column=0, row=3, sticky=W)
        button5_1 = tk.Button(frame5_1, text="Set [Difference Check] Result", font=("", 14), width=50,  command=lambda: self.click_action('5-1'))
        button5_1.grid(column=1, row=4)

        frame5_3 = tk.Frame(tab5, pady=10)
        frame5_3.grid(column=0, row=4, sticky=W)
        label5_3 = tk.Label(frame5_3, font=("", 14), text="    (Optional) Modify the Output File path of PowerPoint", background="#DCE6f1")
        label5_3.grid(column=0, row=5, sticky=W,columnspan = 5)
        label5_3_0 = tk.Label(frame5_3, font=("", 6), text="                                                               ")
        label5_3_0.grid(column=0, row=6, sticky=W)
        self.entry5_3 = tk.Entry(frame5_3, font=("", 10), justify="left", width=100)
        self.entry5_3.grid(column=1, row=6)

        frame5_2 = tk.Frame(tab5, pady=10)
        frame5_2.grid(column=0, row=7, sticky=W)
        label5_2_1 = tk.Label(frame5_2, font=("", 14), text="    (Optional) Input the Template file of PowerPoint( use .pptx so NOT .potx)", background="#DCE6f1")
        label5_2_1.grid(column=0, row=8,columnspan = 5, sticky=W)
        self.entry5_2_1 = tk.Entry(frame5_2, font=("", 10), justify="left", width=100)
        self.entry5_2_1.grid(column=1, row=9,columnspan = 4)
        self.entry5_2_1.insert(tk.END, '<use default template>')
        button5_2_1 = tk.Button(frame5_2, text="Browse...", font=("", 10), width=15,  command=lambda: self.click_action('5-2'))
        button5_2_1.grid(column=5, row=9)
        label5_2_3 = tk.Label(frame5_2,font=("",14),text="\n       Input the Slide Number of Title and Normal from Master Slide viewing in Template PowerPoint\n       *When you use Cisco FY19 template, change [Title Slide] to 2 and [Normal Slide]  to 11.")
        label5_2_3.grid(column=0, row=12 ,columnspan = 5, sticky = W)
        label5_2_4 = tk.Label(frame5_2, font=("", 14), text="       Title Slide")
        label5_2_4.grid(column=0, row=13)
        self.entry5_2_2 = tk.Entry(frame5_2, font=("", 12), justify="center", width=20)
        self.entry5_2_2.grid(column=1, row=13)
        self.entry5_2_2.insert(tk.END, '0')
        label5_2_5 = tk.Label(frame5_2, font=("", 14), text="  Normal Slide")
        label5_2_5.grid(column=2, row=13)
        self.entry5_2_3 = tk.Entry(frame5_2, font=("", 12), justify="center", width=20)
        self.entry5_2_3.grid(column=3, row=13)
        self.entry5_2_3.insert(tk.END, '5')
        label5_2_3 = tk.Label(frame5_2,font=("",14),text="   ")
        label5_2_3.grid(column=0, row=14 ,sticky = W)
        label5_2_6 = tk.Label(frame5_2, font=("", 14), text="    (Optional) Input the Sentences pack if you want to use other sentences or Language(EXCEL)", background="#DCE6f1")
        label5_2_6.grid(column=0, row=15,columnspan = 8, sticky=W)
        self.entry5_2_6 = tk.Entry(frame5_2, font=("", 10), justify="left", width=100)
        self.entry5_2_6.grid(column=1, row=16,columnspan = 4)
        self.entry5_2_6.insert(tk.END, '<use default sentences>')
        button5_2_6 = tk.Button(frame5_2, text="Browse...", font=("", 10), width=15,  command=lambda: self.click_action('5-3'))
        button5_2_6.grid(column=5, row=16)

        frame5_4 = tk.Frame(tab5, pady=10)
        frame5_4.grid(column=0, row=17, sticky=W)
        label5_4_0 = tk.Label(frame5_4, font=("", 6), text="   ")
        label5_4_0.grid(column=0, row=18, sticky=W)
        label5_4 = tk.Label(frame5_4, font=("", 14), text="<STEP2> CREATE PowerPoint ", background="#fff2cc")
        label5_4.grid(column=0, row=19, sticky=W,columnspan = 5)
        label5_4_1 = tk.Label(frame5_4, font=("", 6), text="   ")
        label5_4_1.grid(column=0, row=20, sticky=W)
        button5_4 = tk.Button(frame5_4, text="CREATE the Change procedure with PowerPoint", font=("", 14), width=50,  command=lambda: self.click_action('5-4'))
        button5_4.grid(column=1, row=21)

        #main loop
        root.mainloop()

    def click_action(self,click_value):
        if click_value == '1-1':
            fTyp = [("", ".xlsx")]
            iDir = os.path.abspath(os.path.dirname(__file__))
            filename = tk.filedialog.askopenfilename(filetypes=fTyp, initialdir=iDir)
            self.entry1_1.delete(0, tkinter.END)
            self.entry1_1.insert(tk.END,filename)
            self.entry1_2.delete(0, tkinter.END)
            self.entry1_2.insert(tk.END, iDir + '\\temp_output.xlsx')

        elif click_value == '1-2':
            fTyp = [("", "*")]
            iDir = os.path.abspath(os.path.dirname(__file__))
            filepath = tk.filedialog.askopenfilename(filetypes=fTyp, initialdir=iDir)
            self.entry1_2.delete(0, tkinter.END)
            self.entry1_2.insert(tk.END, filepath)

        elif click_value == '1-3':
            input_excel_gui_tree_file = self.entry1_1.get()  # input excel file name
            output_excel_file = self.entry1_2.get()  # output excel file name
            create_excel_gui_tree(input_excel_gui_tree_file, output_excel_file,'NEW', egt_maker_width_array)
            tkinter.messagebox.showinfo('info', 'Complete to Save the output excel file')

        elif click_value == '2-1':
            fTyp = [("", ".xlsx")]
            iDir = os.path.abspath(os.path.dirname(__file__))
            filename = tk.filedialog.askopenfilename(filetypes=fTyp, initialdir=iDir)
            self.entry2_1.delete(0, tkinter.END)
            self.entry2_1.insert(tk.END,filename)
            #input sheets name to listbox
            self.input_tree_excel = openpyxl.load_workbook(self.entry2_1.get())
            List2_1 = self.input_tree_excel.sheetnames
            self.listbox2_1.delete(0, 'end')
            self.listbox2_2.delete(0, 'end')
            self.listbox2_3.delete(0, 'end')
            self.listbox2_4.delete(0, 'end')
            self.listbox2_5.delete(0, 'end')
            for temp_item in List2_1:
                self.listbox2_1.insert(END, temp_item )

        elif click_value == '2-1-open':
            subprocess.Popen(self.entry2_1.get(), shell=True)

        elif click_value == '2-1-2': # Reload file
            #input sheets name to listbox
            self.input_tree_excel.save(self.entry2_1.get())
            self.input_tree_excel = openpyxl.load_workbook(self.entry2_1.get())
            List2_1 = self.input_tree_excel.sheetnames
            self.entry2_2.delete(0, 'end')
            self.entry2_2.insert(tk.END, '<Cleared>')
            self.entry2_3.delete(0, 'end')
            self.entry2_3.insert(tk.END, '<Cleared>')
            self.entry2_4.delete(0, 'end')
            self.entry2_4.insert(tk.END, '<Cleared>')
            self.listbox2_2.delete(0, 'end')
            self.listbox2_3.delete(0, 'end')
            self.listbox2_4.delete(0, 'end')
            self.listbox2_5.delete(0, 'end')

        elif click_value == '2-2':
            #input sheets name to listbox
            index_1 = self.listbox2_1.curselection()
            self.selected_1 = self.listbox2_1.get(index_1)
            self.input_tree_excel.active = self.input_tree_excel[self.selected_1]
            self.temp_max_row = self.input_tree_excel[self.selected_1].max_row
            self.listbox2_2.delete(0, 'end')
            self.listbox2_3.delete(0, 'end')
            self.listbox2_4.delete(0, 'end')
            self.listbox2_5.delete(0, 'end')
            for temp_row in range(1, self.temp_max_row):
                if second_paragraph_bullet_front in str(self.input_tree_excel.active.cell(temp_row, 2).value):
                    self.listbox2_2.insert(END, self.input_tree_excel.active.cell(temp_row, 2).value)
            #Set STEP3
            self.entry2_2.delete(0, 'end')
            self.entry2_2.insert(tk.END, self.selected_1)
            self.entry2_3.delete(0, 'end')
            self.entry2_3.insert(tk.END, 1)
            self.entry2_4.delete(0, 'end')
            self.entry2_4.insert(tk.END, 1)

        elif click_value == '2-3':
            #get selected name
            index_2= int(self.listbox2_2.curselection()[0])
            self.selected_2 = self.listbox2_2.get(index_2)
            #get selected row
            temp_index_num = -1
            for temp_row in range(1, self.temp_max_row):
                if second_paragraph_bullet_front in str(self.input_tree_excel.active.cell(temp_row, 2).value):
                    temp_index_num += 1
                if temp_index_num == index_2:
                    self.selected_row_2 = temp_row
                    break
            #serch start and end row number
            for temp_selected_row in range(self.selected_row_2+1, self.temp_max_row):
                if second_paragraph_bullet_front in str(self.input_tree_excel.active.cell(temp_selected_row, 2).value):
                    temp_end_row = temp_selected_row
                    break
                elif temp_selected_row == self.temp_max_row-1:
                    temp_end_row = temp_selected_row
                    break
            # serch start and end row number
            self.listbox2_3.delete(0, 'end')
            self.listbox2_4.delete(0, 'end')
            self.listbox2_5.delete(0, 'end')
            for temp_row in range(self.selected_row_2, temp_end_row):
                if third_paragraph_bullet_front in str(self.input_tree_excel.active.cell(temp_row, 3).value):
                    self.listbox2_3.insert(END, self.input_tree_excel.active.cell(temp_row, 3).value)

            #Set STEP3
            self.entry2_2.delete(0, 'end')
            self.entry2_2.insert(tk.END, self.selected_2)
            self.entry2_3.delete(0, 'end')
            self.entry2_3.insert(tk.END, 2)
            self.entry2_4.delete(0, 'end')
            self.entry2_4.insert(tk.END, self.selected_row_2)

        elif click_value == '2-4':
            #get selected name
            index_3 = int(self.listbox2_3.curselection()[0])
            self.selected_3 = self.listbox2_3.get(index_3)
            #get selected row
            temp_index_num = -1
            for temp_row in range(self.selected_row_2+1, self.temp_max_row):
                if third_paragraph_bullet_front in str(self.input_tree_excel.active.cell(temp_row, 3).value):
                    temp_index_num += 1
                if temp_index_num == index_3:
                    self.selected_row_3 = temp_row
                    break
            #serch start and end row number
            for temp_selected_row in range(self.selected_row_3+1, self.temp_max_row):
                if third_paragraph_bullet_front in str(self.input_tree_excel.active.cell(temp_selected_row, 3).value):
                    temp_end_row = temp_selected_row
                    break
                elif temp_selected_row == self.temp_max_row-1:
                    temp_end_row = temp_selected_row
                    break
            # serch start and end row number
            self.listbox2_4.delete(0, 'end')
            self.listbox2_5.delete(0, 'end')
            for temp_row in range(self.selected_row_3, temp_end_row):
                if fourth_paragraph_bullet_front in str(self.input_tree_excel.active.cell(temp_row, 4).value):
                    self.listbox2_4.insert(END, self.input_tree_excel.active.cell(temp_row, 4).value)

            #Set STEP3
            self.entry2_2.delete(0, 'end')
            self.entry2_2.insert(tk.END, self.selected_3)
            self.entry2_3.delete(0, 'end')
            self.entry2_3.insert(tk.END, 3)
            self.entry2_4.delete(0, 'end')
            self.entry2_4.insert(tk.END, self.selected_row_3)

        elif click_value == '2-5':
            #get selected name
            index_4 = int(self.listbox2_4.curselection()[0])
            self.selected_4 = self.listbox2_4.get(index_4)
            #get selected row
            temp_index_num = -1
            for temp_row in range(self.selected_row_3+1, self.temp_max_row):
                if fourth_paragraph_bullet_front in str(self.input_tree_excel.active.cell(temp_row, 4).value):
                    temp_index_num += 1
                if temp_index_num == index_4:
                    self.selected_row_4 = temp_row
                    break
            #serch end row number
            for temp_selected_row in range(self.selected_row_4+1, self.temp_max_row):
                if fourth_paragraph_bullet_front in str(self.input_tree_excel.active.cell(temp_selected_row, 4).value):
                    temp_end_row = temp_selected_row
                    break
                elif temp_selected_row == self.temp_max_row-1:
                    temp_end_row = temp_selected_row
                    break
            # serch start and end row number
            self.listbox2_5.delete(0, 'end')
            for temp_row in range(self.selected_row_4, temp_end_row):
                if fifth_paragraph_bullet_front in str(self.input_tree_excel.active.cell(temp_row, 5).value):
                    self.listbox2_5.insert(END, self.input_tree_excel.active.cell(temp_row, 5).value)

            #Set STEP3
            self.entry2_2.delete(0, 'end')
            self.entry2_2.insert(tk.END, self.selected_4)
            self.entry2_3.delete(0, 'end')
            self.entry2_3.insert(tk.END, 4)
            self.entry2_4.delete(0, 'end')
            self.entry2_4.insert(tk.END, self.selected_row_4)

        elif click_value == '2-6':
            #get selected name
            index_5 = int(self.listbox2_5.curselection()[0])
            self.selected_5 = self.listbox2_5.get(index_5)
            #get selected row
            temp_index_num = -1
            for temp_row in range(self.selected_row_4+1, self.temp_max_row):
                if fifth_paragraph_bullet_front in str(self.input_tree_excel.active.cell(temp_row, 5).value):
                    temp_index_num += 1
                if temp_index_num == index_5:
                    self.selected_row_5 = temp_row
                    break

            #Set STEP3
            self.entry2_2.delete(0, 'end')
            self.entry2_2.insert(tk.END, self.selected_5)
            self.entry2_3.delete(0, 'end')
            self.entry2_3.insert(tk.END, 5)
            self.entry2_4.delete(0, 'end')
            self.entry2_4.insert(tk.END, self.selected_row_5)

        elif click_value == '2-7': #Insert empty row
            start_row = int(self.entry2_4.get())
            #Insert empty row
            self.input_tree_excel = insert_excel_empty_row(self.input_tree_excel, self.selected_1.get(), start_row+1)
            self.input_tree_excel.save(self.entry2_1.get())
            self.click_action('2-1-2')  # reload input file
            tkinter.messagebox.showinfo('info', 'Complete to Insert row')

        elif click_value == '2-31': # insert Basic Table
            worksheet_name = self.selected_1
            start_column = int(self.entry2_3.get())
            start_row = int(self.entry2_4.get())
            input_dict_header = []
            input_dict_num = []
            for n in range (0,int(self.entry2_31.get())):
                input_dict_header.append('<Value>')
                input_dict_num.append(10)
            self.input_tree_excel = write_excel_table(self.input_tree_excel, worksheet_name, start_row + 1, start_column + 1, input_dict_header, input_dict_num, 'TABLE_HEADER', 'INSERT')
            for n in range (0,int(self.entry2_32.get())):
                self.input_tree_excel = write_excel_table(self.input_tree_excel, worksheet_name, start_row + 2 + n, start_column + 1, input_dict_header, input_dict_num, 'TABLE_NORMAL', 'INSERT')
            self.input_tree_excel.save(self.entry2_1.get())
            self.click_action('2-1-2')  # reload input file
            tkinter.messagebox.showinfo('info', 'Complete to Insert Table')

        elif click_value == '2-32':  # insert Custom Table
            worksheet_name = self.selected_1
            start_column = int(self.entry2_3.get())
            start_row = int(self.entry2_4.get())
            custom_table_name = self.entry2_32_2.get()
            self.input_tree_excel = openpyxl.load_workbook(self.entry2_1.get())
            self.input_tree_excel = insert_custom_excel_table(self.input_tree_excel, worksheet_name ,start_row, start_column,custom_table_name)
            self.input_tree_excel.save(self.entry2_1.get())
            self.click_action('2-1-2')  # reload input file
            tkinter.messagebox.showinfo('info', 'Complete to Save the output excel file')

        elif click_value == '2-32-2':  # insert Custom Table file on entry
            fTyp = [("", ".xlsx")]
            iDir = os.path.abspath(os.path.dirname(__file__))
            filename = tk.filedialog.askopenfilename(filetypes=fTyp, initialdir=iDir)
            self.entry2_32_2.delete(0, tkinter.END)
            self.entry2_32_2.insert(tk.END,filename)
            #self.input_tree_excel = openpyxl.load_workbook(self.entry2_32_2.get())

        elif click_value == '3-11':
            fTyp = [("", ".xlsx")]
            iDir = os.path.abspath(os.path.dirname(__file__))
            filename = tk.filedialog.askopenfilename(filetypes=fTyp, initialdir=iDir)
            self.entry3_11.delete(0, tkinter.END)
            self.entry3_11.insert(tk.END,filename)
            self.input_tree_excel = openpyxl.load_workbook(self.entry3_11.get())

        elif click_value == '2-33':
            temp_meta_excel = openpyxl.Workbook()
            temp_meta_excel.active.cell(1, 1).value = self.entry2_36.get()
            temp_meta_excel.save('temp_temp_meta_exel.xlsx')
            self.input_tree_excel = create_excel_gui_tree('temp_temp_meta_exel.xlsx', self.entry2_1.get(),'ADD', egt_maker_width_array)
            os.remove('temp_temp_meta_exel.xlsx')
            tkinter.messagebox.showinfo('info', 'Complete to Save the output excel file')
            self.click_action('2-1-2')  # reload input file

        elif click_value == '2-34':
            temp_meta_excel = openpyxl.Workbook()
            temp_meta_excel.active.cell(1, int(self.entry2_38.get())).value = self.entry2_37.get()
            temp_meta_excel.active.cell(1, 1).value = int(self.entry2_4.get())
            temp_meta_excel.active.title = self.selected_1
            temp_meta_excel.save('temp_temp_meta_exel.xlsx')
            self.input_tree_excel = create_excel_gui_tree('temp_temp_meta_exel.xlsx', self.entry2_1.get(),'PARA', egt_maker_width_array)
            os.remove('temp_temp_meta_exel.xlsx')
            self.click_action('2-1-2')  # reload input file
            tkinter.messagebox.showinfo('info', 'Complete to Save the output excel file')


        elif click_value == '3-21':
            fTyp = [("", "*")]
            iDir = os.path.abspath(os.path.dirname(__file__))
            filename = tk.filedialog.askopenfilename(filetypes=fTyp, initialdir=iDir)
            self.entry3_21.delete(0, tkinter.END)
            self.entry3_21.insert(tk.END,filename)
            #input sheets name to listbox
            self.input_ise_policy_file = self.entry3_21.get()

        elif click_value == '3-22':
            import cisco_ise_module as ise
            '''Add ISE policy to Gui Tree'''
            ### set files name###
            input_tree_excel_file = self.entry3_11.get()  # input gui tree excel file name
            input_ise_policy_xml_file = self.input_ise_policy_file  # input gui tree excel file name
            ### set output number of row and column to excel ###
            get_num_row = get_start_row_or_column(input_tree_excel_file, 'Policy', '⁃Policy Sets', 'row')
            get_num_column = get_start_row_or_column(input_tree_excel_file, 'Policy', '⁃Policy Sets', 'column')
            ### main script ###
            # ise.ISE_PolicySet_to_excel(input_ise_policy_xml_file, input_tree_excel_file, 'Policy',5, get_num_row, get_num_column)
            ise.ISE_All_PolicySets_to_excel(input_ise_policy_xml_file, input_tree_excel_file, 'Policy', get_num_row, get_num_column)
            ''' Add ISE Conditions to Gui Tree'''
            ### set output number of row and column to excel ###
            get_num_row = get_start_row_or_column(input_tree_excel_file, 'Policy', '•Conditions', 'row')
            get_num_column = get_start_row_or_column(input_tree_excel_file, 'Policy', '•Conditions', 'column')
            ### main script ###
            ise.ISE_LibraryConditon_to_excel(input_ise_policy_xml_file, input_tree_excel_file, 'Policy', get_num_row, get_num_column)
            '''Add ISE Results-Auhorization Profiles to Gui Tree'''
            ### set output number of row and column to excel ###
            get_num_row = get_start_row_or_column(input_tree_excel_file, 'Policy', '•Results', 'row')
            get_num_column = get_start_row_or_column(input_tree_excel_file, 'Policy', '•Results', 'column')
            ### main script ###
            ise.ISE_Result_Auhorization_to_excel(input_ise_policy_xml_file, input_tree_excel_file, 'Policy', get_num_row, get_num_column)
            tkinter.messagebox.showinfo('info', 'Complete to Save the output excel file')

        elif click_value == '4-1':
            fTyp = [("", ".xlsx")]
            iDir = os.path.abspath(os.path.dirname(__file__))
            filename = tk.filedialog.askopenfilename(filetypes=fTyp, initialdir=iDir)
            self.entry4_1.delete(0, tkinter.END)
            self.entry4_1.insert(tk.END, filename)
            # input sheets name to listbox
            self.input_tree_excel = openpyxl.load_workbook(self.entry4_1.get())
            List4_1 = self.input_tree_excel.sheetnames
            self.listbox4_1.delete(0, 'end')
            self.listbox4_2.delete(0, 'end')
            for temp_item in List4_1:
                self.listbox4_1.insert(END, temp_item)
                self.listbox4_2.insert(END, temp_item)

        elif click_value == '4-1-2': # Reload file tab4
            #save and reload to tab4
            self.input_tree_excel.save(self.entry4_1.get())
            self.input_tree_excel = openpyxl.load_workbook(self.entry4_1.get())
            List4_1 = self.input_tree_excel.sheetnames
            List4_2 = self.input_tree_excel.sheetnames
            #self.entry4_2.delete(0, 'end')
            #self.entry4_2.insert(tk.END, '<Cleared>')
            #self.entry4_3.delete(0, 'end')
            #self.entry4_3.insert(tk.END, '<Cleared>')

        elif click_value == '4-2':
            # input sheets name to listbox
            index_4_1 = self.listbox4_1.curselection()
            self.selected_4_1 = self.listbox4_1.get(index_4_1)
            self.entry4_2.delete(0, 'end')
            self.entry4_2.insert(tk.END, self.selected_4_1 )

        elif click_value == '4-3':
            # input sheets name to listbox
            index_4_2 = self.listbox4_2.curselection()
            self.selected_4_2 = self.listbox4_2.get(index_4_2)
            self.entry4_3.delete(0, 'end')
            self.entry4_3.insert(tk.END, self.selected_4_2)

        elif click_value == '4-3-1':
            ### check worksheet name ###
            MAS_input_tree_excel = openpyxl.load_workbook(self.entry4_1.get())
            self.input_tree_excel.active = self.input_tree_excel[self.selected_4_1]
            COM_A1_vaule = self.input_tree_excel.active.cell(1, 1).value
            MAS_input_tree_excel.active = MAS_input_tree_excel[self.selected_4_2]
            MAS_A1_vaule = MAS_input_tree_excel.active.cell(1, 1).value
            if COM_A1_vaule != MAS_A1_vaule:
                messagebox.showerror('error', 'Cell A1 in both of Worksheet does not match. \n Please select correct worksheets again.')

            else:
                ###Main ####
                self.text4_1.delete('1.0', 'end')
                self.input_tree_excel = diff_worksheets(self.input_tree_excel,MAS_input_tree_excel,self.text4_1,'OPTION2')
                self.input_tree_excel.save(self.entry4_1.get())
                self.click_action('4-1-2')  # reload input file
                tkinter.messagebox.showinfo('info', 'Complete to Save the output excel file')

        elif click_value == '4-3-2':
            ### check worksheet name ###
            MAS_input_tree_excel = openpyxl.load_workbook(self.entry4_1.get())
            self.input_tree_excel.active = self.input_tree_excel[self.selected_4_1]
            COM_A1_vaule = self.input_tree_excel.active.cell(1, 1).value
            MAS_input_tree_excel.active = MAS_input_tree_excel[self.selected_4_2]
            MAS_A1_vaule = MAS_input_tree_excel.active.cell(1, 1).value
            if COM_A1_vaule != MAS_A1_vaule:
                messagebox.showerror('error', 'Cell A1 in both of Worksheet does not match. \n Please select correct worksheets again.')

            else:
                ###Main ####
                self.text4_1.delete('1.0', 'end')
                self.input_tree_excel = diff_worksheets(self.input_tree_excel,MAS_input_tree_excel,self.text4_1,'ONLY_LOG')
                self.input_tree_excel.save(self.entry4_1.get())
                self.click_action('4-1-2')  # reload input file
                tkinter.messagebox.showinfo('info', 'Complete to Save the output excel file')

        elif click_value == '5-1':
            # OK flag
            self.flag_set_diff = False
            # input DIFF TAB data
            self.diff_file_name = self.entry4_1.get()
            self.diff_worksheet_name = self.entry4_2.get()
            self.diff_log_csv = self.text4_1.get('1.0', 'end -1c')

            # input file path pptx powerpoint
            filename = self.entry4_1.get()
            dirname = os.path.dirname(filename)
            basename = os.path.basename(filename)

            if '.xlsx' not in  basename:
                messagebox.showerror('error', 'Excel file is not selected in Diff Tab. \n Please confirm it again.')
            elif self.diff_worksheet_name == '':
                messagebox.showerror('error', 'Worksheet name is not selected in Diff Tab. \n Please confirm  it again.')
            elif '"Row number","Column number","Master sheet","Compared sheet"' not in self.diff_log_csv:
                messagebox.showerror('error', 'Diff Log is not found in Diff Tab. \n Please confirm  it again.')
            else:
                self.entry5_3.delete(0, tkinter.END)
                self.entry5_3.insert(tk.END, dirname+ '\\Change_procedure_[' + self.diff_worksheet_name + '].pptx')
                tkinter.messagebox.showinfo('Set Diff values', '[Filename]\n   '+ basename + '\n\n[Worksheet name]\n   '+self.diff_worksheet_name+'\n\n[DIFF LOG]\n    OK')
                self.flag_set_diff = True

        elif click_value == '5-2':
            # input sheets name to listbox
            fTyp = [("", ".pptx")]
            iDir = os.path.abspath(os.path.dirname(__file__))
            filename = tk.filedialog.askopenfilename(filetypes=fTyp, initialdir=iDir)
            self.entry5_2_1.delete(0, tkinter.END)
            self.entry5_2_1.insert(tk.END, filename)

        elif click_value == '5-3':
            # input sheets name to listbox
            fTyp = [("", ".xlsx")]
            iDir = os.path.abspath(os.path.dirname(__file__))
            filename = tk.filedialog.askopenfilename(filetypes=fTyp, initialdir=iDir)
            self.entry5_2_6.delete(0, tkinter.END)
            self.entry5_2_6.insert(tk.END, filename)

        elif click_value == '5-4':
            if self.flag_set_diff == True:
                # import module
                import egt_maker_ppt_module as ppt_mod

                ppt_mod.egt_maker_ppt_run.__init__(self)

                #show result
                tkinter.messagebox.showinfo('info', 'Complete to Save the output PowerPoint file')
            else:
                messagebox.showerror('error', '<STEP 1> does not been completed. \n Please confirm  it again.')


if __name__ == '__main__':
    qui_tree_run()

''' Addtional functions '''
def change_string_color(input_tree_excel,ws_name,row,column,color):
    if color == 'GRAY':
        color_num = '808080'
    else:
        color_num = '000000'

    input_tree_excel.active.cell(row, column).font = openpyxl.styles.fonts.Font(name="Futura", size=11,color=color_num)
